"""
File handling utilities for MS Preprocessing Toolkit.

This module provides functions for reading and writing various file formats
commonly used in mass spectrometry data processing.
"""

import logging
import hashlib
from pathlib import Path
from typing import Optional, Union, Tuple, Any
from datetime import datetime
import json

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Color

from ms_core.preprocessing.settings import Settings
from ms_core.utils.intermediate_store import IntermediateStore

logger = logging.getLogger(__name__)


class FileHandler:
    """
    Handles file I/O operations for mass spectrometry data files.

    Supports Excel (.xlsx, .xls), CSV, and TSV formats with
    preservation of formatting information where applicable.
    """

    SUPPORTED_FORMATS = Settings.SUPPORTED_FORMATS

    def __init__(self):
        """Initialize the FileHandler."""
        self._last_loaded_path: Optional[Path] = None

    @staticmethod
    def is_supported_format(file_path: Union[str, Path]) -> bool:
        """Check if the file format is supported."""
        path = Path(file_path)
        return path.suffix.lower() in FileHandler.SUPPORTED_FORMATS

    def load_data(
        self,
        file_path: Union[str, Path],
        sheet_name: Optional[Union[str, int]] = 0,
        header_row: int = 0,
    ) -> Tuple[pd.DataFrame, dict]:
        """
        Load data from a file.

        Args:
            file_path: Path to the input file
            sheet_name: Sheet name or index for Excel files
            header_row: Row index to use as header

        Returns:
            Tuple of (DataFrame, metadata dict)
        """
        path = Path(file_path)
        # Prefer parquet cache only when cache feature is enabled.
        if Settings.SAVE_PARQUET_CACHE and path.suffix.lower() == ".xlsx":
            cached = self._resolve_parquet_cache(path)
            if cached:
                path = cached
        self._last_loaded_path = path
        red_font_rows: set = set()
        metadata = {"source_file": str(path), "load_time": datetime.now().isoformat()}

        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")

        if not self.is_supported_format(path):
            raise ValueError(f"Unsupported file format: {path.suffix}")

        suffix = path.suffix.lower()

        if suffix in {".xlsx", ".xls"}:
            df, red_font_rows = self._load_excel(path, sheet_name, header_row)
            metadata["format"] = "excel"
            metadata["sheet_name"] = sheet_name
        elif suffix == ".csv":
            df = self._load_delimited(path, header_row, sep=",")
            metadata["format"] = "csv"
        elif suffix in {".tsv", ".txt"}:
            df = self._load_delimited(path, header_row, sep="\t")
            metadata["format"] = "tsv"
        elif suffix == ".parquet":
            metadata["format"] = "parquet"
            try:
                df, store_meta = IntermediateStore.load(path)
                metadata.update(store_meta)
                red_font_rows = set(store_meta.get("red_font_rows", []))
            except Exception as exc:
                logger.warning("Intermediate store load failed, falling back to legacy parquet loader: %s", exc)
                df = pd.read_parquet(path)
                meta = self._load_parquet_meta(path)
                if meta:
                    metadata.update(meta)
                    red_font_rows = set(meta.get("red_font_rows", []))
        else:
            raise ValueError(f"Unsupported file format: {suffix}")

        metadata["shape"] = df.shape
        metadata["columns"] = list(df.columns)
        if "red_font_rows" not in metadata:
            metadata["red_font_rows"] = sorted(red_font_rows)
        else:
            metadata["red_font_rows"] = sorted(set(metadata.get("red_font_rows", [])))

        return df, metadata

    @staticmethod
    def _load_excel(
        file_path: Path,
        sheet_name: Optional[Union[str, int]] = 0,
        header_row: int = 0,
    ) -> Tuple[pd.DataFrame, set]:
        """Load data from Excel file with formatting extraction.

        Returns:
            Tuple of (DataFrame, set of red-font row indices).
        """
        # Load with pandas
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, engine="openpyxl")

        # Extract red font information for protection logic
        red_font_rows: set = set()
        try:
            wb = load_workbook(file_path, data_only=False)
            if isinstance(sheet_name, int):
                ws = wb.worksheets[sheet_name]
            else:
                ws = wb[sheet_name] if sheet_name else wb.active

            # Check each row for red font (starting from data rows)
            for row_idx in range(header_row + 2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=1)
                if cell.font and cell.font.color:
                    color = cell.font.color
                    if color.type == "rgb" and color.rgb:
                        # Check for red color (various shades)
                        rgb = color.rgb
                        if isinstance(rgb, str) and len(rgb) >= 6:
                            r = int(rgb[-6:-4], 16)
                            g = int(rgb[-4:-2], 16)
                            b = int(rgb[-2:], 16)
                            if r > 200 and g < 100 and b < 100:
                                # Adjust for pandas DataFrame index (0-based, excluding header)
                                red_font_rows.add(row_idx - header_row - 2)
            wb.close()
        except Exception as exc:
            logger.warning("Failed to extract red-font formatting: %s", exc)

        return df, red_font_rows

    def save_data(
        self,
        df: pd.DataFrame,
        file_path: Union[str, Path],
        sheet_name: str = "Sheet1",
        index: bool = False,
        highlight_rows: Optional[set] = None,
        blue_font_cells: Optional[list] = None,
        red_font_rows: Optional[set] = None,
        extra_sheets: Optional[dict] = None,
        save_parquet_cache: bool = False,
    ) -> Path:
        """
        Save data to a file.

        Args:
            df: DataFrame to save
            file_path: Output file path
            sheet_name: Sheet name for Excel files
            index: Whether to include DataFrame index
            highlight_rows: Set of row indices to highlight with yellow
            blue_font_cells: List of (row, col) tuples for blue font cells

        Returns:
            Path to the saved file
        """
        path = Path(file_path)
        suffix = path.suffix.lower()

        if suffix in {".xlsx", ".xls"}:
            self._save_excel(
                df,
                path,
                sheet_name,
                index,
                highlight_rows,
                blue_font_cells,
                red_font_rows,
                extra_sheets=extra_sheets,
            )
            if save_parquet_cache:
                try:
                    self._save_parquet_cache(
                        df,
                        self._cache_path_for_excel(path),
                        highlight_rows=highlight_rows,
                        blue_font_cells=blue_font_cells,
                        red_font_rows=red_font_rows,
                    )
                except Exception as exc:
                    logger.warning("Parquet cache save failed (non-fatal): %s", exc)
        elif suffix == ".csv":
            df.to_csv(path, index=index)
        elif suffix in {".tsv", ".txt"}:
            df.to_csv(path, sep="\t", index=index)
        elif suffix == ".parquet":
            parquet_meta = {
                "red_font_rows": sorted(red_font_rows) if red_font_rows else [],
                "blue_font_cells": blue_font_cells or [],
                "highlight_rows": sorted(highlight_rows) if highlight_rows else [],
            }
            try:
                IntermediateStore.save(
                    df=df,
                    parquet_path=path,
                    metadata=parquet_meta,
                    index=index,
                )
            except Exception as exc:
                logger.warning("Intermediate store save failed, falling back to raw parquet write: %s", exc)
                df.to_parquet(path, index=index)
        else:
            # Default to Excel format
            path = path.with_suffix(".xlsx")
            self._save_excel(df, path, sheet_name, index, highlight_rows, blue_font_cells, red_font_rows)

        return path

    def _save_excel(
        self,
        df: pd.DataFrame,
        file_path: Path,
        sheet_name: str = "Sheet1",
        index: bool = False,
        highlight_rows: Optional[set] = None,
        blue_font_cells: Optional[list] = None,
        red_font_rows: Optional[set] = None,
        extra_sheets: Optional[dict] = None,
    ) -> None:
        """Save DataFrame to Excel with optional formatting."""
        # First save with pandas
        if extra_sheets:
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=index)
                for sheet, sheet_df in extra_sheets.items():
                    if sheet_df is None:
                        continue
                    sheet_df.to_excel(writer, sheet_name=sheet, index=index)
        else:
            df.to_excel(file_path, sheet_name=sheet_name, index=index)

        # Then apply formatting if needed
        if highlight_rows or blue_font_cells or red_font_rows:
            wb = load_workbook(file_path)
            ws = wb[sheet_name]

            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            blue_font = Font(color="0070C0")
            red_font = Font(color="FF0000")

            # Apply yellow highlight to specified rows
            if highlight_rows:
                for row_idx in highlight_rows:
                    # Account for header row (Excel is 1-indexed, header is row 1)
                    excel_row = row_idx + 2
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=excel_row, column=col).fill = yellow_fill

            # Apply blue font to specified cells
            if blue_font_cells:
                for row_idx, col_idx in blue_font_cells:
                    excel_row = row_idx + 2
                    excel_col = col_idx + 1
                    ws.cell(row=excel_row, column=excel_col).font = blue_font

            # Apply red font to specified rows (Mz/RT column)
            if red_font_rows:
                for row_idx in red_font_rows:
                    excel_row = row_idx + 2
                    ws.cell(row=excel_row, column=1).font = red_font

            wb.save(file_path)
            wb.close()

    @staticmethod
    def generate_output_path(
        input_path: Union[str, Path],
        suffix: str = "_processed",
        output_dir: Optional[Union[str, Path]] = None,
        add_timestamp: bool = True,
    ) -> Path:
        """
        Generate an output file path based on input path.

        Args:
            input_path: Original input file path
            suffix: Suffix to add before file extension
            output_dir: Output directory (defaults to input file directory)
            add_timestamp: Whether to add timestamp to filename

        Returns:
            Generated output path
        """
        input_path = Path(input_path)
        stem = input_path.stem

        if add_timestamp:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_name = f"{stem}{suffix}_{timestamp}{input_path.suffix}"
        else:
            new_name = f"{stem}{suffix}{input_path.suffix}"

        if output_dir:
            output_path = Path(output_dir) / new_name
        else:
            output_path = input_path.parent / new_name

        return output_path

    @staticmethod
    def _parquet_meta_path(parquet_path: Path) -> Path:
        return parquet_path.with_suffix(parquet_path.suffix + ".meta.json")

    @staticmethod
    def _cache_path_for_excel(excel_path: Path) -> Path:
        """Map an excel path to a machine-local cache parquet path."""
        normalized = str(excel_path.resolve(strict=False)).lower()
        digest = hashlib.sha1(normalized.encode("utf-8")).hexdigest()
        cache_root = Settings.get_parquet_cache_root()
        safe_stem = "".join(ch if ch.isalnum() or ch in {"-", "_"} else "_" for ch in excel_path.stem) or "dataset"
        filename = f"{safe_stem}_{digest}.parquet"
        return cache_root / digest[:2] / filename

    def _save_parquet_cache(
        self,
        df: pd.DataFrame,
        parquet_path: Path,
        highlight_rows: Optional[set] = None,
        blue_font_cells: Optional[list] = None,
        red_font_rows: Optional[set] = None,
    ) -> None:
        """Save a parquet cache with metadata sidecar for formatting."""
        parquet_path.parent.mkdir(parents=True, exist_ok=True)
        meta = {
            "red_font_rows": sorted(red_font_rows) if red_font_rows else [],
            "blue_font_cells": blue_font_cells or [],
            "highlight_rows": sorted(highlight_rows) if highlight_rows else [],
        }

        try:
            IntermediateStore.save(
                df=df,
                parquet_path=parquet_path,
                metadata=meta,
                index=False,
            )
            return
        except Exception as exc:
            logger.debug("Intermediate store cache save failed; using legacy fallback: %s", exc)

        if df.columns.duplicated().any():
            logger.debug("Skipping parquet cache because dataframe has duplicate column labels.")
            return

        try:
            df.to_parquet(parquet_path, index=False)
        except Exception as exc:
            logger.debug(
                "Parquet cache raw write failed; retrying with normalized object columns: %s",
                exc,
            )
            normalized_df = self._normalize_for_parquet(df)
            normalized_df.to_parquet(parquet_path, index=False)
        meta_path = self._parquet_meta_path(parquet_path)
        try:
            meta_path.write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")
        except Exception as exc:
            logger.warning("Parquet meta write failed: %s", exc)

    def _load_parquet_meta(self, parquet_path: Path) -> Optional[dict]:
        """Load parquet metadata sidecar if it exists."""
        meta_path = self._parquet_meta_path(parquet_path)
        if not meta_path.exists():
            return None
        try:
            data = json.loads(meta_path.read_text(encoding="utf-8"))
            # Normalize
            return {
                "red_font_rows": data.get("red_font_rows", []),
                "blue_font_cells": data.get("blue_font_cells", []),
                "highlight_rows": data.get("highlight_rows", []),
            }
        except Exception as exc:
            logger.warning("Failed to load parquet meta: %s", exc)
            return None

    def _resolve_parquet_cache(self, excel_path: Path) -> Optional[Path]:
        """Return parquet cache if it exists and is newer than Excel."""
        parquet_path = self._cache_path_for_excel(excel_path)
        meta_path = self._parquet_meta_path(parquet_path)
        if not parquet_path.exists() or not meta_path.exists():
            return None
        try:
            if parquet_path.stat().st_mtime >= excel_path.stat().st_mtime:
                return parquet_path
        except Exception as exc:
            logger.debug("Parquet cache resolution failed: %s", exc)
            return None
        return None

    @staticmethod
    def _normalize_for_parquet(df: pd.DataFrame) -> pd.DataFrame:
        """
        Normalize object columns to parquet-friendly scalar types.

        This keeps numeric-like object columns numeric, preserves pure text
        columns, decodes bytes, and stringifies mixed object columns.
        """
        normalized = df.copy()
        object_positions = [
            idx for idx, dtype in enumerate(normalized.dtypes)
            if dtype == "object"
        ]

        for col_idx in object_positions:
            # Use positional indexing to support duplicate column labels.
            series = normalized.iloc[:, col_idx]
            non_null = series[series.notna()]
            if non_null.empty:
                continue

            if non_null.map(lambda v: isinstance(v, str)).all():
                continue

            if non_null.map(lambda v: isinstance(v, (bytes, bytearray))).all():
                converted = series.map(FileHandler._decode_bytes_value)
                normalized.iloc[:, col_idx] = converted.to_numpy()
                continue

            if non_null.map(lambda v: isinstance(v, (int, float, bool, np.number, np.bool_))).all():
                converted = pd.to_numeric(series, errors="coerce")
                normalized.iloc[:, col_idx] = converted.to_numpy()
                continue

            converted = series.map(FileHandler._stringify_mixed_value)
            normalized.iloc[:, col_idx] = converted.to_numpy()

        return normalized

    @staticmethod
    def _decode_bytes_value(value: Any) -> Any:
        """Decode bytes-like values to UTF-8 strings while preserving nulls."""
        if value is None:
            return np.nan
        try:
            if pd.isna(value):
                return np.nan
        except Exception:
            pass
        if isinstance(value, (bytes, bytearray)):
            return bytes(value).decode("utf-8", errors="replace")
        return value

    @staticmethod
    def _stringify_mixed_value(value: Any) -> Any:
        """Convert mixed-type object values to strings while preserving nulls."""
        if value is None:
            return np.nan
        try:
            if pd.isna(value):
                return np.nan
        except Exception:
            pass
        if isinstance(value, (bytes, bytearray)):
            return bytes(value).decode("utf-8", errors="replace")
        return str(value)

    @staticmethod
    def _load_delimited(path: Path, header_row: int, sep: str) -> pd.DataFrame:
        """Load CSV/TSV using pyarrow engine if available."""
        try:
            import pyarrow  # type: ignore
            engine = "pyarrow"
        except Exception:
            engine = None

        if engine:
            try:
                return pd.read_csv(path, header=header_row, sep=sep, engine=engine)
            except Exception as exc:
                logger.debug(
                    "pyarrow delimited parse failed for %s; falling back to default parser: %s",
                    path,
                    exc,
                )
        return pd.read_csv(path, header=header_row, sep=sep)


def parse_mz_rt_string(value: str) -> Tuple[Optional[float], Optional[float]]:
    """
    Parse a combined m/z and RT string (e.g., "123.456/1.23").

    Args:
        value: String in format "mz/rt"

    Returns:
        Tuple of (mz, rt) as floats, or (None, None) if parsing fails
    """
    try:
        parts = str(value).split("/")
        if len(parts) == 2:
            mz = float(parts[0].strip())
            rt = float(parts[1].strip())
            return mz, rt
    except (ValueError, AttributeError):
        pass
    return None, None


def format_mz_rt_string(mz: float, rt: float, mz_decimals: int = 4, rt_decimals: int = 2) -> str:
    """
    Format m/z and RT values as a combined string.

    Args:
        mz: m/z value
        rt: RT value
        mz_decimals: Decimal places for m/z
        rt_decimals: Decimal places for RT

    Returns:
        Formatted string in "mz/rt" format
    """
    return f"{mz:.{mz_decimals}f}/{rt:.{rt_decimals}f}"
