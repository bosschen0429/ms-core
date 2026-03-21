import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import silhouette_score
from sklearn.metrics.pairwise import manhattan_distances, euclidean_distances
from pycombat import pycombat
import warnings
import os
import datetime
import sys
from openpyxl import load_workbook, Workbook
from copy import copy
from scipy import stats
from matplotlib.patches import Ellipse
from openpyxl.styles import PatternFill, Font
from scipy.stats import chi2, f as f_dist

warnings.filterwarnings('ignore')

# ========== 匯入共用模組 ==========
from ms_core.utils.plotting import setup_matplotlib, plot_pca_comparison_qc_style
from ms_core.utils.constants import FONT_SIZES, COLORBLIND_COLORS, SHEET_NAMES, DATETIME_FORMAT_FULL, NON_SAMPLE_COLUMNS, STAT_COLUMN_KEYWORDS, VALIDATION_THRESHOLDS, COHENS_D_THRESHOLDS, CV_QUALITY_THRESHOLDS
from ms_core.utils.sample_classification import SampleClassifier, identify_sample_columns, normalize_sample_type
from ms_core.utils.file_io import (
    build_output_path,
    build_plots_dir,
    get_output_root,
    generate_output_filename,
)
from ms_core.utils.results import ProcessingResult
from ms_core.utils.console import safe_print as print

# 設定 matplotlib
setup_matplotlib()

# Centralized naming to avoid magic strings in downstream logic
SUMMARY_SHEET_NAME = SHEET_NAMES.get('batch_summary', "Batch_Effect_summary")
PLOT_FOLDER_NAME = "Batch_Effect_plots"

# 確保安裝 scikit-bio (graceful fallback for testing)
SKBIO_AVAILABLE = False
skbio_permanova = None
DistanceMatrix = None

def _has_valid_permanova(d):
    """Check if a PERMANOVA result dict contains real (non-NaN) values."""
    if d is None:
        return False
    return not np.isnan(d.get('pseudo_f', np.nan))

def _safe_get(d, key, default=np.nan):
    """Safely get a value from a dict that may be None or contain NaN."""
    if d is None:
        return default
    val = d.get(key, default)
    if isinstance(val, float) and np.isnan(val):
        return default
    return val

try:
    from skbio.stats.distance import permanova as skbio_permanova
    from skbio import DistanceMatrix
    SKBIO_AVAILABLE = True
except ImportError:
    print("⚠️ 警告: 未安裝 scikit-bio，PERMANOVA 功能將無法使用")
    print("   請執行: pip install scikit-bio")

def select_file():
    raise RuntimeError("input_file is required; GUI must provide the file path.")
    """開啟檔案選擇對話框"""
    root = tk.Tk()
    root.withdraw()
    
    file_path = filedialog.askopenfilename(
        title="選擇要進行批次效應校正的Excel檔案",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    
    if not file_path:
        print("未選擇檔案，程式結束")
        return None
    
    return file_path

def read_excel_data(file_path):
    """
    讀取Excel檔案

    包含完整的防呆措施
    """
    # ===== 防呆1: 文件存在性檢查 =====
    if not os.path.exists(file_path):
        print(f"❌ 錯誤：找不到檔案 '{file_path}'")
        raise FileNotFoundError(f"找不到檔案: {file_path}")

    # ===== 防呆2: 文件格式檢查 =====
    if not (file_path.endswith('.xlsx') or file_path.endswith('.xls')):
        print(f"❌ 錯誤：輸入檔案必須是Excel格式 (.xlsx 或 .xls)，但提供了 {file_path}")
        raise ValueError(f"不支援的檔案格式: {file_path}")

    # ===== 防呆3: 文件大小檢查 =====
    file_size = os.path.getsize(file_path)
    if file_size == 0:
        print(f"❌ 錯誤：檔案大小為 0 bytes，可能是空檔案")
        raise ValueError("檔案大小為 0 bytes")
    elif file_size < 1024:  # 小於 1KB
        print(f"⚠️ 警告：檔案大小僅 {file_size} bytes，可能不是有效的 Excel 檔案")

    print(f"✓ 檔案大小: {file_size / 1024:.2f} KB")

    # ===== 防呆4: Excel 文件有效性檢查 =====
    try:
        xl_file = pd.ExcelFile(file_path)
    except Exception as e:
        print(f"❌ 錯誤：無法讀取 Excel 檔案，可能已損壞或格式不正確")
        print(f"詳細錯誤: {e}")
        raise ValueError(f"無法讀取 Excel 檔案: {e}")

    sheet_names = xl_file.sheet_names
    print(f"✓ 找到工作表: {', '.join(sheet_names)}")

    # ===== 防呆5: 必要工作表檢查 =====
    if SHEET_NAMES['sample_info'] not in sheet_names:
        print(f"❌ 錯誤：找不到 '{SHEET_NAMES['sample_info']}' 工作表")
        print(f"可用的工作表: {', '.join(sheet_names)}")
        raise ValueError(f"找不到 '{SHEET_NAMES['sample_info']}' 工作表")

    # ===== 防呆6: 讀取 SampleInfo =====
    try:
        sample_info = pd.read_excel(file_path, sheet_name=SHEET_NAMES['sample_info'])
    except Exception as e:
        print(f"❌ 錯誤：讀取 '{SHEET_NAMES['sample_info']}' 工作表失敗")
        print(f"詳細錯誤: {e}")
        raise ValueError(f"讀取 '{SHEET_NAMES['sample_info']}' 失敗: {e}")

    # ===== 防呆7: SampleInfo 完整性檢查 =====
    if sample_info.empty:
        print(f"❌ 錯誤：'{SHEET_NAMES['sample_info']}' 工作表為空")
        raise ValueError(f"'{SHEET_NAMES['sample_info']}' 工作表為空")

    required_columns = ['Sample_Name']
    missing_cols = [col for col in required_columns if col not in sample_info.columns]
    if missing_cols:
        print(f"❌ 錯誤：'{SHEET_NAMES['sample_info']}' 缺少必要欄位: {', '.join(missing_cols)}")
        print(f"找到的欄位: {', '.join(sample_info.columns.tolist())}")
        raise ValueError(f"'{SHEET_NAMES['sample_info']}' 缺少必要欄位: {missing_cols}")

    print(f"✓ SampleInfo 包含 {len(sample_info)} 筆樣本資訊")

    # 選擇數據工作表
    sheet_priority = [SHEET_NAMES['qc_lowess'], SHEET_NAMES['istd_correction'], SHEET_NAMES['batch_effect'], SHEET_NAMES['raw_intensity']]
    data_sheet = None

    for sheet in sheet_priority:
        if sheet in sheet_names:
            data_sheet = sheet
            break

    if data_sheet is None:
        for sheet in sheet_names:
            if sheet != SHEET_NAMES['sample_info']:
                data_sheet = sheet
                break

    # ===== 防呆8: 數據工作表檢查 =====
    if data_sheet is None:
        print(f"❌ 錯誤：找不到數據工作表")
        print(f"可用的工作表: {', '.join(sheet_names)}")
        raise ValueError("找不到數據工作表")

    print(f"✓ 選擇數據工作表: '{data_sheet}'")

    # ===== 防呆9: 讀取數據工作表 =====
    try:
        data = pd.read_excel(file_path, sheet_name=data_sheet)
    except Exception as e:
        print(f"❌ 錯誤：讀取 '{data_sheet}' 工作表失敗")
        print(f"詳細錯誤: {e}")
        raise ValueError(f"讀取 '{data_sheet}' 失敗: {e}")

    # ===== 提取 Sample_Type 資訊行（不參與數值計算，保存時回插）=====
    from ms_core.utils.data_helpers import extract_sample_type_row
    feature_col = data.columns[0]  # 通常是 FeatureID 或 Mz/RT
    data, sample_type_row = extract_sample_type_row(data, feature_col)
    if sample_type_row is not None:
        print(f"✓ 偵測到 Sample_Type 資訊行，已提取保存（不參與計算）")

    # ===== 防呆10: 數據基本檢查 =====
    if data.empty:
        print(f"❌ 錯誤：'{data_sheet}' 工作表為空")
        raise ValueError(f"'{data_sheet}' 工作表為空")

    if data.shape[1] < 2:
        print(f"❌ 錯誤：'{data_sheet}' 欄位數不足（至少需要 2 欄）")
        raise ValueError(f"'{data_sheet}' 欄位數不足")

    print(f"✓ 數據維度: {data.shape[0]} 列 × {data.shape[1]} 欄")

    return data, sample_info, data_sheet, sample_type_row

def prepare_data_for_combat(data, sample_info):
    """
    準備Combat所需的數據格式

    包含完整的防呆措施
    """
    # ===== 防呆11: 輸入數據驗證 =====
    if data is None or data.empty:
        print(f"❌ 錯誤：輸入數據為空")
        raise ValueError("輸入數據為空")

    if sample_info is None or sample_info.empty:
        print(f"❌ 錯誤：樣本資訊為空")
        raise ValueError("樣本資訊為空")

    data_columns = data.columns.tolist()
    if len(data_columns) < 2:
        print(f"❌ 錯誤：數據欄位數不足（至少需要 2 欄）")
        raise ValueError("數據欄位數不足")

    feature_col = data_columns[0]
    # 排除已知的非樣本欄位（統計欄位等）
    non_sample = NON_SAMPLE_COLUMNS | {feature_col}
    sample_columns = [col for col in data_columns[1:]
                      if col not in non_sample
                      and not any(kw in str(col).lower() for kw in STAT_COLUMN_KEYWORDS)]

    print(f"\n準備 Combat 數據格式...")
    print(f"  - 特徵欄位: '{feature_col}'")
    print(f"  - 樣本數量: {len(sample_columns)}")

    # ===== 防呆12: Batch 欄位檢查 =====
    if 'Batch' not in sample_info.columns:
        print(f"⚠️ 警告: SampleInfo 缺少 'Batch' 欄位，嘗試推測...")

        # 嘗試從第 4 欄（Column D）推測
        if sample_info.shape[1] >= 4:
            column_d = sample_info.columns[3]
            if 'batch' in str(column_d).lower():
                sample_info['Batch'] = sample_info[column_d]
                print(f"  ✓ 使用 '{column_d}' 作為 Batch 欄位")
            else:
                print(f"❌ 錯誤：無法找到 Batch 資訊")
                print(f"  SampleInfo 欄位: {', '.join(sample_info.columns.tolist())}")
                raise ValueError("SampleInfo 缺少 'Batch' 欄位")
        else:
            print(f"❌ 錯誤：SampleInfo 欄位數不足")
            raise ValueError("SampleInfo 缺少 'Batch' 欄位")

    # ===== 防呆13: Sample_Name 欄位檢查 =====
    if 'Sample_Name' not in sample_info.columns:
        print(f"❌ 錯誤：SampleInfo 缺少 'Sample_Name' 欄位")
        print(f"  SampleInfo 欄位: {', '.join(sample_info.columns.tolist())}")
        raise ValueError("SampleInfo 缺少 'Sample_Name' 欄位")

    # 匹配樣本 - Vectorized (faster than iterrows)
    batch_na_count = sample_info['Batch'].isna().sum()

    # Build sample_to_batch dictionary using vectorized operations
    valid_batch_mask = sample_info['Batch'].notna()
    sample_names = sample_info.loc[valid_batch_mask, 'Sample_Name'].astype(str).str.strip()
    batch_values = sample_info.loc[valid_batch_mask, 'Batch']
    sample_to_batch = dict(zip(sample_names, batch_values))

    if batch_na_count > 0:
        print(f"⚠️ 警告：{batch_na_count} 個樣本的 Batch 資訊為空")

    # ===== 防呆14: 樣本匹配檢查 =====
    if len(sample_to_batch) == 0:
        print(f"❌ 錯誤：所有樣本的 Batch 資訊都為空")
        raise ValueError("無有效的 Batch 資訊")

    valid_samples = []
    valid_batches = []

    for col in sample_columns:
        col_str = str(col).strip()
        if col_str in sample_to_batch:
            valid_samples.append(col)
            valid_batches.append(sample_to_batch[col_str])

    # ===== 防呆15: 有效樣本數檢查 =====
    if len(valid_samples) < 2:
        print(f"⚠️ 警告：直接匹配的樣本數不足 ({len(valid_samples)})，嘗試部分匹配...")

        # 嘗試部分匹配（大小寫不敏感）
        sample_to_batch_lower = {k.lower(): (k, v) for k, v in sample_to_batch.items()}

        for col in sample_columns:
            col_lower = str(col).strip().lower()
            if col_lower in sample_to_batch_lower and col not in valid_samples:
                original_name, batch = sample_to_batch_lower[col_lower]
                valid_samples.append(col)
                valid_batches.append(batch)
                print(f"  ✓ 部分匹配: '{col}' → '{original_name}'")

        # 嘗試按順序對齊（SampleInfo 與數據欄位順序一致）
        if len(valid_samples) < 2:
            all_info_names = sample_info['Sample_Name'].astype(str).str.strip().tolist()
            if len(all_info_names) == len(sample_columns):
                print(f"  ⚠️ 名稱不匹配，嘗試按順序對齊...")
                valid_samples = list(sample_columns)
                valid_batches = []
                for info_name in all_info_names:
                    if info_name in sample_to_batch:
                        valid_batches.append(sample_to_batch[info_name])
                    else:
                        valid_batches.append('Unknown')
                print(f"  ✓ 按順序對齊成功: {len(valid_samples)} 個樣本")

        if len(valid_samples) < 2:
            print(f"❌ 錯誤：有效樣本數不足 ({len(valid_samples)} < 2)")
            print(f"  數據欄位: {', '.join(sample_columns[:5])}...")
            print(f"  SampleInfo 樣本: {', '.join(list(sample_to_batch.keys())[:5])}...")
            raise ValueError(f"有效樣本數不足: {len(valid_samples)}")

    print(f"  ✓ 成功匹配 {len(valid_samples)} 個樣本")

    # ===== 防呆16: 批次數量檢查 =====
    unique_batches = list(set(valid_batches))
    if len(unique_batches) < 2:
        print(f"❌ 錯誤：批次數量不足 ({len(unique_batches)} < 2)")
        print(f"  唯一批次: {unique_batches}")
        raise ValueError(f"批次數量不足，需要至少 2 個批次")

    print(f"  ✓ 批次資訊: {unique_batches} (共 {len(unique_batches)} 個批次)")

    # 統計每個批次的樣本數
    from collections import Counter
    batch_counts = Counter(valid_batches)
    print(f"\n  批次樣本分佈:")
    for batch, count in sorted(batch_counts.items()):
        print(f"    - Batch {batch}: {count} 個樣本")
        if count < 2:
            print(f"      ⚠️ 警告：樣本數過少，可能影響校正效果")

    # ===== 防呆17: 數據矩陣提取 =====
    try:
        # 確保樣本欄位為 numeric dtype（Sample_Type 行移除後可能仍為 object）
        for col in valid_samples:
            data[col] = pd.to_numeric(data[col], errors='coerce')
        data_matrix = data[valid_samples].values.astype(np.float64)
    except Exception as e:
        print(f"❌ 錯誤：提取數據矩陣失敗")
        print(f"詳細錯誤: {e}")
        raise ValueError(f"提取數據矩陣失敗: {e}")

    # ===== 防呆18: 數據矩陣驗證 =====
    if data_matrix.shape[0] == 0:
        print(f"❌ 錯誤：數據矩陣為空（特徵數為 0）")
        raise ValueError("數據矩陣為空")

    if data_matrix.shape[1] != len(valid_samples):
        print(f"❌ 錯誤：數據矩陣維度不一致")
        print(f"  預期樣本數: {len(valid_samples)}")
        print(f"  實際樣本數: {data_matrix.shape[1]}")
        raise ValueError("數據矩陣與批次資訊維度不一致")

    print(f"  ✓ 數據矩陣維度: {data_matrix.shape[0]} 特徵 × {data_matrix.shape[1]} 樣本")

    # ===== 防呆19: NaN 值檢查 =====
    nan_count = np.isnan(data_matrix).sum()
    if nan_count > 0:
        nan_percentage = nan_count / data_matrix.size * 100
        print(f"  ⚠️ 警告：數據包含 {nan_count} 個 NaN 值 ({nan_percentage:.2f}%)")
        print(f"  這些值將在 Combat 校正前被填充為最小非零值的一半")

    # ===== 防呆20: 負值檢查 =====
    negative_count = (data_matrix < 0).sum()
    if negative_count > 0:
        print(f"  ⚠️ 警告：數據包含 {negative_count} 個負值")
        print(f"  這些值將在 Combat 校正前被設為 0")

    feature_ids = data[feature_col].values

    return data_matrix, valid_batches, valid_samples, feature_ids

def check_batch_confounding(sample_info):
    """
    檢測批次與樣本類型之間的混淆（Confounding）

    使用卡方檢定（Chi-square test）和 Cramer's V 檢測 Batch 與 Sample_Type 的關聯性

    Returns:
        dict: 包含檢測結果的字典
            - 'is_confounded': bool, 是否存在高度關聯
            - 'chi2_statistic': float, 卡方統計量
            - 'p_value': float, p值
            - 'cramers_v': float, Cramer's V 係數（0-1）
            - 'contingency_table': DataFrame, 列聯表
            - 'warning_message': str, 警告訊息
    """
    try:
        print("\n🔍 執行批次混淆檢測...")

        # 檢查必要欄位
        if 'Batch' not in sample_info.columns or 'Sample_Type' not in sample_info.columns:
            print("  ⚠️  警告：缺少 'Batch' 或 'Sample_Type' 欄位，跳過混淆檢測")
            return {'is_confounded': False, 'warning_message': '缺少必要欄位'}

        # 排除 QC 樣本（只檢查真實樣本的混淆）
        real_samples = sample_info[~sample_info['Sample_Type'].str.upper().str.contains('QC', na=False)].copy()

        if len(real_samples) < 10:
            print("  ⚠️  警告：真實樣本數量過少（< 10），跳過混淆檢測")
            return {'is_confounded': False, 'warning_message': '樣本數量不足'}

        # 建立列聯表（Contingency Table）
        contingency_table = pd.crosstab(real_samples['Batch'], real_samples['Sample_Type'])
        print(f"\n  列聯表（Batch vs Sample_Type）:")
        print(contingency_table)

        # 執行卡方檢定
        chi2, p_value, dof, expected = stats.chi2_contingency(contingency_table)

        # 計算 Cramer's V（標準化的關聯強度指標，範圍 0-1）
        n = real_samples.shape[0]
        min_dim = min(contingency_table.shape[0] - 1, contingency_table.shape[1] - 1)
        cramers_v = np.sqrt(chi2 / (n * min_dim)) if min_dim > 0 else 0

        print(f"\n  卡方檢定結果:")
        print(f"    - χ² 統計量: {chi2:.4f}")
        print(f"    - p-value: {p_value:.4e}")
        print(f"    - Cramer's V: {cramers_v:.4f}")

        # 判斷標準
        is_confounded = False
        warning_message = ""

        # Cramer's V 解釋（Cohen's 標準）：
        # 0.00-0.10: 微弱關聯
        # 0.10-0.30: 中等關聯
        # 0.30-0.50: 強關聯
        # >0.50: 非常強關聯

        if cramers_v > 0.5:
            is_confounded = True
            warning_message = f"⚠️ 嚴重警告：Batch 與 Sample_Type 高度混淆（Cramer's V = {cramers_v:.4f}）"
        elif cramers_v > 0.3:
            is_confounded = True
            warning_message = f"⚠️ 警告：Batch 與 Sample_Type 存在明顯關聯（Cramer's V = {cramers_v:.4f}）"
        elif cramers_v > 0.1:
            warning_message = f"⚠️ 提示：Batch 與 Sample_Type 存在弱關聯（Cramer's V = {cramers_v:.4f}）"
        else:
            warning_message = f"✓ Batch 與 Sample_Type 無明顯關聯（Cramer's V = {cramers_v:.4f}）"

        print(f"\n  {warning_message}")

        if is_confounded:
            print(f"\n  {'='*70}")
            print(f"  ⚠️⚠️⚠️ 批次混淆警告 ⚠️⚠️⚠️")
            print(f"  {'='*70}")
            print(f"  檢測到 Batch 與 Sample_Type 高度相關！")
            print(f"  這意味著批次效應可能與生物效應混淆。")
            print(f"  ")
            print(f"  範例：")
            print(f"    - Batch 1 幾乎全是 Control 樣本")
            print(f"    - Batch 2 幾乎全是 Treated 樣本")
            print(f"  ")
            print(f"  在這種情況下，ComBat 可能會：")
            print(f"    1. 移除真正的生物訊號（Treatment effect）")
            print(f"    2. 導致假陰性結果（Type II error）")
            print(f"  ")
            print(f"  建議：")
            print(f"    - ❌ 不要使用 ComBat 批次校正")
            print(f"    - ✓ 改用 QC-LOWESS 的全域校正結果")
            print(f"    - ✓ 或在統計分析中將 Batch 作為協變量（covariate）")
            print(f"  {'='*70}\n")

        return {
            'is_confounded': is_confounded,
            'chi2_statistic': float(chi2),
            'p_value': float(p_value),
            'cramers_v': float(cramers_v),
            'contingency_table': contingency_table,
            'warning_message': warning_message
        }

    except Exception as e:
        print(f"  ⚠️  警告：批次混淆檢測失敗: {e}")
        import traceback
        traceback.print_exc()
        return {'is_confounded': False, 'warning_message': f'檢測失敗: {e}'}


def perform_combat_correction(data_matrix, batch_info):
    """
    執行Combat批次效應校正

    包含完整的防呆措施
    """
    # ===== 防呆21: 輸入驗證 =====
    if data_matrix is None or data_matrix.size == 0:
        print(f"❌ 錯誤：數據矩陣為空")
        raise ValueError("數據矩陣為空")

    if batch_info is None or len(batch_info) == 0:
        print(f"❌ 錯誤：批次資訊為空")
        raise ValueError("批次資訊為空")

    # ===== 防呆22: 維度一致性檢查 =====
    if data_matrix.shape[1] != len(batch_info):
        print(f"❌ 錯誤：數據矩陣樣本數與批次資訊數量不一致")
        print(f"  數據矩陣樣本數: {data_matrix.shape[1]}")
        print(f"  批次資訊數量: {len(batch_info)}")
        raise ValueError("數據矩陣與批次資訊維度不一致")

    print(f"\n執行 Combat 批次效應校正...")
    print(f"  - 輸入維度: {data_matrix.shape[0]} 特徵 × {data_matrix.shape[1]} 樣本")

    # ===== 防呆23: 數據類型轉換 =====
    try:
        data_matrix = data_matrix.astype(float)
    except Exception as e:
        print(f"❌ 錯誤：無法將數據轉換為 float 類型")
        print(f"詳細錯誤: {e}")
        raise ValueError(f"數據類型轉換失敗: {e}")

    # ===== 防呆24: 數值處理 =====
    # 處理零值和 NaN
    min_nonzero = np.min(data_matrix[data_matrix > 0]) if np.any(data_matrix > 0) else 1

    if min_nonzero == 0:
        print(f"⚠️ 警告：最小非零值為 0，使用預設值 1")
        min_nonzero = 1

    zero_count = (data_matrix == 0).sum()
    if zero_count > 0:
        print(f"  - 處理 {zero_count} 個零值（填充為 {min_nonzero / 2:.2e}）")
        data_matrix[data_matrix == 0] = min_nonzero / 2

    nan_count = np.isnan(data_matrix).sum()
    if nan_count > 0:
        print(f"  - 處理 {nan_count} 個 NaN 值（填充為 {min_nonzero / 2:.2e}）")
        data_matrix = np.nan_to_num(data_matrix, nan=min_nonzero / 2)

    # 處理負值
    negative_count = (data_matrix < 0).sum()
    if negative_count > 0:
        print(f"  - 處理 {negative_count} 個負值（設為 {min_nonzero / 2:.2e}）")
        data_matrix[data_matrix < 0] = min_nonzero / 2

    # Log2 轉換
    try:
        data_log = np.log2(data_matrix + 1)
    except Exception as e:
        print(f"❌ 錯誤：Log2 轉換失敗")
        print(f"詳細錯誤: {e}")
        raise ValueError(f"Log2 轉換失敗: {e}")

    # ===== 防呆25: Log 轉換後檢查 =====
    if np.any(np.isinf(data_log)):
        inf_count = np.isinf(data_log).sum()
        print(f"⚠️ 警告：Log2 轉換後產生 {inf_count} 個無限值")
        data_log = np.nan_to_num(data_log, nan=0.0, posinf=0.0, neginf=0.0)

    if np.any(np.isnan(data_log)):
        nan_count = np.isnan(data_log).sum()
        print(f"⚠️ 警告：Log2 轉換後產生 {nan_count} 個 NaN 值")
        data_log = np.nan_to_num(data_log, nan=0.0)

    print(f"  ✓ 數據預處理完成")
    print(f"  - Log2 轉換範圍: [{data_log.min():.2f}, {data_log.max():.2f}]")

    # ===== 防呆26: Combat 校正 =====
    try:
        combat_obj = pycombat.Combat()
        print(f"  - 執行 Combat 校正...")
        corrected_data = combat_obj.fit_transform(data_log.T, batch_info)
        corrected_data = corrected_data.T
        print(f"  ✓ Combat 校正完成")
    except Exception as e:
        print(f"❌ 錯誤：Combat 校正失敗")
        print(f"詳細錯誤: {e}")
        print(f"\n可能原因:")
        print(f"  1. 批次資訊格式不正確")
        print(f"  2. 某些批次樣本數過少")
        print(f"  3. 數據中存在常數列或近似常數列")
        raise ValueError(f"Combat 校正失敗: {e}")

    # ===== 防呆27: 反轉 Log2 轉換 =====
    try:
        corrected_data = np.power(2, corrected_data) - 1
    except Exception as e:
        print(f"❌ 錯誤：反轉 Log2 轉換失敗")
        print(f"詳細錯誤: {e}")
        raise ValueError(f"反轉 Log2 轉換失敗: {e}")

    # 處理負值（由於數值精度可能產生）
    negative_count_after = (corrected_data < 0).sum()
    if negative_count_after > 0:
        print(f"  - 修正 {negative_count_after} 個負值（設為 0）")
        corrected_data[corrected_data < 0] = 0

    # ===== 防呆28: 輸出驗證 =====
    if np.any(np.isnan(corrected_data)):
        nan_count = np.isnan(corrected_data).sum()
        print(f"⚠️ 警告：校正後數據包含 {nan_count} 個 NaN 值")
        corrected_data = np.nan_to_num(corrected_data, nan=0.0)

    if np.any(np.isinf(corrected_data)):
        inf_count = np.isinf(corrected_data).sum()
        print(f"⚠️ 警告：校正後數據包含 {inf_count} 個無限值")
        corrected_data = np.nan_to_num(corrected_data, nan=0.0, posinf=0.0, neginf=0.0)

    print(f"  ✓ 輸出維度: {corrected_data.shape[0]} 特徵 × {corrected_data.shape[1]} 樣本")
    print(f"  ✓ 輸出數值範圍: [{corrected_data.min():.2e}, {corrected_data.max():.2e}]")

    return corrected_data
def calculate_permanova(data, batch_labels, distance_metric='manhattan', permutations=999):
    """
    使用 PERMANOVA 評估批次效應

    包含完整的參數驗證
    """
    # ===== 檢查 scikit-bio 是否可用 =====
    if not SKBIO_AVAILABLE:
        print("❌ 錯誤：scikit-bio 未安裝，無法執行 PERMANOVA")
        raise ImportError("scikit-bio is required for PERMANOVA. Install with: pip install scikit-bio")

    # ===== 防呆29: 輸入數據驗證 =====
    if data is None or data.size == 0:
        print(f"❌ 錯誤：輸入數據為空")
        raise ValueError("輸入數據為空")

    if batch_labels is None or len(batch_labels) == 0:
        print(f"❌ 錯誤：批次標籤為空")
        raise ValueError("批次標籤為空")

    # ===== 防呆30: 維度一致性檢查 =====
    if len(data) != len(batch_labels):
        print(f"❌ 錯誤：數據樣本數與批次標籤數量不一致")
        print(f"  數據樣本數: {len(data)}")
        print(f"  批次標籤數量: {len(batch_labels)}")
        raise ValueError("數據與批次標籤維度不一致")

    # ===== 防呆31: 參數驗證 =====
    if distance_metric not in ['manhattan', 'euclidean']:
        print(f"❌ 錯誤：不支援的距離度量 '{distance_metric}'")
        print(f"  支援的度量: 'manhattan', 'euclidean'")
        raise ValueError(f"不支援的距離度量: {distance_metric}")

    if permutations < 0:
        print(f"❌ 錯誤：排列次數必須為非負整數")
        raise ValueError(f"排列次數不能為負: {permutations}")

    if permutations > 100000:
        print(f"⚠️ 警告：排列次數過多 ({permutations})，可能耗時很長")

    # ===== 防呆32: 批次數量檢查 =====
    unique_batches = np.unique(batch_labels)
    if len(unique_batches) < 2:
        print(f"❌ 錯誤：批次數量不足 ({len(unique_batches)} < 2)")
        raise ValueError("批次數量不足，需要至少 2 個批次")

    # ===== 防呆33: 樣本數檢查 =====
    if len(data) < len(unique_batches) + 1:
        print(f"⚠️ 警告：樣本數相對於批次數較少")
        print(f"  樣本數: {len(data)}, 批次數: {len(unique_batches)}")

    print(f"\n🔬 執行 PERMANOVA 批次效應檢驗")
    print(f"   - 距離度量: {distance_metric.capitalize()}")
    print(f"   - 排列次數: {permutations}")
    print(f"   - 樣本數: {len(data)}")
    print(f"   - 批次數: {len(unique_batches)}")

    # 計算距離矩陣
    try:
        if distance_metric == 'manhattan':
            dist_matrix = manhattan_distances(data)
        elif distance_metric == 'euclidean':
            dist_matrix = euclidean_distances(data)
    except Exception as e:
        print(f"❌ 錯誤：計算距離矩陣失敗")
        print(f"詳細錯誤: {e}")
        raise ValueError(f"距離矩陣計算失敗: {e}")
    
    # 轉換為 DistanceMatrix 對象
    sample_ids = [f"S{i}" for i in range(len(data))]
    dm = DistanceMatrix(dist_matrix, ids=sample_ids)
    
    # 執行 PERMANOVA
    result = skbio_permanova(dm, batch_labels, permutations=permutations)
    
    # 計算自由度和 R²
    batch_labels_array = np.array(batch_labels)
    unique_batches = np.unique(batch_labels_array)
    n_samples = len(batch_labels_array)
    k = len(unique_batches)
    
    df_between = k - 1
    df_within = n_samples - k
    
    f_stat = result['test statistic']
    
    # 計算 R²
    if df_within > 0:
        r_squared = (f_stat * df_between) / (f_stat * df_between + df_within)
    else:
        r_squared = 0
    
    r_squared = max(0, min(1, r_squared))
    
    # 🆕 計算 η² (Eta-squared，等同於 R²)
    eta_squared = r_squared
    
    # 🆕 完整報告
    print(f"\n   結果:")
    print(f"   - Pseudo-F statistic: {f_stat:.4f}")
    print(f"   - 自由度: df_between={df_between}, df_within={df_within}")
    print(f"   - 樣本數: n={n_samples}, 批次數: k={k}")
    print(f"   - p-value: {result['p-value']:.4f}", end="")
    
    if result['p-value'] < 0.001:
        print(f" *** (極顯著)")
    elif result['p-value'] < 0.01:
        print(f" ** (非常顯著)")
    elif result['p-value'] < 0.05:
        print(f" * (顯著)")
    else:
        print(f" n.s. (不顯著)")
    
    print(f"   - R² (批次解釋變異): {r_squared*100:.1f}%")
    print(f"   - η² (效應量): {eta_squared:.4f}", end="")
    
    # η² 解釋
    if eta_squared < 0.01:
        print(f" (微弱效應)")
    elif eta_squared < 0.06:
        print(f" (小效應)")
    elif eta_squared < 0.14:
        print(f" (中等效應)")
    else:
        print(f" (大效應)")
    
    return {
        'pseudo_f': f_stat,
        'p_value': result['p-value'],
        'r_squared': r_squared,
        'eta_squared': eta_squared,
        'df_between': df_between,
        'df_within': df_within,
        'sample_size': n_samples,
        'n_batches': k,
        'permutations': result['number of permutations'],
        'distance_metric': distance_metric
    }

def calculate_permdisp(data, batch_labels, distance_metric='manhattan', permutations=999):
    """
    PERMDISP: 檢驗批次間離散度同質性
    (Homogeneity of Multivariate Dispersions)

    用途：確認 PERMANOVA 的顯著性是否來自位置差異（批次效應）
         還是離散度差異（變異不同質）

    Note: Requires scikit-bio to be installed.

    Parameters:
    -----------
    data : np.ndarray
        樣本 x 特徵矩陣
    batch_labels : list or np.ndarray
        批次標籤
    distance_metric : str
        距離度量
    permutations : int
        排列次數

    Returns:
    --------
    dict: PERMDISP 結果
    """
    # ===== 檢查 scikit-bio 是否可用 =====
    if not SKBIO_AVAILABLE:
        print("\n🔍 執行 PERMDISP 檢驗 (檢驗離散度同質性)...")
        print("   ⚠️ scikit-bio 未安裝，無法執行 PERMDISP")
        return {
            'f_statistic': np.nan,
            'p_value': np.nan,
            'permutations': 0
        }

    print(f"\n🔍 執行 PERMDISP 檢驗 (檢驗離散度同質性)...")

    try:
        from skbio.stats.distance import permdisp
        
        # 計算距離矩陣
        if distance_metric == 'manhattan':
            dist_matrix = manhattan_distances(data)
        else:
            dist_matrix = euclidean_distances(data)
        
        # 轉換為 DistanceMatrix 對象
        sample_ids = [f"S{i}" for i in range(len(data))]
        dm = DistanceMatrix(dist_matrix, ids=sample_ids)
        
        # 執行 PERMDISP
        result = permdisp(dm, batch_labels, permutations=permutations)
        
        print(f"   - F-statistic: {result['test statistic']:.4f}")
        print(f"   - p-value: {result['p-value']:.4f}", end="")
        
        if result['p-value'] < 0.05:
            print(f" * (離散度異質，PERMANOVA 結果需謹慎解讀)")
        else:
            print(f" n.s. (離散度同質，PERMANOVA 結果可靠)")
        
        return {
            'f_statistic': result['test statistic'],
            'p_value': result['p-value'],
            'permutations': result['number of permutations']
        }
    
    except ImportError:
        print("   ⚠️ scikit-bio 版本過舊，無法執行 PERMDISP")
        return {
            'f_statistic': np.nan,
            'p_value': np.nan,
            'permutations': 0
        }
    except Exception as e:
        print(f"   ⚠️ PERMDISP 執行失敗: {e}")
        return {
            'f_statistic': np.nan,
            'p_value': np.nan,
            'permutations': 0
        }

def paired_permutation_test(data_before, data_after, batch_labels,
                            metric='permanova', distance_metric='manhattan',
                            n_permutations=1000):
    """
    配對排列檢定：測試校正是否顯著改善批次效應

    H0: 校正前後的批次效應指標無差異
    H1: 校正後的批次效應指標顯著低於校正前

    Parameters:
    -----------
    metric : str
        'permanova' 或 'silhouette'
    """
    # ===== 檢查 scikit-bio 是否可用 (僅 permanova metric 需要) =====
    if metric == 'permanova' and not SKBIO_AVAILABLE:
        print(f"\n🎲 執行配對排列檢定 ({metric.upper()}, {n_permutations} 次排列)...")
        print("   ⚠️ scikit-bio 未安裝，無法執行 PERMANOVA 排列檢定")
        return {
            'observed_improvement': np.nan,
            'p_value': np.nan,
            'null_mean': np.nan,
            'null_std': np.nan,
            'effect_size': np.nan,
            'metric': metric
        }

    print(f"\n🎲 執行配對排列檢定 ({metric.upper()}, {n_permutations} 次排列)...")
    
    # 計算觀察到的改善
    if metric == 'permanova':
        # 計算校正前後的 Pseudo-F
        if distance_metric == 'manhattan':
            dist_before = manhattan_distances(data_before)
            dist_after = manhattan_distances(data_after)
        else:
            dist_before = euclidean_distances(data_before)
            dist_after = euclidean_distances(data_after)
        
        sample_ids = [f"S{i}" for i in range(len(data_before))]
        dm_before = DistanceMatrix(dist_before, ids=sample_ids)
        dm_after = DistanceMatrix(dist_after, ids=sample_ids)
        
        result_before = skbio_permanova(dm_before, batch_labels, permutations=0)
        result_after = skbio_permanova(dm_after, batch_labels, permutations=0)
        
        observed_improvement = result_before['test statistic'] - result_after['test statistic']
        
    elif metric == 'silhouette':
        score_before = silhouette_score(data_before, batch_labels, metric='manhattan' if distance_metric == 'manhattan' else 'euclidean')
        score_after = silhouette_score(data_after, batch_labels, metric='manhattan' if distance_metric == 'manhattan' else 'euclidean')
        observed_improvement = score_before - score_after
    
    print(f"   - 觀察改善值: {observed_improvement:.4f}")
    
    # Permutation Test
    null_distribution = []
    n_samples = len(batch_labels)
    
    for i in range(n_permutations):
        # 隨機交換「校正前」和「校正後」的標籤（配對設計）
        swap_mask = np.random.randint(0, 2, size=n_samples).astype(bool)
        
        # 創建排列後的數據
        perm_before = np.where(swap_mask[:, None], data_after, data_before)
        perm_after = np.where(swap_mask[:, None], data_before, data_after)
        
        # 計算排列後的改善
        if metric == 'permanova':
            if distance_metric == 'manhattan':
                perm_dist_before = manhattan_distances(perm_before)
                perm_dist_after = manhattan_distances(perm_after)
            else:
                perm_dist_before = euclidean_distances(perm_before)
                perm_dist_after = euclidean_distances(perm_after)
            
            perm_dm_before = DistanceMatrix(perm_dist_before, ids=sample_ids)
            perm_dm_after = DistanceMatrix(perm_dist_after, ids=sample_ids)
            
            perm_result_before = skbio_permanova(perm_dm_before, batch_labels, permutations=0)
            perm_result_after = skbio_permanova(perm_dm_after, batch_labels, permutations=0)
            
            perm_improvement = perm_result_before['test statistic'] - perm_result_after['test statistic']
            
        elif metric == 'silhouette':
            perm_score_before = silhouette_score(perm_before, batch_labels, metric='manhattan' if distance_metric == 'manhattan' else 'euclidean')
            perm_score_after = silhouette_score(perm_after, batch_labels, metric='manhattan' if distance_metric == 'manhattan' else 'euclidean')
            perm_improvement = perm_score_before - perm_score_after
        
        null_distribution.append(perm_improvement)
        
        # 進度顯示
        if (i + 1) % 200 == 0:
            print(f"      進度: {i + 1}/{n_permutations}")
    
    null_distribution = np.array(null_distribution)
    
    # 計算 p-value (單尾檢定，右尾)
    p_value = np.sum(null_distribution >= observed_improvement) / n_permutations
    
    print(f"   - p-value: {p_value:.4f}", end="")
    if p_value < 0.001:
        print(f" *** (極顯著)")
    elif p_value < 0.01:
        print(f" ** (非常顯著)")
    elif p_value < 0.05:
        print(f" * (顯著)")
    else:
        print(f" n.s. (不顯著)")
    
    return {
        'observed_improvement': observed_improvement,
        'null_distribution': null_distribution,
        'p_value': p_value,
        'null_mean': np.mean(null_distribution),
        'null_std': np.std(null_distribution)
    }

def calculate_cohens_d_batch_effect(data, batch_labels):
    """計算批次效應的 Cohen's d 效果量"""
    from itertools import combinations
    
    print("\n📏 計算 Cohen's d 效果量...")
    
    unique_batches = sorted(list(set(batch_labels)))
    batch_labels = np.array(batch_labels)
    
    if len(unique_batches) < 2:
        return {
            'overall_cohens_d': 0.0,
            'feature_cohens_d': np.zeros(data.shape[1]),
            'pairwise_cohens_d': {}
        }
    
    n_features = data.shape[1]
    feature_cohens_d_list = []
    pairwise_results = {}
    
    batch_pairs = list(combinations(unique_batches, 2))
    
    for batch1, batch2 in batch_pairs:
        batch1_data = data[batch_labels == batch1]
        batch2_data = data[batch_labels == batch2]
        
        n1 = len(batch1_data)
        n2 = len(batch2_data)
        
        if n1 < 2 or n2 < 2:
            continue
        
        mean1 = np.mean(batch1_data, axis=0)
        mean2 = np.mean(batch2_data, axis=0)
        std1 = np.std(batch1_data, axis=0, ddof=1)
        std2 = np.std(batch2_data, axis=0, ddof=1)
        
        pooled_std = np.sqrt(((n1 - 1) * std1**2 + (n2 - 1) * std2**2) / (n1 + n2 - 2))
        pooled_std = np.where(pooled_std == 0, 1e-10, pooled_std)
        
        cohens_d = np.abs(mean1 - mean2) / pooled_std
        
        feature_cohens_d_list.append(cohens_d)
        pairwise_results[f'Batch{batch1}_vs_Batch{batch2}'] = np.mean(cohens_d)
    
    if len(feature_cohens_d_list) == 0:
        return {
            'overall_cohens_d': 0.0,
            'feature_cohens_d': np.zeros(n_features),
            'pairwise_cohens_d': {}
        }
    
    feature_cohens_d = np.mean(feature_cohens_d_list, axis=0)
    overall_cohens_d = np.mean(feature_cohens_d)
    
    print(f"   - 整體平均 Cohen's d: {overall_cohens_d:.4f}")
    if overall_cohens_d < COHENS_D_THRESHOLDS['small']:
        print(f"     • 小效果 (d < {COHENS_D_THRESHOLDS['small']})")
    elif overall_cohens_d < COHENS_D_THRESHOLDS['medium']:
        print(f"     • 小至中等效果 ({COHENS_D_THRESHOLDS['small']} ≤ d < {COHENS_D_THRESHOLDS['medium']})")
    elif overall_cohens_d < COHENS_D_THRESHOLDS['large']:
        print(f"     • 中等效果 ({COHENS_D_THRESHOLDS['medium']} ≤ d < {COHENS_D_THRESHOLDS['large']})")
    else:
        print(f"     • 大效果 (d ≥ {COHENS_D_THRESHOLDS['large']})")
    
    print(f"\n   批次對之間的 Cohen's d:")
    for pair, d_value in pairwise_results.items():
        print(f"     • {pair}: {d_value:.4f}")
    
    return {
        'overall_cohens_d': overall_cohens_d,
        'feature_cohens_d': feature_cohens_d,
        'pairwise_cohens_d': pairwise_results
    }

def calculate_qc_cv(data, sample_info, sample_columns):
    """計算 QC 樣本的變異係數 (CV%)"""
    print("\n📊 計算 QC 樣本 CV%...")
    
    sample_meta = sample_info.set_index('Sample_Name')
    qc_columns = []
    
    for col in sample_columns:
        if col in sample_meta.index:
            sample_type = sample_meta.loc[col].get('Sample_Type', 'Unknown')
            sample_type_upper = str(sample_type).upper()
            if 'QC' in sample_type_upper:
                qc_columns.append(col)
        else:
            if 'QC' in col.upper():
                qc_columns.append(col)
    
    if len(qc_columns) < 3:
        print(f"   ⚠ QC 樣本數不足 ({len(qc_columns)})，無法計算可靠的 CV%")
        return {
            'qc_cv': np.array([]),
            'median_cv': np.nan,
            'mean_cv': np.nan,
            'cv_below_20': 0.0,
            'cv_below_30': 0.0
        }
    
    print(f"   - 找到 {len(qc_columns)} 個 QC 樣本")
    
    qc_indices = [sample_columns.index(col) for col in qc_columns]
    qc_data = data[:, qc_indices]
    
    qc_mean = np.mean(qc_data, axis=1)
    qc_std = np.std(qc_data, axis=1, ddof=1)
    qc_mean = np.where(qc_mean == 0, 1e-10, qc_mean)
    qc_cv = (qc_std / qc_mean) * 100
    
    median_cv = np.median(qc_cv)
    mean_cv = np.mean(qc_cv)
    cv_below_20 = np.sum(qc_cv < CV_QUALITY_THRESHOLDS['excellent']) / len(qc_cv) * 100
    cv_below_30 = np.sum(qc_cv < CV_QUALITY_THRESHOLDS['acceptable']) / len(qc_cv) * 100

    print(f"   - QC CV% 中位數: {median_cv:.2f}%")
    print(f"   - CV% < {CV_QUALITY_THRESHOLDS['excellent']:.0f}% 的代謝物: {cv_below_20:.1f}%")

    if median_cv < CV_QUALITY_THRESHOLDS['excellent']:
        print(f"   ✅ 技術重現性優良")
    elif median_cv < CV_QUALITY_THRESHOLDS['acceptable']:
        print(f"   ⚠ 技術重現性可接受")
    else:
        print(f"   ❌ 技術重現性較差")
    
    return {
        'qc_cv': qc_cv,
        'median_cv': median_cv,
        'mean_cv': mean_cv,
        'cv_below_20': cv_below_20,
        'cv_below_30': cv_below_30
    }

def calculate_silhouette_overall(data, batch_labels, distance_metric='manhattan'):
    """計算整體 Silhouette Coefficient（簡化版）"""
    print(f"\n📊 計算 Silhouette Coefficient ({distance_metric.capitalize()} 距離)...")
    
    try:
        score = silhouette_score(data, batch_labels, 
                                metric='manhattan' if distance_metric == 'manhattan' else 'euclidean')
        print(f"   - Silhouette Score: {score:.4f}")
        print(f"   ⚠️ 注意：分數越低表示批次分離越弱（越好）")
        return score
    except Exception as e:
        print(f"   ⚠ Silhouette 計算失敗: {e}")
        return np.nan
    
def generate_quality_warnings(permanova_before, permanova_after, 
                             perm_test_results, qc_cv_before, qc_cv_after,
                             cohens_d_before, cohens_d_after,
                             qc_outliers_before, qc_outliers_after,
                             batch_info):
    """
    生成品質檢查警告
    
    Returns:
    --------
    list of dict: 警告訊息列表
    """
    warnings_list = []
    
    # 1. 檢查 PERMANOVA 校正後是否仍顯著
    if _has_valid_permanova(permanova_after):
        if permanova_after['p_value'] < 0.05:
            warnings_list.append({
                'level': 'WARNING',
                'category': '批次效應未完全消除',
                'details': f"PERMANOVA 校正後仍顯著 (p={permanova_after['p_value']:.4f} < 0.05)",
                'suggestion': [
                    "1. 檢查樣本資訊是否正確標記批次",
                    "2. 考慮移除異常批次後重新校正",
                    "3. 或使用其他校正方法 (如 Limma)"
                ]
            })

        # 2. 檢查 R² 是否仍過高
        if permanova_after['r_squared'] > 0.15:
            warnings_list.append({
                'level': 'WARNING',
                'category': '批次仍解釋較多變異',
                'details': f"R² = {permanova_after['r_squared']*100:.1f}% (建議 < 10%)",
                'suggestion': [
                    "1. 批次效應可能與生物學差異混淆",
                    "2. 檢查批次內樣本數是否足夠",
                    "3. 考慮使用更強的校正參數"
                ]
            })

    # 3. 檢查配對檢定是否顯著
    perm_permanova = perm_test_results.get('permanova') if perm_test_results else None
    if perm_permanova and not np.isnan(perm_permanova.get('p_value', np.nan)):
        if perm_permanova['p_value'] > 0.05:
            warnings_list.append({
                'level': 'WARNING',
                'category': '改善統計上不顯著',
                'details': f"配對排列檢定 p={perm_permanova['p_value']:.4f} (p > 0.05)",
                'suggestion': [
                    "1. 增加排列次數 (1000 → 5000) 確認結果",
                    "2. 檢查批次分組是否正確",
                    "3. 考慮使用更強的校正參數"
                ]
            })
    
    # 4. 檢查 QC CV% 是否惡化
    if not np.isnan(qc_cv_before['median_cv']) and not np.isnan(qc_cv_after['median_cv']):
        if qc_cv_after['median_cv'] > qc_cv_before['median_cv']:
            warnings_list.append({
                'level': 'WARNING',
                'category': 'QC 技術重現性未改善',
                'details': f"QC CV% 中位數: {qc_cv_before['median_cv']:.2f}% → {qc_cv_after['median_cv']:.2f}% (惡化)",
                'suggestion': [
                    "1. 檢查 QC 樣本是否為真實技術重複",
                    "2. 考慮先執行 QC 校正，再執行批次校正",
                    "3. 查看 PCA 圖中 QC 聚集情況"
                ]
            })
        
        if qc_cv_after['median_cv'] > CV_QUALITY_THRESHOLDS['acceptable']:
            warnings_list.append({
                'level': 'WARNING',
                'category': 'QC 穩定性差',
                'details': f"QC CV% 中位數 = {qc_cv_after['median_cv']:.2f}% (> {CV_QUALITY_THRESHOLDS['acceptable']:.0f}%)",
                'suggestion': [
                    "1. QC 樣本可能存在品質問題",
                    "2. 檢查儀器穩定性",
                    "3. 考慮移除異常 QC 樣本"
                ]
            })
    
    # 5. 檢查 QC 異常值
    n_outliers_after = np.sum(qc_outliers_after) if len(qc_outliers_after) > 0 else 0
    if n_outliers_after >= 2:
        warnings_list.append({
            'level': 'WARNING',
            'category': 'QC 樣本存在異常值',
            'details': f"校正後仍有 {n_outliers_after} 個 QC 樣本超出 95% CI",
            'suggestion': [
                "1. 檢查這些 QC 樣本的原始數據品質",
                "2. 考慮移除異常 QC 後重新評估",
                "3. 查看這些樣本是否來自特定批次"
            ]
        })
    
    # 6. 檢查批次樣本數
    batch_counts = {}
    for batch in batch_info:
        if batch not in batch_counts:
            batch_counts[batch] = 0
        batch_counts[batch] += 1
    
    small_batches = [b for b, count in batch_counts.items() if count < 3]
    if small_batches:
        warnings_list.append({
            'level': 'WARNING',
            'category': '批次樣本數不足',
            'details': f"批次 {small_batches} 樣本數 < 3",
            'suggestion': [
                "1. 樣本數太少可能影響校正效果",
                "2. 考慮合併相鄰批次",
                "3. 或移除樣本數不足的批次"
            ]
        })
    
    return warnings_list

def print_warnings(warnings_list):
    """打印警告訊息"""
    if len(warnings_list) == 0:
        print("\n✅ 所有品質檢查通過")
        print("   建議: 可以使用校正後數據進行下游分析")
        return
    
    print(f"\n⚠️  品質檢查與警告 (共 {len(warnings_list)} 項)")
    print("="*70)
    
    for i, warning in enumerate(warnings_list, 1):
        print(f"\n[警告 {i}] {warning['category']}")
        print(f"   詳細: {warning['details']}")
        print(f"\n   建議:")
        for suggestion in warning['suggestion']:
            print(f"      {suggestion}")
        print()

def calculate_hotelling_t2_outliers(qc_scores, all_scores, alpha=0.05):
    """
    使用 Hotelling T² 檢測 QC 樣本是否偏離整體中心
    使用「所有樣本」的協方差矩陣
    """
    n_qc, p = qc_scores.shape
    n_all = all_scores.shape[0]
    
    if n_qc < 1 or n_all < 3:
        return np.zeros(n_qc), 0, np.zeros(n_qc, dtype=bool)
    
    # 使用所有樣本的統計量
    mean_all = np.mean(all_scores, axis=0)
    cov_all = np.cov(all_scores, rowvar=False)
    
    # 正則化協方差矩陣
    cov_reg = cov_all + np.eye(p) * 1e-6
    
    try:
        cov_inv = np.linalg.inv(cov_reg)
    except np.linalg.LinAlgError:
        cov_inv = np.linalg.pinv(cov_reg)
    
    # 計算每個 QC 樣本的 Hotelling T² 值
    t2_values = np.zeros(n_qc)
    for i in range(n_qc):
        diff = qc_scores[i] - mean_all
        t2_values[i] = np.dot(np.dot(diff, cov_inv), diff.T)
    
    # 使用卡方分布計算閾值
    threshold = chi2.ppf(1 - alpha, p)
    outliers = t2_values > threshold
    
    return t2_values, threshold, outliers

def calculate_hotelling_t2_outliers_internal(scores, alpha=0.05):
    """
    使用 Hotelling T² 檢測樣本內部的異常值
    使用「內部統計量」
    """
    n, p = scores.shape
    
    if n < 3:
        return np.zeros(n), 0, np.zeros(n, dtype=bool)
    
    # 使用內部統計量
    mean = np.mean(scores, axis=0)
    cov = np.cov(scores, rowvar=False)
    
    # 正則化協方差矩陣
    cov_reg = cov + np.eye(p) * 1e-6
    
    try:
        cov_inv = np.linalg.inv(cov_reg)
    except np.linalg.LinAlgError:
        cov_inv = np.linalg.pinv(cov_reg)
    
    # 計算 Hotelling T² 值
    t2_values = np.zeros(n)
    for i in range(n):
        diff = scores[i] - mean
        t2_values[i] = np.dot(np.dot(diff, cov_inv), diff.T)
    
    # 使用正確的閾值公式
    if n - p - 1 > 0:
        f_critical = f_dist.ppf(1 - alpha, p, n - p - 1)
        threshold = (p * (n + 1) * (n - 1)) / (n * (n - p - 1)) * f_critical
    else:
        threshold = chi2.ppf(1 - alpha, p)
    
    outliers = t2_values > threshold
    
    return t2_values, threshold, outliers

def draw_hotelling_t2_ellipse(ax, scores, alpha=0.05, label=None, 
                               edgecolor='red', linestyle='-', linewidth=2.5):
    """在 2D PCA 圖上繪製 95% Hotelling T² 橢圓"""
    n, p = scores.shape
    
    if n < 3:
        return None
    
    mean = np.mean(scores, axis=0)
    cov = np.cov(scores, rowvar=False)
    
    eigenvalues, eigenvectors = np.linalg.eigh(cov)
    idx = eigenvalues.argsort()[::-1]
    eigenvalues = eigenvalues[idx]
    eigenvectors = eigenvectors[:, idx]
    eigenvalues = np.maximum(eigenvalues, 1e-10)
    
    f_critical = f_dist.ppf(1 - alpha, p, n - p)
    scale_factor = np.sqrt((p * (n - 1) * (n + 1)) / (n * (n - p)) * f_critical)
    
    width = 2 * scale_factor * np.sqrt(eigenvalues[0])
    height = 2 * scale_factor * np.sqrt(eigenvalues[1])
    angle = np.degrees(np.arctan2(eigenvectors[1, 0], eigenvectors[0, 0]))
    
    ellipse = Ellipse(mean, width, height, angle=angle,
                     facecolor='none', edgecolor=edgecolor,
                     linewidth=linewidth, linestyle=linestyle, label=label)
    ax.add_patch(ellipse)
    
    # 計算橢圓邊界
    t = np.linspace(0, 2*np.pi, 100)
    ellipse_x = (width/2) * np.cos(t)
    ellipse_y = (height/2) * np.sin(t)
    
    cos_angle = np.cos(np.radians(angle))
    sin_angle = np.sin(np.radians(angle))
    x_rot = ellipse_x * cos_angle - ellipse_y * sin_angle + mean[0]
    y_rot = ellipse_x * sin_angle + ellipse_y * cos_angle + mean[1]
    
    bounds = (np.min(x_rot), np.max(x_rot), np.min(y_rot), np.max(y_rot))
    
    return bounds

def create_comparison_pca_plot(original_data, corrected_data, sample_info, 
                               sample_columns, batch_info, 
                               title_suffix="", grouping="batch"):
    """創建前後對比的 PCA 圖（左圖：校正前，右圖：校正後）。

    Parameters
    ----------
    grouping : str
        'batch' 或 'sample_type'
    """

    # 識別樣本類型
    sample_meta = sample_info.set_index('Sample_Name') if 'Sample_Name' in sample_info.columns else sample_info

    qc_columns = []
    control_columns = []
    exposed_columns = []
    sample_batches = {}
    sample_type_map = {}

    for col in sample_columns:
        if hasattr(sample_meta, 'index') and col in sample_meta.index:
            raw_type = str(sample_meta.loc[col].get('Sample_Type', 'Unknown'))
            batch = sample_meta.loc[col].get('Batch', 'Unknown')
            if (batch == 'Unknown' or pd.isna(batch)) and isinstance(batch_info, dict):
                batch = batch_info.get(col, batch)
        else:
            raw_type = 'Unknown'
            batch = batch_info.get(col, 'Unknown') if isinstance(batch_info, dict) else 'Unknown'

        norm_type = normalize_sample_type(raw_type)
        sample_type_map[col] = norm_type
        sample_batches[col] = batch

        if norm_type == 'QC':
            qc_columns.append(col)
        elif norm_type == 'Exposure':
            exposed_columns.append(col)
        elif norm_type in ('Control', 'Normal'):
            control_columns.append(col)

    if len(sample_columns) < 3:
        print("   ⚠ 樣本數量不足，無法進行 PCA 分析")
        return None

    # 數據預處理
    def preprocess_data(data):
        data_processed = data.astype(float)
        min_nonzero = np.min(data_processed[data_processed > 0]) if np.any(data_processed > 0) else 1
        data_processed[data_processed == 0] = min_nonzero / 2
        data_processed = np.nan_to_num(data_processed, nan=min_nonzero / 2)
        data_log = np.log2(data_processed + 1)
        scaler = StandardScaler()
        return scaler.fit_transform(data_log)

    original_scaled = preprocess_data(original_data)
    corrected_scaled = preprocess_data(corrected_data)

    # PCA
    pca_original = PCA(n_components=2)
    pca_corrected = PCA(n_components=2)

    scores_original = pca_original.fit_transform(original_scaled)
    scores_corrected = pca_corrected.fit_transform(corrected_scaled)

    var_original = pca_original.explained_variance_ratio_
    var_corrected = pca_corrected.explained_variance_ratio_

    # Hotelling T² 檢測 QC 異常值
    qc_indices = [i for i, col in enumerate(sample_columns) if col in qc_columns]
    if len(qc_indices) >= 3:
        qc_scores_original = scores_original[qc_indices]
        qc_scores_corrected = scores_corrected[qc_indices]

        if grouping == "batch":
            t2_orig, t2_threshold_orig, outliers_orig = calculate_hotelling_t2_outliers(
                qc_scores_original, scores_original, alpha=0.05
            )
            t2_corr, t2_threshold_corr, outliers_corr = calculate_hotelling_t2_outliers(
                qc_scores_corrected, scores_corrected, alpha=0.05
            )
        else:
            t2_orig, t2_threshold_orig, outliers_orig = calculate_hotelling_t2_outliers_internal(
                qc_scores_original, alpha=0.05
            )
            t2_corr, t2_threshold_corr, outliers_corr = calculate_hotelling_t2_outliers_internal(
                qc_scores_corrected, alpha=0.05
            )
    else:
        t2_threshold_orig = np.nan
        t2_threshold_corr = np.nan
        outliers_orig = np.zeros(0, dtype=bool)
        outliers_corr = np.zeros(0, dtype=bool)

    # 統一 PCA 圖樣式（以 QC 子程式風格為主）
    suptitle = (
        f'2D PCA Comparison: Before vs After Correction (Grouped by Batch){title_suffix}'
        if grouping == 'batch'
        else f'2D PCA Comparison: Before vs After Correction (Grouped by Sample Type){title_suffix}'
    )

    sample_types = [sample_type_map.get(col, 'Unknown') for col in sample_columns]
    batch_labels = [sample_batches.get(col, 'Unknown') for col in sample_columns]

    qc_outliers_left = set()
    qc_outliers_right = set()
    if len(qc_indices) >= 3:
        for i, qc_name in enumerate(qc_columns):
            if i < len(outliers_orig) and outliers_orig[i]:
                qc_outliers_left.add(qc_name)
            if i < len(outliers_corr) and outliers_corr[i]:
                qc_outliers_right.add(qc_name)

    fig, _ = plot_pca_comparison_qc_style(
        scores_original,
        scores_corrected,
        var_original,
        var_corrected,
        sample_columns,
        sample_types,
        batch_labels=batch_labels,
        grouping=('batch' if grouping == 'batch' else 'sample_type'),
        suptitle=suptitle,
        left_title='Before Correction',
        right_title='After Correction',
        left_threshold_text=(f'Hotelling T² Threshold: {t2_threshold_orig:.2f}' if len(qc_indices) >= 3 else None),
        right_threshold_text=(f'Hotelling T² Threshold: {t2_threshold_corr:.2f}' if len(qc_indices) >= 3 else None),
        qc_outlier_names_left=qc_outliers_left,
        qc_outlier_names_right=qc_outliers_right,
    )

    return fig, outliers_orig, outliers_corr

def plot_permanova_comparison(permanova_before, permanova_after, perm_test_result):
    """
    繪製 PERMANOVA Pseudo-F 統計量比較圖
    """
    fig, ax = plt.subplots(figsize=(10, 7))
    
    # 數據準備
    categories = ['Before\nCorrection', 'After\nCorrection']
    f_values = [permanova_before['pseudo_f'], permanova_after['pseudo_f']]
    colors = ['#DC143C', '#4169E1']
    
    # 繪製柱狀圖
    bars = ax.bar(categories, f_values, color=colors, alpha=0.7, 
                   edgecolor='black', linewidth=2, width=0.5)
    
    # 添加數值標註
    for i, (bar, f_val, pval) in enumerate(zip(bars, f_values, 
                                                 [permanova_before['p_value'], 
                                                  permanova_after['p_value']])):
        height = bar.get_height()
        
        # 顯著性標記
        if pval < 0.001:
            sig_mark = '***'
        elif pval < 0.01:
            sig_mark = '**'
        elif pval < 0.05:
            sig_mark = '*'
        else:
            sig_mark = 'n.s.'
        
        ax.text(bar.get_x() + bar.get_width()/2., height,
                f'{f_val:.3f}\n{sig_mark}',
                ha='center', va='bottom', fontsize=14, fontweight='bold')
    
    # 設定軸標籤
    ax.set_ylabel('PERMANOVA Pseudo-F Statistic', fontsize=14, fontweight='bold')
    ax.set_title('PERMANOVA Batch Effect Comparison', fontsize=16, fontweight='bold', pad=20)
    
    # 添加網格
    ax.grid(True, alpha=0.3, linestyle='--', axis='y')
    ax.set_axisbelow(True)
    
    # 添加統計摘要框
    r2_before = permanova_before['r_squared']
    r2_after = permanova_after['r_squared']
    r2_reduction = (r2_before - r2_after) / r2_before * 100 if r2_before > 0 else 0
    
    f_reduction = (f_values[0] - f_values[1]) / f_values[0] * 100 if f_values[0] > 0 else 0
    
    stats_text = f"Statistical Summary:\n"
    stats_text += f"  Pseudo-F reduction: {f_reduction:.1f}%\n"
    stats_text += f"  R² (before): {r2_before*100:.1f}%\n"
    stats_text += f"  R² (after): {r2_after*100:.1f}%\n"
    stats_text += f"  R² reduction: {r2_reduction:.1f}%\n"
    stats_text += f"\n"
    stats_text += f"Paired Permutation Test:\n"
    stats_text += f"  Observed improvement: {perm_test_result['observed_improvement']:.4f}\n"
    stats_text += f"  p-value: {perm_test_result['p_value']:.4f}"
    
    if perm_test_result['p_value'] < 0.05:
        stats_text += f" *"
    
    ax.text(0.98, 0.97, stats_text, transform=ax.transAxes,
            fontsize=10, verticalalignment='top', horizontalalignment='right',
            bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.7),
            family='monospace')
    
    plt.tight_layout()
    
    return fig

def plot_cohens_d_forest(cohens_d_before, cohens_d_after):
    """
    繪製 Cohen's d 森林圖（批次對效應量比較）
    
    展示校正前後各批次對之間的 Cohen's d 變化
    """
    # 提取批次對資訊
    pairwise_before = cohens_d_before['pairwise_cohens_d']
    pairwise_after = cohens_d_after['pairwise_cohens_d']
    
    if len(pairwise_before) == 0 or len(pairwise_after) == 0:
        print("   ⚠ 批次對數據不足，無法繪製森林圖")
        return None
    
    # 準備數據
    pairs = list(pairwise_before.keys())
    d_before = [pairwise_before[pair] for pair in pairs]
    d_after = [pairwise_after[pair] for pair in pairs]
    improvements = [d_before[i] - d_after[i] for i in range(len(pairs))]
    
    # 按改善程度排序
    sorted_indices = np.argsort(improvements)[::-1]
    pairs_sorted = [pairs[i] for i in sorted_indices]
    d_before_sorted = [d_before[i] for i in sorted_indices]
    d_after_sorted = [d_after[i] for i in sorted_indices]
    improvements_sorted = [improvements[i] for i in sorted_indices]
    
    # 創建圖表
    fig, ax = plt.subplots(figsize=(10, max(6, len(pairs) * 0.6)))
    
    y_positions = np.arange(len(pairs_sorted))
    
    # 繪製校正前（紅色）
    ax.scatter(d_before_sorted, y_positions, color='#DC143C', s=120, 
               marker='D', label='Before Correction', zorder=3, alpha=0.8)
    
    # 繪製校正後（藍色）
    ax.scatter(d_after_sorted, y_positions, color='#4169E1', s=120, 
               marker='o', label='After Correction', zorder=3, alpha=0.8)
    
    # 繪製連接線（顯示改善方向）
    for i, (before, after) in enumerate(zip(d_before_sorted, d_after_sorted)):
        ax.plot([before, after], [i, i], 'k-', linewidth=1.5, alpha=0.5, zorder=2)
        
        # 在連接線旁邊標註改善值
        improvement = before - after
        mid_x = (before + after) / 2
        
        # 調整文字位置避免重疊
        if improvement > 0:
            ax.text(mid_x, i + 0.25, f'↓ {improvement:.3f}', 
                   fontsize=8, ha='center', color='green', fontweight='bold')
        else:
            ax.text(mid_x, i + 0.25, f'↑ {abs(improvement):.3f}', 
                   fontsize=8, ha='center', color='red', fontweight='bold')
    
    # 設定 y 軸標籤
    ax.set_yticks(y_positions)
    ax.set_yticklabels([pair.replace('Batch', 'B').replace('_vs_', ' vs ') 
                         for pair in pairs_sorted], fontsize=10)
    
    # 設定 x 軸
    ax.set_xlabel("Cohen's d (Effect Size)", fontsize=12, fontweight='bold')
    ax.set_title("Forest Plot: Batch-wise Cohen's d Comparison\n(Before vs After Correction)", 
                 fontsize=14, fontweight='bold', pad=20)
    
    # 添加 Cohen's d 解釋線
    ax.axvline(x=0.2, color='gray', linestyle='--', linewidth=1, alpha=0.5, label='Small (d=0.2)')
    ax.axvline(x=0.5, color='gray', linestyle='--', linewidth=1, alpha=0.5, label='Medium (d=0.5)')
    ax.axvline(x=0.8, color='gray', linestyle='--', linewidth=1, alpha=0.5, label='Large (d=0.8)')
    
    # 網格
    ax.grid(True, alpha=0.3, linestyle='--', axis='x')
    ax.set_axisbelow(True)
    
    # 圖例
    ax.legend(loc='upper right', fontsize=10, frameon=True, fancybox=True, shadow=True)
    
    # 統計摘要框
    avg_before = np.mean(d_before_sorted)
    avg_after = np.mean(d_after_sorted)
    avg_improvement = avg_before - avg_after
    improvement_pct = (avg_improvement / avg_before * 100) if avg_before > 0 else 0
    
    stats_text = f"Summary Statistics:\n"
    stats_text += f"  Avg d (before): {avg_before:.3f}\n"
    stats_text += f"  Avg d (after): {avg_after:.3f}\n"
    stats_text += f"  Avg improvement: {avg_improvement:.3f}\n"
    stats_text += f"  Reduction: {improvement_pct:.1f}%\n"
    stats_text += f"\n"
    stats_text += f"Batch pairs: {len(pairs_sorted)}"
    
    ax.text(0.02, 0.98, stats_text, transform=ax.transAxes,
            fontsize=9, verticalalignment='top', horizontalalignment='left',
            bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.7),
            family='monospace')
    
    plt.tight_layout()
    
    return fig

def plot_batch_residuals(data_before, data_after, batch_labels):
    """
    繪製批次效應校正的殘差圖
    
    用途：檢查校正後是否還有系統性的批次偏差
    """
    # 計算每個批次的中心
    batch_labels_array = np.array(batch_labels)
    unique_batches = sorted(list(set(batch_labels_array)))
    
    # 計算全局均值
    global_mean_before = np.mean(data_before, axis=0)
    global_mean_after = np.mean(data_after, axis=0)
    
    # 創建子圖
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))
    
    colors = COLORBLIND_COLORS[:len(unique_batches)]
    batch_color_map = {batch: colors[i] for i, batch in enumerate(unique_batches)}
    
    # ========== 左圖：校正前 ==========
    for batch in unique_batches:
        batch_mask = batch_labels_array == batch
        batch_data = data_before[batch_mask]
        
        # 計算批次均值與全局均值的偏差
        batch_mean = np.mean(batch_data, axis=0)
        residuals = batch_mean - global_mean_before
        
        # 繪製殘差
        ax1.scatter(np.arange(len(residuals)), residuals, 
                   color=batch_color_map[batch], alpha=0.6, s=30,
                   label=f'Batch {batch}')
    
    ax1.axhline(y=0, color='red', linestyle='--', linewidth=2, label='Zero Line')
    ax1.set_xlabel('Feature Index', fontsize=12, fontweight='bold')
    ax1.set_ylabel('Residual (Batch Mean - Global Mean)', fontsize=12, fontweight='bold')
    ax1.set_title('Before Correction\n(Systematic batch bias visible)', 
                  fontsize=14, fontweight='bold', pad=15)
    ax1.legend(loc='best', fontsize=9)
    ax1.grid(True, alpha=0.3)
    
    # ========== 右圖：校正後 ==========
    for batch in unique_batches:
        batch_mask = batch_labels_array == batch
        batch_data = data_after[batch_mask]
        
        # 計算批次均值與全局均值的偏差
        batch_mean = np.mean(batch_data, axis=0)
        residuals = batch_mean - global_mean_after
        
        # 繪製殘差
        ax2.scatter(np.arange(len(residuals)), residuals, 
                   color=batch_color_map[batch], alpha=0.6, s=30,
                   label=f'Batch {batch}')
    
    ax2.axhline(y=0, color='red', linestyle='--', linewidth=2, label='Zero Line')
    ax2.set_xlabel('Feature Index', fontsize=12, fontweight='bold')
    ax2.set_ylabel('Residual (Batch Mean - Global Mean)', fontsize=12, fontweight='bold')
    ax2.set_title('After Correction\n(Random scatter around zero = good)', 
                  fontsize=14, fontweight='bold', pad=15)
    ax2.legend(loc='best', fontsize=9)
    ax2.grid(True, alpha=0.3)
    
    # 統一 y 軸範圍
    all_residuals_before = []
    all_residuals_after = []
    
    for batch in unique_batches:
        batch_mask = batch_labels_array == batch
        
        batch_data_before = data_before[batch_mask]
        batch_mean_before = np.mean(batch_data_before, axis=0)
        all_residuals_before.extend(batch_mean_before - global_mean_before)
        
        batch_data_after = data_after[batch_mask]
        batch_mean_after = np.mean(batch_data_after, axis=0)
        all_residuals_after.extend(batch_mean_after - global_mean_after)
    
    y_max = max(np.abs(all_residuals_before + all_residuals_after))
    ax1.set_ylim(-y_max * 1.1, y_max * 1.1)
    ax2.set_ylim(-y_max * 1.1, y_max * 1.1)
    
    plt.suptitle('Batch Effect Residual Analysis', fontsize=16, fontweight='bold', y=0.98)
    plt.tight_layout(rect=[0, 0, 1, 0.96])
    
    return fig

def plot_permutation_null_distribution(perm_test_result, metric_name='PERMANOVA F'):
    """
    繪製配對排列檢定的 Null Distribution
    """
    null_dist = perm_test_result['null_distribution']
    observed = perm_test_result['observed_improvement']
    p_value = perm_test_result['p_value']
    
    if len(null_dist) == 0:
        print("   ⚠ Null distribution 為空，無法繪製分佈圖")
        return None
    
    fig, ax = plt.subplots(figsize=(12, 7))
    
    # 繪製直方圖
    ax.hist(null_dist, bins=50, color='lightblue', edgecolor='black', 
            alpha=0.7, label='Null Distribution')
    
    # 繪製觀察值
    ax.axvline(observed, color='red', linestyle='--', linewidth=3,
               label=f'Observed Improvement = {observed:.4f}')
    
    # 標題和標籤
    ax.set_xlabel(f'{metric_name} Improvement (Before - After)', 
                  fontsize=12, fontweight='bold')
    ax.set_ylabel('Frequency', fontsize=12, fontweight='bold')
    ax.set_title(f'Paired Permutation Test: Null Distribution of {metric_name} Improvement',
                 fontsize=14, fontweight='bold', pad=15)
    
    # 圖例
    ax.legend(fontsize=10, loc='upper left', frameon=True, fancybox=True, shadow=True)
    
    # 網格
    ax.grid(True, alpha=0.3, linestyle='--')
    
    # 統計摘要
    percentile_95 = np.percentile(null_dist, 95)
    percentile_99 = np.percentile(null_dist, 99)
    
    # 🆕 計算精確的 p-value（不使用 < 符號）
    n_permutations = len(null_dist)
    n_extreme = np.sum(null_dist >= observed)
    
    stats_text = f"Null Distribution Statistics:\n"
    stats_text += f"  Mean: {np.mean(null_dist):.4f}\n"
    stats_text += f"  Std: {np.std(null_dist):.4f}\n"
    stats_text += f"  95th percentile: {percentile_95:.4f}\n"
    stats_text += f"  99th percentile: {percentile_99:.4f}\n"
    stats_text += f"\n"
    stats_text += f"Permutation Test:\n"
    stats_text += f"  Observed: {observed:.4f}\n"
    stats_text += f"  Extreme values: {n_extreme}/{n_permutations}\n"
    stats_text += f"  p-value: {p_value:.6f}"  # 🔧 顯示完整 p-value
    
    if p_value < 0.001:
        stats_text += f" ***"
    elif p_value < 0.01:
        stats_text += f" **"
    elif p_value < 0.05:
        stats_text += f" *"
    
    stats_text += f"\n\n"
    stats_text += f"Permutations: {n_permutations:,}"  # 🔧 千分位格式
    
    ax.text(0.98, 0.97, stats_text, transform=ax.transAxes,
            fontsize=9, verticalalignment='top', horizontalalignment='right',
            bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.6),
            family='monospace')
    
    plt.tight_layout()
    
    return fig

from ms_core.utils.excel_format import copy_sheet_with_style  # noqa: E302

def color_result_cells(ws):
    """
    對 Batch_effect_result 工作表進行條件格式化
    """
    # 定義顏色
    blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')      # 淺藍色 - PERMANOVA
    yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')    # 淺黃色 - Silhouette
    pink_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')      # 淺粉色 - Cohen's d
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')     # 淺綠色 - QC CV%
    
    # 取得欄位標題
    header = [cell.value for cell in ws[1]]
    col_map = {name: idx+1 for idx, name in enumerate(header) if name is not None}
    
    # 定義欄位分組
    permanova_cols = ['PERMANOVA_F_overall', 'PERMANOVA_F_after_overall', 
                      'PERMANOVA_R2_overall', 'PERMANOVA_R2_after_overall']
    silhouette_cols = ['Silhouette_overall_before', 'Silhouette_overall_after']
    cohens_d_cols = ['Cohens_d_before', 'Cohens_d_after', 'Cohens_d_improvement']
    qc_cv_cols = ['QC_CV%_before', 'QC_CV%_after', 'QC_CV%_improvement']
    
    # 對每個儲存格進行上色
    for row in range(2, ws.max_row + 1):
        for col_name, col_idx in col_map.items():
            cell = ws.cell(row=row, column=col_idx)
            
            if col_name in permanova_cols:
                cell.fill = blue_fill
            elif col_name in silhouette_cols:
                cell.fill = yellow_fill
            elif col_name in cohens_d_cols:
                cell.fill = pink_fill
            elif col_name in qc_cv_cols:
                cell.fill = green_fill
    
    print(f"   ✓ 已對 Batch_effect_result 工作表進行條件格式化")

def create_statistical_summary_sheet(wb, permanova_before, permanova_after,
                                     permdisp_before, permdisp_after,
                                     perm_test_permanova, perm_test_silhouette,
                                     silhouette_before, silhouette_after,
                                     cohens_d_before, cohens_d_after,
                                     qc_cv_before, qc_cv_after):
    """
    創建 Batch_Effect_summary 工作表
    
    Parameters:
    -----------
    wb : openpyxl.Workbook
        Excel 工作簿對象
    permanova_before : dict
        校正前 PERMANOVA 結果
    permanova_after : dict
        校正後 PERMANOVA 結果
    permdisp_before : dict
        校正前 PERMDISP 結果
    permdisp_after : dict
        校正後 PERMDISP 結果
    perm_test_permanova : dict
        PERMANOVA 配對排列檢定結果
    perm_test_silhouette : dict
        Silhouette 配對排列檢定結果
    silhouette_before : float
        校正前 Silhouette 分數
    silhouette_after : float
        校正後 Silhouette 分數
    cohens_d_before : dict
        校正前 Cohen's d 結果
    cohens_d_after : dict
        校正後 Cohen's d 結果
    qc_cv_before : dict
        校正前 QC CV% 結果
    qc_cv_after : dict
        校正後 QC CV% 結果
    """
    from openpyxl.styles import Font, PatternFill
    
    ws = wb.create_sheet(SUMMARY_SHEET_NAME, 0)
    
    # ========== 標題行 ==========
    headers = ['指標類別', '項目', '校正前', '校正後', '改善', '改善%', 'p-value', '顯著性']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        from openpyxl.styles import Alignment
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # ========== 數據行 ==========
    data_rows = []
    
    # ===== 1. PERMANOVA =====
    _perm_available = _has_valid_permanova(permanova_before) and _has_valid_permanova(permanova_after)
    if _perm_available:
        f_improvement = permanova_before['pseudo_f'] - permanova_after['pseudo_f']
        f_improvement_pct = (f_improvement / permanova_before['pseudo_f'] * 100) if permanova_before['pseudo_f'] > 0 else 0

        r2_improvement = permanova_before['r_squared'] - permanova_after['r_squared']
        r2_improvement_pct = (r2_improvement / permanova_before['r_squared'] * 100) if permanova_before['r_squared'] > 0 else 0

        eta2_improvement = permanova_before.get('eta_squared', 0) - permanova_after.get('eta_squared', 0)
        eta2_improvement_pct = (eta2_improvement / permanova_before.get('eta_squared', 1) * 100) if permanova_before.get('eta_squared', 0) > 0 else 0

        _pt_pval = perm_test_permanova['p_value'] if perm_test_permanova and not np.isnan(perm_test_permanova.get('p_value', np.nan)) else np.nan
        _pt_pval_str = f"{_pt_pval:.6f}" if not np.isnan(_pt_pval) else 'N/A'
        _pt_sig = get_significance_mark(_pt_pval) if not np.isnan(_pt_pval) else 'N/A'

        data_rows.append(['PERMANOVA', 'Pseudo-F',
                         f"{permanova_before['pseudo_f']:.4f}",
                         f"{permanova_after['pseudo_f']:.4f}",
                         f"{f_improvement:.4f}",
                         f"{f_improvement_pct:.1f}%",
                         _pt_pval_str,
                         _pt_sig])

        data_rows.append(['', 'R² (變異解釋)',
                         f"{permanova_before['r_squared']:.4f}",
                         f"{permanova_after['r_squared']:.4f}",
                         f"{r2_improvement:.4f}",
                         f"{r2_improvement_pct:.1f}%",
                         '-',
                         '-'])

        data_rows.append(['', 'η² (效應量)',
                         f"{permanova_before.get('eta_squared', 0):.4f}",
                         f"{permanova_after.get('eta_squared', 0):.4f}",
                         f"{eta2_improvement:.4f}",
                         f"{eta2_improvement_pct:.1f}%",
                         '-',
                         '-'])

        data_rows.append(['', 'p-value',
                         f"{permanova_before['p_value']:.4f}",
                         f"{permanova_after['p_value']:.4f}",
                         '-',
                         '-',
                         '-',
                         f"{get_significance_mark(permanova_before['p_value'])} → {get_significance_mark(permanova_after['p_value'])}"])

        data_rows.append(['', '距離度量',
                         permanova_before.get('distance_metric', 'manhattan').capitalize(),
                         permanova_after.get('distance_metric', 'manhattan').capitalize(),
                         '-',
                         '-',
                         '-',
                         '-'])

        data_rows.append(['', '排列次數',
                         f"{permanova_before.get('permutations', 999)}",
                         f"{permanova_after.get('permutations', 999)}",
                         '-',
                         '-',
                         '-',
                         '-'])
    else:
        data_rows.append(['PERMANOVA', '狀態',
                         'N/A (scikit-bio 未安裝)',
                         'N/A (scikit-bio 未安裝)',
                         '-', '-', '-', '-'])
    
    # ===== 2. PERMDISP =====
    if not np.isnan(permdisp_before['f_statistic']) and not np.isnan(permdisp_after['f_statistic']):
        data_rows.append(['PERMDISP', 'F-statistic', 
                         f"{permdisp_before['f_statistic']:.4f}", 
                         f"{permdisp_after['f_statistic']:.4f}",
                         f"{permdisp_before['f_statistic'] - permdisp_after['f_statistic']:.4f}",
                         '-',
                         '-',
                         '-'])
        
        data_rows.append(['', 'p-value', 
                         f"{permdisp_before['p_value']:.4f}", 
                         f"{permdisp_after['p_value']:.4f}",
                         '-',
                         '-',
                         '-',
                         f"{get_significance_mark(permdisp_before['p_value'])} → {get_significance_mark(permdisp_after['p_value'])}"])
        
        data_rows.append(['', '解讀', 
                         'p<0.05 則離散度異質' if permdisp_before['p_value'] < 0.05 else '離散度同質 (良好)',
                         'p<0.05 則離散度異質' if permdisp_after['p_value'] < 0.05 else '離散度同質 (良好)',
                         '-',
                         '-',
                         '-',
                         '-'])
    else:
        data_rows.append(['PERMDISP', '狀態', 
                         'N/A (無法計算)', 
                         'N/A (無法計算)',
                         '-',
                         '-',
                         '-',
                         '-'])
    
    # ===== 3. 配對排列檢定 =====
    if perm_test_permanova and not np.isnan(perm_test_permanova.get('p_value', np.nan)):
        data_rows.append(['配對排列檢定', 'PERMANOVA F',
                         '-',
                         '-',
                         f"{perm_test_permanova['observed_improvement']:.4f}",
                         '-',
                         f"{perm_test_permanova['p_value']:.6f}",
                         get_significance_mark(perm_test_permanova['p_value'])])

        data_rows.append(['', '排列次數',
                         '-',
                         '-',
                         f"{len(perm_test_permanova.get('null_distribution', [])):,}",
                         '-',
                         '-',
                         '-'])
    else:
        data_rows.append(['配對排列檢定', 'PERMANOVA F',
                         '-', '-', 'N/A (跳過)', '-', 'N/A', 'N/A'])

    if perm_test_silhouette and not np.isnan(perm_test_silhouette.get('p_value', np.nan)):
        data_rows.append(['', 'Silhouette',
                         '-',
                         '-',
                         f"{perm_test_silhouette['observed_improvement']:.4f}",
                         '-',
                         f"{perm_test_silhouette['p_value']:.6f}",
                         get_significance_mark(perm_test_silhouette['p_value'])])
    else:
        data_rows.append(['', 'Silhouette',
                         '-', '-', 'N/A', '-', 'N/A', 'N/A'])
    
    # ===== 4. Silhouette =====
    sil_improvement = silhouette_before - silhouette_after
    sil_improvement_pct = (sil_improvement / abs(silhouette_before) * 100) if silhouette_before != 0 else 0
    
    data_rows.append(['Silhouette', 'Overall Score', 
                     f"{silhouette_before:.4f}", 
                     f"{silhouette_after:.4f}",
                     f"{sil_improvement:.4f}",
                     f"{sil_improvement_pct:.1f}%",
                     '-',
                     '-'])
    
    data_rows.append(['', '註解', 
                     '分數越低 = 批次效應越弱', 
                     '分數越低 = 批次效應越弱',
                     '改善 = 分數下降',
                     '-',
                     '-',
                     '-'])
    
    # ===== 5. Cohen's d =====
    d_improvement = cohens_d_before['overall_cohens_d'] - cohens_d_after['overall_cohens_d']
    d_improvement_pct = (d_improvement / cohens_d_before['overall_cohens_d'] * 100) if cohens_d_before['overall_cohens_d'] > 0 else 0
    
    data_rows.append(['Cohen\'s d', 'Average', 
                     f"{cohens_d_before['overall_cohens_d']:.4f}", 
                     f"{cohens_d_after['overall_cohens_d']:.4f}",
                     f"{d_improvement:.4f}",
                     f"{d_improvement_pct:.1f}%",
                     '-',
                     '-'])
    
    # Cohen's d 批次對詳細資訊
    pairwise_before = cohens_d_before['pairwise_cohens_d']
    pairwise_after = cohens_d_after['pairwise_cohens_d']
    
    if len(pairwise_before) > 0:
        for pair_name in sorted(pairwise_before.keys()):
            d_before = pairwise_before[pair_name]
            d_after = pairwise_after.get(pair_name, 0)
            d_imp = d_before - d_after
            
            data_rows.append(['', pair_name.replace('_', ' '), 
                             f"{d_before:.4f}", 
                             f"{d_after:.4f}",
                             f"{d_imp:.4f}",
                             '-',
                             '-',
                             '-'])
    
    # ===== 6. QC CV% =====
    if not np.isnan(qc_cv_before['median_cv']) and not np.isnan(qc_cv_after['median_cv']):
        cv_improvement = qc_cv_before['median_cv'] - qc_cv_after['median_cv']
        cv_improvement_pct = (cv_improvement / qc_cv_before['median_cv'] * 100) if qc_cv_before['median_cv'] > 0 else 0
        
        data_rows.append(['QC CV%', 'Median', 
                         f"{qc_cv_before['median_cv']:.2f}%", 
                         f"{qc_cv_after['median_cv']:.2f}%",
                         f"{cv_improvement:.2f}%",
                         f"{cv_improvement_pct:.1f}%",
                         '-',
                         '-'])
        
        cv20_improvement = qc_cv_after['cv_below_20'] - qc_cv_before['cv_below_20']
        
        data_rows.append(['', 'CV<20% 比例', 
                         f"{qc_cv_before['cv_below_20']:.1f}%", 
                         f"{qc_cv_after['cv_below_20']:.1f}%",
                         f"{cv20_improvement:+.1f}%",
                         '-',
                         '-',
                         '-'])
        
        # QC 品質評級
        qc_grade_before = "優良" if qc_cv_before['median_cv'] < CV_QUALITY_THRESHOLDS['excellent'] else ("可接受" if qc_cv_before['median_cv'] < CV_QUALITY_THRESHOLDS['acceptable'] else "較差")
        qc_grade_after = "優良" if qc_cv_after['median_cv'] < CV_QUALITY_THRESHOLDS['excellent'] else ("可接受" if qc_cv_after['median_cv'] < CV_QUALITY_THRESHOLDS['acceptable'] else "較差")
        
        data_rows.append(['', '品質評級', 
                         qc_grade_before,
                         qc_grade_after,
                         '-',
                         '-',
                         '-',
                         '-'])
    else:
        data_rows.append(['QC CV%', '狀態', 
                         'N/A (QC樣本不足)', 
                         'N/A (QC樣本不足)',
                         '-',
                         '-',
                         '-',
                         '-'])
    
    # ========== 寫入數據 ==========
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 為不同類別添加背景色
            if col_idx == 1 and value in ['PERMANOVA', 'PERMDISP', '配對排列檢定', 'Silhouette', "Cohen's d", 'QC CV%']:
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    # ========== 調整列寬 ==========
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    
    print(f"   ✓ 已創建 {SUMMARY_SHEET_NAME} 工作表")

def get_significance_mark(p_value):
    """根據 p-value 返回顯著性標記"""
    if p_value < 0.001:
        return '***'
    elif p_value < 0.01:
        return '**'
    elif p_value < 0.05:
        return '*'
    else:
        return 'n.s.'

def save_results_to_excel(input_file, output_file, data, sample_info,
                          corrected_df, data_sheet_name,
                          permanova_before, permanova_after,
                          permdisp_before, permdisp_after,
                          perm_test_permanova, perm_test_silhouette,
                          silhouette_before, silhouette_after,
                          cohens_d_before, cohens_d_after,
                          qc_cv_before, qc_cv_after,
                          sample_type_row=None):
    """
    保存結果到 Excel，複製所有原始工作表並新增結果
    """
    print("\n💾 保存結果到 Excel...")

    # 回插 Sample_Type 資訊行（若有）
    if sample_type_row is not None:
        from ms_core.utils.data_helpers import insert_sample_type_row
        corrected_df = insert_sample_type_row(corrected_df, sample_type_row)

    # 首先創建基本結構
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 寫入主要結果
        corrected_df.to_excel(writer, sheet_name=SHEET_NAMES['batch_effect'], index=False)
        sample_info.to_excel(writer, sheet_name=SHEET_NAMES['sample_info'], index=False)
    
    # 加載輸出文件
    output_wb = load_workbook(output_file)
    
    # 創建統計摘要表
    create_statistical_summary_sheet(output_wb, 
                                     permanova_before, permanova_after,
                                     permdisp_before, permdisp_after,  # 🆕 傳遞
                                     perm_test_permanova, perm_test_silhouette,
                                     silhouette_before, silhouette_after,
                                     cohens_d_before, cohens_d_after,
                                     qc_cv_before, qc_cv_after)
    
    # 對 Batch_effect_result 進行條件格式化
    if SHEET_NAMES['batch_effect'] in output_wb.sheetnames:
        ws = output_wb[SHEET_NAMES['batch_effect']]
        color_result_cells(ws)
    
    # 複製原始文件的其他工作表
    try:
        input_wb = load_workbook(input_file, data_only=True)
        
        for sheet_name in input_wb.sheetnames:
            # 跳過已經存在的工作表
            if sheet_name in [SHEET_NAMES['batch_effect'], SHEET_NAMES['sample_info'], SUMMARY_SHEET_NAME]:
                continue
            
            # 對於 QC LOWESS result 保留格式
            if sheet_name == SHEET_NAMES['qc_lowess']:
                if sheet_name in output_wb.sheetnames:
                    std = output_wb[sheet_name]
                    output_wb.remove(std)
                new_sheet = output_wb.create_sheet(sheet_name)
                copy_sheet_with_style(input_wb[sheet_name], new_sheet)
            else:
                # 其他工作表直接複製
                if sheet_name not in output_wb.sheetnames:
                    new_sheet = output_wb.create_sheet(sheet_name)
                    copy_sheet_with_style(input_wb[sheet_name], new_sheet)
        
        print(f"   ✓ 已複製所有原始工作表")
    except Exception as e:
        print(f"   ⚠ 複製原始工作表時發生錯誤: {e}")
    
    # 保存文件
    output_wb.save(output_file)
    print(f"   ✓ Excel 檔案已保存: {os.path.basename(output_file)}")

def main(input_file=None):
    """
    主函數 - 批次效應校正與評估（無母數版本）

    包含完整的防呆措施
    """
    print("="*70)
    print("  批次效應校正程式 v5.1 (完善防呆版本)")
    print("  - PERMANOVA 批次效應檢驗 (Manhattan 距離)")
    print("  - 配對排列檢定 (Paired Permutation Test)")
    print("  - Cohen's d 效果量")
    print("  - QC CV% 技術重現性")
    print("  - Silhouette Coefficient (次要指標)")
    print("="*70)

    # ========== 1. 檔案選擇 ==========
    if input_file is None:
        raise ValueError("input_file is required; GUI must provide the file path.")

    if input_file is None:
        print("\n請選擇要處理的Excel檔案...")
        input_file = select_file()

        if not input_file:
            print("❌ 未選擇檔案，程式結束。")
            return None

    # ===== 防呆34: 文件存在性檢查 =====
    if not os.path.exists(input_file):
        print(f"❌ 錯誤：找不到檔案 '{input_file}'")
        raise FileNotFoundError(f"找不到檔案: {input_file}")

    # ===== 防呆35: 文件讀取權限檢查 =====
    if not os.access(input_file, os.R_OK):
        print(f"❌ 錯誤：無法讀取檔案（權限不足）")
        raise PermissionError(f"無法讀取檔案: {input_file}")

    print(f"\n✓ 選擇的檔案: {os.path.basename(input_file)}")
    print(f"  路徑: {input_file}")

    # ========== 2. 設定輸出路徑 ==========
    output_dir = get_output_root()

    # ===== 防呆36: 輸出目錄創建與權限檢查 =====
    try:
        if not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
            print(f"✓ 已創建 output 資料夾: {output_dir}")
        else:
            print(f"✓ output 資料夾已存在: {output_dir}")
    except Exception as e:
        print(f"❌ 錯誤：無法創建 output 資料夾")
        print(f"詳細錯誤: {e}")
        raise Exception(f"無法創建輸出目錄: {e}")

    # ===== 防呆37: 輸出目錄寫入權限檢查 =====
    try:
        test_file = os.path.join(output_dir, '.write_test')
        with open(test_file, 'w') as f:
            f.write('test')
        os.remove(test_file)
        print(f"✓ output 資料夾寫入權限正常")
    except Exception as e:
        print(f"❌ 錯誤：無法寫入 output 目錄")
        print(f"  請檢查目錄權限: {output_dir}")
        print(f"  詳細錯誤: {e}")
        raise Exception(f"輸出目錄無寫入權限: {output_dir}")

    timestamp = datetime.datetime.now().strftime(DATETIME_FORMAT_FULL)
    output_file = build_output_path("Combat_corrected", timestamp=timestamp)

    # ===== 防呆38: 輸出文件檢查 =====
    if os.path.exists(output_file):
        print(f"⚠️ 警告：輸出檔案已存在，將被覆蓋")
        print(f"  {output_file}")

    plots_dir = build_plots_dir(PLOT_FOLDER_NAME)

    # ===== 防呆39: 圖表目錄創建 =====
    try:
        if not os.path.exists(plots_dir):
            os.makedirs(plots_dir, exist_ok=True)
            print(f"✓ 已創建圖表資料夾: {plots_dir}")
        else:
            print(f"✓ 圖表資料夾已存在: {plots_dir}")
    except Exception as e:
        print(f"❌ 錯誤：無法創建圖表資料夾")
        print(f"詳細錯誤: {e}")
        raise Exception(f"無法創建圖表目錄: {e}")

    run_plot_dir = build_plots_dir(
        PLOT_FOLDER_NAME,
        timestamp=timestamp,
        session_prefix="Batch_Effect"
    )
    print(f"✓ 本次圖表輸出子資料夾: {run_plot_dir}")
    
    try:
        # ========== 3. 讀取數據 ==========
        print("\n正在讀取數據...")
        data, sample_info, sheet_name, sample_type_row = read_excel_data(input_file)
        print(f"✓ 已讀取工作表: {sheet_name}")
        
        # 二次檢查 Batch 資訊
        if sample_info.shape[1] >= 4:
            column_d = sample_info.columns[3]
            if 'batch' in str(column_d).lower():
                print(f"✓ 使用 column D ('{column_d}') 作為 Batch 資訊")
                sample_info['Batch'] = sample_info[column_d]
        
        # ========== 4. 準備數據 ==========
        print("\n準備數據格式...")
        data_matrix, batch_info, sample_columns, feature_ids = prepare_data_for_combat(data, sample_info)
        print(f"✓ 找到 {len(sample_columns)} 個樣本，{len(feature_ids)} 個代謝物")
        
        unique_batches = list(set(batch_info))
        print(f"✓ 批次信息: {unique_batches}")
        print(f"  批次數量: {len(unique_batches)}")
        
        if len(unique_batches) < 2:
            print("\n⚠️ 警告：只有一個批次，無需進行批次效應校正")
            return {
                'file_path': input_file,
                'metabolites': len(feature_ids),
                'samples': len(sample_columns),
                'batches': len(unique_batches),
                'output_path': input_file,
                'plots_dir': run_plot_dir,
                'skipped': True,
                'skip_reason': 'single_batch',
                'success': True,
            }
        
        # ========== 5. 校正前評估 ==========
        print("\n" + "="*70)
        print("📊 校正前批次效應評估")
        print("="*70)
        
        original_data_for_eval = data_matrix.T
        
        # 數據預處理（log + 標準化）
        original_data_log = np.log2(original_data_for_eval + 1)
        scaler_before = StandardScaler()
        original_scaled = scaler_before.fit_transform(original_data_log)
        
        # 5.1 PERMANOVA (optional - requires scikit-bio)
        permanova_before = None
        permdisp_before = None
        try:
            permanova_before = calculate_permanova(original_scaled, batch_info,
                                                distance_metric='manhattan',
                                                permutations=999)
            # 🆕 5.1b PERMDISP (檢驗離散度同質性)
            permdisp_before = calculate_permdisp(original_scaled, batch_info,
                                                distance_metric='manhattan',
                                                permutations=999)
        except (ImportError, Exception) as e:
            print(f"  ⚠️ PERMANOVA/PERMDISP 跳過: {e}")

        # 提供預設值，確保後續程式碼不會因 None 而崩潰
        _default_permanova = {'pseudo_f': np.nan, 'p_value': np.nan, 'r_squared': np.nan,
                              'eta_squared': np.nan, 'distance_metric': 'manhattan', 'permutations': 0}
        _default_permdisp = {'f_statistic': np.nan, 'p_value': np.nan}
        _default_perm_test = {'p_value': np.nan, 'observed_improvement': np.nan, 'null_distribution': []}
        if permanova_before is None:
            permanova_before = _default_permanova
        if permdisp_before is None:
            permdisp_before = _default_permdisp

        # 5.2 Silhouette
        silhouette_before = calculate_silhouette_overall(original_scaled, batch_info, 
                                                         distance_metric='manhattan')
        
        # 5.3 Cohen's d
        cohens_d_before = calculate_cohens_d_batch_effect(original_scaled, batch_info)
        
        # 5.4 QC CV%
        qc_cv_before = calculate_qc_cv(data_matrix, sample_info, sample_columns)
        
        # ========== 6. Combat 校正 ==========
        print("\n" + "="*70)
        print("⚙️ 執行 Combat 批次效應校正")
        print("="*70)

        # 🆕 檢測批次混淆
        confounding_result = check_batch_confounding(sample_info)

        # 如果檢測到高度混淆，完全跳過 ComBat（不輸出任何文件）
        if confounding_result.get('is_confounded', False):
            print("\n" + "="*70)
            print("WARNING: Batch-Sample Type Confounding Detected!")
            print("="*70)
            print("ComBat correction SKIPPED to protect biological signals.")
            print("")
            print("Next Step: Run Concentration Normalization directly.")
            print("           It will automatically use 'QC LOWESS result' sheet.")
            print("="*70)
            # Return a valid result dict so the pipeline/GUI can continue.
            # Use the original input file as output since no correction is applied.
            return {
                'file_path': input_file,
                'metabolites': len(feature_ids),
                'samples': len(sample_columns),
                'batches': len(unique_batches),
                'output_path': input_file,
                'plots_dir': run_plot_dir,
                'skipped': True,
                'skip_reason': 'batch_sample_type_confounded',
                'confounding': confounding_result,
                'success': False,
            }

        corrected_data = perform_combat_correction(data_matrix, batch_info)
        print("✓ 批次效應校正完成")
        
        if corrected_data.shape[1] != len(sample_columns):
            corrected_data = corrected_data.T
        
        # ========== 7. 校正後評估 ==========
        print("\n" + "="*70)
        print("📊 校正後批次效應評估")
        print("="*70)
        
        corrected_data_for_eval = corrected_data.T
        
        # 數據預處理
        corrected_data_log = np.log2(corrected_data_for_eval + 1)
        scaler_after = StandardScaler()
        corrected_scaled = scaler_after.fit_transform(corrected_data_log)
        
        # 7.1 PERMANOVA (optional - requires scikit-bio)
        permanova_after = None
        permdisp_after = None
        try:
            permanova_after = calculate_permanova(corrected_scaled, batch_info,
                                                distance_metric='manhattan',
                                                permutations=999)
            # 🆕 7.1b PERMDISP
            permdisp_after = calculate_permdisp(corrected_scaled, batch_info,
                                            distance_metric='manhattan',
                                            permutations=999)
        except (ImportError, Exception) as e:
            print(f"  ⚠️ PERMANOVA/PERMDISP 跳過: {e}")
        if permanova_after is None:
            permanova_after = _default_permanova
        if permdisp_after is None:
            permdisp_after = _default_permdisp

        # 7.2 Silhouette
        silhouette_after = calculate_silhouette_overall(corrected_scaled, batch_info, 
                                                        distance_metric='manhattan')
        
        # 7.3 Cohen's d
        cohens_d_after = calculate_cohens_d_batch_effect(corrected_scaled, batch_info)
        
        # 7.4 QC CV%
        qc_cv_after = calculate_qc_cv(corrected_data, sample_info, sample_columns)
        
        # ========== 8. 配對排列檢定 ==========
        print("\n" + "="*70)
        print("🎲 配對排列檢定 (顯著性檢驗)")
        print("="*70)

        # 8.1 PERMANOVA F-statistic (增加到 10,000 次)
        perm_test_permanova = None
        try:
            perm_test_permanova = paired_permutation_test(
                original_scaled, corrected_scaled, batch_info,
                metric='permanova', distance_metric='manhattan', n_permutations=10000
            )
        except (ImportError, Exception) as e:
            print(f"  ⚠️ PERMANOVA 排列檢定跳過: {e}")

        # 8.2 Silhouette Coefficient (保持 1000 次即可)
        perm_test_silhouette = None
        try:
            perm_test_silhouette = paired_permutation_test(
                original_scaled, corrected_scaled, batch_info,
                metric='silhouette', distance_metric='manhattan', n_permutations=1000
            )
        except (ImportError, Exception) as e:
            print(f"  ⚠️ Silhouette 排列檢定跳過: {e}")

        # ========== 9. 統計摘要 ==========
        print("\n" + "="*70)
        print("📈 批次效應校正前後對比統計摘要")
        print("="*70)

        _perm_ok = _has_valid_permanova(permanova_before) and _has_valid_permanova(permanova_after)
        print("\n1️⃣ PERMANOVA (主要指標):")
        if _perm_ok:
            print(f"   校正前: F={permanova_before['pseudo_f']:.4f}, p={permanova_before['p_value']:.4f}, R²={permanova_before['r_squared']*100:.1f}%")
            print(f"   校正後: F={permanova_after['pseudo_f']:.4f}, p={permanova_after['p_value']:.4f}, R²={permanova_after['r_squared']*100:.1f}%")

            f_reduction = (permanova_before['pseudo_f'] - permanova_after['pseudo_f']) / permanova_before['pseudo_f'] * 100
            r2_reduction = (permanova_before['r_squared'] - permanova_after['r_squared']) / permanova_before['r_squared'] * 100

            print(f"   改善: Pseudo-F 降低 {f_reduction:.1f}%, R² 降低 {r2_reduction:.1f}%")
        else:
            print("   ⚠️ 跳過（scikit-bio 未安裝）")
        if perm_test_permanova and not np.isnan(perm_test_permanova.get('p_value', np.nan)):
            print(f"   配對檢定 p-value: {perm_test_permanova['p_value']:.4f} {get_significance_mark(perm_test_permanova['p_value'])}")
        
        print("\n2️⃣ Silhouette Coefficient (次要指標):")
        print(f"   校正前: {silhouette_before:.4f}")
        print(f"   校正後: {silhouette_after:.4f}")
        print(f"   改善: {silhouette_before - silhouette_after:.4f}")
        if perm_test_silhouette and not np.isnan(perm_test_silhouette.get('p_value', np.nan)):
            print(f"   配對檢定 p-value: {perm_test_silhouette['p_value']:.4f} {get_significance_mark(perm_test_silhouette['p_value'])}")
        
        print("\n3️⃣ Cohen's d (效果量):")
        print(f"   校正前: {cohens_d_before['overall_cohens_d']:.4f}")
        print(f"   校正後: {cohens_d_after['overall_cohens_d']:.4f}")
        if cohens_d_before['overall_cohens_d'] > 0:
            d_reduction = (cohens_d_before['overall_cohens_d'] - cohens_d_after['overall_cohens_d']) / cohens_d_before['overall_cohens_d'] * 100
            print(f"   減少: {d_reduction:.1f}%")
        
        print("\n4️⃣ QC CV% (技術重現性):")
        if not np.isnan(qc_cv_before['median_cv']) and not np.isnan(qc_cv_after['median_cv']):
            print(f"   校正前: 中位數 {qc_cv_before['median_cv']:.2f}%, CV<20% = {qc_cv_before['cv_below_20']:.1f}%")
            print(f"   校正後: 中位數 {qc_cv_after['median_cv']:.2f}%, CV<20% = {qc_cv_after['cv_below_20']:.1f}%")
            cv_improvement = qc_cv_before['median_cv'] - qc_cv_after['median_cv']
            print(f"   改善: {cv_improvement:+.2f}%")
        else:
            print(f"   ⚠️ QC 樣本數不足，無法計算 CV%")
        
        # ========== 10. 生成視覺化圖表 ==========
        print("\n" + "="*70)
        print("🎨 生成視覺化圖表")
        print("="*70)

        # 圖 1: PERMANOVA 前後對比
        if permanova_before and permanova_after:
            print("\n生成圖 1: PERMANOVA 統計摘要...")
            fig1 = plot_permanova_comparison(permanova_before, permanova_after,
                                            perm_test_permanova)
            if fig1:
                fig1_file = run_plot_dir / generate_output_filename(
                    "Fig1_PERMANOVA_comparison", timestamp=timestamp, extension=".png"
                )
                fig1.savefig(fig1_file, dpi=300, bbox_inches='tight')
                plt.close(fig1)
                print(f"✓ 已儲存: {os.path.basename(fig1_file)}")
        else:
            print("\n⚠️ 跳過圖 1: PERMANOVA 未執行（scikit-bio 未安裝）")

        # 圖 2: PCA 前後對比（按 Batch）
        print("\n生成圖 2: PCA 前後對比（按 Batch 分組）...")
        outliers_orig_batch = np.array([])
        outliers_corr_batch = np.array([])
        result = create_comparison_pca_plot(
            original_data_for_eval, corrected_data_for_eval,
            sample_info, sample_columns, batch_info,
            title_suffix="", grouping="batch"
        )
        if result:
            fig2, outliers_orig_batch, outliers_corr_batch = result
            fig2_file = run_plot_dir / generate_output_filename(
                "Fig2_PCA_by_batch", timestamp=timestamp, extension=".png"
            )
            fig2.savefig(fig2_file, dpi=300, bbox_inches='tight')
            plt.close(fig2)
            print(f"✓ 已儲存: {os.path.basename(fig2_file)}")
        else:
            print("   ⚠ 樣本不足或繪圖失敗，跳過圖 2")

        # 圖 3: Permutation Test Null Distribution
        if perm_test_permanova:
            print("\n生成圖 3: Permutation Test 顯著性檢驗...")
            fig3 = plot_permutation_null_distribution(perm_test_permanova,
                                                    metric_name='PERMANOVA F')
        else:
            fig3 = None
            print("\n⚠️ 跳過圖 3: PERMANOVA 排列檢定未執行")
        if fig3:
            fig3_file = run_plot_dir / generate_output_filename(
                "Fig3_Permutation_Test", timestamp=timestamp, extension=".png"
            )
            fig3.savefig(fig3_file, dpi=300, bbox_inches='tight')
            plt.close(fig3)
            print(f"✓ 已儲存: {os.path.basename(fig3_file)}")

        # 圖 4: PCA 前後對比（按樣本分類）
        print("\n生成圖 4: PCA 前後對比（按樣本分類分組）...")
        result = create_comparison_pca_plot(
            original_data_for_eval, corrected_data_for_eval,
            sample_info, sample_columns, batch_info,
            title_suffix="", grouping="sample_type"
        )
        if result:
            fig4, outliers_orig_type, outliers_corr_type = result
            fig4_file = run_plot_dir / generate_output_filename(
                "Fig4_PCA_by_sample_type", timestamp=timestamp, extension=".png"
            )
            fig4.savefig(fig4_file, dpi=300, bbox_inches='tight')
            plt.close(fig4)
            print(f"✓ 已儲存: {os.path.basename(fig4_file)}")
        else:
            print("   ⚠ 樣本不足或繪圖失敗，跳過圖 4")

        # 🆕 圖 5: 森林圖（批次對效應量比較）
        print("\n生成圖 5: Cohen's d 森林圖...")
        fig5 = plot_cohens_d_forest(cohens_d_before, cohens_d_after)
        if fig5:
            fig5_file = run_plot_dir / generate_output_filename(
                "Fig5_Cohens_d_Forest", timestamp=timestamp, extension=".png"
            )
            fig5.savefig(fig5_file, dpi=300, bbox_inches='tight')
            plt.close(fig5)
            print(f"✓ 已儲存: {os.path.basename(fig5_file)}")

        # 🆕 圖 6: 殘差圖（可選，如果選擇方案 A）
        print("\n生成圖 6: 批次效應殘差分析...")
        fig6 = plot_batch_residuals(original_scaled, corrected_scaled, batch_info)
        if fig6:
            fig6_file = run_plot_dir / generate_output_filename(
                "Fig6_Residual_Analysis", timestamp=timestamp, extension=".png"
            )
            fig6.savefig(fig6_file, dpi=300, bbox_inches='tight')
            plt.close(fig6)
            print(f"✓ 已儲存: {os.path.basename(fig6_file)}")

        print(f"\n✓ 共生成 6 張圖表")
        
        
        # ========== 11. 準備 Excel 輸出 ==========
        print("\n準備輸出結果到 Excel...")
        corrected_df = pd.DataFrame(corrected_data, columns=sample_columns)
        corrected_df.insert(0, data.columns[0], feature_ids)

        # 🔧 修正：只添加「代謝物層級」的統計指標
        # 移除整體層級的 PERMANOVA 和 Silhouette（這些只存在於 Batch_Effect_summary）

        # Cohen's d（每個代謝物）
        corrected_df['Cohens_d_before'] = cohens_d_before['feature_cohens_d']
        corrected_df['Cohens_d_after'] = cohens_d_after['feature_cohens_d']
        corrected_df['Cohens_d_improvement'] = cohens_d_before['feature_cohens_d'] - cohens_d_after['feature_cohens_d']

        # QC CV%（每個代謝物）
        if len(qc_cv_before['qc_cv']) > 0 and len(qc_cv_after['qc_cv']) > 0:
            corrected_df['QC_CV%_before'] = qc_cv_before['qc_cv']
            corrected_df['QC_CV%_after'] = qc_cv_after['qc_cv']
            
            if len(qc_cv_before['qc_cv']) == len(qc_cv_after['qc_cv']):
                corrected_df['QC_CV%_improvement'] = qc_cv_before['qc_cv'] - qc_cv_after['qc_cv']

        print(f"   - 代謝物數量: {len(feature_ids)}")
        print(f"   - 樣本數量: {len(sample_columns)}")
        print(f"   - 添加統計欄位: Cohen's d (前/後/改善), QC CV% (前/後/改善)")
        
        # 保存到 Excel
        save_results_to_excel(input_file, output_file, data, sample_info,
                     corrected_df, sheet_name,
                     permanova_before, permanova_after,
                     permdisp_before, permdisp_after,
                     perm_test_permanova, perm_test_silhouette,
                     silhouette_before, silhouette_after,
                     cohens_d_before, cohens_d_after,
                     qc_cv_before, qc_cv_after,
                     sample_type_row=sample_type_row)
        
        # ========== 12. 品質檢查與警告 ==========
        print("\n" + "="*70)
        print("⚠️  品質檢查與警告")
        print("="*70)
        
        warnings_list = generate_quality_warnings(
            permanova_before, permanova_after,
            {'permanova': perm_test_permanova, 'silhouette': perm_test_silhouette},
            qc_cv_before, qc_cv_after,
            cohens_d_before, cohens_d_after,
            outliers_orig_batch, outliers_corr_batch,
            batch_info
        )
        
        print_warnings(warnings_list)
        
        # ========== 13. 最終摘要 ==========
        print("\n" + "="*70)
        print("✅ 批次效應校正完成！")
        print("="*70)
        
        # 判斷校正成功與否
        success_criteria = []
        if _perm_ok:
            success_criteria.append(permanova_after['p_value'] > 0.05)  # 校正後不顯著
            success_criteria.append(permanova_after['r_squared'] < 0.15)  # R² < 15%
        if perm_test_permanova and not np.isnan(perm_test_permanova.get('p_value', np.nan)):
            success_criteria.append(perm_test_permanova['p_value'] < 0.05)  # 改善顯著

        # 如果沒有 PERMANOVA，用 Silhouette 和 Cohen's d 作為備用判斷
        if not success_criteria:
            if not np.isnan(silhouette_after) and not np.isnan(silhouette_before):
                success_criteria.append(silhouette_after < silhouette_before)
            success_criteria.append(cohens_d_after['overall_cohens_d'] < cohens_d_before['overall_cohens_d'])

        success_count = sum(success_criteria) if success_criteria else 0

        if success_count >= 2:
            print("\n批次效應校正效果: ✅ 成功")
        elif success_count == 1:
            print("\n批次效應校正效果: ⚠️ 部分成功")
        else:
            print("\n批次效應校正效果: ❌ 效果不佳")

        print("\n主要證據:")
        if _perm_ok:
            print(f"  1. PERMANOVA: F={permanova_before['pseudo_f']:.3f} → {permanova_after['pseudo_f']:.3f}, " +
                  f"p={permanova_before['p_value']:.4f} → {permanova_after['p_value']:.4f}")
            print(f"  2. 批次解釋變異: R²={permanova_before['r_squared']*100:.1f}% → {permanova_after['r_squared']*100:.1f}%")
        else:
            print("  1. PERMANOVA: ⚠️ 跳過（scikit-bio 未安裝）")
            print("  2. 批次解釋變異: ⚠️ 跳過")

        if perm_test_permanova and not np.isnan(perm_test_permanova.get('p_value', np.nan)):
            print(f"  3. 配對檢定: p={perm_test_permanova['p_value']:.4f} {get_significance_mark(perm_test_permanova['p_value'])}")
        else:
            print("  3. 配對檢定: ⚠️ 跳過")

        if not np.isnan(qc_cv_before['median_cv']) and not np.isnan(qc_cv_after['median_cv']):
            print(f"  4. QC CV%: {qc_cv_before['median_cv']:.2f}% → {qc_cv_after['median_cv']:.2f}%")

        print(f"  5. Cohen's d: {cohens_d_before['overall_cohens_d']:.3f} → {cohens_d_after['overall_cohens_d']:.3f}")
        
        # 輸出檔案資訊
        print("\n" + "="*70)
        print("📁 輸出檔案")
        print("="*70)
        
        print(f"\nExcel 檔案: {os.path.basename(output_file)}")
        print(f"   路徑: {output_file}")
        print(f"   工作表:")
        print(f"     - {SUMMARY_SHEET_NAME} (統計摘要)")
        print(f"     - Batch_effect_result (校正後數據)")
        print(f"     - SampleInfo (樣本資訊)")
        print(f"     - 其他原始工作表...")
        
        print(f"\n圖表資料夾: {os.path.basename(run_plot_dir)}")
        print(f"   路徑: {run_plot_dir}")
        print(f"   已生成 {4} 張圖表:")
        print(f"     1. PERMANOVA_comparison_{timestamp}.png")
        print(f"     2. PCA_by_batch_{timestamp}.png")
        print(f"     3. PCA_by_sample_type_{timestamp}.png")
        print(f"     4. Permutation_Test_{timestamp}.png")
        
        print("\n" + "="*70 + "\n")
        
        # 返回統計資訊給 GUI
        return ProcessingResult(
            file_path=input_file,
            output_path=str(output_file),
            plots_dir=str(run_plot_dir),
            metabolites=len(feature_ids),
            samples=len(sample_columns),
            extra={
                'batches': len(unique_batches),
                'permanova_f_before': _safe_get(permanova_before, 'pseudo_f'),
                'permanova_f_after': _safe_get(permanova_after, 'pseudo_f'),
                'permanova_p_before': _safe_get(permanova_before, 'p_value'),
                'permanova_p_after': _safe_get(permanova_after, 'p_value'),
                'r2_before': _safe_get(permanova_before, 'r_squared'),
                'r2_after': _safe_get(permanova_after, 'r_squared'),
                'perm_test_pvalue': _safe_get(perm_test_permanova, 'p_value'),
                'qc_cv_before': qc_cv_before['median_cv'],
                'qc_cv_after': qc_cv_after['median_cv'],
                'cohens_d_before': cohens_d_before['overall_cohens_d'],
                'cohens_d_after': cohens_d_after['overall_cohens_d'],
                'success': success_count >= 2
            }
        )
        
    except Exception as e:
        print(f"\n❌ 錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        print("\n請檢查輸入檔案格式是否正確")
        raise


def process_in_memory(data_df, sample_info_df, **kwargs):
    """Pipeline-friendly entry point — bypass file I/O.

    Parameters
    ----------
    data_df : DataFrame
        Features-as-rows with feature-ID first column + sample columns.
    sample_info_df : DataFrame
        Must contain ``Sample_Name`` and ``Batch`` columns.

    Returns
    -------
    DataFrame or None
        ComBat-corrected result with feature-ID + sample columns, or the
        original *data_df* unchanged when there is only a single batch.
    """
    if data_df is None or data_df.empty:
        return None
    if sample_info_df is None or sample_info_df.empty:
        return None

    # Early single-batch check (prepare_data_for_combat raises on < 2 batches)
    if "Batch" in sample_info_df.columns:
        unique_batches = sample_info_df["Batch"].dropna().unique()
        if len(unique_batches) < 2:
            return data_df

    # prepare_data_for_combat expects the raw data_df (feature-col + samples)
    try:
        data_matrix, batch_info, sample_columns, feature_ids = prepare_data_for_combat(
            data_df, sample_info_df
        )
    except ValueError:
        # Missing Batch column, insufficient batches, etc.
        return data_df

    unique_batches = list(set(batch_info))
    if len(unique_batches) < 2:
        # Single batch → nothing to correct; return data as-is
        return data_df

    # Check batch ↔ sample-type confounding
    confound = check_batch_confounding(sample_info_df)
    if confound.get("is_confounded", False):
        # Confounded → correction would remove biological signal; return as-is
        return data_df

    # Core ComBat correction
    corrected_data = perform_combat_correction(data_matrix, batch_info)

    # Reconstruct DataFrame
    feature_col = data_df.columns[0]
    result_df = pd.DataFrame(corrected_data, columns=sample_columns)
    result_df.insert(0, feature_col, feature_ids)
    return result_df


if __name__ == "__main__":
    main()
