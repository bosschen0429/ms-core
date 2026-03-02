import pandas as pd
import numpy as np
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import statsmodels.api as sm
from scipy.stats import levene, kendalltau, wilcoxon
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from scipy.stats import chi2
import scipy.stats as stats
import matplotlib.pyplot as plt
from matplotlib.patches import Ellipse
import warnings
from collections import Counter

warnings.filterwarnings('ignore')

# ========== 匯入共用模組 ==========
from ms_core.utils.data_helpers import get_valid_values
from ms_core.utils.plotting import setup_matplotlib, plot_pca_comparison_qc_style
from ms_core.utils.constants import (
    NON_SAMPLE_COLUMNS,
    SHEET_NAMES,
    DATETIME_FORMAT_FULL,
    FEATURE_ID_COLUMN,
    CV_QUALITY_THRESHOLDS,
)
from ms_core.utils.sample_classification import (
    normalize_sample_name,
    normalize_sample_type,
    identify_sample_columns,
)
from ms_core.utils.file_io import build_output_path, build_plots_dir, get_output_root
from ms_core.utils.results import ProcessingResult
from ms_core.utils.excel_format import copy_sheet_formatting_only
from ms_core.utils.console import safe_print as print

# 設定 matplotlib
setup_matplotlib()

# Sheet name constant
QC_LOWESS_ADVANCED_SHEET = SHEET_NAMES.get('qc_lowess_advanced', "QC_LOWESS_Advanced Statistics")

# For backward compatibility, alias the old constant names
DEFAULT_NON_SAMPLE_COLUMNS = NON_SAMPLE_COLUMNS


def apply_lowess_correction(qc_orders, qc_intensities, all_orders, all_intensities, debug_flag=None, global_qc_median=None):
    """對單一批次特徵執行 QC-LOWESS 校正並回傳詳細統計。

    Args:
        qc_orders: QC 樣本的 injection order
        qc_intensities: QC 樣本的強度值
        all_orders: 所有樣本的 injection order
        all_intensities: 所有樣本的強度值
        debug_flag: 除錯標記
        global_qc_median: 全域 QC 中位數（跨所有批次計算）
    """
    info = {
        'status': 'failed',
        'cv_before': np.nan,
        'cv_after': np.nan,
        'cv_improvement': np.nan,
        'correction_factor_stats': {},
        'trend_validation': {
            'trend_pvalue': np.nan,
            'trend_tau': np.nan,
            'r_squared': np.nan,
            'rmse': np.nan
        },
        'frac_used': np.nan,
        'qc_cv_for_frac': np.nan,
        'frac_strategy': 'unknown',
        'global_median_used': global_qc_median is not None
    }

    if all_orders is None or all_intensities is None:
        return [], info

    all_orders_arr = np.asarray(all_orders, dtype=float)
    all_intensities_arr = np.asarray(all_intensities, dtype=float)
    if all_orders_arr.size == 0:
        return all_intensities_arr.tolist(), info

    qc_orders_arr = np.asarray(qc_orders, dtype=float)
    qc_intensities_arr = np.asarray(qc_intensities, dtype=float)
    valid_mask = np.isfinite(qc_orders_arr) & np.isfinite(qc_intensities_arr) & (qc_intensities_arr > 0)
    valid_x = qc_orders_arr[valid_mask]
    valid_y = qc_intensities_arr[valid_mask]

    if valid_x.size < 3 or np.unique(valid_x).size < 2:
        info['status'] = 'insufficient_qc'
        info['frac_strategy'] = 'insufficient_qc'
        return all_intensities_arr.tolist(), info

    # ===== 動態 frac 策略 =====
    def compute_qc_cv(values):
        values = np.asarray(values, dtype=float)
        values = values[np.isfinite(values) & (values > 0)]
        if values.size < 2:
            return 100.0
        mean_val = np.nanmean(values)
        if not np.isfinite(mean_val) or mean_val <= 0:
            return 100.0
        std_val = np.nanstd(values, ddof=1)
        if not np.isfinite(std_val):
            return 100.0
        return float(std_val / mean_val * 100)

    qc_cv_for_frac = compute_qc_cv(valid_y)
    qc_cv_for_frac = 100.0 if not np.isfinite(qc_cv_for_frac) else qc_cv_for_frac

    if qc_cv_for_frac > CV_QUALITY_THRESHOLDS['acceptable']:
        frac = 0.8
        frac_strategy = 'high_variation'
    elif qc_cv_for_frac > CV_QUALITY_THRESHOLDS['excellent']:
        frac = 0.7
        frac_strategy = 'medium_variation'
    else:
        dynamic_frac = 0.85 - (valid_x.size / 50.0)
        frac = float(np.clip(dynamic_frac, 0.5, 0.75))
        frac_strategy = 'low_variation_dynamic'

    info['frac_used'] = float(frac)
    info['qc_cv_for_frac'] = float(qc_cv_for_frac)
    info['frac_strategy'] = frac_strategy

    lowess_result = sm.nonparametric.lowess(valid_y, valid_x, frac=frac, it=2, return_sorted=True)
    x_fit, y_fit = lowess_result[:, 0], lowess_result[:, 1]

    # 使用全域 QC 中位數（如果有提供），否則使用批次內中位數
    if global_qc_median is not None and np.isfinite(global_qc_median) and global_qc_median > 0:
        median_qc = global_qc_median
    else:
        median_qc = np.nanmedian(y_fit)
        if not np.isfinite(median_qc) or median_qc <= 0:
            median_qc = np.nanmedian(valid_y)
        if not np.isfinite(median_qc) or median_qc <= 0:
            median_qc = 1.0

    def predict(x_new):
        if not np.isfinite(x_new):
            return np.nan
        return float(np.interp(x_new, x_fit, y_fit, left=y_fit[0], right=y_fit[-1]))

    corrected = []
    factors = []
    for order, intensity in zip(all_orders_arr, all_intensities_arr):
        if not np.isfinite(intensity) or intensity <= 0:
            corrected.append(0.0)
            continue
        fitted = predict(order)
        if not np.isfinite(fitted) or fitted <= 0:
            corrected.append(float(intensity))
            continue
        factor = median_qc / fitted
        factors.append(factor)
        corrected.append(float(intensity * factor))

    qc_pred = np.array([predict(x) for x in valid_x], dtype=float)
    qc_pred = np.where(np.isfinite(qc_pred) & (qc_pred > 0), qc_pred, np.nan)
    qc_factors = np.where(np.isfinite(qc_pred), median_qc / qc_pred, 1.0)
    qc_corrected = valid_y * qc_factors

    def calc_cv(values):
        values = np.asarray(values, dtype=float)
        values = values[np.isfinite(values) & (values > 0)]
        if values.size < 2:
            return np.nan
        mean_val = np.mean(values)
        if mean_val == 0:
            return np.nan
        return float(np.std(values, ddof=1) / mean_val * 100)

    original_cv = calc_cv(valid_y)
    corrected_cv = calc_cv(qc_corrected)
    cv_improvement = original_cv - corrected_cv if np.isfinite(original_cv) and np.isfinite(corrected_cv) else np.nan
    factor_array = np.asarray(factors, dtype=float)
    factor_cv = calc_cv(factor_array) if factor_array.size >= 2 else np.nan

    try:
        output_file = str(output_file)
        trend_tau, trend_pvalue = kendalltau(valid_x, valid_y)
    except Exception:
        trend_tau, trend_pvalue = (np.nan, np.nan)

    qc_predicted = np.array([predict(x) for x in valid_x], dtype=float)
    qc_predicted = np.where(np.isfinite(qc_predicted), qc_predicted, np.nanmedian(valid_y))
    ss_res = np.nansum((valid_y - qc_predicted) ** 2)
    ss_tot = np.nansum((valid_y - np.nanmean(valid_y)) ** 2)
    r_squared = 1 - ss_res / ss_tot if ss_tot > 0 else np.nan
    rmse = np.sqrt(np.nanmean((valid_y - qc_predicted) ** 2))

    status = 'success'
    if not np.isfinite(original_cv) or not np.isfinite(corrected_cv):
        status = 'failed'
    elif cv_improvement < -1:
        status = 'overcorrection_detected'
    elif factor_array.size >= 3 and np.isfinite(factor_cv) and factor_cv > 50:
        status = 'unstable_correction_factors'
    elif not np.isfinite(cv_improvement) or cv_improvement < 2:
        status = 'insufficient_improvement'

    info['status'] = status
    info['cv_before'] = original_cv
    info['cv_after'] = corrected_cv
    info['cv_improvement'] = cv_improvement
    info['correction_factor_stats'] = {
        'median': float(np.nanmedian(factor_array)) if factor_array.size else np.nan,
        'cv_percent': factor_cv,
        'min': float(np.nanmin(factor_array)) if factor_array.size else np.nan,
        'max': float(np.nanmax(factor_array)) if factor_array.size else np.nan
    }
    info['trend_validation'] = {
        'trend_pvalue': trend_pvalue,
        'trend_tau': trend_tau,
        'r_squared': r_squared,
        'rmse': rmse
    }

    if debug_flag:
        delta = cv_improvement if np.isfinite(cv_improvement) else float('nan')
        frac_val = info.get('frac_used', np.nan)
        qc_cv_val = info.get('qc_cv_for_frac', np.nan)
        strategy = info.get('frac_strategy', 'unknown')
        print(f"     [DEBUG] Feature {debug_flag}: status={status}, ΔCV={delta:.2f}%")
        print(f"              - QC CV% = {qc_cv_val:.2f}%")
        print(f"              - Frac used = {frac_val:.3f}")
        print(f"              - Strategy = {strategy}")

        # 儲存繪圖數據用於趨勢擬合圖
        info['plot_data'] = {
            'qc_orders': valid_x.tolist(),
            'qc_raw': valid_y.tolist(),
            'qc_corrected': qc_corrected.tolist(),
            'lowess_x': x_fit.tolist(),
            'lowess_y': y_fit.tolist(),
            'all_orders': all_orders_arr.tolist(),
            'all_raw': all_intensities_arr.tolist(),
            'all_corrected': corrected,
            'median_qc': float(median_qc)
        }

    return corrected, info

def perform_lowess_normalization(istd_df, sample_info_df):
    """執行分批次的 QC-LOWESS 正規化流程。"""
    try:
        if istd_df is None or istd_df.empty:
            raise ValueError("ISTD_Correction 數據為空")

        if sample_info_df is None or sample_info_df.empty:
            raise ValueError("SampleInfo 數據為空")

        # 支援 'Mz/RT' 或 'FeatureID' 作為特徵ID欄位
        if FEATURE_ID_COLUMN in istd_df.columns and FEATURE_ID_COLUMN != 'FeatureID':
            istd_df = istd_df.rename(columns={FEATURE_ID_COLUMN: 'FeatureID'})
        elif 'FeatureID' not in istd_df.columns:
            first_col = istd_df.columns[0]
            print(f"⚠️ 未找到 'FeatureID' 欄位，使用第一欄 '{first_col}' 作為特徵ID")
            istd_df = istd_df.rename(columns={first_col: 'FeatureID'})

        if 'Sample_Name' not in sample_info_df.columns or 'Sample_Type' not in sample_info_df.columns:
            raise ValueError("SampleInfo 缺少必要欄位 (Sample_Name, Sample_Type)")

        sample_columns = istd_df.attrs.get('sample_columns')
        if not sample_columns:
            sample_columns, _ = identify_sample_columns(istd_df, sample_info_df)

        sample_columns = [col for col in sample_columns if col in istd_df.columns]
        if not sample_columns:
            raise ValueError("找不到有效的樣本欄位")

        sample_meta = sample_info_df.set_index('Sample_Name')
        missing_meta = [col for col in sample_columns if col not in sample_meta.index]

        if missing_meta and len(missing_meta) == len(sample_columns):
            # 所有欄位都匹配不上 — 名稱格式不同，使用欄位名稱中的關鍵字判斷 QC
            print(f"⚠️  SampleInfo 與 ISTD_Correction 的樣本名稱格式不同，改用欄位名稱推斷樣本類型")
            # 不過濾 sample_columns，直接使用所有數據欄位
        elif missing_meta:
            print("⚠️  警告：以下樣本在 SampleInfo 中找不到對應資訊，將被排除：")
            for name in missing_meta[:5]:
                print(f"     - {name}")
            if len(missing_meta) > 5:
                print(f"     ... 還有 {len(missing_meta) - 5} 個樣本")
            sample_columns = [col for col in sample_columns if col in sample_meta.index]

        if not sample_columns:
            raise ValueError("無法匹配 SampleInfo 與 ISTD_Correction 的樣本欄位")

        # 判斷 QC 樣本：優先從 SampleInfo 查找，如找不到則從欄位名稱關鍵字判斷
        qc_samples = []
        for sample in sample_columns:
            if sample in sample_meta.index:
                if 'QC' in str(sample_meta.loc[sample].get('Sample_Type', '')).upper():
                    qc_samples.append(sample)
            elif 'QC' in sample.upper() or 'POOLED' in sample.upper():
                qc_samples.append(sample)

        if len(qc_samples) < 5:
            raise ValueError(f"QC 樣本不足 ({len(qc_samples)} < 5)，無法進行校正")

        # 建立欄位名稱到 SampleInfo 的映射（支援名稱不完全匹配的情況）
        col_to_meta = {}
        if all(s in sample_meta.index for s in sample_columns):
            col_to_meta = {s: s for s in sample_columns}
        else:
            # 名稱不匹配 — 按順序對齊（SampleInfo 和 ISTD_Correction 的樣本順序一致）
            info_names = sample_info_df['Sample_Name'].tolist()
            if len(info_names) == len(sample_columns):
                col_to_meta = dict(zip(sample_columns, info_names))
            else:
                # 按位置對齊失敗，嘗試用 QC 關鍵字匹配
                for col in sample_columns:
                    for name in info_names:
                        if name in sample_meta.index:
                            col_to_meta[col] = name
                            break

        batch_groups = {}
        missing_order_samples = []
        for sample in sample_columns:
            meta_name = col_to_meta.get(sample, sample)
            if meta_name not in sample_meta.index:
                continue
            meta_row = sample_meta.loc[meta_name]
            batch_name = str(meta_row.get('Batch', 'Batch1'))
            batch_entry = batch_groups.setdefault(
                batch_name,
                {'samples': [], 'qc_samples': [], 'injection_orders': {}}
            )
            batch_entry['samples'].append(sample)

            sample_type = str(meta_row.get('Sample_Type', '')).upper()
            if 'QC' in sample_type:
                batch_entry['qc_samples'].append(sample)

            order = meta_row.get('Injection_Order')
            if pd.isna(order):
                missing_order_samples.append(sample)
                order = len(batch_entry['injection_orders']) + 1
            batch_entry['injection_orders'][sample] = order

        active_batches = {k: v for k, v in batch_groups.items() if v['samples']}
        if not active_batches:
            raise ValueError("找不到可供處理的批次樣本")

        print("\n📊 數據概覽：")
        print(f"  - 特徵數: {len(istd_df)}")
        print(f"  - 樣本總數: {len(sample_columns)}")
        print(f"  - QC 樣本數: {len(qc_samples)}")
        print(f"  - 批次數: {len(active_batches)}")
        for batch, info in active_batches.items():
            print(f"    • Batch {batch}: {len(info['samples'])} samples (QC: {len(info['qc_samples'])})")

        if missing_order_samples:
            print(f"⚠️  提示：{len(missing_order_samples)} 個樣本缺少 Injection_Order，已套用臨時序號")

        print("\n🔍 計算 QC CV% 以選擇調試特徵...")
        # Vectorized CV% calculation - much faster than iterrows()
        from ms_core.utils.safe_math import safe_cv_percent_vectorized

        valid_qc_cols = [c for c in qc_samples if c in istd_df.columns]
        if valid_qc_cols:
            qc_data = istd_df[valid_qc_cols].apply(pd.to_numeric, errors='coerce').values
            cv_values = safe_cv_percent_vectorized(qc_data, axis=1, min_samples=2)
            feature_ids = istd_df['FeatureID'].values
            # Build list of (feature_id, cv_value) for features with valid CV
            feature_cvs = [
                (fid, cv) for fid, cv in zip(feature_ids, cv_values)
                if not np.isnan(cv)
            ]
        else:
            feature_cvs = []

        debug_features = []
        if feature_cvs:
            feature_cvs_sorted = sorted(feature_cvs, key=lambda x: x[1])
            debug_features = [feature_cvs_sorted[0][0]]
            debug_features.append(feature_cvs_sorted[len(feature_cvs_sorted) // 2][0])
            debug_features.append(feature_cvs_sorted[-1][0])
            debug_features = list(dict.fromkeys(debug_features))
            print("   • 調試特徵:")
            for fid in debug_features:
                matching = [cv for cv in feature_cvs if cv[0] == fid]
                if matching:
                    print(f"     - {fid} (CV% = {matching[0][1]:.2f}%)")

        def normalize_feature_for_batch(feature_row, batch_name, batch_info, debug_flag, global_median=None):
            samples = batch_info['samples']
            injection_orders = batch_info['injection_orders']
            valid_samples = [s for s in samples if s in injection_orders]
            if not valid_samples:
                return {
                    'status': 'failed',
                    'corrected_samples': {},
                    'trend_validation': None,
                    'qc_samples': []
                }

            batch_data = []
            for sample in valid_samples:
                order = injection_orders[sample]
                intensity = feature_row.get(sample, 0)
                if pd.isna(intensity) or intensity <= 0:
                    intensity = 0
                batch_data.append((sample, order, float(intensity)))

            batch_data.sort(key=lambda x: x[1])
            all_sample_names = [d[0] for d in batch_data]
            all_orders = [d[1] for d in batch_data]
            all_intensities = [d[2] for d in batch_data]

            qc_batch_samples = [s for s in batch_info['qc_samples'] if s in all_sample_names]
            qc_indices = [all_sample_names.index(s) for s in qc_batch_samples]
            qc_orders = [all_orders[i] for i in qc_indices]
            qc_intensities = [all_intensities[i] for i in qc_indices]

            corrected_intensities, info = apply_lowess_correction(
                qc_orders, qc_intensities, all_orders, all_intensities, debug_flag, global_median
            )

            corrected_map = dict(zip(all_sample_names, corrected_intensities))
            result = {
                'status': info.get('status', 'unknown'),
                'corrected_samples': corrected_map,
                'trend_validation': info.get('trend_validation', None),
                'qc_samples': qc_batch_samples,
                'frac_info': {
                    'frac_used': info.get('frac_used', np.nan),
                    'qc_cv_for_frac': info.get('qc_cv_for_frac', np.nan),
                    'frac_strategy': info.get('frac_strategy', 'unknown')
                }
            }
            # 如果有繪圖數據（僅限 debug 特徵），則加入
            if 'plot_data' in info:
                result['plot_data'] = info['plot_data']
            return result

        def safe_nanmedian(values):
            if not values:
                return np.nan
            arr = np.array(values, dtype=float)
            if arr.size == 0 or np.all(np.isnan(arr)):
                return np.nan
            return float(np.nanmedian(arr))

        def choose_frac_strategy(strategies):
            if not strategies:
                return 'unknown'
            valid_strategies = [s for s in strategies if isinstance(s, str)]
            if not valid_strategies:
                return 'unknown'
            strategy_counter = Counter(valid_strategies)
            priority = ['high_variation', 'medium_variation', 'low_variation_dynamic',
                        'insufficient_qc', 'unknown']
            for key in priority:
                if key in strategy_counter:
                    return key
            return strategy_counter.most_common(1)[0][0]

        status_categories = [
            'success', 'insufficient_qc', 'insufficient_improvement',
            'unstable_correction_factors', 'overcorrection_detected',
            'failed', 'unknown'
        ]
        decision_stats = {key: 0 for key in status_categories}
        decision_stats['partial_success'] = 0
        decision_stats['event_counts'] = {key: 0 for key in status_categories}
        decision_stats['per_batch'] = {}
        decision_stats['total_features'] = len(istd_df)
        decision_stats['total_batches'] = len(active_batches)
        decision_stats['total_feature_batch_tasks'] = len(active_batches) * len(istd_df)

        # ===== 計算全域 QC 中位數（跨所有批次）=====
        print("\n🌐 計算全域 QC 中位數（跨所有批次）...")
        # Vectorized median calculation - much faster than iterrows()
        valid_qc_cols = [c for c in qc_samples if c in istd_df.columns]
        if valid_qc_cols:
            qc_data = istd_df[valid_qc_cols].apply(pd.to_numeric, errors='coerce').values
            # Replace non-positive values with NaN
            qc_data = np.where(qc_data > 0, qc_data, np.nan)
            # Count valid values per row
            valid_counts = np.sum(np.isfinite(qc_data), axis=1)
            # Calculate median for each row
            medians = np.nanmedian(qc_data, axis=1)
            # Build dictionary: None for features with <3 valid QC values
            feature_ids = istd_df['FeatureID'].values
            global_qc_medians = {
                fid: (med if cnt >= 3 else None)
                for fid, med, cnt in zip(feature_ids, medians, valid_counts)
            }
        else:
            global_qc_medians = {fid: None for fid in istd_df['FeatureID']}

        valid_global_medians = sum(1 for v in global_qc_medians.values() if v is not None)
        print(f"  - 成功計算全域中位數的特徵數: {valid_global_medians}/{len(istd_df)}")

        all_results = []
        qc_corrected_values = {}
        trend_stats = []
        feature_all_success = 0
        feature_partial_success = 0
        feature_no_success = 0
        frac_usage_counter = Counter()
        frac_value_list = []
        trend_plot_data = {}  # 儲存趨勢擬合圖的數據

        for idx, row in istd_df.iterrows():
            feature_id = row['FeatureID']
            debug_flag = feature_id if feature_id in debug_features else None
            global_median = global_qc_medians.get(feature_id)

            result_row = {'FeatureID': feature_id}
            for sample in sample_columns:
                result_row[sample] = row[sample]

            qc_corrected_dict = {sample: row[sample] for sample in qc_samples if sample in row.index}
            trend_metric_buffer = []
            batch_statuses = []
            frac_value_buffer = []
            frac_cv_buffer = []
            frac_strategy_buffer = []

            for batch_name, batch_info in active_batches.items():
                batch_result = normalize_feature_for_batch(row, batch_name, batch_info, debug_flag, global_median)
                status = batch_result['status']
                batch_statuses.append(status)

                decision_stats['event_counts'].setdefault(status, 0)
                decision_stats['event_counts'][status] += 1
                batch_entry = decision_stats['per_batch'].setdefault(batch_name, {})
                batch_entry[status] = batch_entry.get(status, 0) + 1

                corrected_map = batch_result['corrected_samples']
                if corrected_map:
                    for sample, value in corrected_map.items():
                        result_row[sample] = value
                        if sample in qc_corrected_dict:
                            qc_corrected_dict[sample] = value

                if batch_result['trend_validation']:
                    trend_metric_buffer.append(batch_result['trend_validation'])

                frac_info = batch_result.get('frac_info') or {}
                frac_value_buffer.append(frac_info.get('frac_used'))
                frac_cv_buffer.append(frac_info.get('qc_cv_for_frac'))
                frac_strategy_buffer.append(frac_info.get('frac_strategy'))

                # 收集趨勢擬合圖數據（僅限 debug 特徵）
                if debug_flag and 'plot_data' in batch_result:
                    trend_plot_data[(feature_id, batch_name)] = batch_result['plot_data']

            success_batches = batch_statuses.count('success')
            if success_batches == len(active_batches):
                decision_stats['success'] += 1
                feature_all_success += 1
            elif success_batches > 0:
                decision_stats['partial_success'] += 1
                feature_partial_success += 1
            else:
                failure_priority = [
                    status for status in status_categories
                    if status != 'success' and status in batch_statuses
                ]
                failure_key = failure_priority[0] if failure_priority else 'failed'
                decision_stats[failure_key] += 1
                feature_no_success += 1

            frac_values_clean = [val for val in frac_value_buffer if val is not None]
            frac_cvs_clean = [val for val in frac_cv_buffer if val is not None]
            feature_frac_used = safe_nanmedian(frac_values_clean)
            feature_qc_cv = safe_nanmedian(frac_cvs_clean)
            feature_frac_strategy = choose_frac_strategy(frac_strategy_buffer)

            if np.isfinite(feature_frac_used):
                frac_value_list.append(feature_frac_used)
            frac_usage_counter[feature_frac_strategy] += 1

            trend_stats.append({
                'FeatureID': feature_id,
                'MK_Trend_pvalue': safe_nanmedian([m.get('trend_pvalue', np.nan) for m in trend_metric_buffer]),
                'Kendall_Tau': safe_nanmedian([m.get('trend_tau', np.nan) for m in trend_metric_buffer]),
                'LOWESS_R2': safe_nanmedian([m.get('r_squared', np.nan) for m in trend_metric_buffer]),
                'LOWESS_RMSE': safe_nanmedian([m.get('rmse', np.nan) for m in trend_metric_buffer]),
                'Frac_Used': feature_frac_used,
                'QC_CV_for_Frac': feature_qc_cv,
                'Frac_Strategy': feature_frac_strategy
            })

            all_results.append(result_row)
            qc_corrected_values[feature_id] = qc_corrected_dict

            if (idx + 1) % 500 == 0:
                print(f"  進度: {idx + 1}/{len(istd_df)} features")

        print("\n  ✓ 批次化 LOWESS 校正完成")
        print("\n  📊 特徵層級統計：")
        print(f"     ✅ 全批次均成功: {feature_all_success} ({feature_all_success/len(istd_df)*100:.1f}%)")
        print(f"     ⚠️ 部分批次成功: {feature_partial_success} ({feature_partial_success/len(istd_df)*100:.1f}%)")
        print(f"     ❌ 無成功批次: {feature_no_success} ({feature_no_success/len(istd_df)*100:.1f}%)")

        print("\n  📊 決策細節 (以批次為單位)：")
        total_tasks = decision_stats['total_feature_batch_tasks'] or 1
        for status, count in decision_stats['event_counts'].items():
            if count == 0:
                continue
            print(f"     • {status}: {count} ({count/total_tasks*100:.1f}%)")

        print("\n  📊 Frac 使用統計：")
        total_features = len(istd_df) or 1
        high_count = frac_usage_counter.get('high_variation', 0)
        medium_count = frac_usage_counter.get('medium_variation', 0)
        low_count = frac_usage_counter.get('low_variation_dynamic', 0)
        other_count = max(total_features - (high_count + medium_count + low_count), 0)
        frac_mean = float(np.nanmean(frac_value_list)) if frac_value_list else np.nan
        frac_median = float(np.nanmedian(frac_value_list)) if frac_value_list else np.nan

        def frac_pct(count):
            return (count / total_features * 100) if total_features else 0

        def fmt_frac_value(value):
            return f"{value:.2f}" if np.isfinite(value) else "N/A"

        print(f"     - 高變異策略 (frac=0.8): {high_count} ({frac_pct(high_count):.1f}%)")
        print(f"     - 中等變異策略 (frac=0.7): {medium_count} ({frac_pct(medium_count):.1f}%)")
        print(f"     - 低變異動態策略 (frac=0.5~0.75): {low_count} ({frac_pct(low_count):.1f}%)")
        if other_count:
            print(f"     - 其他（資料不足）: {other_count} ({frac_pct(other_count):.1f}%)")
        print(f"     - Frac 平均值: {fmt_frac_value(frac_mean)}")
        print(f"     - Frac 中位數: {fmt_frac_value(frac_median)}")

        lowess_df = pd.DataFrame(all_results)
        trend_stats_df = pd.DataFrame(trend_stats)

        decision_stats['frac_usage_counter'] = dict(frac_usage_counter)
        decision_stats['frac_value_list'] = frac_value_list

        return lowess_df, sample_columns, qc_corrected_values, trend_stats_df, decision_stats, trend_plot_data

    except Exception as e:
        print(f"❌ LOWESS 校正失敗: {e}")
        import traceback
        traceback.print_exc()
        raise


    # ========== 數據載入 ==========
def get_valid_values(row, columns):
    """從 DataFrame 的一行中提取有效值（>0 且非 NaN）"""
    values = []
    for col in columns:
        if col in row.index:
            try:
                val = float(row[col])
                if not pd.isna(val) and val > 0:
                    values.append(val)
            except (ValueError, TypeError):
                pass
    return values


def load_and_process_data(file_path):
    """載入並驗證數據（含完整防呆檢查）"""
    try:
        # ===== 防呆1: 文件存在性檢查 =====
        if not os.path.exists(file_path):
            raise ValueError(f"找不到檔案 '{file_path}'")

        # ===== 防呆2: 文件格式檢查 =====
        if not (file_path.endswith('.xlsx') or file_path.endswith('.xls')):
            raise ValueError(f"輸入檔案必須是 Excel 格式 (.xlsx 或 .xls)，但提供了 {file_path}")

        # ===== 防呆3: 文件大小檢查 =====
        file_size = os.path.getsize(file_path)
        if file_size == 0:
            raise ValueError("檔案大小為 0 bytes，可能是空檔案")
        elif file_size < 1024:  # 小於 1KB
            print(f"⚠️  警告：檔案大小僅 {file_size} bytes，可能不是有效的 Excel 檔案")

        print(f"📄 檔案大小: {file_size / 1024:.2f} KB")

        # ===== 防呆4: Excel 文件有效性檢查 =====
        try:
            excel_file = pd.ExcelFile(file_path)
        except Exception as e:
            raise ValueError(f"無法讀取 Excel 檔案，可能已損壞或格式不正確: {e}") from e

        # ===== 防呆5: 必要工作表檢查 =====
        print(f"📋 找到的工作表: {', '.join(excel_file.sheet_names)}")

        required_sheets = [SHEET_NAMES['istd_correction'], SHEET_NAMES['sample_info']]
        missing_sheets = [sheet for sheet in required_sheets if sheet not in excel_file.sheet_names]

        if missing_sheets:
            raise ValueError(
                f"輸入檔案缺少必要的工作表: {', '.join(missing_sheets)}。"
                f" 找到的工作表: {', '.join(excel_file.sheet_names)}。"
                f" QC-LOWESS 校正需要先執行 ISTD_Correction"
            )

        # ===== 防呆6: SampleInfo 完整性檢查 =====
        sample_info_df = pd.read_excel(excel_file, sheet_name=SHEET_NAMES['sample_info'])
        print(f"✓ 成功讀取 '{SHEET_NAMES['sample_info']}' 工作表，包含 {len(sample_info_df)} 筆樣本資訊")

        if sample_info_df.empty:
            raise ValueError(f"'{SHEET_NAMES['sample_info']}' 工作表為空")

        required_columns = ['Sample_Name', 'Sample_Type', 'Injection_Order']
        missing_cols = [col for col in required_columns if col not in sample_info_df.columns]

        if missing_cols:
            raise ValueError(
                f"'{SHEET_NAMES['sample_info']}' 缺少必要欄位: {', '.join(missing_cols)}。"
                f" 找到的欄位: {', '.join(sample_info_df.columns.tolist())}"
            )

        # ===== 防呆6-1: Batch 欄位處理 =====
        if 'Batch' not in sample_info_df.columns:
            sample_info_df['Batch'] = 'Batch1'
            print(f"⚠️  警告：'{SHEET_NAMES['sample_info']}' 缺少 'Batch' 欄位，已建立預設 Batch1")
        else:
            batch_na_mask = sample_info_df['Batch'].isna()
            if batch_na_mask.any():
                fill_value = 'Unknown'
                sample_info_df.loc[batch_na_mask, 'Batch'] = fill_value
                print(f"⚠️  警告：發現 {batch_na_mask.sum()} 個樣本缺少 Batch，已填入 '{fill_value}'")

        batch_summary = sample_info_df['Batch'].astype(str).value_counts().to_dict()

        # ===== 防呆7: 樣本名稱重複檢查 =====
        duplicate_samples = sample_info_df[sample_info_df['Sample_Name'].duplicated()]
        if not duplicate_samples.empty:
            print(f"⚠️  警告：'{SHEET_NAMES['sample_info']}' 中發現重複的樣本名稱:")
            for idx, row in duplicate_samples.iterrows():
                print(f"     - {row['Sample_Name']}")
            print(f"   建議：請檢查樣本名稱是否正確")

        # ===== 防呆8: 樣本類型檢查 =====
        sample_types = sample_info_df['Sample_Type'].unique()
        print(f"📊 樣本類型: {', '.join([str(t) for t in sample_types])}")

        qc_count = sample_info_df[sample_info_df['Sample_Type'].str.upper().str.contains('QC', na=False)].shape[0]
        if qc_count == 0:
            raise ValueError("未找到 QC 樣本（Sample_Type 中無 'QC' 字樣）。QC-LOWESS 校正需要至少 5 個 QC 樣本")
        elif qc_count < 5:
            print(f"⚠️  警告：QC 樣本數量不足 ({qc_count} < 5)")
            print(f"   提示：建議至少有 5 個 QC 樣本以確保校正準確性")
        else:
            print(f"✓ 找到 {qc_count} 個 QC 樣本")

        # ===== 防呆9: Injection_Order 有效性檢查 =====
        if 'Injection_Order' in sample_info_df.columns:
            invalid_orders = sample_info_df[pd.isna(sample_info_df['Injection_Order'])]
            if not invalid_orders.empty:
                print(f"⚠️  警告：發現 {len(invalid_orders)} 個樣本缺少 Injection_Order:")
                for idx, row in invalid_orders.head(5).iterrows():
                    print(f"     - {row['Sample_Name']}")
                if len(invalid_orders) > 5:
                    print(f"     ... 還有 {len(invalid_orders) - 5} 個樣本")

            # 檢查 Injection_Order 是否為數值
            try:
                sample_info_df['Injection_Order'] = pd.to_numeric(sample_info_df['Injection_Order'], errors='coerce')
                invalid_count = sample_info_df['Injection_Order'].isna().sum()
                if invalid_count > 0:
                    print(f"⚠️  警告：{invalid_count} 個樣本的 Injection_Order 無法轉換為數值")
            except Exception as e:
                print(f"⚠️  警告：Injection_Order 數據類型檢查失敗: {e}")

            # 針對缺失的注射順序提供連續的替補值，避免後續流程中止
            order_na_mask = sample_info_df['Injection_Order'].isna()
            if order_na_mask.any():
                existing_max = sample_info_df['Injection_Order'].max()
                if pd.isna(existing_max):
                    existing_max = 0
                filler = np.arange(1, order_na_mask.sum() + 1) + existing_max
                sample_info_df.loc[order_na_mask, 'Injection_Order'] = filler
                print(f"⚠️  警告：已為缺少 Injection_Order 的樣本指派遞增序號，請於 SampleInfo 中確認")

        # ===== 防呆10: ISTD_Correction 基本檢查 =====
        istd_df = pd.read_excel(excel_file, sheet_name=SHEET_NAMES['istd_correction'])
        print(f"✓ 成功讀取 '{SHEET_NAMES['istd_correction']}' 工作表，包含 {len(istd_df)} 個特徵")

        if istd_df.empty:
            raise ValueError(f"'{SHEET_NAMES['istd_correction']}' 工作表為空")

        # 支援 'Mz/RT' 或 'FeatureID' 作為特徵ID欄位
        if FEATURE_ID_COLUMN in istd_df.columns and FEATURE_ID_COLUMN != 'FeatureID':
            istd_df = istd_df.rename(columns={FEATURE_ID_COLUMN: 'FeatureID'})
        elif 'FeatureID' not in istd_df.columns:
            first_col = istd_df.columns[0]
            print(f"⚠️ 未找到 'FeatureID' 欄位，使用第一欄 '{first_col}' 作為特徵ID")
            istd_df = istd_df.rename(columns={first_col: 'FeatureID'})

        # ===== 提取 Sample_Type 資訊行（不參與數值計算，保存時回插）=====
        from ms_core.utils.data_helpers import extract_sample_type_row
        istd_df, sample_type_row = extract_sample_type_row(istd_df, 'FeatureID')
        if sample_type_row is not None:
            print(f"✓ 偵測到 Sample_Type 資訊行，已提取保存（不參與計算）")

        # ===== 防呆11: FeatureID 重複檢查 =====
        duplicate_features = istd_df[istd_df['FeatureID'].duplicated(keep=False)]
        if not duplicate_features.empty:
            print(f"⚠️  警告：'{SHEET_NAMES['istd_correction']}' 中發現重複的 FeatureID:")
            dup_ids = duplicate_features['FeatureID'].unique()
            for fid in dup_ids[:5]:
                print(f"     - {fid}")
            if len(dup_ids) > 5:
                print(f"     ... 還有 {len(dup_ids) - 5} 個重複的 FeatureID")
            print(f"   建議：請檢查數據是否正確，腳本將保留第一次出現的記錄")

        # ===== 防呆12: 樣本欄位檢查 =====
        sample_columns, dropped_columns = identify_sample_columns(istd_df, sample_info_df)

        if len(sample_columns) == 0:
            raise ValueError(f"'{SHEET_NAMES['istd_correction']}' 中沒有匹配 {SHEET_NAMES['sample_info']} 的樣本欄位")

        print(f"✓ 找到 {len(sample_columns)} 個樣本欄位（來自 SampleInfo）")

        if dropped_columns:
            print(f"⚠️  警告：偵測到 {len(dropped_columns)} 個推定統計欄位，已自動排除：")
            for col in dropped_columns[:5]:
                print(f"     - {col}")
            if len(dropped_columns) > 5:
                print(f"     ... 還有 {len(dropped_columns) - 5} 個欄位")

        # ===== 防呆13: 樣本名稱匹配檢查 =====
        sample_names_in_info = set(sample_info_df['Sample_Name'].astype(str).str.strip().str.lower())
        sample_names_in_istd = {normalize_sample_name(col) for col in sample_columns}

        missing_in_istd = sample_names_in_info - sample_names_in_istd
        missing_in_info = sample_names_in_istd - sample_names_in_info

        if missing_in_istd:
            print(f"⚠️  警告：以下樣本在 SampleInfo 中有記錄，但在 ISTD_Correction 中找不到:")
            for name in list(missing_in_istd)[:5]:
                print(f"     - {name}")
            if len(missing_in_istd) > 5:
                print(f"     ... 還有 {len(missing_in_istd) - 5} 個樣本")

        if missing_in_info:
            print(f"⚠️  警告：以下樣本在 ISTD_Correction 中有數據，但在 SampleInfo 中找不到:")
            for name in list(missing_in_info)[:5]:
                print(f"     - {name}")
            if len(missing_in_info) > 5:
                print(f"     ... 還有 {len(missing_in_info) - 5} 個樣本")

        # ===== 防呆14: 強制轉換樣本欄位為數值 =====
        for col in sample_columns:
            istd_df[col] = pd.to_numeric(istd_df[col], errors='coerce')

        # ===== 防呆15: 全為 NaN 或 0 的列檢查 =====
        empty_columns = []
        for col in sample_columns:
            non_zero_count = (istd_df[col] > 0).sum()
            if non_zero_count == 0:
                empty_columns.append(col)

        if empty_columns:
            print(f"⚠️  警告：以下樣本的所有數值都是 0 或 NaN:")
            for col in empty_columns[:5]:
                print(f"     - {col}")
            if len(empty_columns) > 5:
                print(f"     ... 還有 {len(empty_columns) - 5} 個樣本")

        # 填充 NaN 為 0
        istd_df = istd_df.fillna(0)
        print(f"✓ ISTD_Correction 數據類型檢查：樣本欄位已轉換為數值型")

        # ===== 防呆16: 數值範圍檢查 =====
        negative_count = 0
        extreme_high_count = 0

        for col in sample_columns:
            negative_values = (istd_df[col] < 0).sum()
            if negative_values > 0:
                negative_count += 1
                print(f"⚠️  警告：樣本 '{col}' 包含 {negative_values} 個負值")

            # 檢查極端高值（> 1e12）
            extreme_values = (istd_df[col] > 1e12).sum()
            if extreme_values > 0:
                extreme_high_count += 1
                print(f"⚠️  警告：樣本 '{col}' 包含 {extreme_values} 個極端高值 (> 1e12)")

        if negative_count > 0:
            print(f"   提示：已將負值設為 0")
            for col in sample_columns:
                istd_df[col] = istd_df[col].clip(lower=0)

        # 載入 RawIntensity（可選）
        raw_df = None
        if SHEET_NAMES['raw_intensity'] in excel_file.sheet_names:
            try:
                raw_df = pd.read_excel(excel_file, sheet_name=SHEET_NAMES['raw_intensity'])
                print(f"✓ 已載入 '{SHEET_NAMES['raw_intensity']}' 工作表（可選）")
            except Exception as e:
                print(f"⚠️  警告：無法載入 '{SHEET_NAMES['raw_intensity']}' 工作表: {e}")

        print(f"\n{'='*70}")
        print(f"✓ 數據載入完成")
        print(f"  - 特徵數: {len(istd_df)}")
        print(f"  - 樣本數: {len(sample_info_df)}")
        print(f"  - QC 樣本數: {qc_count}")
        if batch_summary:
            print(f"  - 批次分佈: {', '.join([f'{batch}:{count}' for batch, count in batch_summary.items()])}")
        print(f"{'='*70}\n")

        # 將識別出的樣本欄位保存於 DataFrame attrs，供後續流程使用
        istd_df.attrs['sample_columns'] = sample_columns
        istd_df.attrs['excluded_non_sample_columns'] = dropped_columns

        return raw_df, istd_df, sample_info_df, sample_type_row

    except Exception as e:
        print(f"❌ 載入數據失敗（未預期的錯誤）: {e}")
        import traceback
        traceback.print_exc()
        raise


# ========== ✅ 修正：統計檢定（Levene's test + 整體 Wilcoxon test）==========
def calculate_qc_cv_with_statistical_test(istd_df, lowess_df, sample_columns, sample_info_df, qc_corrected_values):
    """計算 QC CV% 並進行正確的統計檢定"""
    
    istd_df = istd_df.reset_index(drop=True)
    lowess_df = lowess_df.reset_index(drop=True)
    
    min_length = min(len(istd_df), len(lowess_df))
    istd_df = istd_df.iloc[:min_length]
    lowess_df = lowess_df.iloc[:min_length]
    
    print(f"\n{'='*70}")
    print(f"🔬 統計檢定")
    print(f"{'='*70}")
    
    # 識別 QC 樣本（優先從 SampleInfo 查找，若名稱不匹配則從欄位名稱判斷）
    if 'Sample_Type' in sample_info_df.columns:
        qc_names_from_info = sample_info_df[
            sample_info_df['Sample_Type'].str.upper().str.contains('QC', na=False)
        ]['Sample_Name'].tolist()
        qc_columns = [col for col in qc_names_from_info if col in sample_columns]
    else:
        qc_columns = []

    # 如果 SampleInfo 名稱匹配不上，回退到從欄位名稱關鍵字判斷
    if not qc_columns:
        qc_columns = [col for col in sample_columns if 'QC' in col.upper() or 'POOLED' in col.upper()]
    print(f"  - QC 樣本數: {len(qc_columns)}")
    
    if len(qc_columns) == 0:
        print(f"  ❌ 錯誤：未找到任何 QC 樣本！")
        return pd.DataFrame()
    
    cv_results = []
    all_cv_improvements = []  # ✅ 收集所有 CV% 改善值（用於整體評估）
    
    print(f"\n  💡 統計方法:")
    print(f"     • Levene's test: 檢測方差是否顯著改變（單一特徵）")
    print(f"     • Wilcoxon test: 檢測 CV% 是否整體顯著降低（所有特徵）")
    
    for idx in range(len(istd_df)):
        try:
            istd_row = istd_df.iloc[idx]
            lowess_row = lowess_df.iloc[idx]
            feature_id = istd_row['FeatureID']
            
            qc_values_istd = get_valid_values(istd_row, qc_columns)
            
            if feature_id in qc_corrected_values:
                qc_corrected_dict = qc_corrected_values[feature_id]
                qc_values_lowess = []
                for qc in qc_columns:
                    if qc in qc_corrected_dict:
                        val = qc_corrected_dict[qc]
                        if not pd.isna(val) and val > 0:
                            qc_values_lowess.append(val)
            else:
                qc_values_lowess = get_valid_values(lowess_row, qc_columns)
            
            min_len = min(len(qc_values_istd), len(qc_values_lowess))
            if min_len < 3:
                cv_results.append({
                    'FeatureID': feature_id,
                    'Original_QC_CV%': np.nan,
                    'Corrected_QC_CV%': np.nan,
                    'CV_Improvement%': np.nan,
                    'Variance_Test_pvalue': np.nan
                })
                continue
            
            qc_values_istd = np.array(qc_values_istd[:min_len])
            qc_values_lowess = np.array(qc_values_lowess[:min_len])
            
            # ========== 計算 CV% ==========
            original_cv = (np.std(qc_values_istd, ddof=1) / np.mean(qc_values_istd)) * 100
            corrected_cv = (np.std(qc_values_lowess, ddof=1) / np.mean(qc_values_lowess)) * 100
            cv_improvement = original_cv - corrected_cv
            
            # ✅ 收集 CV% 改善值（用於整體評估）
            if not np.isnan(cv_improvement):
                all_cv_improvements.append(cv_improvement)
            
            # ========== ✅ Levene's test（單一特徵）==========
            try:
                levene_stat, levene_pvalue = levene(qc_values_istd, qc_values_lowess)
            except Exception:
                levene_pvalue = np.nan
            
            cv_results.append({
                'FeatureID': feature_id,
                'Original_QC_CV%': original_cv,
                'Corrected_QC_CV%': corrected_cv,
                'CV_Improvement%': cv_improvement,
                'Variance_Test_pvalue': levene_pvalue
            })
            
            if (idx + 1) % 100 == 0:
                print(f"  處理進度: {idx + 1}/{len(istd_df)} features")
        
        except Exception as e:
            print(f"  ⚠️ 處理特徵 {idx} 時發生錯誤: {e}")
            continue
    
    cv_results_df = pd.DataFrame(cv_results)
    
    # ========== ✅ 整體評估：Wilcoxon test ==========
    print(f"\n{'='*70}")
    print(f"📊 整體校正效果評估")
    print(f"{'='*70}")
    
    if len(all_cv_improvements) >= 10:
        try:
            # 檢測 CV% 改善是否整體 > 0
            w_stat, w_pvalue = wilcoxon(all_cv_improvements, alternative='greater')
            
            print(f"\n  ✅ Wilcoxon Signed-Rank Test (整體評估):")
            print(f"     H₀: CV% 改善的中位數 = 0")
            print(f"     H₁: CV% 改善的中位數 > 0")
            print(f"     統計量: {w_stat:.2f}")
            print(f"     P-value: {w_pvalue:.4e}")
            
            if w_pvalue < 0.001:
                print(f"     結論: LOWESS 校正顯著降低了 QC CV% (p < 0.001) ✅✅✅")
            elif w_pvalue < 0.01:
                print(f"     結論: LOWESS 校正顯著降低了 QC CV% (p < 0.01) ✅✅")
            elif w_pvalue < 0.05:
                print(f"     結論: LOWESS 校正顯著降低了 QC CV% (p < 0.05) ✅")
            else:
                print(f"     結論: LOWESS 校正未顯著降低 QC CV% (p ≥ 0.05) ❌")
            
            # 描述性統計
            median_improvement = np.median(all_cv_improvements)
            mean_improvement = np.mean(all_cv_improvements)
            positive_improvements = np.sum(np.array(all_cv_improvements) > 0)
            
            print(f"\n  📊 CV% 改善的描述性統計:")
            print(f"     中位數改善: {median_improvement:.2f}%")
            print(f"     平均改善: {mean_improvement:.2f}%")
            print(f"     改善特徵比例: {positive_improvements}/{len(all_cv_improvements)} ({positive_improvements/len(all_cv_improvements)*100:.1f}%)")
            
        except Exception as e:
            print(f"  ⚠️ 整體評估失敗: {e}")
    else:
        print(f"  ⚠️ 有效特徵數不足 ({len(all_cv_improvements)} < 10)，跳過整體評估")
    
    print(f"{'='*70}\n")
    
    return cv_results_df


# ========== ✅ 修正：P 值分佈圖（只繪製 Levene's test）==========
def plot_pvalue_distribution(cv_results_df, plots_dir, timestamp):
    """繪製 Levene's test p 值分佈圖（含防呆檢查）"""
    try:
        # ===== 防呆1: 輸入數據檢查 =====
        if cv_results_df is None or cv_results_df.empty:
            print("  ⚠️  警告：CV 結果數據為空，無法繪製 p 值分佈圖")
            return

        if 'Variance_Test_pvalue' not in cv_results_df.columns:
            print("  ⚠️  警告：找不到 Variance_Test_pvalue 欄位，無法繪製 p 值分佈圖")
            return

        # ===== 防呆2: 輸出目錄檢查 =====
        if plots_dir is None:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            base_dir = os.path.join(script_dir, 'output', 'QC_LOWESS_plots')
            os.makedirs(base_dir, exist_ok=True)
            plots_dir = os.path.join(base_dir, f"QC_LOWESS_{timestamp}")

        try:
            os.makedirs(plots_dir, exist_ok=True)
        except Exception as e:
            print(f"  ⚠️  警告：無法創建輸出目錄: {e}")
            return

        if not os.access(plots_dir, os.W_OK):
            print(f"  ⚠️  警告：沒有寫入權限到目錄: {plots_dir}")
            return
        
        levene_pvalues = cv_results_df['Variance_Test_pvalue'].dropna()
        
        if len(levene_pvalues) < 10:
            print("  ⚠️ 有效 p 值數量不足，跳過 p 值分佈圖")
            return
        
        # 只繪製 Levene's test 的 p 值分佈
        fig, ax = plt.subplots(1, 1, figsize=(10, 7))
        
        ax.hist(levene_pvalues, bins=20, color='steelblue', edgecolor='black', alpha=0.7)
        ax.axhline(y=len(levene_pvalues)/20, color='red', linestyle='--', linewidth=2,
                   label='Uniform Distribution Expected')
        ax.set_xlabel('P-value (Levene\'s Test)', fontsize=12, fontweight='bold')
        ax.set_ylabel('Frequency', fontsize=12, fontweight='bold')
        ax.set_title('P-value Distribution\n(Variance Homogeneity Test)', 
                     fontsize=14, fontweight='bold')
        ax.legend(fontsize=10)
        ax.grid(True, alpha=0.3, linestyle='--')
        
        # Kolmogorov-Smirnov 檢定
        from scipy.stats import kstest
        ks_stat, ks_pvalue = kstest(levene_pvalues, 'uniform')
        
        textstr = f'Kolmogorov-Smirnov Test:\n'
        textstr += f'Statistic = {ks_stat:.4f}\n'
        textstr += f'P-value = {ks_pvalue:.4f}\n'
        if ks_pvalue > 0.05:
            textstr += 'Result: Uniform ✓'
        else:
            textstr += 'Result: Non-uniform ✗'
        
        ax.text(0.98, 0.97, textstr, transform=ax.transAxes,
                fontsize=10, verticalalignment='top', horizontalalignment='right',
                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
        
        plt.tight_layout()
        
        pvalue_plot_path = os.path.join(plots_dir, f'Pvalue_Distribution_Levene_{timestamp}.png')

        # ===== 防呆: 圖表保存檢查 =====
        try:
            plt.savefig(pvalue_plot_path, dpi=300, bbox_inches='tight')
            plt.close()

            # 驗證文件是否成功保存
            if not os.path.exists(pvalue_plot_path):
                print(f"\n⚠️  警告：P 值分佈圖保存失敗，找不到輸出檔案")
                return
            else:
                plot_size = os.path.getsize(pvalue_plot_path)
                if plot_size == 0:
                    print(f"\n⚠️  警告：P 值分佈圖大小為 0 bytes")
                    return

            print(f"\n✓ P 值分佈圖已儲存: {pvalue_plot_path}")
            print(f"  - 圖表大小: {plot_size / 1024:.2f} KB")
        except Exception as e:
            plt.close()
            print(f"\n⚠️  警告：保存 P 值分佈圖時發生錯誤: {e}")
            return
        print(f"  - Kolmogorov-Smirnov 檢定: KS={ks_stat:.4f}, p={ks_pvalue:.4f}")
        if ks_pvalue > 0.05:
            print(f"  - 結論: p 值分佈接近均勻分佈 ✓")
        else:
            print(f"  - 結論: p 值分佈偏離均勻分佈 ✗")
        
    except Exception as e:
        print(f"  ⚠️ 繪製 p 值分佈圖時發生錯誤: {e}")
        import traceback
        traceback.print_exc()


def plot_lowess_trend_fitting(trend_data_dict, plots_dir, timestamp, max_per_page=6):
    """繪製 LOWESS 擬合趨勢圖（同一特徵一張圖）。

    目的：讓同一個 feature 在不同 batch 的趨勢能直接比較。

    Args:
        trend_data_dict: dict, {(feature_id, batch_name): plot_data}
        plots_dir: 輸出目錄
        timestamp: 時間戳記
        max_per_page: 保留參數（舊版多特徵拼頁用）；新版不使用。
    """
    try:
        if not trend_data_dict:
            print("  ⚠️  警告：沒有趨勢擬合數據可繪製")
            return

        # 確保輸出目錄存在
        if plots_dir is None:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            base_dir = os.path.join(script_dir, 'output', 'QC_LOWESS_plots')
            os.makedirs(base_dir, exist_ok=True)
            plots_dir = os.path.join(base_dir, f"QC_LOWESS_{timestamp}")

        try:
            os.makedirs(plots_dir, exist_ok=True)
        except Exception as e:
            print(f"  ⚠️  警告：無法創建輸出目錄: {e}")
            return

        if not os.access(plots_dir, os.W_OK):
            print(f"  ⚠️  警告：沒有寫入權限到目錄: {plots_dir}")
            return

        # 將數據按 feature 分組（同特徵一張圖）
        feature_grouped = {}
        for (feature_id, batch_name), plot_data in trend_data_dict.items():
            feature_grouped.setdefault(feature_id, []).append((batch_name, plot_data))

        total_features = len(feature_grouped)
        print(f"\n📊 繪製 LOWESS 擬合趨勢圖（同特徵一張）：{total_features} 個特徵...")

        feature_count = 0
        for feature_id, batch_items in feature_grouped.items():
            # 依 batch 名稱排序
            batch_items_sorted = sorted(batch_items, key=lambda x: str(x[0]))
            n_batches = len(batch_items_sorted)

            if n_batches == 0:
                continue

            # 每個 batch 一列：左 Raw+Fit、右 Before vs After
            fig, axes = plt.subplots(
                n_batches,
                2,
                figsize=(14, max(4.2, 4.2 * n_batches)),
                squeeze=False,
            )

            for row_idx, (batch_name, plot_data) in enumerate(batch_items_sorted):
                try:
                    qc_orders = np.array(plot_data['qc_orders'])
                    qc_raw = np.array(plot_data['qc_raw'])
                    qc_corrected = np.array(plot_data['qc_corrected'])
                    lowess_x = np.array(plot_data['lowess_x'])
                    lowess_y = np.array(plot_data['lowess_y'])
                    median_qc = plot_data['median_qc']

                    ax1 = axes[row_idx, 0]
                    ax2 = axes[row_idx, 1]

                    # ===== 左圖：Raw vs LOWESS =====
                    ax1.scatter(
                        qc_orders,
                        qc_raw,
                        c='#0173B2',
                        s=55,
                        alpha=0.75,
                        edgecolors='black',
                        linewidth=0.8,
                        label='QC Raw',
                        zorder=3,
                    )
                    ax1.plot(lowess_x, lowess_y, 'r-', linewidth=2, label='LOWESS Fit', zorder=2)
                    ax1.axhline(
                        y=median_qc,
                        color='green',
                        linestyle='--',
                        linewidth=1.5,
                        label=f'Median={median_qc:.1f}',
                        zorder=1,
                    )
                    ax1.set_xlabel('Injection Order', fontsize=10)
                    ax1.set_ylabel('Intensity', fontsize=10)
                    ax1.set_title(f'Batch: {batch_name}  |  Raw + LOWESS Fit', fontsize=11, fontweight='bold')
                    ax1.legend(fontsize=8, loc='best')
                    ax1.grid(True, alpha=0.3, linestyle='--')

                    # ===== 右圖：Before vs After =====
                    ax2.scatter(
                        qc_orders,
                        qc_raw,
                        c='lightgray',
                        s=55,
                        alpha=0.55,
                        edgecolors='gray',
                        linewidth=0.5,
                        label='Raw',
                        zorder=2,
                    )
                    ax2.scatter(
                        qc_orders,
                        qc_corrected,
                        c='orange',
                        s=55,
                        alpha=0.85,
                        edgecolors='black',
                        linewidth=0.8,
                        label='Corrected',
                        zorder=3,
                    )
                    ax2.axhline(y=median_qc, color='green', linestyle='--', linewidth=1.5, label='Target', zorder=1)
                    ax2.set_xlabel('Injection Order', fontsize=10)
                    ax2.set_ylabel('Intensity', fontsize=10)
                    ax2.set_title('Before vs After Correction', fontsize=11, fontweight='bold')
                    ax2.legend(fontsize=8, loc='best')
                    ax2.grid(True, alpha=0.3, linestyle='--')

                except Exception as e:
                    print(f"  ⚠️  警告：繪製 {feature_id} / {batch_name} 時發生錯誤: {e}")
                    continue

            fig.suptitle(f'LOWESS Trend Fitting (Feature = {feature_id})', fontsize=14, fontweight='bold', y=0.99)
            plt.tight_layout(rect=[0, 0, 1, 0.97])

            safe_feature = str(feature_id).replace('/', '_').replace('\\', '_').replace(':', '_')
            plot_path = os.path.join(plots_dir, f'Trend_Fitting_Feature_{safe_feature}_{timestamp}.png')
            plt.savefig(plot_path, dpi=200, bbox_inches='tight')
            plt.close(fig)

            feature_count += 1
            if os.path.exists(plot_path):
                plot_size = os.path.getsize(plot_path)
                if plot_size > 0:
                    print(f"  ✓ 已保存: {safe_feature} ({n_batches} batches) - {plot_size / 1024:.1f} KB")
                else:
                    print(f"  ⚠️  警告：{safe_feature} 圖表大小為 0 bytes")
            else:
                print(f"  ⚠️  警告：{safe_feature} 圖表保存失敗")

        print(f"✓ LOWESS 擬合趨勢圖繪製完成（共 {feature_count} 張，一特徵一張）")

    except Exception as e:
        print(f"  ⚠️ 繪製 LOWESS 擬合趨勢圖時發生錯誤: {e}")
        import traceback
        traceback.print_exc()


# copy_sheet_with_full_format 已移至 utils/excel_format.py (copy_sheet_with_style)


# ========== ✅ 修正：保存結果到 Excel（移除 Wilcoxon_pvalue）==========
def save_results_to_excel(raw_df, istd_df, lowess_df, sample_info_df, sample_columns,
                          output_file, input_file, qc_corrected_values,
                          trend_stats_df, decision_stats, plots_dir=None, trend_plot_data=None,
                          sample_type_row=None):
    """保存結果到 Excel（含完整防呆檢查）"""
    try:
        # ===== 防呆1: 輸入數據有效性檢查 =====
        if istd_df is None or istd_df.empty:
            print(f"❌ 錯誤：ISTD_Correction 數據為空，無法保存")
            return False

        if lowess_df is None or lowess_df.empty:
            print(f"❌ 錯誤：LOWESS 校正結果為空，無法保存")
            return False

        if sample_info_df is None or sample_info_df.empty:
            print(f"❌ 錯誤：SampleInfo 數據為空，無法保存")
            return False

        # ===== 防呆2: 輸出路徑有效性檢查 =====
        output_dir = os.path.dirname(output_file)
        if not os.path.exists(output_dir):
            print(f"⚠️  警告：輸出目錄不存在，嘗試創建: {output_dir}")
            try:
                os.makedirs(output_dir, exist_ok=True)
                print(f"✓ 成功創建輸出目錄")
            except Exception as e:
                print(f"❌ 錯誤：無法創建輸出目錄: {e}")
                return False

        # ===== 防呆3: 輸出目錄可寫性檢查 =====
        if not os.access(output_dir, os.W_OK):
            print(f"❌ 錯誤：沒有寫入權限到目錄: {output_dir}")
            return False

        # ===== 防呆4: 輸入文件有效性檢查 =====
        if not os.path.exists(input_file):
            print(f"❌ 錯誤：找不到輸入檔案: {input_file}")
            return False
        cv_results_df = calculate_qc_cv_with_statistical_test(
            istd_df, 
            lowess_df, 
            sample_columns, 
            sample_info_df, 
            qc_corrected_values
        )
        
        # ✅ 主表：只保留 Levene's test 和 CV%
        lowess_with_cv = lowess_df.merge(cv_results_df, on='FeatureID', how='left')
        
        # ✅ 主表欄位順序（移除 Wilcoxon_pvalue 和 Significant_Improvement）
        cols_order = [
            'Original_QC_CV%', 
            'Corrected_QC_CV%', 
            'CV_Improvement%',
            'Variance_Test_pvalue'  # 只保留 Levene's test
        ]
        other_cols = [col for col in lowess_with_cv.columns if col not in cols_order]
        lowess_with_cv = lowess_with_cv[other_cols + cols_order]
        
        # ✅ 副表：進階統計指標
        advanced_stats_df = lowess_df[['FeatureID']].merge(
            trend_stats_df, on='FeatureID', how='left'
        )
        
        print(f"\n📋 開始處理 Excel 檔案...")
        print(f"  - 載入原始檔案: {os.path.basename(input_file)}")
        
        def sanitize_excel_df(df):
            if df is None:
                return None
            return df.replace([np.inf, -np.inf], np.nan)

        raw_export = sanitize_excel_df(raw_df) if raw_df is not None else None
        istd_export = sanitize_excel_df(istd_df)
        lowess_export = sanitize_excel_df(lowess_with_cv)
        advanced_export = sanitize_excel_df(advanced_stats_df)
        sample_info_export = sanitize_excel_df(sample_info_df)

        # ===== 回插 Sample_Type 資訊行（若有）=====
        if sample_type_row is not None:
            from ms_core.utils.data_helpers import insert_sample_type_row
            istd_export = insert_sample_type_row(istd_export, sample_type_row)
            lowess_export = insert_sample_type_row(lowess_export, sample_type_row)

        sheets_to_write = []
        if raw_export is not None and not raw_export.empty:
            sheets_to_write.append((SHEET_NAMES['raw_intensity'], raw_export))
        sheets_to_write.extend([
            (SHEET_NAMES['istd_correction'], istd_export),
            (SHEET_NAMES['qc_lowess'], lowess_export),
            (QC_LOWESS_ADVANCED_SHEET, advanced_export),
            (SHEET_NAMES['sample_info'], sample_info_export),
        ])

        # 輸出時將內部欄名 'FeatureID' 還原為 FEATURE_ID_COLUMN
        def _rename_feature_col(df):
            if 'FeatureID' in df.columns and FEATURE_ID_COLUMN != 'FeatureID':
                return df.rename(columns={'FeatureID': FEATURE_ID_COLUMN})
            return df

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for sheet_name, df in sheets_to_write:
                _rename_feature_col(df).to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = load_workbook(output_file)

        # 複製輸入檔的原始格式（保留 ISTD 紅色標記等）
        original_wb = load_workbook(input_file)
        for sheet_name in [SHEET_NAMES['raw_intensity'], SHEET_NAMES['istd_correction']]:
            if sheet_name in original_wb.sheetnames and sheet_name in workbook.sheetnames:
                copy_sheet_formatting_only(original_wb[sheet_name], workbook[sheet_name])
        original_wb.close()

        scientific_format = '0.00E+00'
        
        for sheet_name in [SHEET_NAMES['istd_correction'], SHEET_NAMES['qc_lowess'], QC_LOWESS_ADVANCED_SHEET, SHEET_NAMES['sample_info']]:
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                            if cell.number_format == 'General' or cell.number_format == '0':
                                cell.number_format = scientific_format

        # ✅ 顏色標記（簡化版）
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
        light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

        # 主表顏色標記
        if SHEET_NAMES['qc_lowess'] in workbook.sheetnames:
            worksheet = workbook[SHEET_NAMES['qc_lowess']]
            header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
            
            # CV% 相關欄位 - 橘色
            for col_name in ['Original_QC_CV%', 'Corrected_QC_CV%', 'CV_Improvement%']:
                if col_name in header:
                    col_idx = header.index(col_name) + 1
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            cell.fill = orange_fill
            
            # Levene's test - 淺藍色
            if 'Variance_Test_pvalue' in header:
                col_idx = header.index('Variance_Test_pvalue') + 1
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.fill = light_blue_fill
        
        # 副表顏色標記
        if QC_LOWESS_ADVANCED_SHEET in workbook.sheetnames:
            worksheet = workbook[QC_LOWESS_ADVANCED_SHEET]
            header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
            
            # 所有進階指標 - 淺綠色
            for col_name in ['MK_Trend_pvalue', 'Kendall_Tau', 'LOWESS_R2', 'LOWESS_RMSE',
                             'Frac_Used', 'QC_CV_for_Frac', 'Frac_Strategy']:
                if col_name in header:
                    col_idx = header.index(col_name) + 1
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            cell.fill = light_green_fill

        # ===== 防呆5: 文件保存檢查 =====
        try:
            workbook.save(output_file)
            print(f"✓ 成功保存 Excel 檔案")
        except PermissionError:
            print(f"❌ 錯誤：無法保存檔案，可能檔案已被其他程式開啟")
            print(f"   請關閉檔案後重試: {output_file}")
            workbook.close()
            return False
        except Exception as e:
            print(f"❌ 錯誤：保存檔案時發生錯誤: {e}")
            workbook.close()
            return False

        workbook.close()

        # ===== 防呆6: 文件保存驗證 =====
        if not os.path.exists(output_file):
            print(f"❌ 錯誤：檔案保存失敗，找不到輸出檔案: {output_file}")
            return False

        # 檢查文件大小
        output_size = os.path.getsize(output_file)
        if output_size == 0:
            print(f"❌ 錯誤：輸出檔案大小為 0 bytes")
            return False
        elif output_size < 1024:
            print(f"⚠️  警告：輸出檔案大小異常小 ({output_size} bytes)")

        print(f"✓ 輸出檔案大小: {output_size / 1024:.2f} KB")

        # ✅ 統計報告
        print(f"\n{'='*70}")
        print(f"✓ QC LOWESS 結果已保存:")
        print(f"  {output_file}")
        print(f"{'='*70}")
        
        total_count = len(cv_results_df)
        
        print(f"\n📊 核心統計摘要（主表）:")
        print(f"  - 總特徵數: {total_count}")
        
        # CV% 改善統計
        cv_improvements = cv_results_df['CV_Improvement%'].dropna()
        if len(cv_improvements) > 0:
            print(f"\n  📈 CV% 改善:")
            print(f"     平均: {cv_improvements.mean():.2f}%")
            print(f"     中位數: {cv_improvements.median():.2f}%")
            print(f"     範圍: {cv_improvements.min():.2f}% ~ {cv_improvements.max():.2f}%")
            
            improved = (cv_improvements > 5).sum()
            similar = ((cv_improvements >= -5) & (cv_improvements <= 5)).sum()
            worse = (cv_improvements < -5).sum()
            
            print(f"\n  分類:")
            print(f"     顯著改善 (>5%): {improved} ({improved/len(cv_improvements)*100:.1f}%)")
            print(f"     持平 (±5%): {similar} ({similar/len(cv_improvements)*100:.1f}%)")
            print(f"     變差 (<-5%): {worse} ({worse/len(cv_improvements)*100:.1f}%)")
        
        # Levene's test 統計
        levene_valid = cv_results_df['Variance_Test_pvalue'].notna().sum()
        levene_sig = ((cv_results_df['Variance_Test_pvalue'] < 0.05) & 
                      (cv_results_df['Variance_Test_pvalue'].notna())).sum()
        
        print(f"\n  🔬 Levene's Test（方差齊性）:")
        print(f"     成功執行: {levene_valid}/{total_count} ({levene_valid/total_count*100:.1f}%)")
        if levene_valid > 0:
            print(f"     方差顯著改變 (p < 0.05): {levene_sig}/{levene_valid} ({levene_sig/levene_valid*100:.1f}%)")
        
        # 校正決策統計
        feature_total = decision_stats.get('total_features', total_count)
        feature_success = decision_stats.get('success', 0)
        feature_partial = decision_stats.get('partial_success', 0)
        feature_no_success = max(feature_total - feature_success - feature_partial, 0)

        def pct(value, base):
            return (value / base * 100) if base else 0

        print(f"\n📊 校正決策統計:")
        print(f"  ✅ 全批次成功: {feature_success} ({pct(feature_success, feature_total):.1f}%)")
        print(f"  ⚠️ 部分批次成功: {feature_partial} ({pct(feature_partial, feature_total):.1f}%)")
        print(f"  ❌ 無成功批次: {feature_no_success} ({pct(feature_no_success, feature_total):.1f}%)")

        print(f"\n  無成功批次的主要原因:")
        for key, label in [
            ('insufficient_qc', 'QC 樣本不足'),
            ('insufficient_improvement', 'CV% 改善不足 (<2%)'),
            ('unstable_correction_factors', '校正因子不穩定'),
            ('overcorrection_detected', '檢測到過度校正'),
            ('failed', '其他錯誤')
        ]:
            value = decision_stats.get(key, 0)
            if value:
                print(f"    • {label}: {value}")

        event_counts = decision_stats.get('event_counts')
        if event_counts:
            total_events = decision_stats.get('total_feature_batch_tasks', sum(event_counts.values()))
            print(f"\n  批次層級決策 (feature × batch):")
            for status, count in event_counts.items():
                if count:
                    print(f"    • {status}: {count} ({pct(count, total_events):.1f}%)")

        per_batch_stats = decision_stats.get('per_batch')
        if per_batch_stats:
            print(f"\n  各批次摘要:")
            for batch_name, stats_dict in per_batch_stats.items():
                batch_total = sum(stats_dict.values())
                batch_success = stats_dict.get('success', 0)
                print(f"    • Batch {batch_name}: 成功 {batch_success}/{batch_total} ({pct(batch_success, batch_total):.1f}%)")

        frac_counter = decision_stats.get('frac_usage_counter') or {}
        frac_values_global = decision_stats.get('frac_value_list', [])
        if frac_counter:
            total_features = decision_stats.get('total_features', total_count) or 1
            high_count = frac_counter.get('high_variation', 0)
            medium_count = frac_counter.get('medium_variation', 0)
            low_count = frac_counter.get('low_variation_dynamic', 0)
            other_count = max(total_features - (high_count + medium_count + low_count), 0)
            frac_mean = float(np.nanmean(frac_values_global)) if frac_values_global else np.nan
            frac_median = float(np.nanmedian(frac_values_global)) if frac_values_global else np.nan

            def frac_pct(count):
                return (count / total_features * 100) if total_features else 0

            def fmt_frac_value(value):
                return f"{value:.2f}" if np.isfinite(value) else "N/A"

            print(f"\n  📊 Frac 使用統計：")
            print(f"     - 高變異策略 (frac=0.8): {high_count} ({frac_pct(high_count):.1f}%)")
            print(f"     - 中等變異策略 (frac=0.7): {medium_count} ({frac_pct(medium_count):.1f}%)")
            print(f"     - 低變異動態策略 (frac=0.5~0.75): {low_count} ({frac_pct(low_count):.1f}%)")
            if other_count:
                print(f"     - 其他（資料不足）: {other_count} ({frac_pct(other_count):.1f}%)")
            print(f"     - Frac 平均值: {fmt_frac_value(frac_mean)}")
            print(f"     - Frac 中位數: {fmt_frac_value(frac_median)}")
        
        # Mann-Kendall 趨勢統計（副表）
        mk_valid = trend_stats_df['MK_Trend_pvalue'].notna().sum()
        mk_sig = ((trend_stats_df['MK_Trend_pvalue'] < 0.05) & 
                  (trend_stats_df['MK_Trend_pvalue'].notna())).sum()
        
        print(f"\n📊 進階統計（副表）:")
        print(f"  Mann-Kendall 趨勢檢驗:")
        print(f"    成功執行: {mk_valid}/{total_count} ({mk_valid/total_count*100:.1f}%)")
        if mk_valid > 0:
            print(f"    檢測到顯著趨勢 (p < 0.05): {mk_sig}/{mk_valid} ({mk_sig/mk_valid*100:.1f}%)")
        
        # R² 統計
        r2_valid = trend_stats_df['LOWESS_R2'].notna().sum()
        if r2_valid > 0:
            r2_values = trend_stats_df['LOWESS_R2'].dropna()
            print(f"\n  LOWESS 擬合優度 R²:")
            print(f"    中位數: {np.median(r2_values):.4f}")
            print(f"    平均值: {np.mean(r2_values):.4f}")
        
        print(f"\n💡 提示:")
        print(f"  - 主表 (QC LOWESS result): Levene's test + CV%（單一特徵）")
        print(f"  - 副表 ({QC_LOWESS_ADVANCED_SHEET}): Mann-Kendall + R²/RMSE（進階評估）")
        print(f"  - Frac 參數已依代謝物穩定性與 QC 數量動態調整")
        print(f"  - 新增 Frac_Used / QC_CV_for_Frac / Frac_Strategy 可於 {QC_LOWESS_ADVANCED_SHEET} 交叉檢視")
        print(f"  - 整體評估: Wilcoxon test 已在終端機顯示")
        print(f"\n{'='*70}\n")
        
        # P 值分佈圖 (disabled: provides limited diagnostic value)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        # plot_pvalue_distribution(cv_results_df, plots_dir, timestamp)

        # LOWESS 擬合趨勢圖（僅限 debug 特徵）
        if trend_plot_data:
            plot_lowess_trend_fitting(trend_plot_data, plots_dir, timestamp)

        return True

    except Exception as e:
        print(f"❌ 儲存 Excel 時發生錯誤: {e}")
        import traceback
        traceback.print_exc()
        return False


# ========== Hotelling T² 異常值檢測 ==========
def calculate_hotelling_t2_outliers(qc_scores, all_scores=None, alpha=0.05):
    """
    使用 Hotelling T² 檢測 QC 樣本中的異常值
    
    ✅ 正確邏輯：計算每個 QC 樣本與 QC 群組中心的偏離
    """
    n_qc, p = qc_scores.shape
    
    if n_qc < 3:
        print(f"   ⚠️ QC 樣本數不足 ({n_qc} < 3)，無法進行異常值檢測")
        return np.zeros(n_qc), 0, np.zeros(n_qc, dtype=bool)
    
    # ✅ 只使用 QC 群組的統計量
    qc_mean = np.mean(qc_scores, axis=0)
    qc_cov = np.cov(qc_scores, rowvar=False)
    
    # 正則化協方差矩陣
    qc_cov_reg = qc_cov + np.eye(p) * 1e-6
    
    try:
        qc_cov_inv = np.linalg.inv(qc_cov_reg)
    except np.linalg.LinAlgError:
        print("   ⚠️ 警告：QC 協方差矩陣奇異，使用偽逆矩陣")
        qc_cov_inv = np.linalg.pinv(qc_cov_reg)
    
    # ✅ 計算每個 QC 樣本與 QC 中心的 Hotelling T² 值
    t2_values = np.zeros(n_qc)
    for i in range(n_qc):
        diff = qc_scores[i] - qc_mean
        t2_values[i] = np.dot(np.dot(diff, qc_cov_inv), diff.T)
    
    # 計算閾值
    if n_qc - p - 1 > 0:
        f_critical = stats.f.ppf(1 - alpha, p, n_qc - p - 1)
        threshold = (p * (n_qc + 1) * (n_qc - 1)) / (n_qc * (n_qc - p - 1)) * f_critical
    else:
        threshold = chi2.ppf(1 - alpha, p)
    
    outliers = t2_values > threshold
    
    return t2_values, threshold, outliers


# ========== Hotelling T² 橢圓繪製 ==========
def draw_hotelling_t2_ellipse(ax, scores, alpha=0.05, label=None, edgecolor='black', linestyle='-', linewidth=2.5):
    """在 2D PCA 圖上繪製 Hotelling T² 橢圓"""
    n, p = scores.shape
    
    if n < 3:
        print(f"   ⚠️ 樣本數不足 ({n})，無法繪製 Hotelling T² 橢圓")
        return None
    
    mean = np.mean(scores, axis=0)
    cov = np.cov(scores, rowvar=False)
    
    eigenvalues, eigenvectors = np.linalg.eigh(cov)
    
    idx = eigenvalues.argsort()[::-1]
    eigenvalues = eigenvalues[idx]
    eigenvectors = eigenvectors[:, idx]
    
    eigenvalues = np.maximum(eigenvalues, 1e-10)
    
    f_critical = stats.f.ppf(1 - alpha, p, n - p)
    scale_factor = np.sqrt((p * (n - 1) * (n + 1)) / (n * (n - p)) * f_critical)
    
    width = 2 * scale_factor * np.sqrt(eigenvalues[0])
    height = 2 * scale_factor * np.sqrt(eigenvalues[1])
    
    angle = np.degrees(np.arctan2(eigenvectors[1, 0], eigenvectors[0, 0]))
    
    ellipse = Ellipse(mean, width, height, angle=angle,
                     facecolor='none', edgecolor=edgecolor,
                     linewidth=linewidth, linestyle=linestyle, label=label)
    ax.add_patch(ellipse)
    
    t = np.linspace(0, 2*np.pi, 100)
    ellipse_x = (width/2) * np.cos(t)
    ellipse_y = (height/2) * np.sin(t)
    
    cos_angle = np.cos(np.radians(angle))
    sin_angle = np.sin(np.radians(angle))
    x_rot = ellipse_x * cos_angle - ellipse_y * sin_angle + mean[0]
    y_rot = ellipse_x * sin_angle + ellipse_y * cos_angle + mean[1]
    
    bounds = (np.min(x_rot), np.max(x_rot), np.min(y_rot), np.max(y_rot))
    
    return bounds


# ========== PCA 分析 ==========
def perform_pca_analysis(istd_df, lowess_df, sample_columns, sample_info_df,
                         plots_dir=None, grouping='batch'):
    """繪製與 Batch_Effect 相同風格的 PCA 比較圖 (ISTD vs QC-LOWESS)。"""
    try:
        # ===== 防呆1: 輸入數據有效性檢查 =====
        if istd_df is None or istd_df.empty:
            print(f"❌ 錯誤：ISTD_Correction 數據為空，無法進行 PCA 分析")
            return

        if lowess_df is None or lowess_df.empty:
            print(f"❌ 錯誤：LOWESS 校正結果為空，無法進行 PCA 分析")
            return

        if sample_info_df is None or sample_info_df.empty:
            print(f"❌ 錯誤：SampleInfo 數據為空，無法進行 PCA 分析")
            return

        if not sample_columns or len(sample_columns) == 0:
            print(f"❌ 錯誤：樣本欄位為空，無法進行 PCA 分析")
            return
        # ===== 防呆2: 輸出目錄設置和檢查 =====
        if plots_dir is None:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            base_dir = os.path.join(script_dir, "output", "QC_LOWESS_plots")
            os.makedirs(base_dir, exist_ok=True)
            plots_dir = os.path.join(base_dir, f"QC_LOWESS_{datetime.now().strftime('%Y%m%d_%H%M%S')}")

        try:
            os.makedirs(plots_dir, exist_ok=True)
        except Exception as e:
            print(f"❌ 錯誤：無法創建輸出目錄: {e}")
            return

        if not os.access(plots_dir, os.W_OK):
            print(f"❌ 錯誤：沒有寫入權限到目錄: {plots_dir}")
            return

        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        sample_meta = sample_info_df.set_index('Sample_Name')

        sample_columns_attr = istd_df.attrs.get('sample_columns')
        if not sample_columns_attr:
            sample_columns_attr, _ = identify_sample_columns(istd_df, sample_info_df)

        sample_columns_clean = [col for col in sample_columns_attr
                                if col in istd_df.columns and col in lowess_df.columns]
        
        print(f"\n📊 PCA 數據準備:")
        print(f"   - 用於 PCA 的樣本數: {len(sample_columns_clean)}")
        
        if len(sample_columns_clean) < 3:
            print("❌ 錯誤：可用樣本數不足 (<3)，無法進行 PCA 分析")
            return

        # 識別樣本類型與批次（使用 normalize_sample_type 統一分類）
        qc_columns = []
        control_columns = []
        exposed_columns = []
        sample_batches = {}
        sample_type_map = {}

        for col in sample_columns_clean:
            if col in sample_meta.index:
                raw_type = str(sample_meta.loc[col].get('Sample_Type', 'Unknown'))
                batch_value = str(sample_meta.loc[col].get('Batch', 'Unknown'))
            else:
                raw_type = 'Unknown'
                batch_value = 'Unknown'

            norm_type = normalize_sample_type(raw_type)
            sample_type_map[col] = norm_type
            sample_batches[col] = batch_value

            if norm_type == 'QC':
                qc_columns.append(col)
            elif norm_type == 'Exposure':
                exposed_columns.append(col)
            elif norm_type in ('Control', 'Normal'):
                control_columns.append(col)
        

        
        print(f"\n📋 樣本分類:")
        print(f"   - QC: {len(qc_columns)}")
        print(f"   - Control: {len(control_columns)}")
        print(f"   - Exposed: {len(exposed_columns)}")

        if len(qc_columns) < 3:
            print("⚠️ 警告：QC 樣本不足 (<3)，跳過 PCA 分析")
            return

        # 準備數據矩陣
        def prepare_data_matrix(df, columns):
            data_matrix = df[columns].T.values
            data_matrix = np.where(np.isnan(data_matrix), 0, data_matrix)
            data_matrix = np.where(np.isinf(data_matrix), 0, data_matrix)
            data_matrix = np.where(data_matrix < 0, 0, data_matrix)
            non_zero_features = np.any(data_matrix != 0, axis=0)
            data_matrix = data_matrix[:, non_zero_features]
            data_matrix = np.log2(data_matrix + 1)
            return data_matrix

        istd_matrix = prepare_data_matrix(istd_df, sample_columns_clean)
        lowess_matrix = prepare_data_matrix(lowess_df, sample_columns_clean)
        
        if istd_matrix.shape[1] < 2 or lowess_matrix.shape[1] < 2:
            print("❌ 錯誤：有效特徵數不足 (<2)，無法進行 PCA 分析")
            return

        # PCA
        scaler_istd = StandardScaler()
        scaler_lowess = StandardScaler()
        
        istd_scaled = scaler_istd.fit_transform(istd_matrix)
        lowess_scaled = scaler_lowess.fit_transform(lowess_matrix)

        pca_istd = PCA(n_components=2)
        pca_lowess = PCA(n_components=2)
        
        scores_istd = pca_istd.fit_transform(istd_scaled)
        scores_lowess = pca_lowess.fit_transform(lowess_scaled)

        var_istd = pca_istd.explained_variance_ratio_
        var_lowess = pca_lowess.explained_variance_ratio_

        # Hotelling T² 異常值檢測
        qc_indices = [i for i, col in enumerate(sample_columns_clean) if col in qc_columns]
        qc_scores_istd = scores_istd[qc_indices]
        qc_scores_lowess = scores_lowess[qc_indices]

        t2_istd, t2_threshold_istd, outliers_istd = calculate_hotelling_t2_outliers(
            qc_scores_istd, scores_istd, alpha=0.05
        )
        t2_lowess, t2_threshold_lowess, outliers_lowess = calculate_hotelling_t2_outliers(
            qc_scores_lowess, scores_lowess, alpha=0.05
        )

        qc_outlier_map_istd = {qc_columns[i]: bool(outliers_istd[i]) for i in range(len(qc_columns))}
        qc_outlier_map_lowess = {qc_columns[i]: bool(outliers_lowess[i]) for i in range(len(qc_columns))}

        print("\n🔍 Hotelling T² 異常值檢測：")
        print(f"   ISTD: {np.sum(outliers_istd)}/{len(qc_columns)} QC 被標記為異常")
        print(f"   QC-LOWESS: {np.sum(outliers_lowess)}/{len(qc_columns)} QC 被標記為異常")

        # 繪製 2D PCA 圖（統一為 QC 子程式風格的共用函式）
        print("\n🎨 繪製 2D PCA Score Plot...")
        grouping_tag = 'batch' if grouping == 'batch' else 'sample_type'
        pca_plot_path = os.path.join(plots_dir, f'2D_PCA_ISTD_vs_LOWESS_{grouping_tag}_{timestamp}.png')

        suptitle = (
            '2D PCA Comparison: ISTD vs QC-LOWESS (Grouped by Batch)'
            if grouping == 'batch'
            else '2D PCA Comparison: ISTD vs QC-LOWESS (Grouped by Sample Type)'
        )

        sample_types = [sample_type_map.get(col, 'Unknown') for col in sample_columns_clean]
        batch_labels = [sample_batches.get(col, 'Unknown') for col in sample_columns_clean]

        qc_outliers_left = {name for name, is_out in qc_outlier_map_istd.items() if is_out}
        qc_outliers_right = {name for name, is_out in qc_outlier_map_lowess.items() if is_out}

        plot_pca_comparison_qc_style(
            scores_istd,
            scores_lowess,
            var_istd,
            var_lowess,
            sample_columns_clean,
            sample_types,
            batch_labels=batch_labels,
            grouping=grouping_tag,
            suptitle=suptitle,
            left_title='ISTD Corrected',
            right_title='QC-LOWESS Normalized',
            left_threshold_text=f'Hotelling T² Threshold: {t2_threshold_istd:.2f}',
            right_threshold_text=f'Hotelling T² Threshold: {t2_threshold_lowess:.2f}',
            qc_outlier_names_left=qc_outliers_left,
            qc_outlier_names_right=qc_outliers_right,
            output_path=pca_plot_path,
            dpi=300,
        )

        # ===== 防呆3: 圖表保存檢查 =====
        try:
            print(f"   ✓ 2D PCA 圖表已保存 ({grouping_tag} 分類): {pca_plot_path}")

            # 驗證文件是否成功保存
            if not os.path.exists(pca_plot_path):
                print(f"   ⚠️  警告：PCA 圖表保存失敗，找不到輸出檔案")
            else:
                plot_size = os.path.getsize(pca_plot_path)
                if plot_size == 0:
                    print(f"   ⚠️  警告：PCA 圖表大小為 0 bytes")
                else:
                    print(f"   ✓ PCA 圖表大小: {plot_size / 1024:.2f} KB")

        except Exception as e:
            print(f"   ⚠️  警告：保存 PCA 圖表時發生錯誤: {e}")

        # 統計摘要
        print(f"\n{'='*70}")
        print(f"📊 PCA 分析完成")
        print(f"{'='*70}")
        print(f"  解釋變異量:")
        print(f"    ISTD: PC1={var_istd[0]*100:.2f}%, PC2={var_istd[1]*100:.2f}%")
        print(f"    LOWESS: PC1={var_lowess[0]*100:.2f}%, PC2={var_lowess[1]*100:.2f}%")
        print(f"  異常值: ISTD={np.sum(outliers_istd)}, LOWESS={np.sum(outliers_lowess)}")

    except Exception as e:
        print(f"❌ PCA 分析失敗: {e}")
        import traceback
        traceback.print_exc()


# ========== 主程式 ==========
def main(input_file=None):
    """主程式入口"""
    print("="*70)
    print("🔬 QC-LOWESS 批次效應校正工具 v3")
    print("   ✅ 簡化校正邏輯：CV% 改善 ≥ 2%")
    print("   ✅ 統計方法：Levene's test（單一特徵）+ Wilcoxon test（整體評估）")
    print("   ✅ 進階統計：Mann-Kendall + R²/RMSE（副表）")
    print("="*70)
    
    if input_file is None:
        raise ValueError("input_file is required; GUI must provide the file path.")

    output_dir = get_output_root()
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
        print(f"\n✓ 已建立 'output' 資料夾: {output_dir}")
    
    if input_file:
        file_path = input_file
        print(f"\n📂 使用傳入的檔案: {os.path.basename(file_path)}")
    else:
        root = tk.Tk()
        root.withdraw()
        
        file_path = filedialog.askopenfilename(
            title="選擇 Excel 檔案",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not file_path:
            print("❌ 未選擇檔案，程式結束")
            return
        
        print(f"\n📂 選擇的檔案: {os.path.basename(file_path)}")
    
    print(f"\n{'='*70}")
    print(f"📥 載入數據...")
    print(f"{'='*70}")
    
    raw_df, istd_df, sample_info_df, sample_type_row = load_and_process_data(file_path)

    print(f"\n{'='*70}")
    print(f"🔧 執行 QC-LOWESS 校正...")
    print(f"{'='*70}")

    lowess_df, sample_columns, qc_corrected_values, trend_stats_df, decision_stats, trend_plot_data = (
        perform_lowess_normalization(istd_df, sample_info_df)
    )

    print(f"\n{'='*70}")
    print(f"💾 保存結果...")
    print(f"{'='*70}")

    timestamp = datetime.now().strftime(DATETIME_FORMAT_FULL)
    output_file = build_output_path("QC_LOWESS", timestamp=timestamp)
    plots_session_dir = build_plots_dir(
        "QC_LOWESS_plots",
        timestamp=timestamp,
        session_prefix="QC_LOWESS"
    )

    success = save_results_to_excel(
        raw_df, istd_df, lowess_df, sample_info_df,
        sample_columns, output_file, file_path,
        qc_corrected_values, trend_stats_df, decision_stats,
        plots_dir=plots_session_dir, trend_plot_data=trend_plot_data,
        sample_type_row=sample_type_row
    )
    
    if not success:
        print("❌ 結果保存失敗")
        return
    
    print(f"\n{'='*70}")
    print(f"📊 執行 PCA 分析...")
    print(f"{'='*70}")
    
    perform_pca_analysis(
        istd_df, lowess_df, sample_columns, sample_info_df,
        plots_session_dir, grouping='batch'
    )
    perform_pca_analysis(
        istd_df, lowess_df, sample_columns, sample_info_df,
        plots_session_dir, grouping='sample_type'
    )
    
    print(f"\n{'='*70}")
    print(f"✅ 所有分析完成！")
    print(f"{'='*70}")
    print(f"\n📁 輸出內容:")
    print(f"  - Excel 結果: output/{os.path.basename(output_file)}")
    print(f"    ├── ISTD_Correction（保留原格式）")
    print(f"    ├── QC LOWESS result（主表：Levene's test + CV%）")
    print(f"    ├── {QC_LOWESS_ADVANCED_SHEET}（副表：Mann-Kendall + R²/RMSE）")
    print(f"    └── SampleInfo")
    print(f"\n  - 圖表輸出: {plots_session_dir}")
    print(f"    ├── 2D_PCA_ISTD_vs_LOWESS_*.png")
    print(f"    └── Pvalue_Distribution_Levene_*.png")
    print(f"\n  💡 統計方法:")
    print(f"    - Levene's test: 檢測單一特徵方差變化")
    print(f"    - Wilcoxon test: 檢測整體 CV% 是否顯著降低（終端機顯示）")
    print(f"\n{'='*70}\n")
    
    metabolites_count = len(lowess_df)
    samples_count = len(sample_columns)
    
    return ProcessingResult(
        file_path=file_path,
        output_path=str(output_file),
        plots_dir=str(plots_session_dir),
        metabolites=metabolites_count,
        samples=samples_count
    )


if __name__ == "__main__":
    main()
