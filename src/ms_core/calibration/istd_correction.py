import pandas as pd
import numpy as np
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
import matplotlib.pyplot as plt
from matplotlib.patches import Ellipse
import scipy.stats as stats
from scipy.stats import chi2
import warnings

warnings.filterwarnings('ignore')

# ========== 匯入共用模組 ==========
from ms_core.utils.data_helpers import get_valid_values
from ms_core.utils.plotting import setup_matplotlib, plot_pca_comparison_qc_style
from ms_core.utils.constants import FONT_SIZES, COLORBLIND_COLORS, SHEET_NAMES, DATETIME_FORMAT_FULL, FEATURE_ID_COLUMN, VALIDATION_THRESHOLDS, CV_QUALITY_THRESHOLDS
from ms_core.utils.sample_classification import SampleClassifier, identify_sample_columns
from ms_core.utils.file_io import build_output_path, build_plots_dir, get_output_root
from ms_core.utils.results import ProcessingResult
from ms_core.utils.excel_format import copy_sheet_formatting_only
from ms_core.utils.console import safe_print as print

import re as _re

# 設定 matplotlib
setup_matplotlib()


def simplify_column_name(name):
    """Remove redundant DNA/RNA_programN_ prefix from column names.

    Example: 'DNA_program1_TumorBC2257_DNA' → 'TumorBC2257_DNA'
    """
    return _re.sub(r'^(?:DNA|RNA)_program\d+_', '', str(name))

def load_and_process_data(file_path):
    try:
        # ===== 防呆1: 文件存在性检查 =====
        if not os.path.exists(file_path):
            raise ValueError(f"錯誤：找不到檔案 '{file_path}'")

        # ===== 防呆2: 文件格式检查 =====
        if not (file_path.endswith('.xlsx') or file_path.endswith('.xls')):
            raise ValueError(f"錯誤：輸入檔案必須是Excel格式 (.xlsx 或 .xls)，但提供了 {file_path}")

        # ===== 防呆3: 文件大小检查 =====
        file_size = os.path.getsize(file_path)
        if file_size == 0:
            raise ValueError("錯誤：檔案大小為 0 bytes，可能是空檔案")
        elif file_size < 1024:  # 小于 1KB
            print(f"警告：檔案大小僅 {file_size} bytes，可能不是有效的 Excel 檔案")

        print(f"檔案大小: {file_size / 1024:.2f} KB")

        # ===== 防呆4: Excel 文件有效性检查 =====
        try:
            excel_file = pd.ExcelFile(file_path)
        except Exception as e:
            raise ValueError(f"錯誤：無法讀取 Excel 檔案，可能已損壞或格式不正確。詳細錯誤: {e}") from e
        
        # 讀取所有工作表，儲存為字典 {sheet_name: df}
        all_sheets = {sheet: pd.read_excel(excel_file, sheet_name=sheet) for sheet in excel_file.sheet_names}
        print(f"讀取輸入檔案的所有工作表: {list(all_sheets.keys())}")
        
        # ===== 防呆5: 必要工作表检查 =====
        required_sheets = [SHEET_NAMES['raw_intensity'], SHEET_NAMES['sample_info']]
        missing_sheets = [sheet for sheet in required_sheets if sheet not in all_sheets]
        if missing_sheets:
            raise ValueError(f"錯誤：輸入檔案缺少必要的工作表: {', '.join(missing_sheets)}。找到的工作表: {', '.join(all_sheets.keys())}")

        # ===== 防呆6: SampleInfo 完整性检查 =====
        sample_info_df = all_sheets[SHEET_NAMES['sample_info']]
        print(f"成功讀取 '{SHEET_NAMES['sample_info']}' 工作表，包含 {len(sample_info_df)} 筆樣本資訊")

        if sample_info_df.empty:
            raise ValueError(f"錯誤：'{SHEET_NAMES['sample_info']}' 工作表為空")

        required_columns = ['Sample_Name', 'Sample_Type']
        missing_cols = [col for col in required_columns if col not in sample_info_df.columns]
        if missing_cols:
            raise ValueError(f"錯誤：'{SHEET_NAMES['sample_info']}' 缺少必要欄位: {', '.join(missing_cols)}。找到的欄位: {', '.join(sample_info_df.columns.tolist())}")

        # ===== 防呆7: 样本名称重复检查 =====
        duplicate_samples = sample_info_df[sample_info_df['Sample_Name'].duplicated()]
        if not duplicate_samples.empty:
            print(f"警告：'{SHEET_NAMES['sample_info']}' 中發現重複的樣本名稱:")
            for idx, row in duplicate_samples.iterrows():
                print(f"  - {row['Sample_Name']}")
            print(f"  建議：請檢查樣本名稱是否正確")

        # ===== 防呆8: 样本类型检查 =====
        sample_types = sample_info_df['Sample_Type'].unique()
        print(f"樣本類型: {', '.join([str(t) for t in sample_types])}")

        qc_count = sample_info_df[sample_info_df['Sample_Type'].str.upper().str.contains('QC', na=False)].shape[0]
        if qc_count == 0:
            print(f"警告：未找到 QC 樣本（Sample_Type 中無 'QC' 字樣）")
            print(f"  部分統計分析可能無法執行")
        else:
            print(f"找到 {qc_count} 個 QC 樣本")
        
        workbook = load_workbook(file_path)
        worksheet = workbook[SHEET_NAMES['raw_intensity']]
        istd_feature_ids = []  # 收集紅色 FeatureID 的值
        red_colors = ['FFFF0000', 'FF0000']  # 只檢查紅色變體，全大寫
        for row in worksheet.iter_rows(min_row=2):  # 從第 2 行開始
            cell = row[0]  # 第一欄 (FeatureID)
            if cell.font and cell.font.color and cell.font.color.rgb is not None:
                rgb_str = str(cell.font.color.rgb).upper()  # 強制轉 str 並 upper
                if rgb_str in red_colors:
                    if cell.value:  # 確保有值
                        istd_feature_ids.append(str(cell.value).strip())  # 轉 str 以匹配
        workbook.close()
        
        raw_df = all_sheets[SHEET_NAMES['raw_intensity']]

        # ===== 防呆9: RawIntensity 基本检查 =====
        if raw_df.empty:
            raise ValueError(f"錯誤：'{SHEET_NAMES['raw_intensity']}' 工作表為空")

        # 支援 'Mz/RT' 或 'FeatureID' 作為特徵ID欄位名稱
        if FEATURE_ID_COLUMN in raw_df.columns and FEATURE_ID_COLUMN != 'FeatureID':
            raw_df = raw_df.rename(columns={FEATURE_ID_COLUMN: 'FeatureID'})
        elif 'FeatureID' not in raw_df.columns:
            # 嘗試使用第一欄作為特徵ID
            first_col = raw_df.columns[0]
            print(f"⚠️ 未找到 '{FEATURE_ID_COLUMN}' 或 'FeatureID' 欄位，使用第一欄 '{first_col}' 作為特徵ID")
            raw_df = raw_df.rename(columns={first_col: 'FeatureID'})

        # ===== 防呆10: FeatureID 重复检查 =====
        duplicate_features = raw_df[raw_df['FeatureID'].duplicated(keep=False)]
        if not duplicate_features.empty:
            print(f"警告：'{SHEET_NAMES['raw_intensity']}' 中發現重複的 FeatureID:")
            dup_ids = duplicate_features['FeatureID'].unique()
            for fid in dup_ids[:5]:  # 只显示前5个
                print(f"  - {fid}")
            if len(dup_ids) > 5:
                print(f"  ... 還有 {len(dup_ids) - 5} 個重複的 FeatureID")
            print(f"  建議：請檢查數據是否正確，腳本將保留第一次出現的記錄")

        # ===== 防呆11: 样本列检查 =====
        sample_columns = [col for col in raw_df.columns if col != 'FeatureID']
        if len(sample_columns) == 0:
            raise ValueError(f"錯誤：'{SHEET_NAMES['raw_intensity']}' 中沒有樣本欄位")

        print(f"找到 {len(sample_columns)} 個樣本欄位")

        # 修改：不要自動跳過 'sample_type'，改為檢查並保留
        if not raw_df.empty and str(raw_df.iloc[0]['FeatureID']).strip().lower() == 'sample_type':
            print("偵測到 'Sample_Type' 資訊行，已保留作為元數據。")

        # ===== 防呆12: 样本名称匹配检查（支援模糊匹配）=====
        import re
        from ms_core.utils.sample_classification import normalize_sample_type, normalize_sample_name

        def _extract_key_tokens(name):
            """從樣本名稱中提取關鍵字和編號用於模糊匹配。
            例如 'DNA_program1_TumorBC2257_DNA' 和 'Tumor tissue BC2257_DNA'
            都會提取出類似的 token 集合。"""
            s = str(name).strip()
            # 移除常見前綴 (保留原始大小寫做 camelCase 拆分)
            s = re.sub(r'^(?:DNA|RNA|dna|rna)_program\d+_', '', s)
            # camelCase 拆分：'TumorBC2257' → 'Tumor BC 2257'
            s = re.sub(r'([a-z])([A-Z])', r'\1 \2', s)
            s = re.sub(r'([A-Z]+)([A-Z][a-z])', r'\1 \2', s)
            s = s.lower()
            # 用空白、底線、分隔符拆分
            parts = re.split(r'[\s_\-/]+', s)
            tokens = set()
            for part in parts:
                # 將合併字拆開：'bc2257' → 'bc', '2257'
                sub_tokens = re.findall(r'[a-z]+|[0-9]+', part)
                tokens.update(sub_tokens)
                # 也保留「字母+數字」的組合 token（如 bc2257）
                combo_tokens = re.findall(r'[a-z]+\d+', part)
                tokens.update(combo_tokens)
            # 過濾掉純通用詞
            generic = {'tissue', 'cancer', 'breast', 'pooled', 'fat',
                        'dna', 'rna', 'dnaandrna', 'program1', 'and'}
            return tokens - generic

        sample_names_in_info = set(sample_info_df['Sample_Name'].str.strip().str.lower())
        sample_names_in_raw = set([col.strip().lower() for col in sample_columns])

        # 先嘗試精確匹配
        missing_in_raw = sample_names_in_info - sample_names_in_raw
        missing_in_info = sample_names_in_raw - sample_names_in_info

        if missing_in_raw and missing_in_info:
            # 嘗試模糊匹配：提取關鍵 token（類別+編號），若交集夠大則視為匹配
            fuzzy_matched = 0
            unmatched_info = []
            unmatched_raw = list(missing_in_info)
            raw_tokens_map = {name: _extract_key_tokens(name) for name in unmatched_raw}

            for info_name in missing_in_raw:
                info_tokens = _extract_key_tokens(info_name)
                best_match = None
                best_overlap = 0
                for raw_name, raw_tokens in raw_tokens_map.items():
                    overlap = len(info_tokens & raw_tokens)
                    if overlap > best_overlap:
                        best_overlap = overlap
                        best_match = raw_name
                # 至少有 2 個 token 匹配（類別+編號）才算成功
                if best_overlap >= 2 and best_match:
                    fuzzy_matched += 1
                    raw_tokens_map.pop(best_match)
                else:
                    unmatched_info.append(info_name)

            if fuzzy_matched > 0:
                print(f"✓ 樣本名稱模糊匹配成功：{fuzzy_matched} 個樣本（SampleInfo 與 RawIntensity 名稱格式不同但關鍵字匹配）")

            if unmatched_info:
                print(f"⚠️ 警告：以下 {len(unmatched_info)} 個樣本在 SampleInfo 中有記錄，但無法匹配到 RawIntensity:")
                for name in unmatched_info[:5]:
                    print(f"  - {name}")
            if raw_tokens_map:
                print(f"⚠️ 警告：以下 {len(raw_tokens_map)} 個樣本在 RawIntensity 中有數據，但無法匹配到 SampleInfo:")
                for name in list(raw_tokens_map.keys())[:5]:
                    print(f"  - {name}")
        else:
            if missing_in_raw:
                print(f"⚠️ 警告：以下樣本在 SampleInfo 中有記錄，但在 RawIntensity 中找不到:")
                for name in list(missing_in_raw)[:5]:
                    print(f"  - {name}")
                if len(missing_in_raw) > 5:
                    print(f"  ... 還有 {len(missing_in_raw) - 5} 個樣本")
            if missing_in_info:
                print(f"⚠️ 警告：以下樣本在 RawIntensity 中有數據，但在 SampleInfo 中找不到:")
                for name in list(missing_in_info)[:5]:
                    print(f"  - {name}")
                if len(missing_in_info) > 5:
                    print(f"  ... 還有 {len(missing_in_info) - 5} 個樣本")

        # 防呆：強制轉換 RawIntensity 的樣本欄位為數值 (向量化)
        raw_df[sample_columns] = raw_df[sample_columns].apply(pd.to_numeric, errors='coerce')

        # ===== 防呆13: 全为 NaN 或 0 的列检查 (向量化) =====
        non_zero_counts = (raw_df[sample_columns] > 0).sum()
        zero_cols = non_zero_counts[non_zero_counts == 0].index.tolist()
        for col in zero_cols:
            print(f"警告：樣本 '{col}' 的所有數值都是 0 或 NaN")

        raw_df = raw_df.fillna(0)  # 填充 NaN 為 0
        print(f"RawIntensity 數據類型檢查：樣本欄位已轉換為數值型")

        # ===== 防呆16: 数值范围检查 (向量化) =====
        # Check for negative values across all columns at once
        negative_mask = raw_df[sample_columns] < 0
        negative_counts_per_col = negative_mask.sum()
        negative_count = negative_counts_per_col.sum()

        for col in negative_counts_per_col[negative_counts_per_col > 0].index:
            print(f"⚠️ 警告：樣本 '{col}' 有 {negative_counts_per_col[col]} 個負值，已設為 0")

        if negative_count > 0:
            # Clip all columns at once (vectorized)
            raw_df[sample_columns] = raw_df[sample_columns].clip(lower=0)
            print(f"總計修正了 {negative_count} 個負值")

        # 检查极端高值（可能是数据错误）
        max_values = raw_df[sample_columns].max()
        extreme_cols = max_values[max_values > 1e15]
        for col, max_val in extreme_cols.items():
            print(f"⚠️ 警告：樣本 '{col}' 有極端高值 ({max_val:.2e})，請檢查數據是否正確")

        if 'FeatureID' in raw_df.columns:
            def parse_feature_id(fid):
                if isinstance(fid, str) and '/' in fid:
                    parts = fid.split('/')
                    if len(parts) >= 2:
                        return parts[0], parts[1]
                return np.nan, np.nan
            
            parsed = raw_df['FeatureID'].apply(parse_feature_id)
            raw_df['mz'] = pd.to_numeric([p[0] for p in parsed], errors='coerce')
            raw_df['rt'] = pd.to_numeric([p[1] for p in parsed], errors='coerce')
            
            invalid_count = raw_df['mz'].isna().sum()
            if invalid_count > 0:
                print(f"警告：{invalid_count} 筆 'FeatureID' 無法解析為 m/z 和 RT（已設為 NaN）。")
        
        # 基於 FeatureID 值設定 is_ISTD（避免索引偏移）
        raw_df['is_ISTD'] = raw_df['FeatureID'].astype(str).str.strip().isin(istd_feature_ids)

        # ===== 防呆14: ISTD 识别验证 =====
        print(f"\n{'='*70}")
        print(f"ISTD 識別結果:")
        print(f"{'='*70}")
        print(f"識別到 {len(istd_feature_ids)} 個ISTD（紅色標記的 FeatureID）")

        if len(istd_feature_ids) == 0:
            raise ValueError("錯誤：未找到任何 ISTD（請在 RawIntensity 工作表的 FeatureID 欄位中，將內標物質的 FeatureID 標記為紅色字體）")
        elif len(istd_feature_ids) < 3:
            print(f"⚠️ 警告：ISTD 數量較少（{len(istd_feature_ids)} 個），建議至少使用 3 個以上的 ISTD")
            print(f"   以確保校正效果的穩定性")

        # 验证 ISTD 是否存在于 raw_df 中
        istd_in_df = raw_df[raw_df['is_ISTD']]
        if len(istd_in_df) != len(istd_feature_ids):
            print(f"⚠️ 警告：部分 ISTD 在 RawIntensity 中找不到對應的 FeatureID")
            print(f"   標記的 ISTD: {len(istd_feature_ids)} 個")
            print(f"   實際找到: {len(istd_in_df)} 個")

        # ===== 防呆15: ISTD 强度验证 =====
        print(f"\nISTD 強度驗證:")
        for idx, row in istd_in_df.iterrows():
            fid = row['FeatureID']
            values = get_valid_values(row, sample_columns)

            if len(values) == 0:
                raise ValueError(f"錯誤：ISTD '{fid}' 的所有樣本強度都是 0 或 NaN，無法進行校正，請檢查數據")
            elif len(values) < len(sample_columns) * 0.5:
                print(f"⚠️ 警告：ISTD '{fid}' 有效值比例較低 ({len(values)}/{len(sample_columns)})")

            # 计算 CV%
            if len(values) >= 2:
                mean_val = np.mean(values)
                std_val = np.std(values, ddof=1)
                cv_percent = (std_val / mean_val) * 100 if mean_val != 0 else np.nan

                if cv_percent > CV_QUALITY_THRESHOLDS['acceptable']:
                    print(f"⚠️ 警告：ISTD '{fid}' 的 CV% 較高 ({cv_percent:.1f}%)，可能影響校正品質")

        print(f"ISTD 列表: {', '.join(istd_feature_ids[:5])}")
        if len(istd_feature_ids) > 5:
            print(f"           ... 還有 {len(istd_feature_ids) - 5} 個")
        print(f"{'='*70}\n")

        # ===== 建立 col_to_info 映射 =====
        # 將 RawIntensity 欄位名 → SampleInfo 資訊行 (Sample_Type, Batch 等)
        col_to_info = {}
        info_rows = []
        for _, row in sample_info_df.iterrows():
            info_rows.append({
                'Sample_Name': str(row['Sample_Name']).strip(),
                'Sample_Type': str(row.get('Sample_Type', '')).strip(),
                'Batch': str(row.get('Batch', '')) if pd.notna(row.get('Batch')) else '',
                '_norm': normalize_sample_name(row['Sample_Name']),
                '_tokens': _extract_key_tokens(str(row['Sample_Name'])),
            })

        # Pass 1: exact normalized match
        unmatched_cols = []
        for col in sample_columns:
            col_norm = normalize_sample_name(col)
            matched = False
            for info in info_rows:
                if col_norm == info['_norm']:
                    col_to_info[col] = {
                        'Sample_Name': info['Sample_Name'],
                        'Sample_Type': info['Sample_Type'],
                        'Batch': info['Batch'],
                    }
                    matched = True
                    break
            if not matched:
                unmatched_cols.append(col)

        # Pass 2: fuzzy token match (overlap >= 2)
        still_unmatched = []
        available_infos = [info for info in info_rows
                           if info['Sample_Name'] not in {v['Sample_Name'] for v in col_to_info.values()}]
        for col in unmatched_cols:
            col_tokens = _extract_key_tokens(col)
            best_match = None
            best_overlap = 0
            for info in available_infos:
                overlap = len(col_tokens & info['_tokens'])
                if overlap > best_overlap:
                    best_overlap = overlap
                    best_match = info
            if best_overlap >= 2 and best_match is not None:
                col_to_info[col] = {
                    'Sample_Name': best_match['Sample_Name'],
                    'Sample_Type': best_match['Sample_Type'],
                    'Batch': best_match['Batch'],
                }
                available_infos.remove(best_match)
            else:
                still_unmatched.append(col)

        # Pass 3: keyword fallback from column name
        for col in still_unmatched:
            col_upper = col.upper()
            if 'QC' in col_upper or 'POOLED' in col_upper:
                col_to_info[col] = {'Sample_Name': col, 'Sample_Type': 'QC', 'Batch': ''}
            elif any(k in col_upper for k in ('CONTROL', 'CTL', 'CON', 'BENIGN')):
                col_to_info[col] = {'Sample_Name': col, 'Sample_Type': 'Control', 'Batch': ''}
            elif any(k in col_upper for k in ('EXPOSED', 'EXP', 'TREAT', 'TUMOR')):
                col_to_info[col] = {'Sample_Name': col, 'Sample_Type': 'Exposure', 'Batch': ''}
            elif any(k in col_upper for k in ('NORMAL', 'NOR')):
                col_to_info[col] = {'Sample_Name': col, 'Sample_Type': 'Normal', 'Batch': ''}
            else:
                col_to_info[col] = {'Sample_Name': col, 'Sample_Type': 'Unknown', 'Batch': ''}

        matched_count = len(sample_columns) - len(still_unmatched)
        print(f"✓ col_to_info 映射建立完成：{matched_count}/{len(sample_columns)} 個欄位成功匹配到 SampleInfo")

        return raw_df, sample_info_df, all_sheets, col_to_info
    except Exception:
        raise

def identify_istd_signals(df):
    return df[df['is_ISTD']], df[~df['is_ISTD']]

def calculate_istd_cv(istd_signals, sample_columns):
    """
    Calculate CV% for each ISTD signal.

    Vectorized implementation - much faster than iterrows().
    """
    from ms_core.utils.safe_math import safe_cv_percent_vectorized

    # Extract numeric matrix for sample columns
    valid_cols = [c for c in sample_columns if c in istd_signals.columns]
    if not valid_cols:
        return {}

    # Get numeric data matrix
    numeric_data = istd_signals[valid_cols].apply(pd.to_numeric, errors='coerce').values

    # Calculate CV% for each row (vectorized)
    cv_values = safe_cv_percent_vectorized(numeric_data, axis=1, min_samples=2)

    # Build result dictionary
    return dict(zip(istd_signals['FeatureID'], cv_values))

def find_best_istd_for_analyte(analyte_row, istd_signals, istd_cv, 
                                sample_columns,  # ✅ 新增參數
                                rt_weight=0.6, cv_weight=0.25, 
                                intensity_weight=0.1, mz_weight=0.05):
    """
    多因素加權評分的 ISTD 選擇函數
    
    評分公式（越低越好）：
    score = 0.6 × RT差異(標準化) + 0.25 × CV%(標準化) + 
            0.1 × 強度倒數(標準化) + 0.05 × m/z差異(標準化)
    
    Parameters:
    -----------
    analyte_row : pd.Series
        待校正的代謝物資料
    istd_signals : pd.DataFrame
        所有 ISTD 訊號
    istd_cv : dict
        ISTD 的 CV% 字典
    sample_columns : list
        樣本欄位名稱列表
    rt_weight : float
        RT 差異權重（預設 0.6）
    cv_weight : float
        CV% 權重（預設 0.25）
    intensity_weight : float
        強度權重（預設 0.1）
    mz_weight : float
        m/z 差異權重（預設 0.05）
    
    Returns:
    --------
    best_istd : pd.Series or None
        最佳 ISTD
    rt_diff : float
        RT 差異
    """
    # ===== 防呆21: 权重参数验证 =====
    # 检查权重是否为负
    if rt_weight < 0 or cv_weight < 0 or intensity_weight < 0 or mz_weight < 0:
        raise ValueError(
            f"❌ 錯誤：權重不能為負值\n"
            f"   rt_weight={rt_weight}, cv_weight={cv_weight}, "
            f"intensity_weight={intensity_weight}, mz_weight={mz_weight}"
        )

    # ✅ 權重總和檢查
    total_weight = rt_weight + cv_weight + intensity_weight + mz_weight
    if not np.isclose(total_weight, 1.0, atol=1e-6):
        raise ValueError(
            f"❌ 錯誤：權重總和 = {total_weight:.6f}，必須等於 1.0\n"
            f"   rt_weight={rt_weight}, cv_weight={cv_weight}, "
            f"intensity_weight={intensity_weight}, mz_weight={mz_weight}"
        )
    
    analyte_rt = analyte_row.get('rt', np.nan)
    analyte_mz = analyte_row.get('mz', np.nan)
    
    # 檢查 analyte 資訊完整性
    if np.isnan(analyte_rt) or np.isnan(analyte_mz):
        return None, float('inf')
    
    # ========== 步驟 1: 收集所有 ISTD 的指標 (向量化優化) ==========
    # Pre-calculate median intensities for all ISTDs (vectorized)
    valid_sample_cols = [c for c in sample_columns if c in istd_signals.columns]
    if valid_sample_cols:
        # Extract numeric matrix and calculate medians vectorized
        numeric_data = istd_signals[valid_sample_cols].apply(pd.to_numeric, errors='coerce')
        numeric_data = numeric_data.where(numeric_data > 0, np.nan)  # Replace <=0 with NaN
        median_intensities = numeric_data.median(axis=1).values
    else:
        median_intensities = np.zeros(len(istd_signals))

    # Pre-calculate RT and mz differences (vectorized)
    istd_rts = istd_signals['rt'].values
    istd_mzs = istd_signals['mz'].values
    rt_diffs_all = np.abs(analyte_rt - istd_rts)
    mz_diffs_all = np.abs(analyte_mz - istd_mzs) / analyte_mz * 1e6

    candidates = []

    for idx, istd_row in enumerate(istd_signals.itertuples()):
        istd_id = istd_row.FeatureID
        istd_rt = istd_rts[idx]
        istd_mz = istd_mzs[idx]

        # 跳過缺少資訊的 ISTD
        if np.isnan(istd_rt) or np.isnan(istd_mz):
            continue

        # 獲取預先計算的值
        rt_diff = rt_diffs_all[idx]
        mz_diff_ppm = mz_diffs_all[idx]
        median_intensity = median_intensities[idx]

        # 獲取 CV%
        cv = istd_cv.get(istd_id, np.nan)
        if np.isnan(cv):
            cv = 100.0  # 如果沒有 CV%，設為高值

        # 儲存候選資訊 (convert namedtuple row back to Series for compatibility)
        istd_row_series = istd_signals.iloc[idx]
        candidates.append({
            'istd_row': istd_row_series,
            'istd_id': istd_id,
            'rt_diff': rt_diff,
            'cv': cv,
            'intensity': median_intensity if not np.isnan(median_intensity) else 0.0,
            'mz_diff_ppm': mz_diff_ppm
        })
    
    # ========== 步驟 2: 如果沒有候選，返回 None ==========
    if len(candidates) == 0:
        return None, float('inf')
    
    # ========== 步驟 3: 標準化各指標（Min-Max Normalization）==========
    # 提取所有候選的指標
    rt_diffs = np.array([c['rt_diff'] for c in candidates])
    cvs = np.array([c['cv'] for c in candidates])
    intensities = np.array([c['intensity'] for c in candidates])
    mz_diffs = np.array([c['mz_diff_ppm'] for c in candidates])
    
    # 🔧 修正：標準化函數（處理所有值相同的情況）
    def normalize(values):
        """
        Min-Max 標準化到 [0, 1] 範圍
        如果所有值相同，返回 0（表示無差異）
        """
        min_val = np.min(values)
        max_val = np.max(values)
        if np.isclose(max_val, min_val, atol=1e-10):
            # 如果所有值相同，表示無差異，返回 0（不影響評分）
            return np.zeros(len(values))
        return (values - min_val) / (max_val - min_val)
    
    # 標準化各指標
    normalized_rt = normalize(rt_diffs)
    normalized_cv = normalize(cvs)
    
    # 🔧 修正：強度標準化（強度越高 → 分數越低）
    normalized_intensity = normalize(intensities)
    normalized_intensity_inv = 1.0 - normalized_intensity  # 反轉（強度高得分低）
    
    normalized_mz = normalize(mz_diffs)
    
    # ========== 步驟 4: 計算加權評分 ==========
    for i, candidate in enumerate(candidates):
        # 加權評分（越低越好）
        score = (
            rt_weight * normalized_rt[i] +
            cv_weight * normalized_cv[i] +
            intensity_weight * normalized_intensity_inv[i] +
            mz_weight * normalized_mz[i]
        )
        candidate['score'] = score
        
        # 🔧 新增：記錄各項評分（用於調試）
        candidate['score_breakdown'] = {
            'rt_score': rt_weight * normalized_rt[i],
            'cv_score': cv_weight * normalized_cv[i],
            'intensity_score': intensity_weight * normalized_intensity_inv[i],
            'mz_score': mz_weight * normalized_mz[i]
        }
    
    # ========== 步驟 5: 選擇評分最低的 ISTD ==========
    best_candidate = min(candidates, key=lambda x: x['score'])
    
    return best_candidate['istd_row'], best_candidate['rt_diff']

def calculate_istd_medians(istd_signals, sample_columns):
    """
    Calculate median intensity for each ISTD signal.

    Vectorized implementation - much faster than iterrows().
    """
    # Extract numeric matrix for sample columns
    valid_cols = [c for c in sample_columns if c in istd_signals.columns]
    if not valid_cols:
        return {}

    # Get numeric data and calculate median per row
    numeric_data = istd_signals[valid_cols].apply(pd.to_numeric, errors='coerce').values

    # Replace non-positive values with NaN for proper median calculation
    numeric_data = np.where(numeric_data > 0, numeric_data, np.nan)

    # Calculate median for each row (ignoring NaN)
    medians = np.nanmedian(numeric_data, axis=1)

    # Build result dictionary
    return dict(zip(istd_signals['FeatureID'], medians))

def calculate_corrected_ratios(df, sample_info_df):
    """
    计算 ISTD 校正后的比值

    增强了多层防呆检查
    """
    # ===== 防呆19: 基本输入验证 =====
    if df is None or df.empty:
        raise ValueError("錯誤：輸入數據為空")

    if sample_info_df is None or sample_info_df.empty:
        raise ValueError("錯誤：樣本資訊為空")

    istd_signals, analyte_signals = identify_istd_signals(df)

    # ===== 防呆20: ISTD 信号检查 =====
    if len(istd_signals) == 0:
        raise ValueError("錯誤：未找到 ISTD 信號")

    if len(analyte_signals) == 0:
        raise ValueError("錯誤：未找到代謝物信號（所有 FeatureID 都被標記為 ISTD）")

    print(f"\n校正統計:")
    print(f"  - ISTD 數量: {len(istd_signals)}")
    print(f"  - 代謝物數量: {len(analyte_signals)}")

    if 'Sample_Name' not in sample_info_df.columns:
        raise ValueError(f"錯誤：'{SHEET_NAMES['sample_info']}' 缺少 'Sample_Name' 欄位")
    
    sample_names = sample_info_df['Sample_Name'].tolist()

    from ms_core.utils.constants import NON_SAMPLE_COLUMNS
    all_columns = df.columns.tolist()
    # 排除已知的非樣本欄位（FeatureID, mz, rt, is_ISTD 等）
    non_sample = NON_SAMPLE_COLUMNS | {'is_ISTD', 'Sample_Type', 'sample_type'}
    raw_sample_columns = [col for col in all_columns
                          if col not in non_sample and col != 'FeatureID']

    # 修改：標準化名稱（轉小寫、去除空格）以避免不匹配
    normalized_sample_names = [name.strip().lower() for name in sample_names]
    normalized_columns = {col: col.strip().lower() for col in raw_sample_columns}

    sample_columns = []
    name_mapping = {}  # 記錄映射
    for col, norm_col in normalized_columns.items():
        if norm_col in normalized_sample_names:
            sample_columns.append(col)  # 保留原始欄位名
            name_mapping[col] = sample_names[normalized_sample_names.index(norm_col)]

    # 如果精確匹配結果太少（名稱格式不同），回退使用 RawIntensity 中的所有數據欄位
    if len(sample_columns) < len(raw_sample_columns) * 0.5:
        print(f"⚠️ 精確名稱匹配僅找到 {len(sample_columns)}/{len(raw_sample_columns)} 個樣本")
        print(f"  → 回退使用 {SHEET_NAMES['raw_intensity']} 中的所有數據欄位進行校正")
        sample_columns = raw_sample_columns
    else:
        missing_columns = [name for name in sample_names if name.strip().lower() not in normalized_columns.values()]
        if missing_columns:
            print(f"警告：以下樣本名稱在 '{SHEET_NAMES['raw_intensity']}' 中缺少 (即使大小寫不同): {', '.join(missing_columns)}")
    
    istd_cv = calculate_istd_cv(istd_signals, sample_columns)
    istd_medians = calculate_istd_medians(istd_signals, sample_columns)
    
    # ✅ 新增：印出權重設定資訊
    print(f"\n{'='*70}")
    print(f"🎯 ISTD 選擇權重設定:")
    print(f"{'='*70}")
    print(f"  - RT 差異權重:    60%")
    print(f"  - CV% 權重:       25%")
    print(f"  - 強度權重:       10%")
    print(f"  - m/z 差異權重:    5%")
    print(f"  - 總和:          100%")
    print(f"{'='*70}\n")
    
    results = []
    for _, analyte_row in analyte_signals.iterrows():
        # ✅ 傳入 sample_columns
        best_istd, min_rt_diff = find_best_istd_for_analyte(
            analyte_row, istd_signals, istd_cv, 
            sample_columns  # ✅ 新增參數
        )
        
        if best_istd is None: 
            continue
        
        istd_id = best_istd['FeatureID']
        istd_median = istd_medians[istd_id]
        rt_diff = analyte_row['rt'] - best_istd['rt']
        
        result_row = {
            'FeatureID': analyte_row['FeatureID'], 
            'RT': analyte_row['rt'], 
            'ISTD': istd_id, 
            'ISTD_RT': best_istd['rt'], 
            'RT_Difference': rt_diff, 
            'ISTD_Median': istd_median
        }
        
        for col in sample_columns:
            if col not in df.columns: 
                continue
            try:
                analyte_intensity = float(analyte_row[col])
                istd_intensity = float(best_istd[col])
            except (ValueError, KeyError):
                analyte_intensity = np.nan
                istd_intensity = np.nan
            
            corrected = (
                (analyte_intensity / istd_intensity) * istd_median 
                if istd_intensity > 0 and not np.isnan(istd_median) and not np.isnan(analyte_intensity) 
                else np.nan
            )
            result_row[col] = corrected
        
        results.append(result_row)
    
    results_df = pd.DataFrame(results)
    
    # 防呆：強制轉換 results_df 的樣本欄位為數值 (向量化)
    valid_sample_cols = [col for col in sample_columns if col in results_df.columns]
    if valid_sample_cols:
        results_df[valid_sample_cols] = results_df[valid_sample_cols].apply(pd.to_numeric, errors='coerce')
    results_df = results_df.fillna(0)
    print(f"ISTD Correction 數據類型檢查：{results_df.dtypes}")
    
    return results_df, sample_columns



from scipy.stats import wilcoxon, levene

def calculate_qc_cv_with_statistical_test(results_df, sample_columns, sample_info_df, original_df,
                                           col_to_info=None):
    """
    計算 QC 樣本的 CV%，並進行正確的統計檢定

    統計方法：
    1. Wilcoxon 配對符號等級檢定（檢驗中位數偏移，適用於非常態分佈）
    2. Levene's test（檢驗方差齊性）

    Performance optimized:
    - Uses indexed lookup instead of O(N) search per row (was O(N^2), now O(N))
    - Pre-extracts numeric matrices for QC columns
    """
    from ms_core.utils.safe_math import safe_divide
    from ms_core.utils.sample_classification import normalize_sample_type

    # Use col_to_info for QC identification (primary), fallback to direct name match
    if col_to_info:
        qc_columns = [col for col in sample_columns
                       if normalize_sample_type(col_to_info.get(col, {}).get('Sample_Type', '')) == 'QC']
    else:
        qc_samples = sample_info_df[sample_info_df['Sample_Type'].str.upper().str.contains('QC')]['Sample_Name'].tolist()
        qc_columns = [col for col in sample_columns if col in qc_samples]
        # Keyword fallback if no match
        if not qc_columns:
            qc_columns = [col for col in sample_columns
                          if 'QC' in col.upper() or 'POOLED' in col.upper()]

    print(f"\n{'='*70}")
    print(f"🔬 開始統計檢定（Wilcoxon 配對符號等級檢定 + Levene's test）")
    print(f"{'='*70}")
    print(f"  - QC 樣本數: {len(qc_columns)}")
    print(f"  - Feature 總數: {len(results_df)}")

    # ===== PERFORMANCE OPTIMIZATION: Pre-extract all QC data vectorized =====
    # This replaces O(N*M) row-by-row extraction with O(N+M) vectorized operations
    original_indexed = original_df.set_index('FeatureID')

    # Pre-extract QC columns data for faster access
    valid_qc_cols = [c for c in qc_columns if c in results_df.columns and c in original_df.columns]

    # ===== VECTORIZED: Extract all QC data at once =====
    # Convert to numeric and replace <=0 with NaN (vectorized)
    qc_corrected_data = results_df[valid_qc_cols].apply(pd.to_numeric, errors='coerce')
    qc_corrected_data = qc_corrected_data.where(qc_corrected_data > 0, np.nan)

    qc_original_data = original_indexed[valid_qc_cols].apply(pd.to_numeric, errors='coerce')
    qc_original_data = qc_original_data.where(qc_original_data > 0, np.nan)

    cv_results = []
    total_features = len(results_df)

    # Use itertuples for faster iteration (2-3x faster than iterrows)
    for row_idx, row_tuple in enumerate(results_df.itertuples()):
        idx = row_tuple.Index
        feature_id = row_tuple.FeatureID

        # Extract QC values from pre-processed data (vectorized access)
        try:
            qc_values_corrected = qc_corrected_data.loc[idx].dropna().values
        except KeyError:
            qc_values_corrected = np.array([])

        try:
            if feature_id in qc_original_data.index:
                qc_values_original = qc_original_data.loc[feature_id].dropna().values
            else:
                qc_values_original = np.array([])
        except KeyError:
            qc_values_original = np.array([])

        # 確保配對樣本數一致
        min_len = min(len(qc_values_original), len(qc_values_corrected))

        if min_len < 3:
            cv_results.append({
                'FeatureID': feature_id,
                'Original_QC_CV%': np.nan,
                'Corrected_QC_CV%': np.nan,
                'CV_Improvement%': np.nan,
                'Wilcoxon_pvalue': np.nan,
                'Variance_Test_pvalue': np.nan,
                'Significant_Improvement': 'N/A'
            })
            continue

        qc_values_original = np.array(qc_values_original[:min_len])
        qc_values_corrected = np.array(qc_values_corrected[:min_len])

        # 計算 CV% with safe division
        orig_mean = np.mean(qc_values_original)
        corr_mean = np.mean(qc_values_corrected)
        original_cv = safe_divide(np.std(qc_values_original, ddof=1), orig_mean, np.nan) * 100
        corrected_cv = safe_divide(np.std(qc_values_corrected, ddof=1), corr_mean, np.nan) * 100
        cv_improvement = original_cv - corrected_cv

        # ✅ 1. Wilcoxon 配對符號等級檢定（檢驗中位數是否改變）
        try:
            # 計算差異
            differences = qc_values_original - qc_values_corrected

            # 只有當存在非零差異時才進行檢定
            if np.any(differences != 0):
                wilcoxon_stat, wilcoxon_pvalue = wilcoxon(
                    qc_values_original,
                    qc_values_corrected,
                    alternative='two-sided',
                    zero_method='wilcox'  # 處理零差異的方法
                )
            else:
                # 所有值都相同，p-value = 1.0
                wilcoxon_pvalue = 1.0
        except Exception:
            wilcoxon_pvalue = np.nan

        # ✅ 2. Levene's test（檢驗方差齊性）
        try:
            levene_stat, variance_test_pvalue = levene(qc_values_original, qc_values_corrected)
        except Exception:
            variance_test_pvalue = np.nan

        # ✅ 判斷顯著性
        if not np.isnan(variance_test_pvalue) and cv_improvement > 5:
            if variance_test_pvalue < 0.05:
                significant = 'Yes'
            else:
                significant = 'Marginal'
        elif cv_improvement > 10:
            significant = 'Yes (CV% only)'
        else:
            significant = 'No'

        cv_results.append({
            'FeatureID': feature_id,
            'Original_QC_CV%': original_cv,
            'Corrected_QC_CV%': corrected_cv,
            'CV_Improvement%': cv_improvement,
            'Wilcoxon_pvalue': wilcoxon_pvalue,
            'Variance_Test_pvalue': variance_test_pvalue,
            'Significant_Improvement': significant
        })

        if (row_idx + 1) % 500 == 0:
            print(f"  處理進度: {row_idx + 1}/{total_features} features")
    
    print(f"  ✓ 統計檢定完成！")
    
    cv_results_df = pd.DataFrame(cv_results)
    
    # 統計摘要
    sig_yes = (cv_results_df['Significant_Improvement'] == 'Yes').sum()
    sig_marginal = (cv_results_df['Significant_Improvement'] == 'Marginal').sum()
    sig_cv_only = (cv_results_df['Significant_Improvement'] == 'Yes (CV% only)').sum()
    sig_no = (cv_results_df['Significant_Improvement'] == 'No').sum()
    total_count = len(cv_results_df)
    
    print(f"\n📊 統計檢定摘要:")
    print(f"  - 總特徵數: {total_count}")
    print(f"  - 顯著改善 (Yes): {sig_yes} ({sig_yes/total_count*100:.1f}%)")
    print(f"  - 邊緣顯著 (Marginal): {sig_marginal} ({sig_marginal/total_count*100:.1f}%)")
    print(f"  - 僅 CV% 改善: {sig_cv_only} ({sig_cv_only/total_count*100:.1f}%)")
    print(f"  - 無顯著改善 (No): {sig_no} ({sig_no/total_count*100:.1f}%)")
    
    # Wilcoxon 檢定統計
    wilcoxon_valid = cv_results_df['Wilcoxon_pvalue'].notna().sum()
    wilcoxon_sig = ((cv_results_df['Wilcoxon_pvalue'] < 0.05) & 
                    (cv_results_df['Wilcoxon_pvalue'].notna())).sum()
    
    print(f"\n🔬 Wilcoxon 配對符號等級檢定（中位數變化）:")
    print(f"  - 成功執行: {wilcoxon_valid}/{total_count} ({wilcoxon_valid/total_count*100:.1f}%)")
    if wilcoxon_valid > 0:
        print(f"  - 中位數顯著改變 (p < 0.05): {wilcoxon_sig}/{wilcoxon_valid} ({wilcoxon_sig/wilcoxon_valid*100:.1f}%)")
        print(f"  - 中位數無顯著改變: {wilcoxon_valid - wilcoxon_sig}/{wilcoxon_valid} ({(wilcoxon_valid-wilcoxon_sig)/wilcoxon_valid*100:.1f}%)")
    
    # Levene's test 統計
    variance_valid = cv_results_df['Variance_Test_pvalue'].notna().sum()
    variance_sig = ((cv_results_df['Variance_Test_pvalue'] < 0.05) & 
                    (cv_results_df['Variance_Test_pvalue'].notna())).sum()
    
    print(f"\n🔬 Levene's Test（方差齊性）:")
    print(f"  - 成功執行: {variance_valid}/{total_count} ({variance_valid/total_count*100:.1f}%)")
    if variance_valid > 0:
        print(f"  - 方差顯著改變 (p < 0.05): {variance_sig}/{variance_valid} ({variance_sig/variance_valid*100:.1f}%)")
        print(f"  - 方差無顯著改變: {variance_valid - variance_sig}/{variance_valid} ({(variance_valid-variance_sig)/variance_valid*100:.1f}%)")
    
    # CV% 改善統計
    cv_improvement_valid = cv_results_df['CV_Improvement%'].notna()
    if cv_improvement_valid.sum() > 0:
        improvements = cv_results_df.loc[cv_improvement_valid, 'CV_Improvement%']
        improvements_finite = improvements[np.isfinite(improvements)]
        
        if len(improvements_finite) > 0:
            median_improvement = np.median(improvements_finite)
            mean_improvement = np.mean(improvements_finite)
            
            print(f"\n📊 CV% 改善統計:")
            print(f"  - 中位數改善: {median_improvement:.2f}%")
            print(f"  - 平均改善: {mean_improvement:.2f}%")
            print(f"  - 範圍: {improvements_finite.min():.2f}% - {improvements_finite.max():.2f}%")
            
            improved_cv = (improvements_finite > 5).sum()
            similar_cv = ((improvements_finite >= -5) & (improvements_finite <= 5)).sum()
            worse_cv = (improvements_finite < -5).sum()
            
            print(f"\n  改善程度分類:")
            print(f"  - 顯著改善 (>5%): {improved_cv} ({improved_cv/len(improvements_finite)*100:.1f}%)")
            print(f"  - 無明顯變化 (±5%): {similar_cv} ({similar_cv/len(improvements_finite)*100:.1f}%)")
            print(f"  - 變差 (<-5%): {worse_cv} ({worse_cv/len(improvements_finite)*100:.1f}%)")
    
    print(f"\n{'='*70}\n")
    
    return cv_results_df

# ========== ✅ 修正：基於 QC 群組內部的 Hotelling T² 異常值檢測 ==========
def calculate_hotelling_t2_outliers(qc_scores, all_scores=None, alpha=0.05):
    """
    使用 Hotelling T² 檢測 QC 樣本中的異常值

    ✅ 正確邏輯：計算每個 QC 樣本與 QC 群組中心的偏離

    Parameters:
    -----------
    qc_scores : ndarray
        QC 樣本的 PCA 分數 (n_qc, n_components)
    all_scores : ndarray, optional
        所有樣本的 PCA 分數（此參數保留以兼容舊代碼，但不使用）
    alpha : float
        顯著水平（預設 0.05）

    Returns:
    --------
    t2_values : ndarray
        每個 QC 樣本的 Hotelling T² 值
    threshold : float
        T² 閾值
    outliers : ndarray (bool)
        異常值標記
    """
    # ===== 防呆23: alpha 参数验证 =====
    if not (0 < alpha < 1):
        raise ValueError(f"❌ 錯誤：alpha 必須在 (0, 1) 範圍內，當前值: {alpha}")

    n_qc, p = qc_scores.shape
    
    if n_qc < 3:
        print(f"   ⚠️ QC 樣本數不足 ({n_qc} < 3)，無法進行異常值檢測")
        return np.zeros(n_qc), 0, np.zeros(n_qc, dtype=bool)
    
    # ✅ 關鍵：使用 QC 群組的統計量
    qc_mean = np.mean(qc_scores, axis=0)
    qc_cov = np.cov(qc_scores, rowvar=False)
    
    # 正則化協方差矩陣（防止奇異矩陣）
    qc_cov_reg = qc_cov + np.eye(p) * 1e-6
    
    try:
        qc_cov_inv = np.linalg.inv(qc_cov_reg)
    except np.linalg.LinAlgError:
        print("   ⚠️ 警告：QC 協方差矩陣奇異，使用偽逆矩陣")
        qc_cov_inv = np.linalg.pinv(qc_cov_reg)
    
    # ✅ 計算每個 QC 樣本與 QC 中心的 Hotelling T² 值
    t2_values = np.zeros(n_qc)
    for i in range(n_qc):
        diff = qc_scores[i] - qc_mean  # ✅ 與 QC 中心比較
        t2_values[i] = np.dot(np.dot(diff, qc_cov_inv), diff.T)
    
    # ✅ 使用 F 分佈計算閾值（考慮樣本數）
    if n_qc - p - 1 > 0:
        f_critical = stats.f.ppf(1 - alpha, p, n_qc - p - 1)
        threshold = (p * (n_qc + 1) * (n_qc - 1)) / (n_qc * (n_qc - p - 1)) * f_critical
    else:
        # 樣本數太少，使用卡方分布
        threshold = chi2.ppf(1 - alpha, p)
    
    # 識別異常值
    outliers = t2_values > threshold
    
    return t2_values, threshold, outliers

def plot_pvalue_distribution(cv_results_df, plots_dir, timestamp):
    """繪製 p 值分佈圖（驗證統計檢定有效性）"""
    try:
        # ✅ 圖表儲存在指定目錄，必要時建立預設路徑
        if plots_dir is None:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            base_dir = os.path.join(script_dir, 'output', 'ISTD_Correction_plots')
            os.makedirs(base_dir, exist_ok=True)
            plots_dir = os.path.join(base_dir, f"ISTD_Correction_{timestamp}")

        os.makedirs(plots_dir, exist_ok=True)
        
        variance_pvalues = cv_results_df['Variance_Test_pvalue'].dropna()
        
        if len(variance_pvalues) < 10:
            print("  ⚠️ 有效 p 值數量不足，跳過 p 值分佈圖")
            return
        
        # ✅ 只繪製直方圖（移除 Q-Q Plot）
        fig, ax = plt.subplots(1, 1, figsize=(10, 6))
        
        # 直方圖
        ax.hist(variance_pvalues, bins=20, color='steelblue', edgecolor='black', alpha=0.7)
        ax.axhline(y=len(variance_pvalues)/20, color='red', linestyle='--', linewidth=2,
                   label='Uniform Distribution Expected')
        ax.set_xlabel('P-value (Levene\'s Test)', fontsize=12, fontweight='bold')
        ax.set_ylabel('Frequency', fontsize=12, fontweight='bold')
        ax.set_title('P-value Distribution (Variance Homogeneity Test)', 
                     fontsize=14, fontweight='bold')
        ax.legend(fontsize=10)
        ax.grid(True, alpha=0.3, linestyle='--')
        
        # Kolmogorov-Smirnov 檢定
        from scipy.stats import kstest
        ks_stat, ks_pvalue = kstest(variance_pvalues, 'uniform')
        
        # 統計摘要
        p_below_005 = (variance_pvalues < 0.05).sum()
        p_below_001 = (variance_pvalues < 0.01).sum()
        total = len(variance_pvalues)
        
        textstr = f'📊 Statistical Summary:\n'
        textstr += f'Total features: {total}\n'
        textstr += f'p < 0.05: {p_below_005} ({p_below_005/total*100:.1f}%)\n'
        textstr += f'p < 0.01: {p_below_001} ({p_below_001/total*100:.1f}%)\n\n'
        textstr += f'Kolmogorov-Smirnov Test:\n'
        textstr += f'KS statistic = {ks_stat:.4f}\n'
        textstr += f'P-value = {ks_pvalue:.4f}\n\n'
        
        if ks_pvalue < 0.05:
            textstr += '✅ Result: Non-uniform\n'
            textstr += '→ ISTD correction significantly\n'
            textstr += '   reduced QC variance'
            bgcolor = 'lightgreen'
        else:
            textstr += '⚠️ Result: Uniform\n'
            textstr += '→ Limited effect of\n'
            textstr += '   ISTD correction'
            bgcolor = 'lightyellow'
        
        ax.text(0.98, 0.97, textstr, transform=ax.transAxes,
                fontsize=10, verticalalignment='top', horizontalalignment='right',
                bbox=dict(boxstyle='round', facecolor=bgcolor, alpha=0.8))
        
        plt.tight_layout()
        
        # ✅ 儲存到 output/ISTD_Correction_plots/
        pvalue_plot_path = os.path.join(plots_dir, f'Pvalue_Distribution_{timestamp}.png')
        plt.savefig(pvalue_plot_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        print(f"\n✓ P 值分佈圖已儲存: {pvalue_plot_path}")
        print(f"  - Kolmogorov-Smirnov 檢定: KS={ks_stat:.4f}, p={ks_pvalue:.4f}")
        if ks_pvalue < 0.05:
            print(f"  - ✅ 結論: ISTD 校正顯著改善 QC 方差穩定性")
        else:
            print(f"  - ⚠️ 結論: ISTD 校正效果有限")
        
    except Exception as e:
        print(f"  ⚠️ 繪製 p 值分佈圖時發生錯誤: {e}")
        import traceback
        traceback.print_exc()

# ========== Hotelling T² 橢圓繪製函數 ==========
def draw_hotelling_t2_ellipse(ax, scores, alpha=0.05, label=None, edgecolor='black', linestyle='-', linewidth=2.5):
    """
    在 2D PCA 圖上繪製 Hotelling T² 橢圓
    
    🔧 修正：橢圓中心使用樣本實際均值（而非強制為原點）
    
    Parameters:
    -----------
    ax : matplotlib.axes.Axes
        繪圖軸
    scores : ndarray
        PCA 分數矩陣 (n_samples, 2)
    alpha : float
        顯著水平（預設 0.05）
    label : str
        圖例標籤
    edgecolor : str
        橢圓邊框顏色
    linestyle : str
        線條樣式
    linewidth : float
        線條寬度
    
    Returns:
    --------
    bounds : tuple
        橢圓邊界 (x_min, x_max, y_min, y_max)
    """
    n, p = scores.shape
    
    if n < 3:
        print(f"   ⚠ 樣本數不足 ({n})，無法繪製 Hotelling T² 橢圓")
        return None
    
    # 🔧 關鍵修正：使用樣本實際均值（而非強制為原點）
    mean = np.mean(scores, axis=0)
    
    # 計算協方差矩陣
    cov = np.cov(scores, rowvar=False)
    
    # 計算特徵值和特徵向量，並按特徵值降序排列
    eigenvalues, eigenvectors = np.linalg.eigh(cov)
    
    # 按特徵值降序排序
    idx = eigenvalues.argsort()[::-1]
    eigenvalues = eigenvalues[idx]
    eigenvectors = eigenvectors[:, idx]
    
    # 防止負值或零值
    eigenvalues = np.maximum(eigenvalues, 1e-10)
    
    # 計算 F 臨界值
    f_critical = stats.f.ppf(1 - alpha, p, n - p)
    
    # 計算橢圓的縮放因子
    scale_factor = np.sqrt((p * (n - 1) * (n + 1)) / (n * (n - p)) * f_critical)
    
    # width 對應最大特徵值（主軸），height 對應次要軸
    width = 2 * scale_factor * np.sqrt(eigenvalues[0])
    height = 2 * scale_factor * np.sqrt(eigenvalues[1])
    
    # 旋轉角度使用第一個特徵向量
    angle = np.degrees(np.arctan2(eigenvectors[1, 0], eigenvectors[0, 0]))
    
    # 🔧 繪製橢圓（中心為實際均值）
    ellipse = Ellipse(mean, width, height, angle=angle,
                     facecolor='none', edgecolor=edgecolor,
                     linewidth=linewidth, linestyle=linestyle, label=label)
    ax.add_patch(ellipse)
    
    # 計算橢圓邊界
    t = np.linspace(0, 2*np.pi, 100)
    ellipse_x = (width/2) * np.cos(t)
    ellipse_y = (height/2) * np.sin(t)
    
    # 旋轉橢圓
    cos_angle = np.cos(np.radians(angle))
    sin_angle = np.sin(np.radians(angle))
    x_rot = ellipse_x * cos_angle - ellipse_y * sin_angle + mean[0]
    y_rot = ellipse_x * sin_angle + ellipse_y * cos_angle + mean[1]
    
    bounds = (np.min(x_rot), np.max(x_rot), np.min(y_rot), np.max(y_rot))
    
    return bounds


# ========== 修改：2D PCA 分析（Hotelling T² 異常值檢測 + 橢圓）==========
def perform_pca_analysis_2d(raw_df, corrected_df, lowess_df, sample_columns, sample_info_df, plots_dir,
                            col_to_info=None):
    """
    執行 2D PCA 分析
    - 🔧 使用 Hotelling T² 檢測異常值
    - 使用 Hotelling T² 繪製橢圓（中心固定為原點）
    - 🎨 不同組別使用不同形狀：控制組=方形（無邊框）、暴露組=三角形（無邊框）、QC=圓形（黑邊框）
    """
    # ✅ 修正：圖表儲存在 output/ISTD_Correction_plots/ 下的指定子資料夾
    if plots_dir is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        base_dir = os.path.join(script_dir, "output", "ISTD_Correction_plots")
        os.makedirs(base_dir, exist_ok=True)
        plots_dir = os.path.join(base_dir, f"ISTD_Correction_{datetime.now().strftime('%Y%m%d_%H%M%S')}")

    if not os.path.exists(plots_dir):
        os.makedirs(plots_dir, exist_ok=True)
        print(f"已建立 'ISTD_Correction_plots' 資料夾: {plots_dir}")

    timestamp = datetime.now().strftime('%Y%m%d_%H%M')

    from ms_core.utils.sample_classification import normalize_sample_type
    from ms_core.utils.constants import SAMPLE_TYPE_COLORS, SAMPLE_TYPE_MARKERS

    # 使用 col_to_info 識別樣本類型（精確映射，不再靠名稱比對）
    qc_columns = []
    control_columns = []
    exposed_columns = []
    sample_type_map = {}

    for col in sample_columns:
        if col_to_info and col in col_to_info:
            raw_type = col_to_info[col].get('Sample_Type', 'Unknown')
        else:
            # Fallback: try direct match against sample_info_df
            col_norm = str(col).strip().lower()
            sample_info_norm = sample_info_df.copy()
            sample_info_norm['_norm'] = sample_info_norm['Sample_Name'].astype(str).str.strip().str.lower()
            match = sample_info_norm[sample_info_norm['_norm'] == col_norm]
            raw_type = match.iloc[0]['Sample_Type'] if not match.empty else 'Unknown'

        norm_type = normalize_sample_type(raw_type)
        sample_type_map[col] = norm_type

        if norm_type == 'QC':
            qc_columns.append(col)
        elif norm_type == 'Exposure':
            exposed_columns.append(col)
        elif norm_type in ('Control', 'Normal'):
            control_columns.append(col)

    if len(qc_columns) < 3:
        print("警告：QC 樣本不足 (<3)，跳過 PCA 分析")
        return

    type_counts = {}
    for v in sample_type_map.values():
        type_counts[v] = type_counts.get(v, 0) + 1
    print(f"識別到樣本分組:")
    for t, c in sorted(type_counts.items()):
        print(f"  - {t}: {c} 個")
    print(f"  - 總計: {len(sample_columns)} 個")

    # 🎨 顏色和形狀映射（使用共用常數）
    color_map = {}
    marker_map = {}

    for col in sample_columns:
        s_type = sample_type_map.get(col, 'Unknown')
        color_map[col] = SAMPLE_TYPE_COLORS.get(s_type, SAMPLE_TYPE_COLORS.get('Unknown', '#808080'))
        marker_map[col] = SAMPLE_TYPE_MARKERS.get(s_type, SAMPLE_TYPE_MARKERS.get('Unknown', 'x'))

    def prepare_matrix(df, cols, feature_col='FeatureID'):
        try:
            matrix = df.set_index(feature_col)[cols].T
            matrix = matrix.apply(pd.to_numeric, errors='coerce').fillna(0)
            if matrix.shape[1] < 2:
                raise ValueError("特徵數不足 2，無法進行 PCA")
            matrix = np.log2(matrix + 1)
            scaler = StandardScaler()
            matrix = scaler.fit_transform(matrix)
            return matrix
        except Exception as e:
            print(f"警告：準備 PCA 矩陣失敗 ({e})")
            return None

    datasets = {
        'Raw Data': raw_df,
        'ISTD Corrected': corrected_df,
        'LOWESS Normalized': lowess_df
    }

    comparisons = [('Raw Data', 'ISTD Corrected')]
    if lowess_df is not None:
        comparisons.append(('ISTD Corrected', 'LOWESS Normalized'))

    for left_name, right_name in comparisons:
        left_df = datasets.get(left_name)
        right_df = datasets.get(right_name)
        
        if left_df is None or right_df is None:
            continue

        # 準備所有樣本的矩陣
        left_matrix = prepare_matrix(left_df, sample_columns)
        right_matrix = prepare_matrix(right_df, sample_columns)

        if left_matrix is None or right_matrix is None:
            continue

        # 執行 PCA（2 個主成分）        
        pca_left = PCA(n_components=2)
        scores_left = pca_left.fit_transform(left_matrix)
        var_left = pca_left.explained_variance_ratio_

        pca_right = PCA(n_components=2)
        scores_right = pca_right.fit_transform(right_matrix)
        var_right = pca_right.explained_variance_ratio_

        # 🔧 使用 Hotelling T² 檢測 QC 異常值
        qc_indices = [i for i, col in enumerate(sample_columns) if col in qc_columns]
        qc_scores_left = scores_left[qc_indices]
        qc_scores_right = scores_right[qc_indices]

        t2_left, t2_threshold_left, outliers_left = calculate_hotelling_t2_outliers(
            qc_scores_left, scores_left, alpha=0.05
        )
        t2_right, t2_threshold_right, outliers_right = calculate_hotelling_t2_outliers(
            qc_scores_right, scores_right, alpha=0.05
        )

        print(f"\n🔍 Hotelling T² 異常值檢測:")
        print(f"   {left_name}:")
        print(f"   - T² 閾值: {t2_threshold_left:.2f}")
        print(f"   - 異常值數量: {np.sum(outliers_left)}/{len(qc_columns)}")
        print(f"   {right_name}:")
        print(f"   - T² 閾值: {t2_threshold_right:.2f}")
        print(f"   - 異常值數量: {np.sum(outliers_right)}/{len(qc_columns)}")

        # 統一 PCA 圖樣式（使用 sample_type_map 保留所有類型）
        sample_types = [sample_type_map.get(col, 'Unknown') for col in sample_columns]

        qc_outliers_left = {qc_columns[i] for i in range(len(qc_columns)) if outliers_left[i]}
        qc_outliers_right = {qc_columns[i] for i in range(len(qc_columns)) if outliers_right[i]}

        output_path = os.path.join(plots_dir, f"2D_PCA_{left_name.replace(' ', '_')}_vs_{right_name.replace(' ', '_')}_{timestamp}.png")
        plot_pca_comparison_qc_style(
            scores_left,
            scores_right,
            var_left,
            var_right,
            sample_columns,
            sample_types,
            batch_labels=None,
            grouping='sample_type',
            suptitle=f'2D PCA Comparison: {left_name} vs {right_name}',
            left_title=left_name,
            right_title=right_name,
            left_threshold_text=f'Hotelling T² Threshold: {t2_threshold_left:.2f}',
            right_threshold_text=f'Hotelling T² Threshold: {t2_threshold_right:.2f}',
            qc_outlier_names_left=qc_outliers_left,
            qc_outlier_names_right=qc_outliers_right,
            output_path=output_path,
            dpi=300,
        )

        plt.close('all')
        print(f"✓ 2D PCA 圖已儲存: {output_path}")

        # ===== 輸出異常值摘要 =====
        print(f"\n{'='*70}")
        print(f"📊 {left_name} - Hotelling T² 異常值檢測結果")
        print(f"{'='*70}")
        print(f"QC 異常值數量: {np.sum(outliers_left)}/{len(outliers_left)}")
        if np.sum(outliers_left) > 0:
            outlier_samples = [qc_columns[i] for i in range(len(outliers_left)) if outliers_left[i]]
            outlier_t2_values = [t2_left[i] for i in range(len(outliers_left)) if outliers_left[i]]
            print(f"異常樣本:")
            for sample, t2_val in zip(outlier_samples, outlier_t2_values):
                print(f"  - {sample}: T² = {t2_val:.2f} (閾值 = {t2_threshold_left:.2f})")
        else:
            print("  ✓ 無異常樣本")
        
        print(f"\n📊 {right_name} - Hotelling T² 異常值檢測結果")
        print(f"{'='*70}")
        print(f"QC 異常值數量: {np.sum(outliers_right)}/{len(outliers_right)}")
        if np.sum(outliers_right) > 0:
            outlier_samples = [qc_columns[i] for i in range(len(outliers_right)) if outliers_right[i]]
            outlier_t2_values = [t2_right[i] for i in range(len(outliers_right)) if outliers_right[i]]
            print(f"異常樣本:")
            for sample, t2_val in zip(outlier_samples, outlier_t2_values):
                print(f"  - {sample}: T² = {t2_val:.2f} (閾值 = {t2_threshold_right:.2f})")
        else:
            print("  ✓ 無異常樣本")
        print(f"{'='*70}\n")

def apply_fdr_correction(pvalues):
    """
    使用 Benjamini-Hochberg 方法進行 FDR 校正

    Args:
        pvalues: p 值的 array 或 Series

    Returns:
        qvalues: 校正後的 q 值（FDR-adjusted p-values）
    """
    # 移除 NaN 值
    pvalues_array = np.array(pvalues)
    valid_mask = ~np.isnan(pvalues_array)

    # 初始化 q-values 為 NaN
    qvalues = np.full_like(pvalues_array, np.nan, dtype=float)

    if np.sum(valid_mask) == 0:
        return qvalues

    # 提取有效的 p-values
    valid_pvalues = pvalues_array[valid_mask]
    n = len(valid_pvalues)

    # Benjamini-Hochberg 方法
    # 1. 對 p 值排序，記錄原始索引
    sorted_indices = np.argsort(valid_pvalues)
    sorted_pvalues = valid_pvalues[sorted_indices]

    # 2. 計算 q 值：q_i = p_i * n / rank_i
    ranks = np.arange(1, n + 1)
    sorted_qvalues = sorted_pvalues * n / ranks

    # 3. 確保單調性（從後往前取最小值）
    for i in range(n - 2, -1, -1):
        sorted_qvalues[i] = min(sorted_qvalues[i], sorted_qvalues[i + 1])

    # 4. q 值不能超過 1
    sorted_qvalues = np.minimum(sorted_qvalues, 1.0)

    # 5. 恢復原始順序
    unsorted_qvalues = np.empty_like(sorted_qvalues)
    unsorted_qvalues[sorted_indices] = sorted_qvalues

    # 6. 將結果填回包含 NaN 的陣列
    qvalues[valid_mask] = unsorted_qvalues

    return qvalues


# ========== 🔧 修改：save_results_to_excel==========
def save_results_to_excel(original_df, results_df, sample_info_df, output_file,
                          all_sheets, sample_columns, original_workbook, plots_dir=None,
                          col_to_info=None):
    """
    儲存結果到 Excel，使用 Wilcoxon 配對符號等級檢定 + Levene's test
    """
    # ===== 防呆22: 结果数据验证 =====
    if results_df is None or results_df.empty:
        print("❌ 錯誤：校正結果為空，無法儲存")
        raise ValueError("校正結果為空")

    if len(results_df) == 0:
        print("❌ 錯誤：沒有成功校正的代謝物")
        raise ValueError("沒有成功校正的代謝物")

    print(f"\n準備儲存結果:")
    print(f"  - 校正成功的代謝物數量: {len(results_df)}")
    print(f"  - 樣本數量: {len(sample_columns)}")

    # 检查必要欄位
    required_columns = ['FeatureID', 'ISTD']
    missing_cols = [col for col in required_columns if col not in results_df.columns]
    if missing_cols:
        print(f"❌ 錯誤：結果缺少必要欄位: {', '.join(missing_cols)}")
        raise ValueError(f"結果缺少必要欄位: {missing_cols}")

    # ✅ 圖表輸出基底（可覆蓋為特定目錄）
    plot_output_dir = plots_dir or os.path.dirname(output_file)
    
    # ✅ 使用新的統計檢定函數
    cv_results_df = calculate_qc_cv_with_statistical_test(
        results_df, sample_columns, sample_info_df, original_df,
        col_to_info=col_to_info
    )
    
    # 合併結果
    results_with_cv = results_df.merge(cv_results_df, on='FeatureID', how='left')

    # 🆕 應用 FDR 校正（Benjamini-Hochberg 方法）
    print("\n📊 應用 FDR 校正（Benjamini-Hochberg 方法）...")

    # 對 Wilcoxon p-value 進行 FDR 校正
    if 'Wilcoxon_pvalue' in results_with_cv.columns:
        results_with_cv['Wilcoxon_qvalue'] = apply_fdr_correction(results_with_cv['Wilcoxon_pvalue'])
        wilcoxon_valid = results_with_cv['Wilcoxon_pvalue'].notna().sum()
        wilcoxon_sig_p = ((results_with_cv['Wilcoxon_pvalue'] < 0.05) &
                          (results_with_cv['Wilcoxon_pvalue'].notna())).sum()
        wilcoxon_sig_q = ((results_with_cv['Wilcoxon_qvalue'] < 0.05) &
                          (results_with_cv['Wilcoxon_qvalue'].notna())).sum()
        print(f"  Wilcoxon test:")
        print(f"    - 有效檢定數: {wilcoxon_valid}")
        print(f"    - p < 0.05: {wilcoxon_sig_p} ({wilcoxon_sig_p/wilcoxon_valid*100:.1f}%)")
        print(f"    - q < 0.05 (FDR 校正後): {wilcoxon_sig_q} ({wilcoxon_sig_q/wilcoxon_valid*100:.1f}%)")

    # 對 Variance Test p-value 進行 FDR 校正
    if 'Variance_Test_pvalue' in results_with_cv.columns:
        results_with_cv['Variance_Test_qvalue'] = apply_fdr_correction(results_with_cv['Variance_Test_pvalue'])
        variance_valid = results_with_cv['Variance_Test_pvalue'].notna().sum()
        variance_sig_p = ((results_with_cv['Variance_Test_pvalue'] < 0.05) &
                          (results_with_cv['Variance_Test_pvalue'].notna())).sum()
        variance_sig_q = ((results_with_cv['Variance_Test_qvalue'] < 0.05) &
                          (results_with_cv['Variance_Test_qvalue'].notna())).sum()
        print(f"  Levene's test:")
        print(f"    - 有效檢定數: {variance_valid}")
        print(f"    - p < 0.05: {variance_sig_p} ({variance_sig_p/variance_valid*100:.1f}%)")
        print(f"    - q < 0.05 (FDR 校正後): {variance_sig_q} ({variance_sig_q/variance_valid*100:.1f}%)")

    print("  ✓ FDR 校正完成\n")

    # ✅ 調整欄位順序（加入 q-value 欄位）
    cols_order = [
        'Original_QC_CV%', 'Corrected_QC_CV%', 'CV_Improvement%',
        'Wilcoxon_pvalue', 'Wilcoxon_qvalue',
        'Variance_Test_pvalue', 'Variance_Test_qvalue',
        'Significant_Improvement'
    ]
    other_cols = [col for col in results_with_cv.columns if col not in cols_order]
    results_with_cv = results_with_cv[other_cols + cols_order]
    
    # 寫入 Excel（將內部欄名 'FeatureID' 還原為 FEATURE_ID_COLUMN）
    def _rename_feature_col(df):
        if 'FeatureID' in df.columns and FEATURE_ID_COLUMN != 'FeatureID':
            return df.rename(columns={'FeatureID': FEATURE_ID_COLUMN})
        return df

    # ===== Fix 3: 簡化欄位名稱 (移除 DNA/RNA_programN_ prefix) =====
    rename_map = {}
    for col in sample_columns:
        simplified = simplify_column_name(col)
        if simplified != col:
            rename_map[col] = simplified

    if rename_map:
        print(f"✓ 簡化 {len(rename_map)} 個欄位名稱（移除 DNA/RNA_programN_ 前綴）")
        # Rename in results
        results_with_cv = results_with_cv.rename(columns=rename_map)
        # Rename in all_sheets
        for sheet_name in all_sheets:
            all_sheets[sheet_name] = all_sheets[sheet_name].rename(columns=rename_map)

        # Update SampleInfo Sample_Name to match simplified column names
        # so downstream processors can match columns to SampleInfo directly
        if SHEET_NAMES['sample_info'] in all_sheets and col_to_info:
            si = all_sheets[SHEET_NAMES['sample_info']]
            if 'Sample_Name' in si.columns:
                # Build reverse mapping: original SampleInfo name → simplified column name
                info_name_to_col = {}
                for orig_col, info in col_to_info.items():
                    simplified = rename_map.get(orig_col, orig_col)
                    info_name_to_col[info['Sample_Name'].strip()] = simplified

                def _update_sample_name(name):
                    name_stripped = str(name).strip()
                    return info_name_to_col.get(name_stripped, name_stripped)

                si['Sample_Name'] = si['Sample_Name'].apply(_update_sample_name)

    # ===== 在 ISTD_Correction 中插入 Sample_Type 資訊行 =====
    # 讓下游步驟可直接從資料 sheet 讀取分組資訊，無需另查 SampleInfo
    if col_to_info:
        from ms_core.utils.sample_classification import normalize_sample_type
        sample_type_row = {'FeatureID': 'Sample_Type'}
        for col in results_with_cv.columns:
            if col in sample_type_row:
                continue
            # 先查原始欄名（rename 前），再查 rename 後的名稱
            orig_col = col
            if rename_map:
                # rename_map: original -> simplified，需要反查
                reverse_map = {v: k for k, v in rename_map.items()}
                orig_col = reverse_map.get(col, col)
            info = col_to_info.get(orig_col, col_to_info.get(col))
            if info and 'Sample_Type' in info:
                sample_type_row[col] = normalize_sample_type(info['Sample_Type'])
            else:
                sample_type_row[col] = ''
        type_row_df = pd.DataFrame([sample_type_row], columns=results_with_cv.columns)
        results_with_cv = pd.concat([type_row_df, results_with_cv], ignore_index=True)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, df in all_sheets.items():
            _rename_feature_col(df).to_excel(writer, sheet_name=sheet_name, index=False)
        _rename_feature_col(results_with_cv).to_excel(writer, sheet_name=SHEET_NAMES['istd_correction'], index=False)
    
    # 格式設定
    workbook = load_workbook(original_workbook)
    new_workbook = load_workbook(output_file)
    
    # 複製 RawIntensity 格式（font, border, fill, number_format, protection, alignment）
    if SHEET_NAMES['raw_intensity'] in workbook.sheetnames and SHEET_NAMES['raw_intensity'] in new_workbook.sheetnames:
        copy_sheet_formatting_only(workbook[SHEET_NAMES['raw_intensity']], new_workbook[SHEET_NAMES['raw_intensity']])
    
    # ISTD_Correction 格式
    scientific_format = '0.00E+00'
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    light_pink_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    
    if SHEET_NAMES['istd_correction'] in new_workbook.sheetnames:
        worksheet = new_workbook[SHEET_NAMES['istd_correction']]
        header = [cell.value for cell in worksheet[1]]
        
        # CV% 欄位塗橙色
        for col_name in ['Original_QC_CV%', 'Corrected_QC_CV%', 'CV_Improvement%']:
            if col_name in header:
                col_idx = header.index(col_name) + 1
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.fill = orange_fill
        
        # ✅ 統計檢定欄位塗淡藍色（包含 p-value 和 q-value）
        for col_name in ['Wilcoxon_pvalue', 'Wilcoxon_qvalue', 'Variance_Test_pvalue', 'Variance_Test_qvalue']:
            if col_name in header:
                col_idx = header.index(col_name) + 1
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.fill = light_blue_fill
        
        # ✅ 顯著性標記
        if 'Significant_Improvement' in header:
            col_idx = header.index('Significant_Improvement') + 1
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value == 'Yes' or cell.value == 'Yes (CV% only)':
                        cell.fill = green_fill
                    elif cell.value == 'Marginal':
                        cell.fill = yellow_fill
                    elif cell.value == 'No':
                        cell.fill = light_pink_fill
        
        # 數字格式
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value is not None:
                    cell.number_format = scientific_format
    
    new_workbook.save(output_file)
    
    print(f"\n{'='*70}")
    print(f"✓ ISTD Correction 結果已保存:")
    print(f"  {output_file}")
    print(f"{'='*70}\n")
    
    # P 值分佈圖 (disabled: provides limited diagnostic value)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    # plot_pvalue_distribution(cv_results_df, plot_output_dir, timestamp)

# ========== main 函數 ==========
def main(input_file=None):
    """
    主函數 - 修改為與 GUI 配合
    
    Parameters:
    -----------
    input_file : str, optional
        輸入檔案路徑（由 GUI 傳入）
        如果為 None，則顯示檔案選擇對話框
    
    Returns:
    --------
    dict or None
        - None: 用戶取消檔案選擇
        - dict: 執行成功，包含統計資訊
    """
    if input_file is None:
        raise ValueError("input_file is required; GUI must provide the file path.")

    
    # 🔧 建立 output 資料夾
    output_dir = get_output_root()
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
        print(f"已建立 'output' 資料夾: {output_dir}")
    
    # 🔧 關鍵修正：如果沒有提供 input_file，則顯示對話框
    if input_file is None:
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        
        input_file = filedialog.askopenfilename(
            title="選擇原始 Excel 檔案", 
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        root.destroy()
        
        # 如果用戶取消選擇，返回 None
        if not input_file:
            print("⚠️ 用戶取消了檔案選擇")
            return None
    
    # 驗證檔案是否存在
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"找不到檔案: {input_file}")
    
    print("\n" + "="*70)
    print("🔬 開始 ISTD Correction 分析")
    print(f"📁 輸入檔案: {os.path.basename(input_file)}")
    print("="*70 + "\n")
    
    # 載入數據 (raises ValueError on failure)
    original_df, sample_info_df, all_sheets, col_to_info = load_and_process_data(input_file)

    # 計算校正結果 (raises ValueError on failure)
    results_df, sample_columns = calculate_corrected_ratios(original_df, sample_info_df)
    
    # 🔧 修改：儲存結果到 output 資料夾
    run_timestamp = datetime.now().strftime(DATETIME_FORMAT_FULL)
    output_file = build_output_path("ISTD_Results", timestamp=run_timestamp)
    plots_session_dir = build_plots_dir(
        "ISTD_Correction_plots",
        timestamp=run_timestamp,
        session_prefix="ISTD_Correction"
    )

    # ===== 防呆17: 输出目录权限检查 =====
    try:
        # 测试写入权限
        test_file = os.path.join(output_dir, '.write_test')
        with open(test_file, 'w') as f:
            f.write('test')
        os.remove(test_file)
    except Exception as e:
        print(f"❌ 錯誤：無法寫入 output 目錄")
        print(f"   請檢查目錄權限: {output_dir}")
        print(f"   詳細錯誤: {e}")
        raise Exception(f"輸出目錄無寫入權限: {output_dir}")

    # ===== 防呆18: 输出文件检查 =====
    if os.path.exists(output_file):
        print(f"⚠️ 警告：輸出檔案已存在，將被覆蓋")
        print(f"   {output_file}")

    save_results_to_excel(
        original_df, results_df, sample_info_df,
        output_file, all_sheets, sample_columns, input_file,
        plots_dir=plots_session_dir, col_to_info=col_to_info
    )
    
    # ✅ 執行 2D PCA 分析（傳入 output_dir）
    print("\n" + "="*70)
    print("📊 開始 2D PCA 分析（Hotelling T² 異常值檢測 + 橢圓 + 固定原點）")
    print("="*70 + "\n")
    perform_pca_analysis_2d(
        original_df, results_df, None,
        sample_columns, sample_info_df, plots_session_dir,
        col_to_info=col_to_info
    )
    
    print("\n" + "="*70)
    print("✅ ISTD Correction 完成！")
    print("="*70)
    print("\n📁 輸出檔案:")
    print(f"  1. Excel 結果: output/{os.path.basename(output_file)}")
    print(f"  2. PCA 圖表: {plots_session_dir}")
    print("\n💡 請使用輸出的檔案進行後續 QC LOWESS 處理。\n")
    
    # 🎯 返回統計資訊給 GUI
    return ProcessingResult(
        file_path=input_file,
        output_path=str(output_file),
        plots_dir=str(plots_session_dir),
        metabolites=len(original_df),
        samples=len(sample_columns)
    )


def process_in_memory(data_df, sample_info_df, **kwargs):
    """Pipeline-friendly entry point — bypass file I/O.

    Parameters
    ----------
    data_df : DataFrame
        Features-as-rows with ``FeatureID`` first column and sample columns.
    sample_info_df : DataFrame
        Must contain ``Sample_Name`` and ``Sample_Type`` columns.
    **kwargs
        istd_feature_ids : list[str], optional
            Explicit list of FeatureID values that are ISTDs.
            If not provided and ``is_ISTD`` column is absent, returns *None*
            so the wrapper falls back to the temp-file path (which can read
            red-font ISTD markers from Excel).

    Returns
    -------
    DataFrame or None
        Result with ``FeatureID`` + sample columns, or None if ISTD info
        is unavailable.
    """
    if data_df is None or data_df.empty:
        return None

    df = data_df.copy()

    # Ensure first column is named FeatureID
    if df.columns[0] != "FeatureID":
        df.rename(columns={df.columns[0]: "FeatureID"}, inplace=True)

    # --- Parse mz / rt from FeatureID (format "mz/rt") ---
    if "mz" not in df.columns or "rt" not in df.columns:
        def _parse_fid(fid):
            if isinstance(fid, str) and "/" in fid:
                parts = fid.split("/")
                if len(parts) >= 2:
                    return parts[0], parts[1]
            return np.nan, np.nan

        parsed = df["FeatureID"].apply(_parse_fid)
        df["mz"] = pd.to_numeric([p[0] for p in parsed], errors="coerce")
        df["rt"] = pd.to_numeric([p[1] for p in parsed], errors="coerce")

    # --- Determine is_ISTD ---
    istd_ids = kwargs.get("istd_feature_ids")
    if istd_ids is not None:
        df["is_ISTD"] = df["FeatureID"].astype(str).str.strip().isin(istd_ids)
    elif "is_ISTD" not in df.columns:
        # Cannot identify ISTDs without file-level font info → fallback
        return None

    # Convert sample columns to numeric, fill NaN with 0
    from ms_core.utils.constants import NON_SAMPLE_COLUMNS

    non_sample = NON_SAMPLE_COLUMNS | {"is_ISTD", "Sample_Type", "sample_type", "mz", "rt"}
    sample_columns = [c for c in df.columns if c not in non_sample and c != "FeatureID"]
    df[sample_columns] = df[sample_columns].apply(pd.to_numeric, errors="coerce").fillna(0)
    df[sample_columns] = df[sample_columns].clip(lower=0)

    # --- Core processing ---
    results_df, result_sample_cols = calculate_corrected_ratios(df, sample_info_df)

    # Return only FeatureID + sample columns (drop ISTD metadata columns)
    keep_cols = ["FeatureID"] + [c for c in result_sample_cols if c in results_df.columns]
    return results_df[keep_cols]


if __name__ == "__main__":
    # 🔧 獨立運行時不傳入 input_file，會顯示對話框
    main()
