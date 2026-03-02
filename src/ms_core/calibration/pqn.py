import pandas as pd
import numpy as np
from pathlib import Path
import warnings
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns
from copy import copy
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler

from ms_core.utils.plotting import plot_pca_comparison_qc_style, setup_matplotlib
from ms_core.utils.constants import FONT_SIZES, SHEET_NAMES, DATETIME_FORMAT_FULL, VALIDATION_THRESHOLDS, COHENS_D_THRESHOLDS, CV_QUALITY_THRESHOLDS
from ms_core.utils.sample_classification import SampleClassifier, normalize_sample_type
from ms_core.utils.file_io import (
    build_plots_dir,
    get_output_root,
    generate_output_filename,
)
from ms_core.utils.results import ProcessingResult
from ms_core.utils.console import safe_print as print

warnings.filterwarnings('ignore')

# ========== Matplotlib Global Settings ==========
# Use centralized setup
setup_matplotlib()

# Centralized summary metadata to avoid magic strings and ease maintenance
SUMMARY_SHEET_NAME = SHEET_NAMES.get('concentration', "ConcNormalization_Summary")
SUMMARY_REPORT_SEPARATOR = "-" * 80

def _lookup_sample_type(sample, sample_info_df, col_to_info_row=None, default='UNKNOWN'):
    """Helper: look up sample type using col_to_info_row mapping or fallback."""
    if col_to_info_row and sample in col_to_info_row:
        return str(col_to_info_row[sample].get('Sample_Type', default)).upper()
    # Direct lookup fallback
    rows = sample_info_df[sample_info_df.iloc[:, 0] == sample]
    if not rows.empty:
        return str(rows.iloc[0].get('Sample_Type', default)).upper()
    # Column-name keyword fallback
    s_upper = str(sample).upper()
    if any(kw in s_upper for kw in ['QC', 'POOLED']):
        return 'QC'
    return default.upper()

# ==================== 標準化方法 ====================

def enhanced_pqn_normalization(data_matrix, sample_info_df, sample_columns, reference_values,
                               col_to_info_row=None):
    """
    改進版 PQN 標準化：優先使用 QC 樣本作為參考

    Parameters:
    -----------
    col_to_info_row : dict, optional
        Mapping from data column name to SampleInfo row (Series).
        Used when column names don't match SampleInfo names.
    """
    print("\n執行改進版混合標準化方法：")

    # ========== 🔍 除錯輸出 ==========
    print(f"\n【除錯資訊】")
    print(f"  總樣本數: {len(sample_columns)}")

     # ========== Step 1: 分離 QC 和真實樣本（改進版）==========
    sample_types = {}
    for sample in sample_columns:
        info_row = col_to_info_row.get(sample) if col_to_info_row else None
        if info_row is not None:
            sample_type = str(info_row.get('Sample_Type', ''))
            sample_types[sample] = sample_type.upper()
        else:
            # Fallback: try exact match
            sample_row = sample_info_df[sample_info_df.iloc[:, 0] == sample]
            if not sample_row.empty:
                sample_type = str(sample_row.iloc[0].get('Sample_Type', ''))
                sample_types[sample] = sample_type.upper()
            else:
                # Fallback: detect QC from column name keywords
                if any(kw in str(sample).upper() for kw in ['QC', 'POOLED']):
                    sample_types[sample] = 'QC'
                else:
                    sample_types[sample] = 'UNKNOWN'
    
    qc_indices = [i for i, s in enumerate(sample_columns) if sample_types[s] == 'QC']
    real_indices = [i for i, s in enumerate(sample_columns) if sample_types[s] != 'QC']
    
    # 🔍 除錯：顯示樣本類型分佈
    print(f"  樣本類型統計:")
    from collections import Counter
    type_counts = Counter(sample_types.values())
    for stype, count in type_counts.items():
        print(f"    - {stype}: {count}")
    
    # 🔍 除錯：顯示 QC 樣本名稱
    qc_samples = [s for s, t in sample_types.items() if t == 'QC']
    print(f"  QC 樣本列表: {qc_samples}")
    
    qc_indices = [i for i, s in enumerate(sample_columns) if sample_types[s] == 'QC']
    real_indices = [i for i, s in enumerate(sample_columns) if sample_types[s] != 'QC']
    
    qc_count = len(qc_indices)
    real_count = len(real_indices)
    
    print(f"  樣本分類:")
    print(f"    - QC 樣本數量: {qc_count}")
    print(f"    - 真實樣本數量: {real_count}")
    # ========== 除錯輸出結束 ==========
    
    # ========== Step 2: 評估 QC 樣本質量 ==========
    qc_cv_median = np.nan
    reference_strategy = 'NONE'
    
    if qc_count > 0:
        qc_data = data_matrix[:, qc_indices]
        qc_cv = calculate_rsd(qc_data)
        qc_cv_median = np.nanmedian(qc_cv)
        
        print(f"  QC 質量評估:")
        print(f"    - QC 中位數 CV%: {qc_cv_median:.2f}%")
        
        if qc_count >= VALIDATION_THRESHOLDS['min_qc_samples'] and qc_cv_median < CV_QUALITY_THRESHOLDS['acceptable']:
            reference_strategy = 'QC'
            print(f"    - ✓ QC 樣本質量良好，使用 QC 作為 PQN 參考")
        elif qc_count >= 1:
            reference_strategy = 'QC_LIMITED'
            print(f"    - ⚠ QC 樣本數量有限（{qc_count}），但仍使用 QC 作為參考")
        else:
            reference_strategy = 'ROBUST_MEDIAN'
            print(f"    - ⚠ QC 樣本不足，使用穩健中位數作為參考")
    else:
        reference_strategy = 'ROBUST_MEDIAN'
        print(f"  ⚠ 無 QC 樣本，使用穩健中位數作為參考")
    
    # ========== Step 3: 肌酐校正（僅針對真實樣本）==========
    print("\n  步驟1: Sample-specific Normalization (肌酐校正)")
    
    real_data = data_matrix[:, real_indices]
    real_reference_values = reference_values[real_indices]
    
    # 過濾有效的參考值
    valid_ref_mask = ~np.isnan(real_reference_values) & (real_reference_values > 0)
    
    if np.sum(valid_ref_mask) < len(real_reference_values) * 0.5:
        print(f"    ⚠ 警告：有效肌酐值不足 50% ({np.sum(valid_ref_mask)}/{len(real_reference_values)})")
    
    # 肌酐校正
    median_ref = np.nanmedian(real_reference_values[valid_ref_mask])
    real_data_corrected = real_data.copy()
    real_data_corrected[:, valid_ref_mask] = (real_data[:, valid_ref_mask] / 
                                               real_reference_values[valid_ref_mask]) * median_ref
    
    print(f"    - 肌酐中位數: {median_ref:.2f}")
    print(f"    - 肌酐範圍: {np.nanmin(real_reference_values):.2f} - {np.nanmax(real_reference_values):.2f}")
    print(f"    - 校正樣本數: {np.sum(valid_ref_mask)}/{len(real_reference_values)}")
    
    # ========== Step 4: PQN 標準化 ==========
    print("\n  步驟2: Probabilistic Quotient Normalization (PQN)")
    
    # 決定參考樣本
    if reference_strategy == 'QC' or reference_strategy == 'QC_LIMITED':
        # 使用 QC 樣本中位數
        reference_sample = np.nanmedian(data_matrix[:, qc_indices], axis=1)
        print(f"    - 使用 QC 樣本中位數作為參考")
    else:
        # 使用真實樣本的穩健中位數（排除極端 10%）
        sorted_totals = np.argsort(np.nansum(real_data_corrected, axis=0))
        n_exclude = max(1, int(len(sorted_totals) * 0.1))
        robust_indices = sorted_totals[n_exclude:-n_exclude]
        reference_sample = np.nanmedian(real_data_corrected[:, robust_indices], axis=1)
        print(f"    - 使用穩健中位數作為參考（排除極端 {n_exclude*2} 個樣本）")
    
    # 4a. 真實樣本的 PQN
    quotients_real = real_data_corrected / reference_sample[:, np.newaxis]
    quotients_real = np.where(np.isfinite(quotients_real), quotients_real, np.nan)
    normalization_factors_real = np.nanmedian(quotients_real, axis=0)
    
    real_data_final = real_data_corrected / normalization_factors_real
    
    print(f"    - 真實樣本標準化因子範圍: {np.nanmin(normalization_factors_real):.4f} - {np.nanmax(normalization_factors_real):.4f}")
    
    # 4b. QC 樣本的 PQN（不做肌酐校正）
    normalization_factors_qc = None
    qc_data_final = None
    
    if qc_count > 0:
        qc_data = data_matrix[:, qc_indices]
        quotients_qc = qc_data / reference_sample[:, np.newaxis]
        quotients_qc = np.where(np.isfinite(quotients_qc), quotients_qc, np.nan)
        normalization_factors_qc = np.nanmedian(quotients_qc, axis=0)
        
        qc_data_final = qc_data / normalization_factors_qc
        
        print(f"    - QC 樣本標準化因子範圍: {np.nanmin(normalization_factors_qc):.4f} - {np.nanmax(normalization_factors_qc):.4f}")
    
    # ========== Step 5: 合併結果 ==========
    final_data = np.full_like(data_matrix, np.nan)
    final_data[:, real_indices] = real_data_final
    if qc_count > 0:
        final_data[:, qc_indices] = qc_data_final
    
    print("  ✓ 混合標準化完成")
    
    # 返回資訊
    pqn_info = {
        'reference_strategy': reference_strategy,
        'qc_count': qc_count,
        'qc_cv': qc_cv_median,
        'real_count': real_count,
        'normalization_factors_real': normalization_factors_real,
        'normalization_factors_qc': normalization_factors_qc,
        'creatinine_median': median_ref,
        'creatinine_valid_count': np.sum(valid_ref_mask)
    }
    
    return final_data, pqn_info

def get_all_sample_columns(df, sample_info_df):
    """
    獲取所有樣本欄位（包含 QC，但排除統計欄位）
    """
    # 排除的統計欄位關鍵字
    exclude_keywords = [
        'CV', 'Silhouette', 'Permutation', 'Correlation', 
        'R_squared', 'Improvement', 'Before', 'After', 
        'Original', 'Corrected', 'pvalue', 'p_value'
    ]
    
    sample_columns = []
    
    for col in df.columns:
        if col == df.columns[0]:  # 跳過第一欄（特徵ID）
            continue
        
        # 檢查是否包含排除關鍵字
        is_stat_column = any(keyword.lower() in str(col).lower() for keyword in exclude_keywords)
        
        if not is_stat_column:
            sample_columns.append(col)  # ← 不排除 QC
    
    return sample_columns

def calculate_cohens_d(group1, group2):
    """
    計算 Cohen's d (effect size)
    
    Parameters:
    -----------
    group1 : np.ndarray
        第一組數據
    group2 : np.ndarray
        第二組數據
    
    Returns:
    --------
    float : Cohen's d 值
    """
    n1 = len(group1)
    n2 = len(group2)
    
    if n1 < 2 or n2 < 2:
        return np.nan
    
    mean1 = np.mean(group1)
    mean2 = np.mean(group2)
    
    var1 = np.var(group1, ddof=1)
    var2 = np.var(group2, ddof=1)
    
    # Pooled standard deviation
    pooled_std = np.sqrt(((n1 - 1) * var1 + (n2 - 1) * var2) / (n1 + n2 - 2))
    
    if pooled_std == 0:
        return np.nan
    
    # Cohen's d
    d = (mean1 - mean2) / pooled_std
    
    return d

def evaluate_group_difference_preservation(original_data, normalized_data,
                                           sample_info_df, sample_columns,
                                           col_to_info_row=None):
    """
    評估標準化對組間差異的影響

    基於 Sample_Type: CONTROL vs EXPOSURE

    Parameters:
    -----------
    original_data : np.ndarray
        原始數據矩陣 (特徵 x 樣本)
    normalized_data : np.ndarray
        標準化後數據矩陣 (特徵 x 樣本)
    sample_info_df : pd.DataFrame
        樣本資訊表
    sample_columns : list
        樣本名稱列表
    
    Returns:
    --------
    dict or None : 組間差異評估結果
    """
    
    # 1. 提取組別資訊
    sample_groups = np.array([
        _lookup_sample_type(s, sample_info_df, col_to_info_row)
        for s in sample_columns
    ])

    # 2. 識別 CONTROL 和 EXPOSURE
    control_indices = np.where(sample_groups == 'CONTROL')[0]
    exposure_indices = np.where(sample_groups == 'EXPOSURE')[0]
    
    if len(control_indices) == 0 or len(exposure_indices) == 0:
        print(f"\n【組間差異評估】")
        print(f"  ⚠ 警告: 未找到 CONTROL 或 EXPOSURE 組別")
        print(f"    找到的組別類型: {np.unique(sample_groups)}")
        print(f"  跳過組間差異評估")
        return None
    
    print(f"\n【組間差異評估】")
    print(f"  Control 組樣本數: {len(control_indices)}")
    print(f"  Exposure 組樣本數: {len(exposure_indices)}")
    
    # 3. 計算 Cohen's d
    n_features = original_data.shape[0]
    cohens_d_before = []
    cohens_d_after = []
    
    print(f"  計算 Effect Size (Cohen's d)...", end="")
    
    for i in range(n_features):
        # 標準化前
        control_before = original_data[i, control_indices]
        exposure_before = original_data[i, exposure_indices]
        
        control_before = control_before[~np.isnan(control_before)]
        exposure_before = exposure_before[~np.isnan(exposure_before)]
        
        d_before = calculate_cohens_d(control_before, exposure_before)
        cohens_d_before.append(d_before)
        
        # 標準化後
        control_after = normalized_data[i, control_indices]
        exposure_after = normalized_data[i, exposure_indices]
        
        control_after = control_after[~np.isnan(control_after)]
        exposure_after = exposure_after[~np.isnan(exposure_after)]
        
        d_after = calculate_cohens_d(control_after, exposure_after)
        cohens_d_after.append(d_after)
        
        if (i + 1) % 100 == 0:
            print(f"\r  計算 Effect Size... {i+1}/{n_features}", end="")
    
    print(f"\r  ✓ 計算完成 ({n_features}/{n_features})")
    
    cohens_d_before = np.array(cohens_d_before)
    cohens_d_after = np.array(cohens_d_after)

    # 4. 分析 Effect Size 變化
    valid_mask = ~(np.isnan(cohens_d_before) | np.isnan(cohens_d_after))

    d_before_valid = cohens_d_before[valid_mask]
    d_after_valid = cohens_d_after[valid_mask]

    # 計算相對變化（百分比）
    d_change = np.abs(d_after_valid) - np.abs(d_before_valid)
    d_change_pct = (d_change / (np.abs(d_before_valid) + 1e-10)) * 100

    # 分類
    enhanced = np.sum(d_change > 0)  # Effect size 增強
    stable = np.sum(np.abs(d_change_pct) <= 10)  # 變化 < 10%
    mild_reduction = np.sum((d_change_pct < -10) & (d_change_pct >= -30))
    severe_reduction = np.sum(d_change_pct < -30)

    total = len(d_change)

    # 平均保留率
    avg_preservation = np.mean(np.abs(d_after_valid) / (np.abs(d_before_valid) + 1e-10)) * 100

    # ========== 新增：統計檢驗與 FDR 校正 ==========
    wilcoxon_stat = np.nan
    wilcoxon_pvalue = np.nan
    q_values = None
    flagged_features = []
    flagged_ratio = 0.0

    # 檢查是否有足夠的有效特徵進行統計檢驗
    if total >= 3:
        try:
            from scipy.stats import wilcoxon

            # Wilcoxon signed-rank test (配對雙尾檢驗)
            # 檢驗標準化前後 Cohen's d 絕對值是否有顯著差異
            try:
                wilcoxon_stat, wilcoxon_pvalue = wilcoxon(
                    np.abs(d_before_valid),
                    np.abs(d_after_valid),
                    alternative='two-sided'
                )

                # FDR 校正（Benjamini-Hochberg）
                # 為每個特徵計算個別的 p 值（這裡我們使用配對差異的符號檢驗作為簡化）
                # 實際上，對於 Cohen's d 的變化，我們關注的是整體趨勢
                # 因此這裡使用 Wilcoxon 檢驗的 p 值作為全局顯著性指標

                # 為每個特徵分配相同的校正 p 值（因為是全局檢驗）
                # 在實際應用中，如果需要特徵級別的 FDR，需要對每個特徵進行獨立檢驗
                q_values = np.full(total, wilcoxon_pvalue)

                # 識別關鍵特徵：q < 0.05 且 Cohen's d 下降超過 30%
                if wilcoxon_pvalue < 0.05:
                    flagged_mask = d_change_pct < -30
                    flagged_indices = np.where(valid_mask)[0][flagged_mask]

                    for idx in flagged_indices:
                        flagged_features.append({
                            'feature_index': idx,
                            'cohens_d_before': cohens_d_before[idx],
                            'cohens_d_after': cohens_d_after[idx],
                            'change_pct': ((np.abs(cohens_d_after[idx]) - np.abs(cohens_d_before[idx])) /
                                         (np.abs(cohens_d_before[idx]) + 1e-10)) * 100,
                            'q_value': wilcoxon_pvalue
                        })

                    flagged_ratio = len(flagged_features) / total

            except Exception as e:
                print(f"  ⚠ Wilcoxon 檢驗警告: {e}")
                # 如果檢驗失敗（例如所有差異為0），保持 NaN 值
                pass

        except ImportError as e:
            print(f"  ⚠ 統計檢驗套件導入失敗: {e}")
            print(f"     請確保已安裝 scipy 和 statsmodels")
    else:
        print(f"  ⚠ 有效特徵數 ({total}) 不足，跳過統計檢驗（至少需要 3 個）")

    print(f"\n  【Effect Size 變化統計】")
    print(f"  分析特徵數: {total}")
    print(f"  - 增強: {enhanced} ({enhanced/total*100:.1f}%)")
    print(f"  - 穩定 (±10%): {stable} ({stable/total*100:.1f}%)")
    print(f"  - 輕度減弱 (-10% ~ -30%): {mild_reduction} ({mild_reduction/total*100:.1f}%)")
    print(f"  - 顯著減弱 (< -30%): {severe_reduction} ({severe_reduction/total*100:.1f}%)")
    print(f"\n  平均 Effect Size 保留率: {avg_preservation:.1f}%")

    # 輸出統計檢驗結果
    if not np.isnan(wilcoxon_pvalue):
        print(f"\n  【統計檢驗】")
        print(f"  Wilcoxon signed-rank test:")
        print(f"  - 統計量: {wilcoxon_stat:.2f}")
        print(f"  - p-value: {wilcoxon_pvalue:.4f}")

        if wilcoxon_pvalue < 0.05:
            median_change_pct = np.median(d_change_pct)
            if median_change_pct < 0:
                print(f"  - 結論: ✗ Cohen's d 中位數顯著下降 ({median_change_pct:.1f}%)")
            else:
                print(f"  - 結論: ✓ Cohen's d 中位數顯著上升 ({median_change_pct:.1f}%)")
        else:
            print(f"  - 結論: ○ Cohen's d 中位數變化不顯著")

        if len(flagged_features) > 0:
            print(f"\n  【標記特徵】")
            print(f"  顯著改變的特徵數: {len(flagged_features)} ({flagged_ratio*100:.1f}%)")
            print(f"  (標準: q < 0.05 且 |Cohen's d| 下降 > 30%)")

    # 評估
    if severe_reduction / total > 0.1:
        print(f"\n  結論: ⚠⚠ 超過 10% 的特徵顯著減弱，需要檢查")
    elif severe_reduction / total > 0.05:
        print(f"  結論: ⚠ 約 5-10% 的特徵顯著減弱，建議關注")
    elif avg_preservation > 90:
        print(f"  結論: ✓✓ 組間差異保留優秀")
    elif avg_preservation > 80:
        print(f"  結論: ✓ 組間差異保留良好")
    else:
        print(f"  結論: ○ 組間差異保留尚可")

    return {
        'cohens_d_before': cohens_d_before,
        'cohens_d_after': cohens_d_after,
        'enhanced': enhanced,
        'stable': stable,
        'mild_reduction': mild_reduction,
        'severe_reduction': severe_reduction,
        'avg_preservation': avg_preservation,
        'total': total,
        'control_count': len(control_indices),
        'exposure_count': len(exposure_indices),
        # 新增的統計檢驗結果
        'wilcoxon_stat': wilcoxon_stat,
        'wilcoxon_pvalue': wilcoxon_pvalue,
        'q_values': q_values,
        'flagged_features': flagged_features,
        'flagged_ratio': flagged_ratio
    }

def plot_qc_variability(original_qc, normalized_qc, qc_names, output_path):
    """
    Fig4 - QC Variability Assessment (Improved)

    包含：
    1. CV% 分佈對比
    2. CV% 改善散點圖
    3. 盒鬚圖對比
    4. 客觀標準評級（取代主觀星級）

    Parameters:
    -----------
    original_qc : np.ndarray
        原始 QC 數據 (特徵 x QC樣本)
    normalized_qc : np.ndarray
        標準化後 QC 數據
    qc_names : list
        QC 樣本名稱
    output_path : Path
        輸出路徑
    """
    # 移除底部大型備註/解釋框，讓主圖占比更高
    fig = plt.figure(figsize=(18, 6.5))
    gs = fig.add_gridspec(1, 3, hspace=0.25, wspace=0.3)

    cv_before = calculate_rsd(original_qc)
    cv_after = calculate_rsd(normalized_qc)

    # === 子圖 1: CV% 分佈對比 ===
    ax1 = fig.add_subplot(gs[0, 0])

    ax1.hist(cv_before, bins=30, alpha=0.6, color='blue', label='Before', edgecolor='black')
    ax1.hist(cv_after, bins=30, alpha=0.6, color='red', label='After', edgecolor='black')

    ax1.axvline(x=np.median(cv_before), color='blue', linestyle='--', linewidth=2,
                label=f'Median Before: {np.median(cv_before):.1f}%')
    ax1.axvline(x=np.median(cv_after), color='red', linestyle='--', linewidth=2,
                label=f'Median After: {np.median(cv_after):.1f}%')
    ax1.axvline(x=20, color='orange', linestyle=':', linewidth=2, label='20% threshold')

    ax1.set_xlabel('CV%', fontsize=10, fontweight='bold')
    ax1.set_ylabel('Frequency', fontsize=10, fontweight='bold')
    ax1.set_title('QC Sample CV% Distribution', fontsize=12, fontweight='bold')
    ax1.legend(fontsize=8)
    ax1.grid(True, alpha=0.3)

    # === 子圖 5: CV% 改善散點圖 ===
    ax5 = fig.add_subplot(gs[0, 1])

    ax5.scatter(cv_before, cv_after, alpha=0.5, s=20, color='steelblue')

    max_cv = max(np.max(cv_before), np.max(cv_after))
    ax5.plot([0, max_cv], [0, max_cv], 'r--', linewidth=2, label='No change')
    ax5.plot([0, max_cv], [0, max_cv*0.8], 'orange', linestyle='--',
             linewidth=1, alpha=0.5, label='-20%')

    ax5.set_xlabel('CV% Before', fontsize=10, fontweight='bold')
    ax5.set_ylabel('CV% After', fontsize=10, fontweight='bold')
    ax5.set_title('Feature-wise CV% Change', fontsize=12, fontweight='bold')
    ax5.legend(fontsize=8)
    ax5.grid(True, alpha=0.3)

    # === 子圖 6: 盒鬚圖對比 ===
    ax6 = fig.add_subplot(gs[0, 2])

    box_data = [cv_before, cv_after]
    bp = ax6.boxplot(box_data, labels=['Before', 'After'], patch_artist=True,
                     showfliers=True, widths=0.6)

    bp['boxes'][0].set_facecolor('lightblue')
    bp['boxes'][1].set_facecolor('lightcoral')

    ax6.set_ylabel('CV%', fontsize=10, fontweight='bold')
    ax6.set_title('QC CV% Distribution Comparison', fontsize=12, fontweight='bold')
    ax6.grid(True, alpha=0.3, axis='y')
    ax6.axhline(y=20, color='orange', linestyle='--', linewidth=1, alpha=0.5, label='20% threshold')
    ax6.legend(fontsize=8)

    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"  ✓ QC Variability 圖已儲存 (Fig4 - Improved)")


def plot_qc_reproducibility(original_qc, normalized_qc, qc_names, output_path):
    """
    Fig5 - QC Reproducibility Assessment (Improved)

    包含：
    1. QC 樣本總強度
    2. QC 相關性熱圖（Before）
    3. QC 相關性熱圖（After）

    Parameters:
    -----------
    original_qc : np.ndarray
        原始 QC 數據 (特徵 x QC樣本)
    normalized_qc : np.ndarray
        標準化後 QC 數據
    qc_names : list
        QC 樣本名稱
    output_path : Path
        輸出路徑
    """
    # 移除底部大型備註/解釋框，讓主圖占比更高
    fig = plt.figure(figsize=(18, 6.5))
    gs = fig.add_gridspec(1, 3, hspace=0.25, wspace=0.3)

    # === 子圖 2: QC 樣本總強度 ===
    ax2 = fig.add_subplot(gs[0, 0])

    total_before = np.sum(original_qc, axis=0)
    total_after = np.sum(normalized_qc, axis=0)

    x_pos = np.arange(len(qc_names))
    width = 0.35

    ax2.bar(x_pos - width/2, total_before, width, label='Before', alpha=0.7, color='blue')
    ax2.bar(x_pos + width/2, total_after, width, label='After', alpha=0.7, color='red')

    ax2.axhline(y=np.median(total_before), color='blue', linestyle='--', linewidth=1, alpha=0.5)
    ax2.axhline(y=np.median(total_after), color='red', linestyle='--', linewidth=1, alpha=0.5)

    ax2.set_ylabel('Total Intensity', fontsize=10, fontweight='bold')
    ax2.set_title('QC Sample Total Intensity', fontsize=12, fontweight='bold')
    ax2.set_xticks(x_pos)
    ax2.set_xticklabels(qc_names, rotation=45, fontsize=8, ha='right')
    ax2.legend(fontsize=8)
    ax2.grid(True, alpha=0.3, axis='y')

    # === 子圖 3: QC 相關性熱圖（標準化前）===
    ax3 = fig.add_subplot(gs[0, 1])

    corr_before = np.corrcoef(original_qc.T)
    im3 = ax3.imshow(corr_before, cmap='coolwarm', vmin=0.9, vmax=1, aspect='auto')
    ax3.set_xticks(range(len(qc_names)))
    ax3.set_yticks(range(len(qc_names)))
    ax3.set_xticklabels(qc_names, rotation=45, fontsize=8, ha='right')
    ax3.set_yticklabels(qc_names, fontsize=8)
    ax3.set_title('QC Correlation (Before)', fontsize=12, fontweight='bold')
    plt.colorbar(im3, ax=ax3, label='Correlation', fraction=0.046)

    # 在格子中顯示數值
    for i in range(len(qc_names)):
        for j in range(len(qc_names)):
            text = ax3.text(j, i, f'{corr_before[i, j]:.2f}',
                           ha="center", va="center", color="black", fontsize=7)

    # === 子圖 4: QC 相關性熱圖（標準化後）===
    ax4 = fig.add_subplot(gs[0, 2])

    corr_after = np.corrcoef(normalized_qc.T)
    im4 = ax4.imshow(corr_after, cmap='coolwarm', vmin=0.9, vmax=1, aspect='auto')
    ax4.set_xticks(range(len(qc_names)))
    ax4.set_yticks(range(len(qc_names)))
    ax4.set_xticklabels(qc_names, rotation=45, fontsize=8, ha='right')
    ax4.set_yticklabels(qc_names, fontsize=8)
    ax4.set_title('QC Correlation (After)', fontsize=12, fontweight='bold')
    plt.colorbar(im4, ax=ax4, label='Correlation', fraction=0.046)

    for i in range(len(qc_names)):
        for j in range(len(qc_names)):
            text = ax4.text(j, i, f'{corr_after[i, j]:.2f}',
                           ha="center", va="center", color="black", fontsize=7)

        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"  ✓ QC Reproducibility 圖已儲存 (Fig5 - Improved)")

# ==================== 輔助函數 ====================

def is_numeric_value(value):
    """檢查值是否為有效的數值"""
    if pd.isna(value):
        return False
    try:
        float_val = float(value)
        return float_val > 0
    except (ValueError, TypeError):
        return False

def get_non_qc_columns(df, sample_info_df):
    """獲取非QC樣本的欄位列表"""
    non_qc_columns = []

    # Vectorized sample type dict building (faster than iterrows)
    sample_names = sample_info_df.iloc[:, 0].astype(str)
    sample_types = sample_info_df.get('Sample_Type', pd.Series([''] * len(sample_info_df))).fillna('').astype(str).str.upper()
    sample_type_dict = dict(zip(sample_names, sample_types))

    for col in df.columns:
        if col != df.columns[0]:
            if col in sample_type_dict:
                if sample_type_dict[col] != 'QC':
                    non_qc_columns.append(col)
            else:
                if not col.upper().startswith('QC'):
                    non_qc_columns.append(col)
    
    return non_qc_columns

def get_sample_columns_only(df, sample_info_df):
    """
    獲取純樣本欄位（排除QC和統計欄位）
    """
    # 排除的統計欄位關鍵字
    exclude_keywords = [
        'CV', 'Silhouette', 'Permutation', 'Correlation', 
        'R_squared', 'Improvement', 'Before', 'After', 
        'Original', 'Corrected', 'pvalue', 'p_value'
    ]
    
    sample_columns = []

    # Vectorized sample type dict building (faster than iterrows)
    sample_names = sample_info_df.iloc[:, 0].astype(str)
    sample_types = sample_info_df.get('Sample_Type', pd.Series(['Unknown'] * len(sample_info_df))).fillna('Unknown').astype(str).str.upper()
    sample_type_dict = dict(zip(sample_names, sample_types))

    for col in df.columns:
        if col == df.columns[0]:  # 跳過第一欄（特徵ID）
            continue
        
        # 檢查是否包含排除關鍵字
        is_stat_column = any(keyword.lower() in str(col).lower() for keyword in exclude_keywords)
        
        if not is_stat_column:
            # 檢查是否為QC樣本
            if col in sample_type_dict:
                if sample_type_dict[col] != 'QC':
                    sample_columns.append(col)
            else:
                if not col.upper().startswith('QC'):
                    sample_columns.append(col)
    
    return sample_columns

def calculate_cv_per_feature(data_matrix):
    """
    計算每個特徵的CV%
    
    Parameters:
    -----------
    data_matrix : np.ndarray
        數據矩陣 (特徵 x 樣本)
    
    Returns:
    --------
    cv_values : np.ndarray
        每個特徵的CV%
    """
    mean = np.nanmean(data_matrix, axis=1)
    std = np.nanstd(data_matrix, axis=1)
    cv = (std / mean) * 100
    return cv

def calculate_rsd(data_matrix):
    """計算相對標準偏差 (RSD%)"""
    mean = np.nanmean(data_matrix, axis=1, keepdims=True)
    std = np.nanstd(data_matrix, axis=1, keepdims=True)
    rsd = (std / mean) * 100
    return rsd.flatten()

def calculate_sample_correlation(data_matrix):
    """計算樣本間的相關性"""
    # 移除含有NaN的特徵
    valid_features = ~np.isnan(data_matrix).any(axis=1)
    clean_data = data_matrix[valid_features, :]
    
    if clean_data.shape[0] < 3:
        return np.nan, np.nan
    
    corr_matrix = np.corrcoef(clean_data.T)
    mask = ~np.eye(corr_matrix.shape[0], dtype=bool)
    correlations = corr_matrix[mask]
    return np.mean(correlations), np.std(correlations)

# ==================== 視覺化函數 ====================

def plot_density_comparison(original_data, normalized_data, sample_names, output_path, method_name):
    """繪製標準化前後的密度圖對比"""
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))
    
    # 標準化前
    for i in range(min(original_data.shape[1], 30)):  # 最多顯示30個樣本
        sample_data = original_data[:, i]
        sample_data = sample_data[sample_data > 0]
        if len(sample_data) > 0:
            ax1.hist(np.log10(sample_data), bins=50, alpha=0.3, density=True)
    
    ax1.set_xlabel('Log10(Intensity)', fontsize=12, fontweight='bold')
    ax1.set_ylabel('Density', fontsize=12, fontweight='bold')
    ax1.set_title('Before Normalization', fontsize=14, fontweight='bold')
    ax1.grid(True, alpha=0.3)
    
    # 標準化後
    for i in range(min(normalized_data.shape[1], 30)):
        sample_data = normalized_data[:, i]
        sample_data = sample_data[sample_data > 0]
        if len(sample_data) > 0:
            ax2.hist(np.log10(sample_data), bins=50, alpha=0.3, density=True)
    
    ax2.set_xlabel('Log10(Intensity)', fontsize=12, fontweight='bold')
    ax2.set_ylabel('Density', fontsize=12, fontweight='bold')
    ax2.set_title(f'After Normalization ({method_name})', fontsize=14, fontweight='bold')
    ax2.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"  ✓ 密度圖已儲存")

def plot_boxplot_comparison(original_data, normalized_data, sample_names, output_path, method_name, sample_info_df, col_to_info_row=None):
    """
    繪製標準化前後的盒鬚圖與樣本總強度 (Fig 1 - Integrated)

    改進項目：
    - 標題包含樣本數統計
    - x 軸標籤用顏色區分 (Control/Exposed/QC)
    - 統計摘要框加入總強度指標
    - 同一圖表內整合樣本總強度長條圖（取代舊 Fig2）
    """
    from scipy.stats import levene

    COLOR_SCHEME = {
        'CONTROL': '#3498DB',
        'EXPOSURE': '#E74C3C',
        'QC': '#F39C12',
        'Unknown': '#95A5A6'
    }

    sample_groups = [
        _lookup_sample_type(s, sample_info_df, col_to_info_row, default='Unknown')
        for s in sample_names
    ]

    from collections import Counter
    group_counts = Counter(sample_groups)
    n_control = group_counts.get('CONTROL', 0)
    n_exposure = group_counts.get('EXPOSURE', 0)
    n_qc = group_counts.get('QC', 0)

    original_log = np.log10(original_data + 1)
    normalized_log = np.log10(normalized_data + 1)

    median_before = np.median([np.median(original_log[:, i]) for i in range(len(sample_names))])
    median_after = np.median([np.median(normalized_log[:, i]) for i in range(len(sample_names))])

    sample_medians_before = np.array([np.median(original_log[:, i]) for i in range(len(sample_names))])
    sample_medians_after = np.array([np.median(normalized_log[:, i]) for i in range(len(sample_names))])

    rsd_before = (np.std(sample_medians_before, ddof=1) / np.mean(sample_medians_before)) * 100
    rsd_after = (np.std(sample_medians_after, ddof=1) / np.mean(sample_medians_after)) * 100
    rsd_reduction = ((rsd_before - rsd_after) / rsd_before) * 100

    original_totals = np.nansum(original_data, axis=0)
    normalized_totals = np.nansum(normalized_data, axis=0)
    total_median_before = np.nanmedian(original_totals)
    total_median_after = np.nanmedian(normalized_totals)

    total_cv_before = (np.nanstd(original_totals) / np.nanmean(original_totals)) * 100 if np.nanmean(original_totals) > 0 else np.nan
    total_cv_after = (np.nanstd(normalized_totals) / np.nanmean(normalized_totals)) * 100 if np.nanmean(normalized_totals) > 0 else np.nan
    if np.isfinite(total_cv_before) and total_cv_before != 0:
        total_cv_reduction = ((total_cv_before - total_cv_after) / total_cv_before) * 100
    else:
        total_cv_reduction = np.nan

    try:
        levene_stat, levene_p = levene(*[original_log[:, i] for i in range(len(sample_names))])
        if levene_p < 0.001:
            sig_mark = '***'
        elif levene_p < 0.01:
            sig_mark = '**'
        elif levene_p < 0.05:
            sig_mark = '*'
        else:
            sig_mark = 'n.s.'
    except Exception:
        levene_p = np.nan
        sig_mark = 'N/A'

    fig = plt.figure(figsize=(max(18, len(sample_names) * 0.55), 14))
    gs = fig.add_gridspec(3, 1, height_ratios=[2.2, 2.2, 1.7], hspace=0.35)
    ax1 = fig.add_subplot(gs[0, 0])
    ax2 = fig.add_subplot(gs[1, 0])
    ax3 = fig.add_subplot(gs[2, 0])

    positions = np.arange(len(sample_names))
    bp1 = ax1.boxplot([original_log[:, i] for i in range(len(sample_names))],
                      positions=positions,
                      widths=0.6,
                      patch_artist=True,
                      showfliers=False)

    for patch, group in zip(bp1['boxes'], sample_groups):
        patch.set_facecolor(COLOR_SCHEME.get(group, '#95A5A6'))
        patch.set_alpha(0.7)

    ax1.axhline(y=median_before, color='red', linestyle='--', linewidth=2,
                label=f'Median: {median_before:.2f}', alpha=0.7)

    ax1.set_ylabel('Log10(Intensity + 1)', fontsize=12, fontweight='bold')
    ax1.set_title(
        f'Sample Intensity Distribution: PQN Normalization Effect\n(n={len(sample_names)}: {n_control} Control, {n_exposure} Exposed, {n_qc} QC) - Before',
        fontsize=14, fontweight='bold')
    ax1.set_xticks(positions)
    ax1.set_xticklabels(sample_names, rotation=90, fontsize=8)
    for tick_label, group in zip(ax1.get_xticklabels(), sample_groups):
        tick_label.set_color(COLOR_SCHEME.get(group, '#95A5A6'))
        tick_label.set_fontweight('bold')

    ax1.grid(True, alpha=0.3, axis='y')
    ax1.legend(loc='upper right', fontsize=10)

    bp2 = ax2.boxplot([normalized_log[:, i] for i in range(len(sample_names))],
                      positions=positions,
                      widths=0.6,
                      patch_artist=True,
                      showfliers=False)

    for patch, group in zip(bp2['boxes'], sample_groups):
        patch.set_facecolor(COLOR_SCHEME.get(group, '#95A5A6'))
        patch.set_alpha(0.7)

    ax2.axhline(y=median_after, color='red', linestyle='--', linewidth=2,
                label=f'Median: {median_after:.2f}', alpha=0.7)

    ax2.set_ylabel('Log10(Intensity + 1)', fontsize=12, fontweight='bold')
    ax2.set_title(f'After Normalization ({method_name})', fontsize=14, fontweight='bold')
    ax2.set_xticks(positions)
    ax2.set_xticklabels(sample_names, rotation=90, fontsize=8)
    for tick_label, group in zip(ax2.get_xticklabels(), sample_groups):
        tick_label.set_color(COLOR_SCHEME.get(group, '#95A5A6'))
        tick_label.set_fontweight('bold')

    ax2.grid(True, alpha=0.3, axis='y')
    ax2.legend(loc='upper right', fontsize=10)

    bar_width = 0.42
    ax3.bar(positions - bar_width / 2, original_totals, width=bar_width,
            color='#4C72B0', alpha=0.75, label='Before normalization')
    ax3.bar(positions + bar_width / 2, normalized_totals, width=bar_width,
            color='#ED553B', alpha=0.75, label=f'After normalization ({method_name})')

    ax3.axhline(y=total_median_before, color='#4C72B0', linestyle='--', linewidth=1.5,
                alpha=0.8, label=f'Before median: {total_median_before:.2e}')
    ax3.axhline(y=total_median_after, color='#ED553B', linestyle='--', linewidth=1.5,
                alpha=0.8, label=f'After median: {total_median_after:.2e}')

    ax3.set_ylabel('Total Intensity', fontsize=12, fontweight='bold')
    ax3.set_title('Sample Total Intensity Overview (linear scale)', fontsize=14, fontweight='bold')
    ax3.set_xticks(positions)
    ax3.set_xticklabels(sample_names, rotation=90, fontsize=8)
    for tick_label, group in zip(ax3.get_xticklabels(), sample_groups):
        tick_label.set_color(COLOR_SCHEME.get(group, '#95A5A6'))
        tick_label.set_fontweight('bold')

    ax3.grid(True, alpha=0.25, axis='y')
    ax3.legend(loc='upper right', fontsize=9, ncol=2)

    stats_text = f"""Statistical Summary:
Median intensity:
  Before: {median_before:.2f}
  After:  {median_after:.2f}

Inter-sample RSD:
  Before: {rsd_before:.2f}%
  After:  {rsd_after:.2f}%
  Reduction: {rsd_reduction:.1f}%

Sample total CV:
  Before: {total_cv_before:.2f}%
  After:  {total_cv_after:.2f}%
  Reduction: {total_cv_reduction:.1f}%

Levene's test:
  p = {levene_p:.4f} {sig_mark}
"""

    ax1.text(1.02, 0.5, stats_text, transform=ax1.transAxes,
             fontsize=10, verticalalignment='center',
             bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8, edgecolor='black', linewidth=1.5),
             family='monospace')

    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"  ✓ 盒鬚圖已儲存 (Fig 1 - Integrated)")


def plot_rle(original_data, normalized_data, sample_names, output_path, method_name):
    """
    繪製 RLE (Relative Log Expression) Plot

    RLE 是組學數據正規化品質評估的黃金標準
    顯示每個樣本相對於中位數的偏差分佈

    Args:
        original_data: 原始數據矩陣 (features × samples)
        normalized_data: 正規化後數據矩陣 (features × samples)
        sample_names: 樣本名稱列表
        output_path: 輸出路徑
        method_name: 正規化方法名稱
    """
    print(f"\n  繪製 RLE Plot...")

    # 計算 RLE
    def calculate_rle(data):
        """
        計算 Relative Log Expression

        對每個 feature，計算 log2(sample_intensity / median_intensity)
        """
        # 移除全為 0 或 NaN 的 features
        valid_features = ~np.all((data == 0) | np.isnan(data), axis=1)
        data_valid = data[valid_features, :]

        if data_valid.shape[0] == 0:
            print("    ⚠️  警告：沒有有效的 features，無法計算 RLE")
            return None

        # 對數據進行 log2 轉換（加 1 避免 log(0)）
        data_log = np.log2(data_valid + 1)

        # 計算每個 feature 的中位數（跨所有樣本）
        median_per_feature = np.median(data_log, axis=1, keepdims=True)

        # 計算 RLE：log2(sample) - log2(median)
        rle = data_log - median_per_feature

        return rle

    original_rle = calculate_rle(original_data)
    normalized_rle = calculate_rle(normalized_data)

    if original_rle is None or normalized_rle is None:
        print("    ⚠️  警告：RLE 計算失敗，跳過繪圖")
        return

    # 創建圖表
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(max(20, len(sample_names) * 0.6), 16))

    # ===== 上圖：Original RLE =====
    bp1 = ax1.boxplot([original_rle[:, i] for i in range(original_rle.shape[1])],
                       labels=sample_names, patch_artist=True,
                       widths=0.6, showfliers=False)

    # 設定箱型圖顏色
    for patch in bp1['boxes']:
        patch.set_facecolor('lightcoral')
        patch.set_alpha(0.7)
        patch.set_edgecolor('darkred')
        patch.set_linewidth(1.5)

    for whisker in bp1['whiskers']:
        whisker.set(color='darkred', linewidth=1.5, linestyle='-')

    for cap in bp1['caps']:
        cap.set(color='darkred', linewidth=1.5)

    for median in bp1['medians']:
        median.set(color='red', linewidth=2.5)

    # 添加基準線（理想狀態應該在 0）
    ax1.axhline(y=0, color='green', linestyle='--', linewidth=2, label='理想基準 (RLE = 0)', zorder=1)

    ax1.set_xlabel('Sample', fontsize=14, fontweight='bold')
    ax1.set_ylabel('RLE (Log2 Ratio)', fontsize=14, fontweight='bold')
    ax1.set_title(f'RLE Plot - Original Data\n(Before {method_name} Normalization)',
                  fontsize=16, fontweight='bold', pad=20)
    ax1.tick_params(axis='x', rotation=90, labelsize=10)
    ax1.tick_params(axis='y', labelsize=12)
    ax1.legend(fontsize=12, loc='upper right')
    ax1.grid(True, alpha=0.3, linestyle='--', axis='y')

    # 計算 RLE 中位數絕對偏差 (MAD) - 品質指標
    original_mad = np.median([np.median(np.abs(original_rle[:, i])) for i in range(original_rle.shape[1])])
    ax1.text(0.02, 0.98, f'Median Absolute Deviation: {original_mad:.4f}',
             transform=ax1.transAxes, fontsize=12, verticalalignment='top',
             bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))

    # ===== 下圖：Normalized RLE =====
    bp2 = ax2.boxplot([normalized_rle[:, i] for i in range(normalized_rle.shape[1])],
                       labels=sample_names, patch_artist=True,
                       widths=0.6, showfliers=False)

    # 設定箱型圖顏色
    for patch in bp2['boxes']:
        patch.set_facecolor('lightblue')
        patch.set_alpha(0.7)
        patch.set_edgecolor('darkblue')
        patch.set_linewidth(1.5)

    for whisker in bp2['whiskers']:
        whisker.set(color='darkblue', linewidth=1.5, linestyle='-')

    for cap in bp2['caps']:
        cap.set(color='darkblue', linewidth=1.5)

    for median in bp2['medians']:
        median.set(color='blue', linewidth=2.5)

    # 添加基準線
    ax2.axhline(y=0, color='green', linestyle='--', linewidth=2, label='理想基準 (RLE = 0)', zorder=1)

    ax2.set_xlabel('Sample', fontsize=14, fontweight='bold')
    ax2.set_ylabel('RLE (Log2 Ratio)', fontsize=14, fontweight='bold')
    ax2.set_title(f'RLE Plot - Normalized Data\n(After {method_name} Normalization)',
                  fontsize=16, fontweight='bold', pad=20)
    ax2.tick_params(axis='x', rotation=90, labelsize=10)
    ax2.tick_params(axis='y', labelsize=12)
    ax2.legend(fontsize=12, loc='upper right')
    ax2.grid(True, alpha=0.3, linestyle='--', axis='y')

    # 計算 Normalized RLE 的 MAD
    normalized_mad = np.median([np.median(np.abs(normalized_rle[:, i])) for i in range(normalized_rle.shape[1])])
    improvement = ((original_mad - normalized_mad) / original_mad * 100) if original_mad > 0 else 0

    ax2.text(0.02, 0.98, f'Median Absolute Deviation: {normalized_mad:.4f}\n'
                          f'Improvement: {improvement:.1f}%',
             transform=ax2.transAxes, fontsize=12, verticalalignment='top',
             bbox=dict(boxstyle='round', facecolor='lightgreen' if improvement > 0 else 'wheat', alpha=0.8))

    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"    ✓ RLE Plot 已儲存")
    print(f"    - Original MAD: {original_mad:.4f}")
    print(f"    - Normalized MAD: {normalized_mad:.4f}")
    print(f"    - Improvement: {improvement:.1f}%")


def plot_cv_comparison(original_cv, normalized_cv, output_path, method_name):
    """
    繪製CV%分佈對比圖 (Fig 2 - Improved)

    改進項目：
    - 標題加警示信息
    - 增加閾值線 (20%, 30%, 50%)
    - 增加統計檢驗框 (Wilcoxon, Cohen's d)
    - 增加解釋框
    """
    from scipy.stats import wilcoxon

    # 創建圖形（移除底部備註框，讓主圖占比更高）
    fig, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(20, 6))

    # 移除 NaN 值
    original_cv_clean = original_cv[~np.isnan(original_cv)]
    normalized_cv_clean = normalized_cv[~np.isnan(normalized_cv)]

    # 確保兩者長度一致（用於配對檢驗）
    min_len = min(len(original_cv_clean), len(normalized_cv_clean))
    original_cv_clean = original_cv_clean[:min_len]
    normalized_cv_clean = normalized_cv_clean[:min_len]

    # 計算統計量
    median_improvement = np.median(original_cv_clean) - np.median(normalized_cv_clean)

    # Wilcoxon 配對檢驗
    try:
        w_stat, p_value = wilcoxon(original_cv_clean, normalized_cv_clean)
        if p_value < 0.001:
            sig_mark = '***'
        elif p_value < 0.01:
            sig_mark = '**'
        elif p_value < 0.05:
            sig_mark = '*'
        else:
            sig_mark = 'n.s.'
    except (ValueError, TypeError) as e:
        # ValueError: sample too small or all values identical
        # TypeError: invalid input types
        w_stat, p_value = np.nan, np.nan
        sig_mark = 'N/A'

    # Cohen's d (效應量)
    pooled_std = np.sqrt((np.var(original_cv_clean, ddof=1) + np.var(normalized_cv_clean, ddof=1)) / 2)
    cohens_d = (np.mean(original_cv_clean) - np.mean(normalized_cv_clean)) / pooled_std

    # 效應量解讀
    abs_d = abs(cohens_d)
    if abs_d < COHENS_D_THRESHOLDS['small']:
        effect_interpretation = 'Negligible'
    elif abs_d < COHENS_D_THRESHOLDS['medium']:
        effect_interpretation = 'Small'
    elif abs_d < COHENS_D_THRESHOLDS['large']:
        effect_interpretation = 'Medium'
    else:
        effect_interpretation = 'Large'

    # 閾值定義
    THRESHOLDS = {
        'good': 20,        # 綠色
        'acceptable': 30,  # 橙色
        'poor': 50         # 紅色
    }

    # === 1. 標準化前CV%分佈 ===
    ax1.hist(original_cv_clean, bins=50, color='steelblue', alpha=0.7, edgecolor='black')
    ax1.axvline(x=np.median(original_cv_clean), color='red', linestyle='--',
                linewidth=2, label=f'Median: {np.median(original_cv_clean):.2f}%')

    # 增加閾值線
    ax1.axvline(x=THRESHOLDS['good'], color='#27AE60', linestyle=':', linewidth=2,
                label=f'Good (<{THRESHOLDS["good"]}%)', alpha=0.7)
    ax1.axvline(x=THRESHOLDS['acceptable'], color='#F39C12', linestyle=':', linewidth=2,
                label=f'Acceptable (<{THRESHOLDS["acceptable"]}%)', alpha=0.7)
    ax1.axvline(x=THRESHOLDS['poor'], color='#E74C3C', linestyle=':', linewidth=2,
                label=f'Poor (>{THRESHOLDS["poor"]}%)', alpha=0.7)

    ax1.set_xlabel('CV%', fontsize=12, fontweight='bold')
    ax1.set_ylabel('Frequency', fontsize=12, fontweight='bold')
    ax1.set_title('Before Normalization', fontsize=14, fontweight='bold')
    ax1.legend(fontsize=8)
    ax1.grid(True, alpha=0.3)

    # === 2. 標準化後CV%分佈 ===
    ax2.hist(normalized_cv_clean, bins=50, color='coral', alpha=0.7, edgecolor='black')
    ax2.axvline(x=np.median(normalized_cv_clean), color='red', linestyle='--',
                linewidth=2, label=f'Median: {np.median(normalized_cv_clean):.2f}%')

    # 增加閾值線
    ax2.axvline(x=THRESHOLDS['good'], color='#27AE60', linestyle=':', linewidth=2,
                label=f'Good (<{THRESHOLDS["good"]}%)', alpha=0.7)
    ax2.axvline(x=THRESHOLDS['acceptable'], color='#F39C12', linestyle=':', linewidth=2,
                label=f'Acceptable (<{THRESHOLDS["acceptable"]}%)', alpha=0.7)
    ax2.axvline(x=THRESHOLDS['poor'], color='#E74C3C', linestyle=':', linewidth=2,
                label=f'Poor (>{THRESHOLDS["poor"]}%)', alpha=0.7)

    ax2.set_xlabel('CV%', fontsize=12, fontweight='bold')
    ax2.set_ylabel('Frequency', fontsize=12, fontweight='bold')
    ax2.set_title(f'After Normalization ({method_name})', fontsize=14, fontweight='bold')
    ax2.legend(fontsize=8)
    ax2.grid(True, alpha=0.3)

    # === 3. CV%改善分佈 ===
    cv_improvement = original_cv_clean - normalized_cv_clean
    ax3.hist(cv_improvement, bins=50, color='green', alpha=0.7, edgecolor='black')
    ax3.axvline(x=0, color='black', linestyle='-', linewidth=2, label='No change')
    ax3.axvline(x=np.median(cv_improvement), color='red', linestyle='--',
                linewidth=2, label=f'Median: {np.median(cv_improvement):.2f}%')

    ax3.set_xlabel('CV% Improvement', fontsize=12, fontweight='bold')
    ax3.set_ylabel('Frequency', fontsize=12, fontweight='bold')
    ax3.set_title('CV% Improvement Distribution', fontsize=14, fontweight='bold')
    ax3.legend(fontsize=10)
    ax3.grid(True, alpha=0.3)

    # === 統計檢驗框 ===
    improved_count = np.sum(cv_improvement > 0)
    total_count = len(cv_improvement)
    improvement_rate = (improved_count / total_count) * 100

    stats_text = f"""Statistical Tests:
Wilcoxon test:
  statistic = {w_stat:.1f}
  p = {p_value:.4f} {sig_mark}

Cohen's d:
  d = {cohens_d:.2f}
  ({effect_interpretation})

Features improved:
  {improved_count}/{total_count}
  ({improvement_rate:.1f}%)
"""

    ax3.text(0.05, 0.95, stats_text, transform=ax3.transAxes,
            fontsize=9, verticalalignment='top',
            bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8, edgecolor='black', linewidth=1.5),
            family='monospace')

    fig.suptitle(
        f'Coefficient of Variation (CV%) Comparison ({method_name})',
        fontsize=16,
        y=0.98,
        fontweight='bold',
    )
    plt.tight_layout(rect=[0, 0, 1, 0.96])
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"  ✓ CV%分佈圖已儲存 (Fig 2 - Improved)")

def plot_pca_with_confidence_ellipse(
    original_data,
    normalized_data,
    sample_names,
    sample_info_df,
    output_path,
    method_name,
    exclude_qc=True,
    col_to_info_row=None,
):
    """
    繪製 PCA 對比圖 (Fig 3 - Improved)

    改進項目：
    - 簡化橢圓：只繪製 Control、Exposed、QC 三種（刪除 "All Samples"）
    - 增加 PERMANOVA 統計框
    - 增加 QC 聚集度指標
    - QC 樣本標記加粗、增大
    """
    from scipy.stats import f as f_dist
    from matplotlib.patches import Ellipse

    # 數據預處理（轉置：樣本 x 特徵）
    original_transposed = original_data.T
    normalized_transposed = normalized_data.T

    # 移除含有NaN的樣本
    valid_samples_orig = ~np.isnan(original_transposed).any(axis=1)
    valid_samples_norm = ~np.isnan(normalized_transposed).any(axis=1)
    valid_samples = valid_samples_orig & valid_samples_norm

    original_clean = original_transposed[valid_samples]
    normalized_clean = normalized_transposed[valid_samples]
    sample_names_clean = [sample_names[i] for i in range(len(sample_names)) if valid_samples[i]]

    # 獲取樣本分組信息
    sample_groups = np.array([
        _lookup_sample_type(s, sample_info_df, col_to_info_row, default='Unknown')
        for s in sample_names_clean
    ])

    # 依需求：濃度校正的 PCA 可排除 QC 樣本
    if exclude_qc:
        non_qc_mask_all = sample_groups != 'QC'
        if np.sum(non_qc_mask_all) < 3:
            print("  ⚠ 非 QC 樣本不足，略過 PCA 對比圖")
            return

        original_clean = original_clean[non_qc_mask_all]
        normalized_clean = normalized_clean[non_qc_mask_all]
        sample_names_clean = [
            sample_names_clean[i] for i in range(len(sample_names_clean)) if non_qc_mask_all[i]
        ]
        sample_groups = sample_groups[non_qc_mask_all]

    # 顏色映射
    color_palette = {
        'QC': '#9B59B6',
        'CONTROL': '#3498DB',
        'EXPOSURE': '#E74C3C',
        'UNKNOWN': '#95A5A6'
    }
    marker_palette = {
        'QC': 'o',          # circle
        'CONTROL': 's',     # square
        'EXPOSURE': '^',    # triangle
        'UNKNOWN': 'd'
    }

    # 標準化（用於PCA）
    scaler_orig = StandardScaler()
    scaler_norm = StandardScaler()

    original_scaled = scaler_orig.fit_transform(original_clean)
    normalized_scaled = scaler_norm.fit_transform(normalized_clean)

    # PCA
    pca_orig = PCA(n_components=2)
    pc_original = pca_orig.fit_transform(original_scaled)
    var_original = pca_orig.explained_variance_ratio_

    pca_norm = PCA(n_components=2)
    pc_normalized = pca_norm.fit_transform(normalized_scaled)
    var_normalized = pca_norm.explained_variance_ratio_

    # === 計算 PERMANOVA (簡化版 - 使用歐氏距離) ===
    def simple_permanova(X, groups, n_permutations=999):
        """簡化版 PERMANOVA"""
        from scipy.spatial.distance import pdist, squareform

        # 計算距離矩陣
        dist_matrix = squareform(pdist(X, metric='euclidean'))

        # 計算 F 統計量
        def calculate_f_stat(dist_mat, grps):
            unique_groups = np.unique(grps)
            n_total = len(grps)

            # Total sum of squares
            grand_centroid = X.mean(axis=0)
            ss_total = np.sum([np.sum((X[i] - grand_centroid)**2) for i in range(len(X))])

            # Within-group sum of squares
            ss_within = 0
            for group in unique_groups:
                mask = grps == group
                if np.sum(mask) > 1:
                    group_centroid = X[mask].mean(axis=0)
                    ss_within += np.sum([np.sum((X[i] - group_centroid)**2) for i in range(len(X)) if mask[i]])

            # Between-group sum of squares
            ss_between = ss_total - ss_within

            # Degrees of freedom
            df_between = len(unique_groups) - 1
            df_within = n_total - len(unique_groups)

            # F statistic
            if df_within > 0 and ss_within > 0:
                f_stat = (ss_between / df_between) / (ss_within / df_within)
                # R² (proportion of variance explained)
                r_squared = ss_between / ss_total
            else:
                f_stat = 0
                r_squared = 0

            return f_stat, r_squared

        obs_f, obs_r2 = calculate_f_stat(dist_matrix, groups)

        # Permutation test
        perm_f_stats = []
        for _ in range(n_permutations):
            perm_groups = np.random.permutation(groups)
            perm_f, _ = calculate_f_stat(dist_matrix, perm_groups)
            perm_f_stats.append(perm_f)

        # Calculate p-value
        p_value = np.sum(np.array(perm_f_stats) >= obs_f) / n_permutations

        return obs_f, obs_r2, p_value

    # 計算 PERMANOVA (Control vs Exposure，排除 QC)
    non_qc_mask = (sample_groups == 'CONTROL') | (sample_groups == 'EXPOSURE')

    if np.sum(non_qc_mask) > 2:
        f_before, r2_before, p_before = simple_permanova(
            pc_original[non_qc_mask], sample_groups[non_qc_mask], n_permutations=999
        )
        f_after, r2_after, p_after = simple_permanova(
            pc_normalized[non_qc_mask], sample_groups[non_qc_mask], n_permutations=999
        )
    else:
        f_before, r2_before, p_before = np.nan, np.nan, np.nan
        f_after, r2_after, p_after = np.nan, np.nan, np.nan

    # === 計算 QC 聚集度指標（若排除 QC，則不計算） ===
    if exclude_qc:
        qc_t2_before, qc_t2_after, qc_t2_reduction_pct = np.nan, np.nan, np.nan
        qc_mean_dist_before, qc_mean_dist_after, qc_dist_reduction_pct = np.nan, np.nan, np.nan
    else:
        qc_mask = sample_groups == 'QC'

        if np.sum(qc_mask) >= 3:
            # Hotelling T² (相對於 QC 中心的平均距離平方)
            qc_centroid_before = pc_original[qc_mask].mean(axis=0)
            qc_centroid_after = pc_normalized[qc_mask].mean(axis=0)

            qc_t2_before = np.mean([np.sum((pt - qc_centroid_before)**2) for pt in pc_original[qc_mask]])
            qc_t2_after = np.mean([np.sum((pt - qc_centroid_after)**2) for pt in pc_normalized[qc_mask]])
            qc_t2_reduction_pct = ((qc_t2_before - qc_t2_after) / qc_t2_before) * 100 if qc_t2_before > 0 else 0

            # Mean distance to centroid
            qc_mean_dist_before = np.mean([np.linalg.norm(pt - qc_centroid_before) for pt in pc_original[qc_mask]])
            qc_mean_dist_after = np.mean([np.linalg.norm(pt - qc_centroid_after) for pt in pc_normalized[qc_mask]])
            qc_dist_reduction_pct = ((qc_mean_dist_before - qc_mean_dist_after) / qc_mean_dist_before) * 100 if qc_mean_dist_before > 0 else 0
        else:
            qc_t2_before, qc_t2_after, qc_t2_reduction_pct = np.nan, np.nan, np.nan
            qc_mean_dist_before, qc_mean_dist_after, qc_dist_reduction_pct = np.nan, np.nan, np.nan

    # ========== 繪圖（統一為 QC 子程式 PCA 風格）==========
    sample_types = [normalize_sample_type(group) for group in sample_groups]

    plot_pca_comparison_qc_style(
        pc_original,
        pc_normalized,
        var_original,
        var_normalized,
        sample_names_clean,
        sample_types,
        batch_labels=None,
        grouping='sample_type',
        suptitle=f'2D PCA Comparison: Before vs After Normalization ({method_name})',
        left_title='Before Normalization',
        right_title=f'After Normalization ({method_name})',
        output_path=output_path,
        dpi=300,
    )

    plt.close('all')

    print(f"  ✓ PCA 對比圖已儲存 (Fig 4)")


def plot_confidence_ellipse(points, ax, color='blue', label=None, n_std=2.447, 
                            linestyle='--', linewidth=2.5):
    """
    繪製 Hotelling's T² 95% 信賴橢圓
    
    Parameters:
    -----------
    points : np.ndarray
        二維數據點 (n_samples, 2)
    ax : matplotlib.axes.Axes
        繪圖軸
    color : str
        橢圓顏色
    label : str
        標籤
    n_std : float
        標準差倍數（2.447 對應 95% 信賴區間）
    linestyle : str
        線條樣式
    linewidth : float
        線條寬度
    """
    from scipy.stats import f as f_dist
    from matplotlib.patches import Ellipse
    
    if len(points) < 3:
        return
    
    # 計算均值和協方差矩陣
    mean = np.mean(points, axis=0)
    cov = np.cov(points.T)
    
    # 計算特徵值和特徵向量
    eigenvalues, eigenvectors = np.linalg.eigh(cov)
    
    # 排序（從大到小）
    order = eigenvalues.argsort()[::-1]
    eigenvalues = eigenvalues[order]
    eigenvectors = eigenvectors[:, order]
    
    # 計算橢圓的角度
    angle = np.degrees(np.arctan2(eigenvectors[1, 0], eigenvectors[0, 0]))
    
    # 計算 Hotelling's T² 的臨界值（95% 信賴區間）
    n = len(points)
    p = 2  # 維度（PC1 和 PC2）
    
    # F 分佈臨界值
    f_critical = f_dist.ppf(0.95, p, n - p)
    
    # Hotelling's T² 臨界值
    chi2_critical = (n - 1) * p / (n - p) * f_critical
    
    # 橢圓的寬度和高度
    width, height = 2 * np.sqrt(eigenvalues * chi2_critical)
    
    # 繪製橢圓
    ellipse = Ellipse(mean, width, height, angle=angle,
                     facecolor='none', edgecolor=color, 
                     linewidth=linewidth, linestyle=linestyle, 
                     alpha=0.6, zorder=2, label=label)
    
    ax.add_patch(ellipse)
    

def plot_oplsda_comparison(original_data, normalized_data, sample_names,
                           sample_info_df, output_path, method_name):
    """
    佔位函式：OPLS-DA/PLS-DA 功能已停用，僅保留 PCA 提示。
    """
    print("  ℹ OPLS-DA/PLS-DA 視覺化已停用；此版本專注於 PCA 分析。")
    print("    若需此功能，請於版本控管中啟用對應模組。")


def _plot_simple_plsda_fallback(*args, **kwargs):
    """保留函式簽名，提供簡短提示以維持相容性。"""
    print("  ℹ PLS-DA 備援邏輯已停用，請參考 PCA 圖以評估組間差異。")


def plot_correlation_heatmap(original_data, normalized_data, sample_names, output_path, method_name):
    """繪製樣本相關性熱圖（標準化前後）"""
    # 計算相關性矩陣
    corr_original = np.corrcoef(original_data.T)
    corr_normalized = np.corrcoef(normalized_data.T)
    
    # 繪圖
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(20, 8))
    
    # 標準化前
    im1 = ax1.imshow(corr_original, cmap='coolwarm', vmin=-1, vmax=1, aspect='auto')
    ax1.set_xticks(range(len(sample_names)))
    ax1.set_yticks(range(len(sample_names)))
    ax1.set_xticklabels(sample_names, rotation=90, fontsize=8)
    ax1.set_yticklabels(sample_names, fontsize=8)
    ax1.set_title('Before Normalization', fontsize=14, fontweight='bold')
    plt.colorbar(im1, ax=ax1, label='Correlation')
    
    # 標準化後
    im2 = ax2.imshow(corr_normalized, cmap='coolwarm', vmin=-1, vmax=1, aspect='auto')
    ax2.set_xticks(range(len(sample_names)))
    ax2.set_yticks(range(len(sample_names)))
    ax2.set_xticklabels(sample_names, rotation=90, fontsize=8)
    ax2.set_yticklabels(sample_names, fontsize=8)
    ax2.set_title(f'After Normalization ({method_name})', fontsize=14, fontweight='bold')
    plt.colorbar(im2, ax=ax2, label='Correlation')
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"  ✓ 相關性熱圖已儲存")

# ==================== 評估函數 ====================

def evaluate_normalization_quality(original_data, normalized_data):
    """綜合評估標準化質量"""
    results = {}
    
    # 1. CV% 評估
    original_cv = calculate_rsd(original_data)
    normalized_cv = calculate_rsd(normalized_data)
    
    results['median_cv_before'] = np.nanmedian(original_cv)
    results['median_cv_after'] = np.nanmedian(normalized_cv)
    results['mean_cv_before'] = np.nanmean(original_cv)
    results['mean_cv_after'] = np.nanmean(normalized_cv)
    results['cv_improvement'] = results['median_cv_before'] - results['median_cv_after']
    results['cv_improvement_pct'] = (results['cv_improvement'] / results['median_cv_before']) * 100 if results['median_cv_before'] > 0 else 0
    
    # CV%改善的特徵比例
    cv_improved = np.sum((original_cv - normalized_cv) > 0)
    cv_total = len(original_cv[~np.isnan(original_cv)])
    results['cv_improved_ratio'] = (cv_improved / cv_total) * 100 if cv_total > 0 else 0
    
    # 2. 樣本總強度變異
    original_totals = np.nansum(original_data, axis=0)
    normalized_totals = np.nansum(normalized_data, axis=0)
    
    results['total_cv_before'] = (np.std(original_totals) / np.mean(original_totals)) * 100
    results['total_cv_after'] = (np.std(normalized_totals) / np.mean(normalized_totals)) * 100
    results['total_cv_improvement'] = results['total_cv_before'] - results['total_cv_after']
    
    # 3. 樣本間相關性
    corr_mean_before, corr_std_before = calculate_sample_correlation(original_data)
    corr_mean_after, corr_std_after = calculate_sample_correlation(normalized_data)
    
    results['sample_corr_mean_before'] = corr_mean_before
    results['sample_corr_mean_after'] = corr_mean_after
    results['sample_corr_std_before'] = corr_std_before
    results['sample_corr_std_after'] = corr_std_after
    
    # 4. 數據範圍評估
    results['data_range_before'] = np.nanmax(original_data) - np.nanmin(original_data)
    results['data_range_after'] = np.nanmax(normalized_data) - np.nanmin(normalized_data)
    
    return results

def create_normalization_summary_report(quality_metrics, method_name, n_features, n_samples,
                                       pqn_info=None, group_diff_results=None):
    """
    建立增強版標準化摘要報告（基於非參數統計方法）

    Parameters:
    -----------
    quality_metrics : dict
        標準化質量指標
    method_name : str
        標準化方法名稱
    n_features : int
        特徵數量
    n_samples : int
        樣本數量
    pqn_info : dict, optional
        PQN 相關資訊
    group_diff_results : dict, optional
        組間差異評估結果

    Returns:
    --------
    str : 格式化的報告文字
    """
    report = []
    report.append(SUMMARY_REPORT_SEPARATOR)
    report.append(f"標準化效果摘要報告 - {method_name}")
    report.append(f"報告生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report.append(SUMMARY_REPORT_SEPARATOR)
    
    # ========== 基本資訊 ==========
    report.append("")
    report.append("【基本資訊】")
    report.append(f"標準化方法: {method_name}")
    report.append(f"特徵數量: {n_features}")
    report.append(f"樣本數量: {n_samples}")
    
    # ========== PQN 參考樣本資訊（新增）==========
    if pqn_info:
        report.append("")
        report.append("【PQN 參考樣本資訊】")
        report.append(f"參考策略: {pqn_info['reference_strategy']}")
        report.append(f"QC 樣本數量: {pqn_info['qc_count']}")
        report.append(f"真實樣本數量: {pqn_info['real_count']}")
        
        if pqn_info['qc_count'] > 0 and not np.isnan(pqn_info['qc_cv']):
            report.append(f"QC 中位數 CV%: {pqn_info['qc_cv']:.2f}%")
            if pqn_info['qc_cv'] < CV_QUALITY_THRESHOLDS['excellent']:
                report.append("QC 質量評估: ✓✓ 優秀")
            elif pqn_info['qc_cv'] < CV_QUALITY_THRESHOLDS['acceptable']:
                report.append("QC 質量評估: ✓ 良好")
            else:
                report.append("QC 質量評估: ⚠ 需改進")
        
        report.append(f"肌酐校正:")
        report.append(f"  - 肌酐中位數: {pqn_info['creatinine_median']:.2f}")
        report.append(f"  - 有效樣本數: {pqn_info['creatinine_valid_count']}/{pqn_info['real_count']}")
    
    # ========== CV% 評估 ==========
    report.append("")
    report.append("【CV% 評估】")
    report.append(f"標準化前:")
    report.append(f"  - 中位數CV%: {quality_metrics['median_cv_before']:.2f}%")
    report.append(f"  - 平均CV%: {quality_metrics['mean_cv_before']:.2f}%")
    report.append(f"標準化後:")
    report.append(f"  - 中位數CV%: {quality_metrics['median_cv_after']:.2f}%")
    report.append(f"  - 平均CV%: {quality_metrics['mean_cv_after']:.2f}%")
    report.append(f"改善:")
    report.append(f"  - CV%降低: {quality_metrics['cv_improvement']:.2f}%")
    report.append(f"  - 改善百分比: {quality_metrics['cv_improvement_pct']:.2f}%")
    report.append(f"  - CV%改善的特徵比例: {quality_metrics['cv_improved_ratio']:.1f}%")
    
    # ========== 樣本總強度變異 ==========
    report.append("")
    report.append("【樣本總強度變異】")
    report.append(f"標準化前總強度CV%: {quality_metrics['total_cv_before']:.2f}%")
    report.append(f"標準化後總強度CV%: {quality_metrics['total_cv_after']:.2f}%")
    report.append(f"總強度CV%改善: {quality_metrics['total_cv_improvement']:.2f}%")
    
    # ========== 樣本間相關性 ==========
    report.append("")
    report.append("【樣本間相關性】")
    if not np.isnan(quality_metrics['sample_corr_mean_before']):
        report.append(f"標準化前:")
        report.append(f"  - 平均相關性: {quality_metrics['sample_corr_mean_before']:.4f}")
        report.append(f"  - 相關性標準差: {quality_metrics['sample_corr_std_before']:.4f}")
        report.append(f"標準化後:")
        report.append(f"  - 平均相關性: {quality_metrics['sample_corr_mean_after']:.4f}")
        report.append(f"  - 相關性標準差: {quality_metrics['sample_corr_std_after']:.4f}")
    else:
        report.append("  - 數據不足，無法計算樣本間相關性")

    # ========== 組間差異保留評估 ==========
    if group_diff_results:
        report.append("")
        report.append("【組間差異保留評估】(Control vs Exposure)")
        report.append(f"Control 組樣本數: {group_diff_results['control_count']}")
        report.append(f"Exposure 組樣本數: {group_diff_results['exposure_count']}")
        report.append(f"分析特徵數: {group_diff_results['total']}")
        report.append(f"Effect Size 變化統計:")
        report.append(f"  - 增強: {group_diff_results['enhanced']} ({group_diff_results['enhanced']/group_diff_results['total']*100:.1f}%)")
        report.append(f"  - 穩定 (±10%): {group_diff_results['stable']} ({group_diff_results['stable']/group_diff_results['total']*100:.1f}%)")
        report.append(f"  - 輕度減弱 (-10% ~ -30%): {group_diff_results['mild_reduction']} ({group_diff_results['mild_reduction']/group_diff_results['total']*100:.1f}%)")
        report.append(f"  - 顯著減弱 (< -30%): {group_diff_results['severe_reduction']} ({group_diff_results['severe_reduction']/group_diff_results['total']*100:.1f}%)")
        report.append(f"平均 Effect Size 保留率: {group_diff_results['avg_preservation']:.1f}%")
        
        if group_diff_results['severe_reduction'] / group_diff_results['total'] > 0.1:
            report.append("評估: ⚠⚠ 部分特徵差異顯著減弱，需檢查")
        elif group_diff_results['avg_preservation'] > 90:
            report.append("評估: ✓✓ 組間差異保留優秀")
        elif group_diff_results['avg_preservation'] > 80:
            report.append("評估: ✓ 組間差異保留良好")
        else:
            report.append("評估: ○ 組間差異保留尚可")
    
    # ========== 評估結論 ==========
    report.append("")
    report.append("【評估結論】")
    
    # CV%評估
    if quality_metrics['cv_improvement'] > 0:
        if quality_metrics['cv_improvement_pct'] > 20:
            report.append("✓✓✓ CV%顯著降低，標準化效果極佳")
        elif quality_metrics['cv_improvement_pct'] > 10:
            report.append("✓✓ CV%明顯降低，標準化效果良好")
        else:
            report.append("✓ CV%略有降低，標準化效果尚可")
    else:
        report.append("⚠ CV%未降低，建議檢查數據或嘗試其他標準化方法")
    
    # 總強度變異評估
    if quality_metrics['total_cv_improvement'] > 0:
        if quality_metrics['total_cv_improvement'] > 10:
            report.append("✓✓ 樣本總強度變異顯著降低")
        else:
            report.append("✓ 樣本總強度變異有所降低")
    else:
        report.append("⚠ 樣本總強度變異未改善")
    
    # 樣本間相關性評估
    if not np.isnan(quality_metrics['sample_corr_std_before']):
        if quality_metrics['sample_corr_std_after'] < quality_metrics['sample_corr_std_before']:
            report.append("✓ 樣本間相關性更一致")
        else:
            report.append("⚠ 樣本間相關性一致性未改善")
    
    # 特徵改善比例評估
    if quality_metrics['cv_improved_ratio'] > 70:
        report.append(f"✓✓ 大多數特徵({quality_metrics['cv_improved_ratio']:.1f}%)的CV%得到改善")
    elif quality_metrics['cv_improved_ratio'] > 50:
        report.append(f"✓ 超過半數特徵({quality_metrics['cv_improved_ratio']:.1f}%)的CV%得到改善")
    else:
        report.append(f"⚠ 僅{quality_metrics['cv_improved_ratio']:.1f}%的特徵CV%得到改善")
    
    # ========== 整體評分（更新評分邏輯）==========
    report.append("")
    report.append("【整體評分】")
    score = 0
    max_score = 100
    
    # CV% 改善 (25 分)
    if quality_metrics['cv_improvement'] > 0:
        if quality_metrics['cv_improvement_pct'] > 20:
            score += 25
        elif quality_metrics['cv_improvement_pct'] > 10:
            score += 20
        else:
            score += 15
    
    # 總強度變異改善 (20 分)
    if quality_metrics['total_cv_improvement'] > 0:
        if quality_metrics['total_cv_improvement'] > 10:
            score += 20
        else:
            score += 15
    
    # 樣本間相關性改善 (15 分)
    if not np.isnan(quality_metrics['sample_corr_std_before']) and quality_metrics['sample_corr_std_after'] < quality_metrics['sample_corr_std_before']:
        score += 15

    # 組間差異保留 (30 分) - 包含統計檢驗評估
    if group_diff_results:
        # 基本分：Effect Size 保留率 (20 分)
        if group_diff_results['avg_preservation'] > 90:
            score += 20
        elif group_diff_results['avg_preservation'] > 80:
            score += 15
        elif group_diff_results['avg_preservation'] > 70:
            score += 10

        # 統計檢驗額外分 (10 分)
        # 1. Wilcoxon 檢驗結果 (5 分)
        if 'wilcoxon_pvalue' in group_diff_results and not np.isnan(group_diff_results['wilcoxon_pvalue']):
            if group_diff_results['wilcoxon_pvalue'] >= 0.05:
                # p >= 0.05 表示 Cohen's d 中位數無顯著變化，這是好的
                score += 5
            # 如果 p < 0.05 但中位數變化不大，給予部分分數
            elif group_diff_results['avg_preservation'] > 85:
                score += 2

        # 2. 標記特徵比例 (5 分)
        if 'flagged_ratio' in group_diff_results:
            if group_diff_results['flagged_ratio'] < 0.05:
                # 少於 5% 的特徵被標記為顯著改變
                score += 5
            elif group_diff_results['flagged_ratio'] < 0.10:
                # 5-10% 的特徵被標記
                score += 3

        # 嚴重減弱特徵的懲罰
        if group_diff_results['severe_reduction'] / group_diff_results['total'] > 0.1:
            score -= 10
            report.append("  ⚠ 警告：超過 10% 的特徵 Effect Size 顯著減弱（-10 分）")
    
    report.append(f"標準化質量評分: {score}/{max_score}")
    if score >= 85:
        report.append("評級: 優秀 ★★★★★")
    elif score >= 70:
        report.append("評級: 良好 ★★★★")
    elif score >= 50:
        report.append("評級: 尚可 ★★★")
    else:
        report.append("評級: 需改進 ★★")
    
    # ========== 建議與注意事項（新增）==========
    report.append("")
    report.append("【建議與注意事項】")
    
    # QC 相關建議
    if pqn_info:
        if pqn_info['qc_count'] < 3:
            report.append("⚠ 建議：QC 樣本數量較少，建議至少使用 3 個 QC 樣本以提高標準化穩定性")
        
        if not np.isnan(pqn_info['qc_cv']) and pqn_info['qc_cv'] > 25:
            report.append("⚠ 建議：QC CV% 較高，可能需要檢查實驗技術重現性")
        
        if pqn_info['creatinine_valid_count'] < pqn_info['real_count'] * 0.9:
            report.append(f"⚠ 注意：有 {pqn_info['real_count'] - pqn_info['creatinine_valid_count']} 個樣本缺少有效肌酐值")

    # 組間差異相關建議
    if group_diff_results:
        if group_diff_results['severe_reduction'] > 0:
            report.append(f"⚠ 注意：有 {group_diff_results['severe_reduction']} 個特徵的組間差異顯著減弱")
            report.append("  建議：檢查這些特徵是否與肌酐代謝相關，可能需要特別處理")
        
        if group_diff_results['avg_preservation'] < 80:
            report.append("⚠ 建議：組間差異保留率較低，建議檢查標準化方法是否適合您的數據")
    
    # CV% 改善相關建議
    if quality_metrics['cv_improved_ratio'] < 50:
        report.append("⚠ 建議：僅不到一半的特徵 CV% 得到改善，可能需要考慮其他標準化方法")
    
    # 如果所有指標都良好
    if (quality_metrics['cv_improvement_pct'] > 20 and
        quality_metrics['total_cv_improvement'] > 10 and
        (group_diff_results is None or group_diff_results['avg_preservation'] > 90)):
        report.append("✓✓✓ 恭喜！所有評估指標均表現優異，標準化效果極佳")
    
    report.append("")
    report.append("=" * 80)
    
    return "\n".join(report)

# ==================== 檔案處理函數 ====================

def select_file():
    raise RuntimeError("input_file is required; GUI must provide the file path.")
    """讓使用者選擇Excel檔案"""
    root = tk.Tk()
    root.withdraw()
    
    file_path = filedialog.askopenfilename(
        title="請選擇Excel檔案",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    
    if not file_path:
        print("未選擇檔案，程式結束。")
        return None
    
    return file_path

def load_excel_sheets(file_path):
    """載入Excel檔案的所有工作表"""
    try:
        xl_file = pd.ExcelFile(file_path)
        sheet_names = xl_file.sheet_names
        
        sheets = {}
        for sheet_name in sheet_names:
            sheets[sheet_name] = pd.read_excel(file_path, sheet_name=sheet_name)
            
        return sheets, sheet_names
    except Exception as e:
        print(f"讀取Excel檔案時發生錯誤: {e}")
        return None, None

def determine_correction_sheet(sheets):
    """按指定順序確定要標準化的資料工作表"""
    priority_sheets = [
        SHEET_NAMES['batch_effect'],
        SHEET_NAMES['qc_lowess'],
        SHEET_NAMES['istd_correction'],
        SHEET_NAMES['raw_intensity']
    ]
    
    for sheet_name in priority_sheets:
        if sheet_name in sheets:
            print(f"✓ 依優先順序選擇工作表: {sheet_name}")
            return sheets[sheet_name], sheet_name
    
    print("警告：未找到指定的資料工作表")
    return None, None

def find_sample_info_sheet(sheets):
    """尋找包含樣本資訊的工作表"""
    possible_names = [SHEET_NAMES['sample_info'], 'Sample_Info', 'sample_info', 'Sample Info']
    
    for name in possible_names:
        if name in sheets:
            return sheets[name], name
    
    for sheet_name, df in sheets.items():
        if any(col for col in df.columns if 'normalization' in str(col).lower() or 'creatinine' in str(col).lower()):
            return df, sheet_name
    
    return None, None

def find_correction_column(df):
    """在樣本資訊工作表尋找可用於校正的欄位"""
    if df.shape[1] < 6:
        print("警告：樣本資訊工作表欄位不足")
        return None, None
    
    # 預設使用第 F 欄（索引 5）
    correction_col = df.columns[5]
    valid_values = df[correction_col].dropna()
    
    if len(valid_values) > 0:
        # 檢查是否為數值
        numeric_count = valid_values.apply(lambda x: str(x).replace('.','').replace('-','').replace('e','').replace('E','').replace('+','').isnumeric()).sum()
        if numeric_count / len(valid_values) > 0.5:
            if 'creatinine' in str(correction_col).lower():
                correction_type = 'Creatinine'
            else:
                correction_type = 'Normalization_adduct'
            
            print(f"✓ 偵測到校正欄位: {correction_col} (類型: {correction_type})")
            return correction_col, correction_type
    
    # 如果第 F 欄不可用，嘗試找其他可用的數值欄位
    for col in df.columns[6:]:
        valid_values = df[col].dropna()
        if len(valid_values) > 0:
            numeric_count = valid_values.apply(lambda x: str(x).replace('.','').replace('-','').replace('e','').replace('E','').replace('+','').isnumeric()).sum()
            if numeric_count / len(valid_values) > 0.5:
                if 'creatinine' in str(col).lower():
                    correction_type = 'Creatinine'
                else:
                    correction_type = 'Normalization_adduct'
                
                print(f"✓ 偵測到校正欄位: {col} (類型: {correction_type})")
                return col, correction_type
    
    print("錯誤：找不到可用於校正的欄位")
    return None, None

def clean_dataframe_for_excel(df):
    """清理DataFrame以避免Excel格式問題"""
    cleaned_df = df.copy()
    
    for col in cleaned_df.columns:
        str_series = cleaned_df[col].astype(str)
        
        # 移除公式
        mask_formula = str_series.str.startswith('=')
        cleaned_df.loc[mask_formula, col] = ''
        
        # 移除Excel錯誤值
        error_values = ['#REF!', '#VALUE!', '#NAME?', '#DIV/0!', '#N/A', '#NULL!', '#NUM!']
        for error_val in error_values:
            cleaned_df[col] = cleaned_df[col].replace(error_val, '')
        
        # 嘗試轉換數值欄位
        if col != cleaned_df.columns[0]:
            cleaned_df[col] = pd.to_numeric(cleaned_df[col], errors='ignore')
    
    return cleaned_df

from ms_core.utils.excel_format import copy_cell_style  # noqa: E302

# ==================== 主要處理函數 ====================

def perform_normalization(data_df, sample_info_df, correction_col, file_path):
    """
    執行 PQN + Sample-specific 混合標準化處理（增強版）
    """
    print(f"\n" + "="*70)
    print("開始執行標準化處理...")
    print("="*70)
    
    method_name = "PQN_SampleSpecific"
    
    # 獲取純樣本欄位（排除統計欄位）
    sample_columns = get_all_sample_columns(data_df, sample_info_df)  # ← 改用新函數
    print(f"✓ 樣本數量（含QC）: {len(sample_columns)}")
    
    if len(sample_columns) == 0:
        print("錯誤：未找到有效的樣本欄位")
        return None
    
    # 準備數據矩陣 (特徵 x 樣本) - Vectorized (much faster than iterrows)
    feature_ids = data_df[data_df.columns[0]].tolist()

    # Extract sample columns and convert to numeric matrix directly
    valid_sample_cols = [c for c in sample_columns if c in data_df.columns]
    data_matrix = data_df[valid_sample_cols].apply(pd.to_numeric, errors='coerce').values
    print(f"✓ 數據矩陣形狀: {data_matrix.shape} (特徵 x 樣本)")
    
    # 保存原始數據用於對比
    original_data = data_matrix.copy()
    
    # 獲取參考值（肌酐/DNA 濃度）
    # 先嘗試精確匹配，若失敗則用位置對齊
    info_names = sample_info_df.iloc[:, 0].tolist()
    exact_match_count = sum(1 for s in sample_columns if s in info_names)

    if exact_match_count >= len(sample_columns) * 0.5:
        # 精確匹配
        col_to_info_row = {}
        for s in sample_columns:
            rows = sample_info_df[sample_info_df.iloc[:, 0] == s]
            if not rows.empty:
                col_to_info_row[s] = rows.iloc[0]
        print(f"  名稱匹配: 精確匹配 {len(col_to_info_row)}/{len(sample_columns)}")
    else:
        # 位置對齊 fallback
        col_to_info_row = {}
        if len(info_names) == len(sample_columns):
            for i, col in enumerate(sample_columns):
                col_to_info_row[col] = sample_info_df.iloc[i]
            print(f"  名稱匹配: 位置對齊 {len(col_to_info_row)}/{len(sample_columns)}")
        else:
            # 嘗試模糊匹配（使用 ISTD 中的 token 方法）
            import re
            def _extract_tokens(name):
                s = str(name).strip().lower()
                s = re.sub(r'^(dna|rna)_program\d+_', '', s)
                parts = re.split(r'[\s_\-/]+', s)
                tokens = set()
                numbers = set()
                for part in parts:
                    sub = re.findall(r'[a-z]+|[0-9]+', part)
                    tokens.update(sub)
                    combo = re.findall(r'[a-z]+\d+', part)
                    tokens.update(combo)
                    # Extract pure numbers (specimen IDs)
                    nums = re.findall(r'\d{3,}', part)
                    numbers.update(nums)
                generic = {'tissue', 'cancer', 'breast', 'pooled', 'fat', 'dna', 'rna', 'and', 'program1'}
                return tokens - generic, numbers

            for col in sample_columns:
                col_tokens, col_nums = _extract_tokens(col)
                best_match = None
                best_score = 0
                for idx, info_name in enumerate(info_names):
                    info_tokens, info_nums = _extract_tokens(info_name)
                    # Numeric ID match is strongest signal
                    num_overlap = len(col_nums & info_nums)
                    token_overlap = len(col_tokens & info_tokens)
                    if num_overlap > 0:
                        score = 0.8 + 0.2 * (token_overlap / max(len(col_tokens), len(info_tokens), 1))
                    elif col_tokens and info_tokens:
                        score = token_overlap / max(len(col_tokens), len(info_tokens))
                    else:
                        score = 0
                    if score > best_score:
                        best_score = score
                        best_match = idx
                if best_match is not None and best_score >= 0.5:
                    col_to_info_row[col] = sample_info_df.iloc[best_match]
            print(f"  名稱匹配: 模糊匹配 {len(col_to_info_row)}/{len(sample_columns)}")

    reference_values = []
    for sample in sample_columns:
        info_row = col_to_info_row.get(sample)
        if info_row is not None and pd.notna(info_row.get(correction_col)):
            try:
                ref_val = float(info_row[correction_col])
                if ref_val > 0:
                    reference_values.append(ref_val)
                else:
                    reference_values.append(np.nan)
            except (ValueError, TypeError, KeyError):
                reference_values.append(np.nan)
        else:
            reference_values.append(np.nan)

    reference_values = np.array(reference_values)

    # 計算有效參考值（排除 QC 樣本，因為 QC 不需要濃度校正）
    non_qc_mask = np.array([
        not (str(col_to_info_row[s].get('Sample_Type', '')).upper() == 'QC')
        if s in col_to_info_row else
        not any(kw in str(s).upper() for kw in ['QC', 'POOLED'])
        for s in sample_columns
    ])
    non_qc_count = np.sum(non_qc_mask)
    valid_count = np.sum(~np.isnan(reference_values) & (reference_values > 0))
    print(f"✓ 有效參考值數量: {valid_count}/{non_qc_count} (排除 {len(sample_columns) - non_qc_count} 個 QC 樣本)")

    if non_qc_count > 0 and valid_count < non_qc_count * 0.3:
        print("警告：有效參考值不足30%")
        return None
    
    # ========== 🔧 關鍵修正：正確呼叫 enhanced_pqn_normalization ==========
    normalized_data, pqn_info = enhanced_pqn_normalization(
        data_matrix,           # 完整的數據矩陣
        sample_info_df,        # 樣本資訊表
        sample_columns,        # 樣本名稱列表
        reference_values,      # 參考值（肌酐濃度）
        col_to_info_row=col_to_info_row  # 名稱映射
    )
    
    print(f"✓ 標準化完成")
    
    # 統一輸出目錄與圖表路徑
    output_dir = get_output_root()
    run_timestamp = datetime.now().strftime(DATETIME_FORMAT_FULL)
    method_slug = method_name.replace(' ', '_')
    figures_dir = build_plots_dir(
        "Normalization_Figures",
        timestamp=run_timestamp,
        session_prefix=method_slug
    )
    print(f"✓ Excel 將輸出到: {output_dir}")
    print(f"✓ 本次圖表輸出目錄: {figures_dir}")

    figure_paths = {
        "boxplot": figures_dir / generate_output_filename(
            f"Fig1_Boxplot_{method_slug}", timestamp=run_timestamp, extension=".png"
        ),
        "cv": figures_dir / generate_output_filename(
            f"Fig2_CV_{method_slug}", timestamp=run_timestamp, extension=".png"
        ),
        "rle": figures_dir / generate_output_filename(
            f"Fig3_RLE_{method_slug}", timestamp=run_timestamp, extension=".png"
        ),
        "pca": figures_dir / generate_output_filename(
            f"Fig4_PCA_{method_slug}", timestamp=run_timestamp, extension=".png"
        ),
        "qc_variability": figures_dir / generate_output_filename(
            f"Fig5_QC_Variability_{method_slug}", timestamp=run_timestamp, extension=".png"
        ),
        "qc_reproducibility": figures_dir / generate_output_filename(
            f"Fig6_QC_Reproducibility_{method_slug}", timestamp=run_timestamp, extension=".png"
        ),
        "correlation": figures_dir / generate_output_filename(
            f"Fig7_Correlation_{method_slug}", timestamp=run_timestamp, extension=".png"
        ),
    }
    # 分離有效樣本用於評估
    valid_sample_mask = ~np.isnan(normalized_data[0, :])
    original_data_valid = original_data[:, valid_sample_mask]
    normalized_data_valid = normalized_data[:, valid_sample_mask]
    sample_columns_valid = [sample_columns[i] for i in range(len(sample_columns)) if valid_sample_mask[i]]
    
    # ========== 評估標準化質量 ==========
    print("\n評估標準化質量...")
    quality_metrics = evaluate_normalization_quality(original_data_valid, normalized_data_valid)

    # ========== 組間差異保留評估 ==========
    try:
        group_diff_results = evaluate_group_difference_preservation(
            original_data_valid, normalized_data_valid, sample_info_df, sample_columns_valid,
            col_to_info_row=col_to_info_row
        )
    except Exception as e:
        print(f"  ⚠ 組間差異評估失敗: {e}")
        group_diff_results = None
    
    # ========== 生成視覺化圖表 ==========
    print("\n生成視覺化圖表...")

    

    # 2. 盒鬚圖（含樣本總強度資訊）
    plot_boxplot_comparison(
        original_data_valid, normalized_data_valid, sample_columns_valid,
        figure_paths["boxplot"],
        method_name,
        sample_info_df,
        col_to_info_row=col_to_info_row
    )
    
    # 3. CV%分佈圖
    original_cv = calculate_cv_per_feature(original_data_valid)
    normalized_cv = calculate_cv_per_feature(normalized_data_valid)
    plot_cv_comparison(
        original_cv, normalized_cv,
        figure_paths["cv"],
        method_name
    )

    # 3b. RLE Plot（組學正規化品質評估黃金標準）
    try:
        plot_rle(
            original_data_valid, normalized_data_valid, sample_columns_valid,
            figure_paths["rle"],
            method_name
        )
    except Exception as e:
        print(f"  ⚠ RLE Plot 生成失敗: {e}")

    # 4. PCA對比圖（改進版：加入信賴橢圓）
    try:
        plot_pca_with_confidence_ellipse(
            original_data_valid, normalized_data_valid, sample_columns_valid, sample_info_df,
            figure_paths["pca"],
            method_name,
            exclude_qc=True,
            col_to_info_row=col_to_info_row,
        )
    except Exception as e:
        print(f"  ⚠ PCA對比圖生成失敗: {e}")
    
    

    # 6. 相關性熱圖
    try:
        plot_correlation_heatmap(
            original_data_valid, normalized_data_valid, sample_columns_valid,
            figure_paths["correlation"],
            method_name
        )
    except Exception as e:
        print(f"  ⚠ 相關性熱圖生成失敗: {e}")

    # ========== QC 質量評估圖 ==========
    if pqn_info['qc_count'] > 0:
        try:
            # 提取 QC 樣本索引
            qc_indices = []
            qc_names = []
            for i, sample in enumerate(sample_columns_valid):
                sample_type = _lookup_sample_type(sample, sample_info_df, col_to_info_row)
                if sample_type == 'QC':
                    qc_indices.append(i)
                    qc_names.append(sample)
            
            if len(qc_indices) > 0:
                original_qc = original_data_valid[:, qc_indices]
                normalized_qc = normalized_data_valid[:, qc_indices]

                # Fig4 - QC Variability
                plot_qc_variability(
                    original_qc, normalized_qc, qc_names,
                    output_path=figure_paths["qc_variability"]
                )

                # Fig5 - QC Reproducibility
                plot_qc_reproducibility(
                    original_qc, normalized_qc, qc_names,
                    output_path=figure_paths["qc_reproducibility"]
                )
        except Exception as e:
            print(f"  ⚠ QC 質量評估圖生成失敗: {e}")
    
    # 創建結果DataFrame（只包含樣本數據和CV%）
    normalized_df = pd.DataFrame()
    normalized_df[data_df.columns[0]] = feature_ids
    
    # 添加標準化後的樣本數據
    for i, col in enumerate(sample_columns):
        normalized_df[col] = normalized_data[:, i]
    
    # 只添加CV%欄位
    normalized_df['Original_CV%'] = calculate_cv_per_feature(original_data)
    normalized_df['Normalized_CV%'] = calculate_cv_per_feature(normalized_data)
    normalized_df['CV_Improvement%'] = normalized_df['Original_CV%'] - normalized_df['Normalized_CV%']
    
    # ========== 生成摘要報告 ==========
    summary_report = create_normalization_summary_report(
        quality_metrics, method_name, len(feature_ids), len(sample_columns_valid),
        pqn_info=pqn_info,
        group_diff_results=group_diff_results
    )
    
    print(f"\n{summary_report}")
    
    return normalized_df, summary_report, method_name, output_dir, quality_metrics, figures_dir

def save_normalization_results(normalized_df, summary_report, file_path, method_name, original_sheets, output_dir):
    """儲存標準化結果到Excel檔案"""
    try:
        timestamp = datetime.now().strftime(DATETIME_FORMAT_FULL)
        output_filename = generate_output_filename(
            f"Normalized_{method_name}", timestamp=timestamp, extension=".xlsx"
        )
        output_path = output_dir / output_filename
        
        # 載入原始工作簿
        wb_original = load_workbook(file_path, data_only=False)
        
        # 創建新工作簿
        wb_new = Workbook()
        wb_new.remove(wb_new.active)
        
        # 1. 儲存標準化後的資料
        ws_normalized = wb_new.create_sheet(title=f'{method_name}_Result')
        
        cleaned_normalized_df = clean_dataframe_for_excel(normalized_df)
        for r_idx, row in enumerate(dataframe_to_rows(cleaned_normalized_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_normalized.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 自動調整列寬
        for column in ws_normalized.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    # TypeError: cell.value is None or non-stringable
                    # AttributeError: cell has no value attribute
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_normalized.column_dimensions[column_letter].width = adjusted_width
        
        # 2. 儲存摘要報告
        ws_summary = wb_new.create_sheet(title=SUMMARY_SHEET_NAME)
        summary_rows = summary_report.split('\n')
        for r_idx, row_text in enumerate(summary_rows, 1):
            cell_value = row_text
            if row_text.startswith('='):
                cell_value = f" {row_text}"
            cell = ws_summary.cell(row=r_idx, column=1, value=cell_value)
            if '=' in row_text and r_idx <= 3:
                cell.font = Font(bold=True, size=12)
            elif '【' in row_text:
                cell.font = Font(bold=True, size=11, color='0000FF')
            elif '✓' in row_text:
                cell.font = Font(color='008000')
            elif '⚠' in row_text:
                cell.font = Font(color='FF6600')
        
        ws_summary.column_dimensions['A'].width = 80
        
        # 3. 複製原始工作表並保留格式
        for sheet_name in wb_original.sheetnames:
            if sheet_name in [f'{method_name}_Result', SUMMARY_SHEET_NAME]:
                continue
            
            ws_original = wb_original[sheet_name]
            ws_new = wb_new.create_sheet(title=sheet_name[:31])
            
            # 設置列寬
            for col in ws_original.column_dimensions:
                ws_new.column_dimensions[col].width = ws_original.column_dimensions[col].width
            
            # 設置行高
            for row_idx, row_dim in ws_original.row_dimensions.items():
                ws_new.row_dimensions[row_idx].height = row_dim.height
            
            # 複製儲存格內容和格式
            for row in ws_original.iter_rows():
                for cell in row:
                    new_cell = ws_new.cell(row=cell.row, column=cell.column, value=cell.value)
                    copy_cell_style(cell, new_cell)
        
        # 儲存新工作簿
        wb_new.save(output_path)
        
        print(f"\n✓ 結果已儲存至: {output_path}")
        print(f"\n包含工作表:")
        print(f"  1. {method_name}_Result (標準化後資料)")
        print(f"  2. {SUMMARY_SHEET_NAME} (摘要報告)")
        for sheet_name in wb_original.sheetnames:
            if sheet_name not in [f'{method_name}_Result', SUMMARY_SHEET_NAME]:
                print(f"  3. {sheet_name} (原始資料)")
        
        return str(output_path)
        
    except Exception as e:
        print(f"儲存檔案時發生錯誤: {e}")
        import traceback
        traceback.print_exc()
        return None

# ==================== 主程式 ====================

def main(input_file=None):
    """
    主函數 - 支援 GUI 和獨立運行
    
    Parameters:
    -----------
    input_file : str, optional
        輸入檔案路徑（由 GUI 傳入）
        如果為 None，則顯示檔案選擇對話框
    
    Returns:
    --------
    dict or None
        - None: 用戶取消檔案選擇
        - dict: 執行成功，包含統計資訊
    """
    print("=" * 80)
    print("  代謝體學標準化程式 v3.0")
    print("  標準化方法: PQN + Sample-specific (肌酐校正)")
    print("  - 步驟1: Sample-specific Normalization (肌酐校正)")
    print("  - 步驟2: Probabilistic Quotient Normalization (PQN)")
    print("  - 視覺化評估工具（密度圖、盒鬚圖、PCA、CV%分佈等）")
    print("=" * 80)
    
    # 🔧 關鍵修正：如果沒有提供 input_file，則顯示對話框
    if input_file is None:
        raise ValueError("input_file is required; GUI must provide the file path.")

    if input_file is None:
        file_path = select_file()
        
        # 如果用戶取消選擇，返回 None
        if not file_path:
            print("❌ 未選擇檔案，程式結束。")
            return None
        
        input_file = file_path
    
    # 驗證檔案是否存在
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"找不到檔案: {input_file}")
    
    print(f"\n✓ 選擇的檔案: {Path(input_file).name}")
    
    # 載入Excel工作表
    sheets, sheet_names = load_excel_sheets(input_file)
    if not sheets:
        raise Exception("無法載入 Excel 工作表")
    
    print(f"✓ 找到 {len(sheet_names)} 個工作表")
    
    # 尋找樣本資訊工作表
    sample_info_df, sample_info_sheet_name = find_sample_info_sheet(sheets)
    if sample_info_df is None:
        print("❌ 錯誤：找不到包含樣本資訊的工作表")
        raise Exception("找不到樣本資訊工作表")
    
    print(f"✓ 使用樣本資訊工作表: {sample_info_sheet_name}")
    
    # 尋找校正欄位（肌酐濃度）
    correction_col, correction_type = find_correction_column(sample_info_df)
    if not correction_col:
        print("❌ 錯誤：找不到校正用的欄位")
        raise Exception("找不到校正欄位")
    
    # 確定要標準化的資料工作表
    data_df, data_sheet_name = determine_correction_sheet(sheets)
    if data_df is None:
        print("❌ 錯誤：找不到要標準化的資料工作表")
        raise Exception("找不到資料工作表")
    
    print(f"✓ 使用資料工作表: {data_sheet_name}")

    # 提取 Sample_Type 資訊行（不參與數值計算，保存時回插）
    from ms_core.utils.data_helpers import extract_sample_type_row
    feature_col = data_df.columns[0]
    data_df, sample_type_row = extract_sample_type_row(data_df, feature_col)
    if sample_type_row is not None:
        print(f"✓ 偵測到 Sample_Type 資訊行，已提取保存（不參與計算）")

    # 執行標準化
    result = perform_normalization(data_df, sample_info_df, correction_col, input_file)
    
    if result is None:
        print("\n❌ 標準化失敗")
        raise Exception("標準化失敗")
    
    normalized_df, summary_report, method_name, output_dir, quality_metrics, figures_dir = result
    
   
    
    # 回插 Sample_Type 資訊行到 normalized_df（若有）
    if sample_type_row is not None:
        from ms_core.utils.data_helpers import insert_sample_type_row
        normalized_df = insert_sample_type_row(normalized_df, sample_type_row, feature_col)

    # 儲存結果
    print("\n儲存結果...")
    output_path = save_normalization_results(
        normalized_df, summary_report, input_file, method_name, sheets, output_dir
    )
    
    if not output_path:
        raise Exception("儲存結果失敗")
    
    print("\n" + "=" * 80)
    print("✅ 標準化完成！")
    print("=" * 80)
    print(f"\n📁 輸出資料夾: {output_dir}")
    print(f"📄 Excel檔案: {Path(output_path).name}")
    
    print("\n📊 生成的視覺化圖表:")
    print(f"  - 儲存路徑: {figures_dir}")
    print("  - 圖檔: Fig1_Boxplot_*.png, Fig2_CV_*.png, Fig3_PCA_*.png, fig4~6 QC 與相關性評估")
    
    print("\n" + "=" * 80)
    print("📈 標準化質量評估摘要:")
    print("=" * 80)
    print(f"  中位數CV%改善: {quality_metrics['median_cv_before']:.2f}% → {quality_metrics['median_cv_after']:.2f}%")
    print(f"  CV%降低: {quality_metrics['cv_improvement']:.2f}% ({quality_metrics['cv_improvement_pct']:.1f}%)")
    print(f"  改善特徵比例: {quality_metrics['cv_improved_ratio']:.1f}%")
    print(f"  樣本總強度CV%: {quality_metrics['total_cv_before']:.2f}% → {quality_metrics['total_cv_after']:.2f}%")
    
    if not np.isnan(quality_metrics['sample_corr_mean_before']):
        print(f"  樣本間平均相關性: {quality_metrics['sample_corr_mean_before']:.4f} → {quality_metrics['sample_corr_mean_after']:.4f}")
    
    # 計算整體評分
    score = 0
    if quality_metrics['cv_improvement'] > 0:
        score += 25
    if quality_metrics['total_cv_improvement'] > 0:
        score += 25
    if not np.isnan(quality_metrics['sample_corr_std_before']) and quality_metrics['sample_corr_std_after'] < quality_metrics['sample_corr_std_before']:
        score += 25
    if quality_metrics['cv_improved_ratio'] > 50:
        score += 25
    
    print(f"\n  整體評分: {score}/100", end=" ")
    if score >= 75:
        print("★★★★★ (優秀)")
    elif score >= 50:
        print("★★★★ (良好)")
    elif score >= 25:
        print("★★★ (尚可)")
    else:
        print("★★ (需改進)")
    
    print("=" * 80)
    
    print("\n✅ 所有檔案已成功儲存！")
    print(f"✅ 完整路徑: {output_path}")
    
    # 🎯 返回統計資訊給 GUI
    # 從 normalized_df 中提取樣本數量（排除第一列 FeatureID）
    sample_count = len(normalized_df.columns) - 1
    metabolite_count = len(normalized_df)
    
    return ProcessingResult(
        file_path=input_file,
        output_path=str(output_path),
        plots_dir=str(figures_dir),
        metabolites=metabolite_count,
        samples=sample_count
    )


if __name__ == "__main__":
    # 🔧 獨立運行時不傳入 input_file，會顯示對話框
    main()

